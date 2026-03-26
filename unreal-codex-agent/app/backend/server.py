#!/usr/bin/env python3
"""
UEFN Codex Agent - Flask Backend Server

Provides REST API for tool discovery, execution, orchestration, and AI research.
"""

import ast
import base64
import copy
import csv as csv_lib
import html as html_lib
import hashlib
import os
import re
import sys
import io
import json
import threading
import time
import logging
import mimetypes
import random
import math
from datetime import datetime, UTC
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import socket
import asyncio
import aiohttp
import urllib.request
import urllib.parse
import urllib.error
import subprocess
import tempfile
import zipfile
import shutil
from uuid import uuid4
from html.parser import HTMLParser
import xml.etree.ElementTree as ET

from flask import Flask, jsonify, request
from flask_cors import CORS

# OpenAI-compatible client (works with Ollama, LM Studio, or OpenAI)
try:
    from openai import OpenAI as _OpenAI
    _HAS_OPENAI_PKG = True
except ImportError:
    _HAS_OPENAI_PKG = False
    _OpenAI = None

try:
    from rapidocr import RapidOCR as _RapidOCR
    _HAS_RAPIDOCR = True
except ImportError:
    _HAS_RAPIDOCR = False
    _RapidOCR = None

try:
    from pypdf import PdfReader as _PdfReader
    _HAS_PYPDF = True
except ImportError:
    _HAS_PYPDF = False
    _PdfReader = None

try:
    from PIL import Image as _PILImage, ImageEnhance as _ImageEnhance, ImageFilter as _ImageFilter, ImageOps as _ImageOps
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False
    _PILImage = None
    _ImageEnhance = None
    _ImageFilter = None
    _ImageOps = None

try:
    import trafilatura as _trafilatura
    _HAS_TRAFILATURA = True
except ImportError:
    _HAS_TRAFILATURA = False
    _trafilatura = None

try:
    import pymupdf as _pymupdf
    _HAS_PYMUPDF = True
except ImportError:
    _HAS_PYMUPDF = False
    _pymupdf = None

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    _openpyxl_load_workbook = None

try:
    from transformers import (
        pipeline as _hf_pipeline,
        TrOCRProcessor as _TrOCRProcessor,
        VisionEncoderDecoderModel as _VisionEncoderDecoderModel,
    )
    _HAS_TRANSFORMERS = True
except ImportError:
    _HAS_TRANSFORMERS = False
    _hf_pipeline = None
    _TrOCRProcessor = None
    _VisionEncoderDecoderModel = None

try:
    from transformers import (
        AutoProcessor as _AutoProcessor,
        Florence2ForConditionalGeneration as _Florence2ForConditionalGeneration,
    )
    _HAS_FLORENCE2 = True
except ImportError:
    _HAS_FLORENCE2 = False
    _AutoProcessor = None
    _Florence2ForConditionalGeneration = None

try:
    import torch as _torch
    _HAS_TORCH = True
except ImportError:
    _HAS_TORCH = False
    _torch = None

try:
    from defusedxml import ElementTree as _SafeElementTree
    _HAS_DEFUSEDXML = True
except ImportError:
    _HAS_DEFUSEDXML = False
    _SafeElementTree = ET

# Load .env from workspace root (manual parser — no dependency on python-dotenv)
_ws_root = Path(__file__).parent.parent.parent
_env_file = _ws_root / ".env"

def _load_env_file(path: Path):
    """Parse a .env file and inject into os.environ (only sets keys not already set)."""
    if not path.exists():
        return
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip()
            # Remove surrounding quotes if present
            if len(value) >= 2 and value[0] == value[-1] and value[0] in ('"', "'"):
                value = value[1:-1]
            # Only set if not already in environment (env vars take precedence)
            if key not in os.environ or not os.environ[key].strip():
                os.environ[key] = value
    except Exception:
        pass

_load_env_file(_env_file)

# ── AI Provider configuration ─────────────────────────────────────────────
# Supported free providers (all OpenAI-compatible, free API key):
#   groq     - Fastest inference, 14400 req/day free, get key at console.groq.com
#   cerebras - Ultra-fast, 1M tokens/day free, get key at cloud.cerebras.ai
#   gemini   - Smartest (Gemini 2.5 Pro), 1M context, get key at aistudio.google.dev
#   ollama   - Local, unlimited, no key needed, install from ollama.com

AI_PROVIDERS = {
    "groq": {
        "base_url": "https://api.groq.com/openai/v1",
        "default_model": "llama-3.3-70b-versatile",
        "models": [
            "llama-3.3-70b-versatile",
            "deepseek-r1-distill-llama-70b",
            "llama-3.1-8b-instant",
            "gemma2-9b-it",
        ],
        "key_env": "GROQ_API_KEY",
        "label": "Groq (free, fastest)",
        "hint": "Get free key at console.groq.com — 14,400 requests/day",
    },
    "cerebras": {
        "base_url": "https://api.cerebras.ai/v1",
        "default_model": "qwen-3-235b-a22b-instruct-2507",
        "models": ["qwen-3-235b-a22b-instruct-2507", "llama3.1-8b"],
        "key_env": "CEREBRAS_API_KEY",
        "label": "Cerebras (free, ultra-fast)",
        "hint": "Get free key at cloud.cerebras.ai — 1M tokens/day",
    },
    "gemini": {
        "base_url": "https://generativelanguage.googleapis.com/v1beta/openai/",
        "default_model": "gemini-2.5-flash",
        "models": ["gemini-2.5-pro", "gemini-2.5-flash", "gemini-2.5-flash-lite"],
        "key_env": "GEMINI_API_KEY",
        "label": "Google Gemini (free, smartest)",
        "hint": "Get free key at aistudio.google.dev — 1M token context",
    },
    "ollama": {
        "base_url": os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434") + "/v1",
        "default_model": "llama3.1",
        "models": [],
        "key_env": None,
        "label": "Ollama (local, unlimited)",
        "hint": "Install from ollama.com, then: ollama pull llama3.1",
    },
}

# Active provider config
AI_PROVIDER = os.environ.get("AI_PROVIDER", "")  # auto-detect if empty
AI_MODEL = os.environ.get("AI_MODEL", "")

OLLAMA_BASE_URL = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
OLLAMA_MODEL = "llama3.1"

PROVIDER_MODEL_ENV_KEYS = {
    "groq": "GROQ_MODEL",
    "cerebras": "CEREBRAS_MODEL",
    "gemini": "GEMINI_MODEL",
    "ollama": "OLLAMA_MODEL",
    "openai": "OPENAI_MODEL",
}
VISION_CAPABLE_PROVIDERS = {"gemini", "openai"}
GEMINI_GENERATE_CONTENT_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
LOCAL_STRUCTURED_VLM_ENABLED = os.environ.get("LOCAL_STRUCTURED_VLM_ENABLED", "1").strip().lower() not in {"0", "false", "no", "off"}
LOCAL_STRUCTURED_VLM_MODEL_ID = os.environ.get("LOCAL_STRUCTURED_VLM_MODEL_ID", "florence-community/Florence-2-base-ft").strip() or "florence-community/Florence-2-base-ft"
LOCAL_VLM_ENABLED = os.environ.get("LOCAL_VLM_ENABLED", "1").strip().lower() not in {"0", "false", "no", "off"}
LOCAL_VLM_MODEL_ID = os.environ.get("LOCAL_VLM_MODEL_ID", "Salesforce/blip-image-captioning-base").strip() or "Salesforce/blip-image-captioning-base"
LOCAL_HTR_ENABLED = os.environ.get("LOCAL_HTR_ENABLED", "1").strip().lower() not in {"0", "false", "no", "off"}
LOCAL_HTR_MODEL_ID = os.environ.get("LOCAL_HTR_MODEL_ID", "microsoft/trocr-base-handwritten").strip() or "microsoft/trocr-base-handwritten"
LOCAL_MODEL_PREWARM_ENABLED = os.environ.get("LOCAL_MODEL_PREWARM_ENABLED", "1").strip().lower() not in {"0", "false", "no", "off"}
LOCAL_MODEL_PREWARM_DELAY_SECONDS = max(0.0, float(os.environ.get("LOCAL_MODEL_PREWARM_DELAY_SECONDS", "1.5").strip() or "1.5"))
HOSTED_LLM_MAX_RETRIES = max(0, int(os.environ.get("HOSTED_LLM_MAX_RETRIES", "2").strip() or "2"))
HOSTED_LLM_RETRY_BASE_SECONDS = max(0.2, float(os.environ.get("HOSTED_LLM_RETRY_BASE_SECONDS", "1.0").strip() or "1.0"))
HOSTED_LLM_RETRY_MAX_SECONDS = max(HOSTED_LLM_RETRY_BASE_SECONDS, float(os.environ.get("HOSTED_LLM_RETRY_MAX_SECONDS", "8.0").strip() or "8.0"))

_VISUAL_COLOR_TERMS = {
    "red", "blue", "green", "yellow", "orange", "purple", "pink", "white", "black", "gray",
    "grey", "brown", "gold", "silver", "teal", "cyan",
}
_VISUAL_REVIEW_PHRASES = (
    "does this match",
    "match the request",
    "what is wrong",
    "what's wrong",
    "what should change",
    "what should i change",
    "does it look right",
    "does this look right",
    "does this look good",
    "does it look good",
    "is this correct",
    "is it correct",
    "is the building red",
    "is this red",
    "is it red",
)
_VISUAL_COMPARE_PHRASES = (
    "what changed",
    "what is different",
    "what's different",
    "compare",
    "difference",
    "before and after",
    "before/after",
    "compare these",
)
_VISUAL_STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "better", "between", "by", "can", "change",
    "changes", "check", "compare", "comparison", "correct", "describe", "difference", "do",
    "does", "file", "for", "from", "good", "how", "i", "if", "image", "images", "in", "into",
    "is", "it", "its", "look", "looks", "make", "match", "matches", "my", "of", "on", "or",
    "please", "request", "right", "screenshot", "see", "should", "show", "tell", "than", "that",
    "the", "them", "these", "this", "to", "what", "whats", "when", "why", "with", "wrong",
}
_ATTACHMENT_SUMMARY_PHRASES = (
    "summarize",
    "summarise",
    "summary",
    "what is this about",
    "what's this about",
    "what is it about",
    "what's it about",
    "tell me what this is about",
    "tell me what it is about",
    "read this and tell me",
    "read it and tell me",
    "give me the overview",
    "give me an overview",
    "overview of this",
    "overview of it",
    "high level overview",
    "main takeaways",
    "key takeaways",
)
_ATTACHMENT_FOLLOWUP_REFERENCE_PHRASES = (
    "fix it",
    "fix this",
    "fix that",
    "could you fix it",
    "can you fix it",
    "change it",
    "change this",
    "change that",
    "improve it",
    "improve this",
    "improve that",
    "edit it",
    "edit this",
    "use that",
    "use this",
    "summarize it",
    "summarise it",
    "read it",
    "describe it",
    "analyze it",
    "analyse it",
    "compare them",
    "what about this",
    "what about that",
)
_ATTACHMENT_REFERENCE_TOKENS = {"it", "this", "that", "them", "these", "those"}
_ATTACHMENT_REFERENCE_INTENT_TOKENS = {
    "fix", "change", "improve", "edit", "adjust", "update", "make", "read", "summarize",
    "summarise", "summary", "describe", "analyze", "analyse", "review", "compare", "tell",
    "explain", "use", "apply", "match", "wrong", "about", "what", "why", "how",
}


def _check_ollama_running() -> bool:
    """Check if Ollama server is reachable."""
    try:
        req = urllib.request.Request(f"{OLLAMA_BASE_URL}/api/tags", method="GET")
        with urllib.request.urlopen(req, timeout=2) as resp:
            return resp.status == 200
    except Exception:
        return False


def _get_ollama_models() -> list:
    """List locally available Ollama models."""
    try:
        req = urllib.request.Request(f"{OLLAMA_BASE_URL}/api/tags", method="GET")
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = json.loads(resp.read().decode())
            return [m.get("name", "") for m in data.get("models", [])]
    except Exception:
        return []


def _detect_provider() -> tuple:
    """Auto-detect which AI provider is configured. Returns (provider_name, api_key)."""
    # Check explicit setting
    explicit = os.environ.get("AI_PROVIDER", "").strip().lower()
    if explicit and explicit in AI_PROVIDERS:
        prov = AI_PROVIDERS[explicit]
        key_env = prov.get("key_env")
        key = os.environ.get(key_env, "").strip() if key_env else "ollama"
        if explicit == "ollama":
            if _check_ollama_running():
                return ("ollama", "ollama")
        elif key:
            return (explicit, key)

    # Auto-detect: check each provider's key
    for name in ["groq", "cerebras", "gemini"]:
        prov = AI_PROVIDERS[name]
        key = os.environ.get(prov["key_env"], "").strip()
        if key:
            return (name, key)

    # Fall back to OpenAI if key exists
    openai_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if openai_key:
        return ("openai", openai_key)

    # Fall back to Ollama
    if _check_ollama_running():
        return ("ollama", "ollama")

    return (None, None)

# ============================================================================
# LOGGING
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================

UEFN_MCP_HOST = "127.0.0.1"
UEFN_MCP_PORT = 8765
UEFN_MCP_MAX_PORT = 8770
BACKEND_HOST = os.environ.get("BACKEND_HOST", "127.0.0.1").strip() or "127.0.0.1"
BACKEND_PORT = int(os.environ.get("BACKEND_PORT", "8000").strip() or "8000")

# Paths
WORKSPACE_ROOT = Path(__file__).parent.parent.parent
if str(WORKSPACE_ROOT) not in sys.path:
    sys.path.insert(0, str(WORKSPACE_ROOT))

VENDOR_UEFN = WORKSPACE_ROOT / "vendor" / "uefn-toolbelt"
DATA_DIR = WORKSPACE_ROOT / "data"
CHAT_STORE_FILE = DATA_DIR / "chat_sessions.json"
KNOWLEDGE_STORE_FILE = DATA_DIR / "knowledge_store.json"
ATTACHMENT_CACHE_FILE = DATA_DIR / "attachment_analysis_cache.json"
RUNTIME_IMPORT_DIR = DATA_DIR / "runtime_imports"

from apps.integrations.uefn_backend import choose_action_backend
from apps.integrations.uefn_mcp import (
    _delete_actors_by_paths,
    apply_action_via_mcp,
    collect_scene_state,
    inspect_actor,
    set_actor_material,
)
from apps.mcp_extensions.scene_tools import enrich_scene_state
from apps.mcp_extensions.uefn_tools import _stray_tool_generated_paths_for_zone
from apps.orchestrator.state_store import SessionStateStore
from apps.placement.assembly_builder import (
    HouseSpec,
    StructureSpec,
    SUPPORTED_GENERATIVE_STRUCTURE_TYPES,
    build_house_actions,
    build_house_structure_plan,
    build_structure_actions,
    build_structure_plan,
    canonical_structure_type,
    plan_house_spec,
    plan_structure_spec,
)
from apps.placement.managed_registry import managed_records_for_zone, release_slot
from apps.placement.structure_validation import validate_structure_plan
from apps.uefn.verse_export import apply_action_via_verse_export

# ============================================================================
# TOOL CATALOG - COMPREHENSIVE DESCRIPTIONS
# ============================================================================

TOOL_DESCRIPTIONS = {
    # ── API Explorer ──────────────────────────────────────────────────────
    "api_crawl_selection": {
        "name": "API Crawl Selection",
        "category": "API Explorer",
        "description": "Reads the deeply nested exposed properties of the currently selected actors. Produces a full JSON property map including components, materials, and nested structs — ideal for AI analysis of what an actor can do.",
        "short": "Deep-inspect selected actor properties",
        "tags": ["api", "inspector", "selection", "properties"]
    },
    "api_crawl_level_classes": {
        "name": "API Crawl Level Classes",
        "category": "API Explorer",
        "description": "Headlessly scans every actor in the map, aggregates them by class, and produces a JSON report of all unique actor classes with counts and property schemas. Great for understanding what's in your level.",
        "short": "Scan all actors by class",
        "tags": ["api", "scan", "level", "classes"]
    },
    "api_search": {
        "name": "API Search",
        "category": "API Explorer",
        "description": "Search the Unreal Python API by keyword. Finds classes, functions, and properties matching your query — useful for discovering what's available in the editor scripting API.",
        "short": "Search the UE Python API",
        "tags": ["api", "search", "documentation"]
    },
    "api_inspect": {
        "name": "API Inspect",
        "category": "API Explorer",
        "description": "Inspect a specific Unreal Python class or object, listing all its methods, properties, and inheritance chain with type information.",
        "short": "Inspect a UE class/object",
        "tags": ["api", "inspect", "reflection"]
    },
    "api_generate_stubs": {
        "name": "API Generate Stubs",
        "category": "API Explorer",
        "description": "Generate Python type stub files (.pyi) for Unreal classes, enabling IDE autocompletion and type checking for editor scripting.",
        "short": "Generate Python type stubs",
        "tags": ["api", "stubs", "typing"]
    },
    "api_list_subsystems": {
        "name": "API List Subsystems",
        "category": "API Explorer",
        "description": "List all available editor subsystems (EditorActorSubsystem, LevelEditorSubsystem, etc.) with their methods. Shows what automation entry points are available.",
        "short": "List editor subsystems",
        "tags": ["api", "subsystems", "editor"]
    },
    "api_export_full": {
        "name": "API Full Export",
        "category": "API Explorer",
        "description": "Export a comprehensive JSON dump of the entire available Unreal Python API surface — every class, method, and property discoverable at runtime.",
        "short": "Export full API to JSON",
        "tags": ["api", "export", "reference"]
    },

    # ── Arena Generator ───────────────────────────────────────────────────
    "arena_generate": {
        "name": "Arena Generator",
        "category": "Procedural",
        "description": "Procedurally generates a symmetrical competitive arena with configurable size, wall height, spawn points, and floor patterns. Supports multiple arena styles (box, hexagon, circle) with automatic team-colored zones.",
        "short": "Generate symmetrical arenas",
        "tags": ["arena", "procedural", "competitive", "generation"]
    },

    # ── Asset Importer ────────────────────────────────────────────────────
    "import_image_from_url": {
        "name": "Import Image from URL",
        "category": "Pipeline",
        "description": "Downloads an image from a URL and imports it as a Texture2D asset into the UEFN content browser. Supports PNG, JPG, and BMP formats with automatic naming.",
        "short": "Import image from web URL",
        "tags": ["import", "image", "url", "texture"]
    },
    "import_image_from_clipboard": {
        "name": "Import Image from Clipboard",
        "category": "Pipeline",
        "description": "Grabs the current clipboard image and imports it directly as a Texture2D asset. Perfect for quick screenshots or reference images.",
        "short": "Import clipboard image",
        "tags": ["import", "image", "clipboard", "texture"]
    },
    "import_fbx": {
        "name": "Import FBX",
        "category": "Assets",
        "description": "Import a single FBX file as a Static Mesh with configurable LOD generation, collision setup, and material assignment.",
        "short": "Import FBX mesh",
        "tags": ["import", "fbx", "mesh", "3d"]
    },
    "import_fbx_folder": {
        "name": "Import FBX Folder",
        "category": "Assets",
        "description": "Batch-import all FBX files from a folder with consistent LOD and collision settings. Automatically organizes imported assets by type.",
        "short": "Batch import FBX folder",
        "tags": ["import", "fbx", "batch", "folder"]
    },
    "organize_assets": {
        "name": "Organize Assets",
        "category": "Assets",
        "description": "Automatically sorts assets from a source folder into organized subfolders (Meshes/, Materials/, Textures/) based on asset type. Follows Epic naming conventions.",
        "short": "Auto-organize by type",
        "tags": ["organize", "assets", "folders", "cleanup"]
    },

    # ── Asset Renamer ─────────────────────────────────────────────────────
    "rename_dry_run": {
        "name": "Rename Dry Run",
        "category": "Assets",
        "description": "Prints every asset that violates Epic naming conventions without renaming anything. Use this to preview what would change before committing.",
        "short": "Preview naming violations",
        "tags": ["rename", "audit", "conventions", "preview"]
    },
    "rename_enforce_conventions": {
        "name": "Rename Enforce Conventions",
        "category": "Assets",
        "description": "Renames every violating asset to match Epic naming conventions (SM_ for static meshes, MI_ for material instances, T_ for textures, etc.). Always run rename_dry_run first.",
        "short": "Auto-rename to conventions",
        "tags": ["rename", "conventions", "enforce"]
    },
    "rename_strip_prefix": {
        "name": "Rename Strip Prefix",
        "category": "Assets",
        "description": "Removes a specified prefix from every asset in the scan path that starts with it. Useful for cleaning up imported asset names.",
        "short": "Strip prefix from names",
        "tags": ["rename", "prefix", "cleanup"]
    },
    "rename_report": {
        "name": "Rename Report",
        "category": "Assets",
        "description": "Full naming audit — doesn't rename anything, writes a comprehensive JSON report of all naming violations, suggestions, and statistics.",
        "short": "Generate naming report",
        "tags": ["rename", "report", "audit"]
    },

    # ── Asset Tagger ──────────────────────────────────────────────────────
    "tag_add": {
        "name": "Tag Add",
        "category": "Asset Tagger",
        "description": "Apply a custom metadata tag (TB:key=value) to all assets currently selected in the Content Browser. Tags persist with the asset and can be searched later.",
        "short": "Add tag to selected assets",
        "tags": ["tag", "metadata", "organize"]
    },
    "tag_remove": {
        "name": "Tag Remove",
        "category": "Asset Tagger",
        "description": "Remove a tag key from all assets selected in the Content Browser.",
        "short": "Remove tag from assets",
        "tags": ["tag", "remove", "metadata"]
    },
    "tag_show": {
        "name": "Tag Show",
        "category": "Asset Tagger",
        "description": "Print all TB: tags on every asset currently selected in the Content Browser.",
        "short": "Show tags on selection",
        "tags": ["tag", "show", "inspect"]
    },
    "tag_search": {
        "name": "Tag Search",
        "category": "Asset Tagger",
        "description": "Find all assets under a folder where a specific tag key matches a value. Like a database query for your content.",
        "short": "Search assets by tag",
        "tags": ["tag", "search", "filter"]
    },
    "tag_list_all": {
        "name": "Tag List All",
        "category": "Asset Tagger",
        "description": "Print all unique TB: tag keys used anywhere under a folder, with asset counts per tag.",
        "short": "List all tag keys",
        "tags": ["tag", "list", "summary"]
    },
    "tag_export": {
        "name": "Tag Export",
        "category": "Asset Tagger",
        "description": "Scan all assets under a folder, collect every TB: tag, and write the results to a JSON file for external tools or AI analysis.",
        "short": "Export tags to JSON",
        "tags": ["tag", "export", "json"]
    },

    # ── Bulk Operations ───────────────────────────────────────────────────
    "bulk_align": {
        "name": "Bulk Align",
        "category": "Bulk Ops",
        "description": "Align selected actors along an axis (X, Y, or Z) to the first selected actor, the average position, or a custom value. Supports min/max/center alignment modes.",
        "short": "Align actors on axis",
        "tags": ["align", "bulk", "transform"]
    },
    "bulk_distribute": {
        "name": "Bulk Distribute",
        "category": "Bulk Ops",
        "description": "Evenly distribute selected actors along an axis with equal spacing between them. Perfect for creating uniform rows or grids.",
        "short": "Distribute actors evenly",
        "tags": ["distribute", "spacing", "layout"]
    },
    "bulk_randomize": {
        "name": "Bulk Randomize",
        "category": "Bulk Ops",
        "description": "Add random offset to position, rotation, and/or scale of selected actors. Great for making placed props look natural and organic.",
        "short": "Randomize transforms",
        "tags": ["randomize", "variation", "organic"]
    },
    "bulk_snap_to_grid": {
        "name": "Bulk Snap to Grid",
        "category": "Bulk Ops",
        "description": "Snap all selected actors to the nearest grid point with a configurable grid size. Ensures clean, aligned placement.",
        "short": "Snap to grid",
        "tags": ["snap", "grid", "alignment"]
    },
    "bulk_reset": {
        "name": "Bulk Reset",
        "category": "Bulk Ops",
        "description": "Reset rotation and scale of selected actors to defaults (0,0,0 rotation, 1,1,1 scale) while keeping their positions.",
        "short": "Reset rotation & scale",
        "tags": ["reset", "transform", "cleanup"]
    },
    "bulk_stack": {
        "name": "Bulk Stack",
        "category": "Bulk Ops",
        "description": "Stack selected actors vertically on top of each other with optional spacing. Useful for building towers or layered structures.",
        "short": "Stack actors vertically",
        "tags": ["stack", "vertical", "layout"]
    },
    "bulk_mirror": {
        "name": "Bulk Mirror",
        "category": "Bulk Ops",
        "description": "Mirror (reflect) selected actors across an axis plane. Creates symmetrical layouts by duplicating and flipping positions.",
        "short": "Mirror actors across axis",
        "tags": ["mirror", "symmetry", "duplicate"]
    },
    "bulk_normalize_scale": {
        "name": "Bulk Normalize Scale",
        "category": "Bulk Ops",
        "description": "Set all selected actors to the same uniform scale, normalizing size differences while keeping positions.",
        "short": "Normalize all scales",
        "tags": ["scale", "normalize", "uniform"]
    },
    "bulk_face_camera": {
        "name": "Bulk Face Camera",
        "category": "Bulk Ops",
        "description": "Rotate all selected actors to face the current viewport camera position. Useful for billboards or oriented signage.",
        "short": "Rotate to face camera",
        "tags": ["rotate", "camera", "billboard"]
    },

    # ── Foliage / Scatter ─────────────────────────────────────────────────
    "scatter_props": {
        "name": "Scatter Props",
        "category": "Procedural",
        "description": "Randomly scatter copies of a mesh across a circular area with configurable radius, count, random rotation, and scale variation. Perfect for foliage, debris, or decorations.",
        "short": "Scatter objects randomly",
        "tags": ["scatter", "procedural", "foliage"]
    },
    "scatter_hism": {
        "name": "Scatter HISM",
        "category": "Procedural",
        "description": "Places all instances inside a single Hierarchical Instanced Static Mesh actor for much better performance. Use this instead of scatter_props for large counts (100+).",
        "short": "High-performance scatter",
        "tags": ["scatter", "hism", "performance"]
    },
    "scatter_along_path": {
        "name": "Scatter Along Path",
        "category": "Procedural",
        "description": "Drops clusters of instances around each point along a spline path. Great for placing fence posts along a road or trees along a trail.",
        "short": "Scatter along spline path",
        "tags": ["scatter", "spline", "path"]
    },
    "scatter_clear": {
        "name": "Scatter Clear",
        "category": "Procedural",
        "description": "Delete all actors in the scatter output folder, cleaning up a previous scatter operation.",
        "short": "Clear scattered actors",
        "tags": ["scatter", "clear", "cleanup"]
    },
    "scatter_export_manifest": {
        "name": "Scatter Export Manifest",
        "category": "Procedural",
        "description": "Exports a JSON manifest of every actor in the scatter folder with positions, rotations, and scales for external tools.",
        "short": "Export scatter data",
        "tags": ["scatter", "export", "manifest"]
    },

    # ── Level Snapshot ────────────────────────────────────────────────────
    "snapshot_save": {
        "name": "Snapshot Save",
        "category": "Level Snapshot",
        "description": "Capture all actor transforms (position, rotation, scale) and save them to a named snapshot file. Like a save point you can restore later.",
        "short": "Save level snapshot",
        "tags": ["snapshot", "save", "backup"]
    },
    "snapshot_restore": {
        "name": "Snapshot Restore",
        "category": "Level Snapshot",
        "description": "Restore all actor transforms from a saved snapshot, undoing any changes made since the snapshot was taken.",
        "short": "Restore from snapshot",
        "tags": ["snapshot", "restore", "undo"]
    },
    "snapshot_list": {
        "name": "Snapshot List",
        "category": "Level Snapshot",
        "description": "Print all saved snapshots in the snapshots folder with timestamps and actor counts.",
        "short": "List saved snapshots",
        "tags": ["snapshot", "list"]
    },
    "snapshot_diff": {
        "name": "Snapshot Diff",
        "category": "Level Snapshot",
        "description": "Diff two snapshots and print what changed — which actors moved, were added, or deleted between saves.",
        "short": "Compare two snapshots",
        "tags": ["snapshot", "diff", "compare"]
    },
    "snapshot_delete": {
        "name": "Snapshot Delete",
        "category": "Level Snapshot",
        "description": "Delete a single snapshot file by name.",
        "short": "Delete a snapshot",
        "tags": ["snapshot", "delete"]
    },
    "snapshot_export": {
        "name": "Snapshot Export",
        "category": "Level Snapshot",
        "description": "Export a snapshot to any file path for sharing with other creators or backing up externally.",
        "short": "Export snapshot to file",
        "tags": ["snapshot", "export", "share"]
    },
    "snapshot_import": {
        "name": "Snapshot Import",
        "category": "Level Snapshot",
        "description": "Import a snapshot JSON from any file path into the snapshots folder.",
        "short": "Import snapshot file",
        "tags": ["snapshot", "import"]
    },
    "snapshot_compare_live": {
        "name": "Snapshot Compare Live",
        "category": "Level Snapshot",
        "description": "Diff a saved snapshot against the current live level state — see what changed without needing a second snapshot.",
        "short": "Compare snapshot vs live",
        "tags": ["snapshot", "compare", "live"]
    },

    # ── LOD Tools ─────────────────────────────────────────────────────────
    "lod_auto_generate_selection": {
        "name": "LOD Auto-Generate (Selection)",
        "category": "Assets",
        "description": "Automatically generate LOD levels for selected static mesh assets using Unreal's built-in LOD system with configurable reduction percentages.",
        "short": "Generate LODs for selection",
        "tags": ["lod", "optimization", "mesh"]
    },
    "lod_auto_generate_folder": {
        "name": "LOD Auto-Generate (Folder)",
        "category": "Assets",
        "description": "Batch-generate LOD levels for all static meshes in a folder. Essential for performance optimization.",
        "short": "Generate LODs for folder",
        "tags": ["lod", "batch", "optimization"]
    },
    "lod_set_collision_folder": {
        "name": "LOD Set Collision (Folder)",
        "category": "Assets",
        "description": "Set collision complexity for all meshes in a folder (simple box, convex, or use complex as simple).",
        "short": "Set collision for folder",
        "tags": ["collision", "lod", "physics"]
    },
    "lod_audit_folder": {
        "name": "LOD Audit Folder",
        "category": "Assets",
        "description": "Prints a report and saves JSON showing which meshes have LODs, how many levels, and which are missing LODs.",
        "short": "Audit LOD coverage",
        "tags": ["lod", "audit", "report"]
    },

    # ── Materials ─────────────────────────────────────────────────────────
    "material_list_presets": {
        "name": "Material List Presets",
        "category": "Materials",
        "description": "List all available material presets (built-in and custom) with their color values and properties.",
        "short": "List material presets",
        "tags": ["materials", "presets", "list"]
    },
    "material_apply_preset": {
        "name": "Material Apply Preset",
        "category": "Materials",
        "description": "Apply a named material preset (like 'team_red', 'neon_green', 'metallic_gold') to selected actors, creating and assigning material instances automatically.",
        "short": "Apply preset to selection",
        "tags": ["materials", "preset", "apply"]
    },
    "material_randomize_colors": {
        "name": "Material Randomize Colors",
        "category": "Materials",
        "description": "Assign random colors from a palette or fully random hues to each selected actor's material. Great for colorful, varied environments.",
        "short": "Randomize material colors",
        "tags": ["materials", "randomize", "color"]
    },
    "material_gradient_painter": {
        "name": "Material Gradient Painter",
        "category": "Materials",
        "description": "Paint a smooth color gradient across selected actors based on their position along an axis. Creates beautiful transitions.",
        "short": "Paint color gradient",
        "tags": ["materials", "gradient", "color"]
    },
    "material_team_color_split": {
        "name": "Material Team Color Split",
        "category": "Materials",
        "description": "Split actors into two team colors based on their X position relative to a midpoint. Actors left of center get one color, right get another.",
        "short": "Split into team colors",
        "tags": ["materials", "teams", "competitive"]
    },
    "material_pattern_painter": {
        "name": "Material Pattern Painter",
        "category": "Materials",
        "description": "Apply repeating color patterns (checkerboard, stripes, alternating) across selected actors based on their grid position.",
        "short": "Paint material patterns",
        "tags": ["materials", "pattern", "checkerboard"]
    },
    "material_glow_pulse_preview": {
        "name": "Material Glow Pulse Preview",
        "category": "Materials",
        "description": "Preview an emissive glow/pulse effect on selected actors. Sets up material parameters for glowing, pulsating surfaces.",
        "short": "Preview glow effects",
        "tags": ["materials", "glow", "emissive"]
    },
    "material_color_harmony": {
        "name": "Material Color Harmony",
        "category": "Materials",
        "description": "Generate harmonious color schemes (complementary, analogous, triadic) and apply them across selected actors for aesthetically pleasing palettes.",
        "short": "Apply color harmony",
        "tags": ["materials", "harmony", "palette"]
    },
    "material_save_preset": {
        "name": "Material Save Preset",
        "category": "Materials",
        "description": "Read the current material parameters from the first selected actor and save them as a named preset for reuse later.",
        "short": "Save current as preset",
        "tags": ["materials", "preset", "save"]
    },
    "material_bulk_swap": {
        "name": "Material Bulk Swap",
        "category": "Materials",
        "description": "Swap one material for another on all actors in scope. Useful for replacing placeholder materials with final ones.",
        "short": "Swap materials in bulk",
        "tags": ["materials", "swap", "replace"]
    },

    # ── MCP Bridge ────────────────────────────────────────────────────────
    "mcp_start": {
        "name": "MCP Start",
        "category": "MCP Bridge",
        "description": "Start the UEFN Toolbelt MCP HTTP listener, enabling external AI tools (Claude, Codex) to control the editor remotely.",
        "short": "Start MCP listener",
        "tags": ["mcp", "bridge", "start"]
    },
    "mcp_stop": {
        "name": "MCP Stop",
        "category": "MCP Bridge",
        "description": "Stop the MCP HTTP listener. External AI tools will no longer be able to control the editor.",
        "short": "Stop MCP listener",
        "tags": ["mcp", "bridge", "stop"]
    },
    "mcp_restart": {
        "name": "MCP Restart",
        "category": "MCP Bridge",
        "description": "Restart the MCP listener — use after hot-reloading the toolbelt to pick up new tools.",
        "short": "Restart MCP listener",
        "tags": ["mcp", "bridge", "restart"]
    },
    "mcp_status": {
        "name": "MCP Status",
        "category": "MCP Bridge",
        "description": "Print the current MCP listener status (running/stopped, port, uptime, request count) to the Output Log.",
        "short": "Check MCP status",
        "tags": ["mcp", "bridge", "status"]
    },

    # ── Memory Profiler ───────────────────────────────────────────────────
    "memory_scan": {
        "name": "Memory Scan",
        "category": "Optimization",
        "description": "Scans all asset types and prints a summary dashboard showing total memory usage, asset counts, and the biggest consumers.",
        "short": "Full memory scan",
        "tags": ["memory", "optimization", "audit"]
    },
    "memory_scan_textures": {
        "name": "Memory Scan Textures",
        "category": "Optimization",
        "description": "Detailed scan of all textures showing resolution, format, estimated memory, and compression settings. Finds oversized textures.",
        "short": "Scan texture memory",
        "tags": ["memory", "textures", "optimization"]
    },
    "memory_scan_meshes": {
        "name": "Memory Scan Meshes",
        "category": "Optimization",
        "description": "Detailed scan of all meshes showing poly count, LOD levels, and estimated memory. Identifies meshes that need optimization.",
        "short": "Scan mesh memory",
        "tags": ["memory", "meshes", "polycount"]
    },
    "memory_top_offenders": {
        "name": "Memory Top Offenders",
        "category": "Optimization",
        "description": "Lists the heaviest assets by estimated memory usage — your top targets for optimization.",
        "short": "Find heaviest assets",
        "tags": ["memory", "optimization", "offenders"]
    },
    "memory_autofix_lods": {
        "name": "Memory Auto-Fix LODs",
        "category": "Optimization",
        "description": "Convenience wrapper: finds meshes with 1 or fewer LOD levels and automatically generates LODs to reduce rendering cost.",
        "short": "Auto-generate missing LODs",
        "tags": ["memory", "lod", "autofix"]
    },

    # ── Prop Patterns ─────────────────────────────────────────────────────
    "pattern_grid": {
        "name": "Pattern Grid",
        "category": "Prop Patterns",
        "description": "Place a mesh in a rectangular N×M grid with configurable spacing. Great for walls, floors, fences, or any repeating layout.",
        "short": "Place in grid pattern",
        "tags": ["pattern", "grid", "layout"]
    },
    "pattern_circle": {
        "name": "Pattern Circle",
        "category": "Prop Patterns",
        "description": "Place props evenly spaced around a full 360° circle with configurable radius and count.",
        "short": "Place in circle",
        "tags": ["pattern", "circle", "radial"]
    },
    "pattern_arc": {
        "name": "Pattern Arc",
        "category": "Prop Patterns",
        "description": "Place props along a partial arc (e.g., 90° or 180°) with configurable sweep angle and radius.",
        "short": "Place along arc",
        "tags": ["pattern", "arc", "curved"]
    },
    "pattern_spiral": {
        "name": "Pattern Spiral",
        "category": "Prop Patterns",
        "description": "Place props along an Archimedean spiral that expands outward. Creates interesting visual patterns.",
        "short": "Place in spiral",
        "tags": ["pattern", "spiral", "procedural"]
    },
    "pattern_line": {
        "name": "Pattern Line",
        "category": "Prop Patterns",
        "description": "Place props in a straight line from start to end points with even spacing.",
        "short": "Place in straight line",
        "tags": ["pattern", "line", "linear"]
    },
    "pattern_wave": {
        "name": "Pattern Wave",
        "category": "Prop Patterns",
        "description": "Place props along a sine wave path with configurable amplitude, frequency, and length.",
        "short": "Place in wave pattern",
        "tags": ["pattern", "wave", "sine"]
    },
    "pattern_helix": {
        "name": "Pattern Helix",
        "category": "Prop Patterns",
        "description": "Place props along a 3D helix (corkscrew / spiral staircase) with configurable radius, height, and turns.",
        "short": "Place in 3D helix",
        "tags": ["pattern", "helix", "3d"]
    },
    "pattern_radial_rows": {
        "name": "Pattern Radial Rows",
        "category": "Prop Patterns",
        "description": "Place props in concentric rings with increasing density on outer rings. Creates amphitheater or stadium seating patterns.",
        "short": "Concentric ring pattern",
        "tags": ["pattern", "radial", "concentric"]
    },
    "pattern_clear": {
        "name": "Pattern Clear",
        "category": "Prop Patterns",
        "description": "Delete all actors created by pattern tools, cleaning up a previous pattern placement.",
        "short": "Clear pattern actors",
        "tags": ["pattern", "clear", "cleanup"]
    },

    # ── Procedural Geometry ───────────────────────────────────────────────
    "procedural_wire_create": {
        "name": "Procedural Wire",
        "category": "Procedural",
        "description": "Create a wire/cable mesh between two points using procedural geometry. Supports sag, thickness, and segment count.",
        "short": "Create wire between points",
        "tags": ["procedural", "wire", "cable"]
    },
    "procedural_volume_scatter": {
        "name": "Procedural Volume Scatter",
        "category": "Procedural",
        "description": "Scatter objects within a 3D volume (box or sphere) rather than on a surface. Creates floating debris, particles, or volumetric props.",
        "short": "Scatter in 3D volume",
        "tags": ["procedural", "scatter", "volume"]
    },

    # ── Project Scaffold ──────────────────────────────────────────────────
    "scaffold_list_templates": {
        "name": "Scaffold List Templates",
        "category": "Project",
        "description": "List all available project folder templates — both built-in (BR island, creative hub) and custom saved templates.",
        "short": "List project templates",
        "tags": ["scaffold", "templates", "project"]
    },
    "scaffold_preview": {
        "name": "Scaffold Preview",
        "category": "Project",
        "description": "Zero-change preview — shows exactly what scaffold_generate would create without making any changes. Review the folder tree first.",
        "short": "Preview folder structure",
        "tags": ["scaffold", "preview", "safe"]
    },
    "scaffold_generate": {
        "name": "Scaffold Generate",
        "category": "Project",
        "description": "Creates the full folder tree for the chosen template (Meshes/, Materials/, Audio/, Verse/, etc.). Sets up a professional project structure.",
        "short": "Generate project folders",
        "tags": ["scaffold", "generate", "folders"]
    },
    "scaffold_save_template": {
        "name": "Scaffold Save Template",
        "category": "Project",
        "description": "Save your current folder structure as a named template for reuse in future projects.",
        "short": "Save custom template",
        "tags": ["scaffold", "template", "save"]
    },
    "scaffold_delete_template": {
        "name": "Scaffold Delete Template",
        "category": "Project",
        "description": "Delete a saved custom project template by name.",
        "short": "Delete template",
        "tags": ["scaffold", "template", "delete"]
    },
    "scaffold_organize_loose": {
        "name": "Scaffold Organize Loose",
        "category": "Project",
        "description": "Scans /Game (top level) for assets sitting loose outside folders and organizes them into proper directories by type.",
        "short": "Organize loose assets",
        "tags": ["scaffold", "organize", "cleanup"]
    },
    "organize_smart_categorize": {
        "name": "Smart Categorize",
        "category": "Project",
        "description": "Scans a folder, detects asset types, guesses functional categories using regex keyword matching on names, and moves assets accordingly.",
        "short": "AI-categorize assets",
        "tags": ["organize", "categorize", "smart"]
    },

    # ── Reference Auditor ─────────────────────────────────────────────────
    "ref_audit_orphans": {
        "name": "Audit Orphan Assets",
        "category": "Reference Auditor",
        "description": "Find all assets under a path that nothing else references — candidates for deletion to reduce project size.",
        "short": "Find unreferenced assets",
        "tags": ["audit", "orphans", "cleanup"]
    },
    "ref_audit_redirectors": {
        "name": "Audit Redirectors",
        "category": "Reference Auditor",
        "description": "Find all ObjectRedirector assets (leftover from moves/renames) that should be cleaned up.",
        "short": "Find redirectors",
        "tags": ["audit", "redirectors", "cleanup"]
    },
    "ref_audit_duplicates": {
        "name": "Audit Duplicates",
        "category": "Reference Auditor",
        "description": "Find assets with the same base name living in different folders — potential duplicates wasting memory.",
        "short": "Find duplicate assets",
        "tags": ["audit", "duplicates", "cleanup"]
    },
    "ref_audit_unused_textures": {
        "name": "Audit Unused Textures",
        "category": "Reference Auditor",
        "description": "Find textures with zero referencers — prime deletion candidates to save memory.",
        "short": "Find unused textures",
        "tags": ["audit", "textures", "unused"]
    },
    "ref_fix_redirectors": {
        "name": "Fix Redirectors",
        "category": "Reference Auditor",
        "description": "Resolve all ObjectRedirectors under the scan path, updating references to point directly to the real assets.",
        "short": "Fix all redirectors",
        "tags": ["audit", "redirectors", "fix"]
    },
    "ref_delete_orphans": {
        "name": "Delete Orphan Assets",
        "category": "Reference Auditor",
        "description": "Delete assets with no referencers. ⚠️ Destructive — run ref_audit_orphans first to review.",
        "short": "Delete unreferenced assets",
        "tags": ["audit", "orphans", "delete"]
    },
    "ref_full_report": {
        "name": "Full Reference Report",
        "category": "Reference Auditor",
        "description": "Run every audit check (orphans, redirectors, duplicates, unused textures) and write a comprehensive JSON report.",
        "short": "Complete audit report",
        "tags": ["audit", "report", "comprehensive"]
    },

    # ── Screenshot Tools ──────────────────────────────────────────────────
    "screenshot_take": {
        "name": "Screenshot Take",
        "category": "Screenshot",
        "description": "Take a high-resolution screenshot of the current editor viewport and save it to the screenshots folder.",
        "short": "Capture viewport screenshot",
        "tags": ["screenshot", "capture", "viewport"]
    },
    "screenshot_focus_selection": {
        "name": "Screenshot Focus Selection",
        "category": "Screenshot",
        "description": "Automatically frame all selected actors, capture a screenshot, and save it. Perfect for documenting specific assets or areas.",
        "short": "Screenshot selected actors",
        "tags": ["screenshot", "selection", "focus"]
    },
    "screenshot_timed_series": {
        "name": "Screenshot Timed Series",
        "category": "Screenshot",
        "description": "Capture a series of screenshots at regular intervals from the current camera. Useful for time-lapse documentation.",
        "short": "Timed screenshot series",
        "tags": ["screenshot", "series", "timelapse"]
    },
    "screenshot_open_folder": {
        "name": "Screenshot Open Folder",
        "category": "Screenshot",
        "description": "Print the screenshot output folder path so you can find your captured shots quickly.",
        "short": "Show screenshot folder",
        "tags": ["screenshot", "folder", "path"]
    },

    # ── Spline Tools ──────────────────────────────────────────────────────
    "spline_place_props": {
        "name": "Spline Place Props",
        "category": "Procedural",
        "description": "Place props along a spline actor's path with configurable spacing, rotation, and scale variation. Select a spline actor first.",
        "short": "Place props along spline",
        "tags": ["spline", "props", "path"]
    },
    "spline_clear_props": {
        "name": "Spline Clear Props",
        "category": "Procedural",
        "description": "Clear all props that were placed along a spline path.",
        "short": "Clear spline props",
        "tags": ["spline", "clear", "cleanup"]
    },

    # ── Spline to Verse ───────────────────────────────────────────────────
    "spline_to_verse_points": {
        "name": "Spline to Verse Points",
        "category": "Verse Helpers",
        "description": "Extract points from a selected spline actor and generate Verse code defining those positions as a vector array.",
        "short": "Export spline as Verse points",
        "tags": ["spline", "verse", "export"]
    },
    "spline_to_verse_patrol": {
        "name": "Spline to Verse Patrol",
        "category": "Verse Helpers",
        "description": "Generate a complete Verse creative_device that makes an NPC or object patrol along the spline path with configurable speed and looping.",
        "short": "Generate patrol device",
        "tags": ["spline", "verse", "patrol", "npc"]
    },
    "spline_to_verse_zone_boundary": {
        "name": "Spline to Verse Zone Boundary",
        "category": "Verse Helpers",
        "description": "Generate a Verse array of boundary points for custom zone detection — useful for safe zones, capture areas, or trigger regions.",
        "short": "Generate zone boundary",
        "tags": ["spline", "verse", "zone", "boundary"]
    },
    "spline_export_json": {
        "name": "Spline Export JSON",
        "category": "Verse Helpers",
        "description": "Export spline data as JSON including all points, tangents, and metadata for external tools.",
        "short": "Export spline to JSON",
        "tags": ["spline", "export", "json"]
    },

    # ── Text & Signs ──────────────────────────────────────────────────────
    "text_place": {
        "name": "Text Place",
        "category": "Text & Signs",
        "description": "Create and position a 3D text actor in the level with configurable font size, color, and orientation.",
        "short": "Place 3D text",
        "tags": ["text", "signs", "3d"]
    },
    "text_label_selection": {
        "name": "Text Label Selection",
        "category": "Text & Signs",
        "description": "Automatically create floating text labels above each selected actor showing their name. Great for zone tagging, debugging, or creating visual guides.",
        "short": "Label selected actors",
        "tags": ["text", "labels", "debugging"]
    },
    "text_paint_grid": {
        "name": "Text Paint Grid",
        "category": "Text & Signs",
        "description": "Generate grid labels like A1, A2 … D4 spaced evenly from origin. Perfect for map grids and coordinate systems.",
        "short": "Paint coordinate grid",
        "tags": ["text", "grid", "coordinates"]
    },
    "text_color_cycle": {
        "name": "Text Color Cycle",
        "category": "Text & Signs",
        "description": "Place a row of text actors, each one a different color from a palette. Creates rainbow or themed text displays.",
        "short": "Rainbow text row",
        "tags": ["text", "color", "rainbow"]
    },
    "text_save_style": {
        "name": "Text Save Style",
        "category": "Text & Signs",
        "description": "Save the current text settings (font, size, color) as a named style preset for reuse.",
        "short": "Save text style preset",
        "tags": ["text", "style", "preset"]
    },
    "text_list_styles": {
        "name": "Text List Styles",
        "category": "Text & Signs",
        "description": "List all saved text style presets with their settings.",
        "short": "List text presets",
        "tags": ["text", "styles", "list"]
    },
    "text_clear_folder": {
        "name": "Text Clear Folder",
        "category": "Text & Signs",
        "description": "Delete all text actors in the text output folder, cleaning up previous text placements.",
        "short": "Clear text actors",
        "tags": ["text", "clear", "cleanup"]
    },

    # ── Generative Text ───────────────────────────────────────────────────
    "text_render_texture": {
        "name": "Text Render Texture",
        "category": "Generative",
        "description": "Render text into a texture asset that can be applied to surfaces. Creates custom signage, labels, and UI elements.",
        "short": "Render text to texture",
        "tags": ["text", "texture", "render"]
    },
    "text_voxelize_3d": {
        "name": "Text Voxelize 3D",
        "category": "Generative",
        "description": "Convert text characters into 3D voxel sculptures made of cubes. Creates Minecraft-style 3D text displays.",
        "short": "3D voxel text",
        "tags": ["text", "voxel", "3d"]
    },

    # ── Verse Helpers ─────────────────────────────────────────────────────
    "verse_list_devices": {
        "name": "Verse List Devices",
        "category": "Verse Helpers",
        "description": "List all Verse-compatible creative devices in the current level with their class names and properties.",
        "short": "List Verse devices",
        "tags": ["verse", "devices", "list"]
    },
    "verse_bulk_set_property": {
        "name": "Verse Bulk Set Property",
        "category": "Verse Helpers",
        "description": "Set a property value on multiple Verse devices at once. Useful for batch-configuring timers, triggers, or scoring devices.",
        "short": "Bulk set device property",
        "tags": ["verse", "property", "bulk"]
    },
    "verse_select_by_name": {
        "name": "Verse Select by Name",
        "category": "Verse Helpers",
        "description": "Select all actors whose name matches a pattern. Supports wildcards for flexible selection.",
        "short": "Select actors by name",
        "tags": ["verse", "select", "name"]
    },
    "verse_select_by_class": {
        "name": "Verse Select by Class",
        "category": "Verse Helpers",
        "description": "Select all actors of a specific class type (e.g., all PointLights, all TriggerBoxes).",
        "short": "Select actors by class",
        "tags": ["verse", "select", "class"]
    },
    "verse_export_report": {
        "name": "Verse Export Report",
        "category": "Verse Helpers",
        "description": "Write a JSON file listing all devices, their classes, properties, and connections for documentation or AI analysis.",
        "short": "Export device report",
        "tags": ["verse", "export", "report"]
    },
    "verse_gen_custom": {
        "name": "Verse Generate Custom",
        "category": "Verse Helpers",
        "description": "Write arbitrary Verse code to the snippets folder. Quickly scaffold custom scripts without leaving the editor.",
        "short": "Write custom Verse code",
        "tags": ["verse", "generate", "custom"]
    },
    "verse_list_snippets": {
        "name": "Verse List Snippets",
        "category": "Verse Helpers",
        "description": "List all saved Verse code snippets in the snippets folder.",
        "short": "List Verse snippets",
        "tags": ["verse", "snippets", "list"]
    },
    "verse_open_snippets_folder": {
        "name": "Verse Open Snippets",
        "category": "Verse Helpers",
        "description": "Open the Verse snippets folder in your file explorer for easy access.",
        "short": "Open snippets folder",
        "tags": ["verse", "snippets", "folder"]
    },
    "verse_gen_device_declarations": {
        "name": "Verse Gen Declarations",
        "category": "Verse Helpers",
        "description": "Inspect selected actors and generate typed @editable Verse declarations for each device — saves manual typing of device references.",
        "short": "Generate @editable declarations",
        "tags": ["verse", "declarations", "codegen"]
    },
    "verse_gen_game_skeleton": {
        "name": "Verse Gen Game Skeleton",
        "category": "Verse Helpers",
        "description": "Generate a complete Verse game mode skeleton with round management, scoring, and player tracking boilerplate.",
        "short": "Generate game skeleton",
        "tags": ["verse", "skeleton", "gamemode"]
    },
    "verse_gen_elimination_handler": {
        "name": "Verse Gen Elimination Handler",
        "category": "Verse Helpers",
        "description": "Generate a Verse device that handles player eliminations with scoring, respawning, and elimination feed.",
        "short": "Generate elimination handler",
        "tags": ["verse", "elimination", "combat"]
    },
    "verse_gen_scoring_tracker": {
        "name": "Verse Gen Scoring Tracker",
        "category": "Verse Helpers",
        "description": "Generate a Verse scoring tracker device with configurable point values, team scoring, and win conditions.",
        "short": "Generate scoring tracker",
        "tags": ["verse", "scoring", "tracker"]
    },
    "verse_gen_prop_spawner": {
        "name": "Verse Gen Prop Spawner",
        "category": "Verse Helpers",
        "description": "Generate a Verse device that dynamically spawns props at runtime with configurable timing and locations.",
        "short": "Generate prop spawner",
        "tags": ["verse", "spawner", "runtime"]
    },

    # ── System / Utilities ────────────────────────────────────────────────
    "system_optimize_background_cpu": {
        "name": "Optimize Background CPU",
        "category": "System",
        "description": "Toggle 'Use Less CPU when in Background' setting. Reduces CPU usage when the editor is not in focus.",
        "short": "Toggle background CPU mode",
        "tags": ["system", "cpu", "performance"]
    },
    "plugin_validate_all": {
        "name": "Plugin Validate All",
        "category": "Utilities",
        "description": "Check all registered tools for schema and description requirements. Reports any tools with missing or invalid metadata.",
        "short": "Validate all tool plugins",
        "tags": ["plugin", "validate", "quality"]
    },
    "plugin_list_custom": {
        "name": "Plugin List Custom",
        "category": "Utilities",
        "description": "List tools that were loaded from the Custom_Plugins directory — user-created extensions to the toolbelt.",
        "short": "List custom plugins",
        "tags": ["plugin", "custom", "extensions"]
    },
    "toolbelt_integration_test": {
        "name": "Integration Test",
        "category": "Tests",
        "description": "Run the full suite of automated integration tests — spawns fixtures, runs tools, verifies results, and cleans up. 103/103 sections must pass.",
        "short": "Run all integration tests",
        "tags": ["test", "integration", "verify"]
    },
}

# ============================================================================
# TOOL SOURCE PARSING
# ============================================================================

def _safe_literal_eval(node: Optional[ast.AST]) -> Any:
    """Best-effort literal evaluation for decorator metadata and defaults."""
    if node is None:
        return None
    try:
        return ast.literal_eval(node)
    except Exception:
        return None


def _ast_to_text(node: Optional[ast.AST]) -> str:
    """Convert an AST node back to source text when literal eval is not possible."""
    if node is None:
        return ""
    try:
        return ast.unparse(node)
    except Exception:
        return ""


def _infer_param_type(annotation: str, default_value: Any) -> str:
    """Infer a simple parameter type label for UI and tool prompts."""
    if annotation:
        return annotation
    if isinstance(default_value, bool):
        return "bool"
    if isinstance(default_value, int) and not isinstance(default_value, bool):
        return "int"
    if isinstance(default_value, float):
        return "float"
    if isinstance(default_value, list):
        return "list"
    if isinstance(default_value, dict):
        return "dict"
    if default_value is None:
        return "string"
    return type(default_value).__name__


def _parse_param_docs(docstring: str) -> Dict[str, str]:
    """Extract per-parameter descriptions from a simple Google-style Args block."""
    docs: Dict[str, str] = {}
    if not docstring:
        return docs

    lines = docstring.splitlines()
    in_args = False
    current_name = None

    for raw_line in lines:
        line = raw_line.rstrip()
        stripped = line.strip()
        if stripped in {"Args:", "Arguments:", "Parameters:"}:
            in_args = True
            current_name = None
            continue
        if not in_args:
            continue
        if not stripped:
            continue
        if ":" in stripped and not stripped.startswith("-"):
            name, desc = stripped.split(":", 1)
            param_name = name.strip().lstrip("*")
            if param_name:
                docs[param_name] = desc.strip()
                current_name = param_name
                continue
        if current_name:
            docs[current_name] = f"{docs[current_name]} {stripped}".strip()

    return docs


def _extract_function_parameters(node: ast.FunctionDef | ast.AsyncFunctionDef) -> List[Dict[str, Any]]:
    """Convert a registered tool function signature into frontend/tool-call metadata."""
    parameters: List[Dict[str, Any]] = []
    docstring = ast.get_docstring(node) or ""
    param_docs = _parse_param_docs(docstring)

    positional_args = [arg for arg in node.args.args if arg.arg != "self"]
    positional_defaults: List[Optional[ast.AST]] = [None] * (len(positional_args) - len(node.args.defaults))
    positional_defaults.extend(node.args.defaults)

    for arg, default_node in zip(positional_args, positional_defaults):
        if arg.arg == "kwargs":
            continue
        default_value = _safe_literal_eval(default_node)
        default_repr = default_value if default_value is not None or isinstance(default_node, ast.Constant) else _ast_to_text(default_node)
        annotation = _ast_to_text(arg.annotation)
        parameters.append({
            "name": arg.arg,
            "type": _infer_param_type(annotation, default_value),
            "required": default_node is None,
            "default": default_repr,
            "description": param_docs.get(arg.arg, ""),
        })

    for arg, default_node in zip(node.args.kwonlyargs, node.args.kw_defaults):
        if arg.arg == "kwargs":
            continue
        default_value = _safe_literal_eval(default_node)
        default_repr = default_value if default_value is not None or isinstance(default_node, ast.Constant) else _ast_to_text(default_node)
        annotation = _ast_to_text(arg.annotation)
        parameters.append({
            "name": arg.arg,
            "type": _infer_param_type(annotation, default_value),
            "required": default_node is None,
            "default": default_repr,
            "description": param_docs.get(arg.arg, ""),
        })

    return parameters


def _parse_registered_tools_from_file(py_file: Path, python_root: Path) -> List[Dict[str, Any]]:
    """Parse @register_tool decorators from source without importing Unreal-only modules."""
    try:
        source = py_file.read_text(encoding="utf-8")
        tree = ast.parse(source)
    except Exception as exc:
        logger.warning("Failed to parse tool source %s: %s", py_file, exc)
        return []

    default_category = py_file.parent.name.replace("_", " ").title() if py_file.parent != python_root else "Core"
    discovered: List[Dict[str, Any]] = []

    # Walk entire AST to find @register_tool decorators at any nesting level
    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue

        for decorator in node.decorator_list:
            if not isinstance(decorator, ast.Call):
                continue

            decorator_name = ""
            if isinstance(decorator.func, ast.Name):
                decorator_name = decorator.func.id
            elif isinstance(decorator.func, ast.Attribute):
                decorator_name = decorator.func.attr

            if decorator_name != "register_tool":
                continue

            meta = {kw.arg: _safe_literal_eval(kw.value) for kw in decorator.keywords if kw.arg}
            tool_id = str(meta.get("name") or node.name.removeprefix("run_")).strip()
            if not tool_id:
                continue

            catalog_info = TOOL_DESCRIPTIONS.get(tool_id, {})
            description = str(meta.get("description") or catalog_info.get("description") or (ast.get_docstring(node) or "").strip())
            category = str(meta.get("category") or catalog_info.get("category") or default_category)
            tags = meta.get("tags")
            if not isinstance(tags, list):
                tags = catalog_info.get("tags", [category.lower()])

            discovered.append({
                "id": tool_id,
                "name": catalog_info.get("name", tool_id.replace("_", " ").title()),
                "category": category,
                "description": description,
                "short_description": catalog_info.get("short", description.split(".")[0] if description else "Tool"),
                "tags": [str(tag) for tag in tags],
                "parameters": _extract_function_parameters(node),
                "enabled": True,
                "source": str(py_file.relative_to(WORKSPACE_ROOT)).replace("\\", "/"),
            })
            break

    return discovered

# ============================================================================
# FLASK APP
# ============================================================================

app = Flask(__name__)
CORS(app)

# ============================================================================
# TOOL REGISTRY
# ============================================================================

class ToolRegistry:
    """Manages tool discovery and execution with AI awareness."""

    def __init__(self):
        self.tools: Dict[str, Dict[str, Any]] = {}
        self.categories: Dict[str, List[str]] = {}
        self.scan_tools()

    def scan_tools(self):
        """Scan for available tools."""
        logger.info("Scanning for UEFN-TOOLBELT tools...")
        self.tools = {}
        self.categories = {}

        # Load from manifest if exists
        manifest_file = VENDOR_UEFN / "tool_manifest.json"
        if manifest_file.exists():
            try:
                with open(manifest_file, encoding="utf-8") as f:
                    manifest = json.load(f)
                    self._load_manifest_tools(manifest)
                    logger.info(f"Loaded {len(self.tools)} tools from manifest")
            except Exception as e:
                logger.warning(f"Could not load manifest: {e}")

        if not self.tools:
            self._scan_registered_tools()
        if not self.tools:
            self._add_catalog_tools()

    def _load_manifest_tools(self, manifest: Dict[str, Any]):
        """Load tools from manifest."""
        tools = manifest.get("tools", [])
        for tool in tools:
            raw_name = str(tool.get("name") or tool.get("id") or "Unknown")
            tool_id = str(tool.get("id") or raw_name.lower().replace(" ", "_"))
            category = str(tool.get("category") or TOOL_DESCRIPTIONS.get(tool_id, {}).get("category") or "Other")
            catalog_info = TOOL_DESCRIPTIONS.get(tool_id, {})
            self._register_tool_meta({
                "id": tool_id,
                "name": catalog_info.get("name", raw_name),
                "category": category,
                "description": catalog_info.get("description", tool.get("description", "")),
                "short_description": catalog_info.get("short", "Tool"),
                "tags": catalog_info.get("tags", tool.get("tags", [])),
                "parameters": tool.get("parameters", []),
                "enabled": tool.get("enabled", True),
            })

    def _register_tool_meta(self, tool: Dict[str, Any]):
        """Insert a normalized tool record and maintain category indexes."""
        tool_id = str(tool.get("id") or "").strip()
        if not tool_id:
            return

        existing = self.tools.get(tool_id)
        if existing:
            prev_category = existing.get("category", "Other")
            if prev_category in self.categories:
                self.categories[prev_category] = [name for name in self.categories[prev_category] if name != tool_id]

        normalized = {
            "id": tool_id,
            "name": str(tool.get("name") or tool_id.replace("_", " ").title()),
            "category": str(tool.get("category") or "Other"),
            "description": str(tool.get("description") or ""),
            "short_description": str(tool.get("short_description") or "Tool"),
            "tags": [str(tag) for tag in (tool.get("tags") or [])],
            "parameters": list(tool.get("parameters") or []),
            "enabled": bool(tool.get("enabled", True)),
            "source": tool.get("source", ""),
        }
        self.tools[tool_id] = normalized
        self.categories.setdefault(normalized["category"], []).append(tool_id)

    def _scan_registered_tools(self):
        """Parse real @register_tool definitions from source files."""
        python_dir = VENDOR_UEFN / "Content" / "Python" / "UEFN_Toolbelt"
        if not python_dir.exists():
            logger.warning(f"UEFN-TOOLBELT directory not found")
            return

        try:
            discovered_count = 0
            for py_file in sorted(python_dir.glob("**/*.py")):
                if py_file.name.startswith("_"):
                    continue
                for tool in _parse_registered_tools_from_file(py_file, python_dir):
                    self._register_tool_meta(tool)
                    discovered_count += 1

            logger.info("Discovered %s registered tools from source", discovered_count)
        except Exception as e:
            logger.error(f"Error scanning tools: {e}")

    def _add_catalog_tools(self):
        """Add tools from catalog."""
        for tool_id, info in TOOL_DESCRIPTIONS.items():
            self._register_tool_meta({
                "id": tool_id,
                "name": info.get("name", tool_id.title()),
                "category": info.get("category", "Other"),
                "description": info.get("description", ""),
                "short_description": info.get("short", "Tool"),
                "tags": info.get("tags", []),
                "parameters": [],
                "enabled": True,
            })

    def get_all_tools(self) -> List[Dict[str, Any]]:
        """Get all tools with descriptions."""
        return sorted(self.tools.values(), key=lambda tool: (tool["category"], tool["name"]))

    def get_tool(self, tool_name_or_id: str) -> Optional[Dict[str, Any]]:
        """Return a single tool by id or display name."""
        key = self._resolve_tool_key(tool_name_or_id)
        if key in self.tools:
            return copy.deepcopy(self.tools[key])
        return None

    def get_tools_by_category(self, category: str) -> List[Dict[str, Any]]:
        """Get tools in category."""
        tool_names = self.categories.get(category, [])
        return [self.tools[name] for name in tool_names if name in self.tools]

    def search_tools(self, query: str) -> List[Dict[str, Any]]:
        """Search tools by name, description, tags, and intent keywords."""
        query_lower = (query or "").strip().lower()
        if not query_lower:
            return []

        stop_words = {
            "a", "an", "and", "for", "from", "i", "in", "into", "it", "me",
            "my", "of", "on", "please", "that", "the", "this", "to", "with",
            "want",
        }
        tokens = [token for token in re.findall(r"[a-z0-9]+", query_lower) if token not in stop_words]
        token_set = set(tokens)
        scored_results = []

        for tool in self.tools.values():
            haystack_parts = [
                tool.get("id", ""),
                tool.get("name", ""),
                tool.get("category", ""),
                tool.get("description", ""),
                " ".join(tool.get("tags", [])),
                " ".join(param.get("name", "") for param in tool.get("parameters", [])),
                " ".join(param.get("description", "") for param in tool.get("parameters", [])),
            ]
            haystack = " ".join(part for part in haystack_parts if part).lower()
            if not haystack:
                continue

            score = 0
            if query_lower == tool.get("id", "").lower():
                score += 140
            if query_lower == tool.get("name", "").lower():
                score += 120
            if query_lower in tool.get("id", "").lower():
                score += 90
            if query_lower in tool.get("name", "").lower():
                score += 80
            if query_lower in tool.get("description", "").lower():
                score += 50
            if query_lower in tool.get("category", "").lower():
                score += 35
            if any(query_lower in tag.lower() for tag in tool.get("tags", [])):
                score += 45

            matched_tokens = 0
            for token in tokens:
                token_score = 0
                if token in tool.get("id", "").lower():
                    token_score += 30
                if token in tool.get("name", "").lower():
                    token_score += 26
                if token in tool.get("category", "").lower():
                    token_score += 12
                if token in tool.get("description", "").lower():
                    token_score += 10
                if any(token in tag.lower() for tag in tool.get("tags", [])):
                    token_score += 14
                if token_score:
                    matched_tokens += 1
                    score += token_score

            if tokens:
                coverage = matched_tokens / max(len(tokens), 1)
                score += int(coverage * 25)
                if matched_tokens == len(tokens):
                    score += 20

            tool_id = str(tool.get("id") or "").lower()
            if token_set & {"import", "fbx", "mesh", "model", "asset"}:
                if tool_id in {"import_fbx", "import_fbx_folder", "organize_assets"}:
                    score += 70
            if token_set & {"curve", "curved", "curvature", "arc", "circle", "spiral", "helix"}:
                if tool_id in {"pattern_arc", "pattern_circle", "pattern_spiral", "pattern_helix"}:
                    score += 70
            if token_set & {"terrain", "landscape", "spline", "road", "path", "follow"}:
                if tool_id in {"spline_place_props", "scatter_road_edge", "scatter_along_path", "spline_measure"}:
                    score += 80
            if token_set & {"terrain", "landscape", "road"} and tool_id == "scatter_road_edge":
                score += 40

            if score > 0:
                scored_results.append((score, tool))

        scored_results.sort(key=lambda item: (-item[0], item[1]["category"], item[1]["name"]))
        return [tool for _, tool in scored_results]

    def _resolve_tool_key(self, tool_name_or_id: str) -> str:
        """Resolve a tool id, normalized id, or display name to a registry key."""
        candidate = (tool_name_or_id or "").strip()
        if not candidate:
            return ""
        if candidate in self.tools:
            return candidate

        normalized = candidate.lower().replace(" ", "_")
        for key, meta in self.tools.items():
            if meta.get("id", "").lower() == candidate.lower():
                return key
            if meta.get("id", "").lower() == normalized:
                return key
            if meta.get("name", "").lower() == candidate.lower():
                return key
        return candidate

    def execute_tool(self, tool_name: str, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Execute a tool via the UEFN MCP bridge (real execution)."""
        key = self._resolve_tool_key(tool_name)
        if key not in self.tools:
            return {
                "success": False,
                "error": f"Tool not found: {tool_name}",
                "result": None
            }
        if parameters is None:
            parameters = {}
        if not isinstance(parameters, dict):
            return {
                "success": False,
                "error": "Tool parameters must be a JSON object",
                "result": None,
                "tool": key,
            }

        # Try to execute via UEFN MCP bridge
        port = discover_uefn_listener_port()
        if port:
            try:
                # Use execute_python to run toolbelt tools inside UEFN
                code = (
                    "import UEFN_Toolbelt as tb\n"
                    "try:\n"
                    "    tb.register_all_tools()\n"
                    "except Exception:\n"
                    "    pass\n"
                    f"result = tb.run({key!r}, **{parameters!r})"
                )
                mcp_result = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)
                if mcp_result.get("success"):
                    inner = mcp_result.get("result", {})
                    return {
                        "success": True,
                        "result": {
                            "status": "completed",
                            "message": f"Executed {key} via UEFN MCP",
                            "mcp_result": inner.get("result"),
                            "stdout": inner.get("stdout", ""),
                            "stderr": inner.get("stderr", ""),
                            "timestamp": datetime.now().isoformat(),
                            "parameters": parameters,
                            "executed_via": "uefn_mcp"
                        },
                        "tool": key
                    }
                else:
                    return {
                        "success": False,
                        "error": mcp_result.get("error", "MCP command failed"),
                        "result": mcp_result,
                        "tool": key
                    }
            except Exception as e:
                logger.error(f"MCP tool execution failed for {key}: {e}")
                return {
                    "success": False,
                    "error": f"MCP execution error: {str(e)}",
                    "result": None,
                    "tool": key
                }
        else:
            # UEFN not connected — return informative error
            return {
                "success": False,
                "error": "UEFN MCP listener not running. Open UEFN with your project first.",
                "result": None,
                "tool": key
            }

    def describe_tool(self, tool_name_or_id: str) -> Dict[str, Any]:
        """Return a descriptive record for a tool lookup."""
        tool = self.get_tool(tool_name_or_id)
        if tool:
            return {"success": True, "tool": tool}
        return {"success": False, "error": f"Tool not found: {tool_name_or_id}"}

    def get_tools_for_ai(self) -> str:
        """Get tools in a format AI can understand."""
        tools_info = []
        for tool in self.get_all_tools():
            params = tool.get("parameters") or []
            params_text = ""
            if params:
                signature = ", ".join(
                    f"{param['name']}:{param.get('type', 'any')}{'' if param.get('required') else '?'}"
                    for param in params[:6]
                )
                if len(params) > 6:
                    signature += ", ..."
                params_text = f" Params: {signature}."
            tools_info.append(
                f"- {tool['id']} [{tool['category']}]: {tool['description']}{params_text}"
            )
        return "\n".join(tools_info)


# ============================================================================
# UEFN BRIDGE  (HTTP listener inside UEFN — same protocol as vendor/uefn-mcp-server)
# ============================================================================

def discover_uefn_listener_port() -> int | None:
    """Find the UEFN MCP HTTP listener (scans 8765..8770, matches GET / JSON health)."""
    for port in range(UEFN_MCP_PORT, UEFN_MCP_MAX_PORT + 1):
        try:
            with urllib.request.urlopen(f"http://{UEFN_MCP_HOST}:{port}/", timeout=1.2) as resp:
                body = json.loads(resp.read().decode())
            if body.get("status") == "ok":
                return port
        except Exception:
            continue
    return None


def mcp_listener_get_json(port: int) -> Dict[str, Any]:
    with urllib.request.urlopen(f"http://{UEFN_MCP_HOST}:{port}/", timeout=3.0) as resp:
        return json.loads(resp.read().decode())


def mcp_listener_post_command(port: int, command: str, params: Dict[str, Any] | None = None, timeout: float = 45.0) -> Dict[str, Any]:
    payload = json.dumps({"command": command, "params": params or {}}).encode("utf-8")
    req = urllib.request.Request(
        f"http://{UEFN_MCP_HOST}:{port}/",
        data=payload,
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode())


class UEFNBridge:
    """Tracks the UEFN in-editor MCP HTTP listener (not the stdio Claude MCP process)."""

    def __init__(self):
        self.is_connected = False
        self.discovered_port: int | None = None
        self.last_info: Dict[str, Any] = {}
        self.check_connection()

    def check_connection(self) -> bool:
        """True if the UEFN listener responds with JSON health on any scanned port."""
        port = discover_uefn_listener_port()
        self.discovered_port = port
        self.is_connected = port is not None
        if self.is_connected and port is not None:
            try:
                self.last_info = mcp_listener_get_json(port)
            except Exception:
                self.last_info = {}
            logger.info("✅ UEFN MCP HTTP listener on port %s", port)
        return self.is_connected

    def get_status(self) -> Dict[str, Any]:
        return {
            "connected": self.is_connected,
            "host": UEFN_MCP_HOST,
            "port": self.discovered_port or UEFN_MCP_PORT,
            "listener_port": self.discovered_port,
            "protocol_version": self.last_info.get("version"),
            "command_count": len(self.last_info.get("commands", [])),
            "timestamp": datetime.now().isoformat(),
        }


# ============================================================================
# GLOBAL INSTANCES
# ============================================================================

tool_registry = ToolRegistry()


def _iso_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _compact_text(text: str, limit: int = 140) -> str:
    compact = re.sub(r"\s+", " ", (text or "")).strip()
    if len(compact) <= limit:
        return compact
    return compact[: limit - 1].rstrip() + "…"


def _tokenize_query(text: str) -> List[str]:
    """Tokenize free-form text for lightweight in-process retrieval."""
    return [token for token in re.findall(r"[a-z0-9_./-]+", (text or "").lower()) if len(token) >= 2]


def _score_text_match(query: str, *haystacks: str) -> float:
    """Compute a simple relevance score for retrieval without external indexes."""
    query_text = (query or "").strip().lower()
    if not query_text:
        return 0.0

    tokens = _tokenize_query(query_text)
    combined = "\n".join(haystacks).lower()
    score = 0.0

    if query_text in combined:
        score += 8.0

    for token in tokens:
        count = combined.count(token)
        if count:
            score += min(3.0, 0.8 + (count * 0.35))

    return score


def _title_from_message(text: str) -> str:
    title = _compact_text(text, limit=54).strip(" .-:;,")
    return title or "New Chat"


def _build_session_memory(messages: List[Dict[str, Any]]) -> str:
    """Create a compact, deterministic memory summary from a session transcript."""
    if not messages:
        return ""

    recent_requests: List[str] = []
    recent_actions: List[str] = []

    for message in messages:
        content = message.get("content", "")
        if message.get("role") == "user" and content:
            recent_requests.append(_compact_text(content, limit=180))

        tool_result = message.get("toolResult")
        if tool_result:
            tool_name = tool_result.get("tool", "tool")
            output = tool_result.get("output")
            if isinstance(output, dict):
                if output.get("success") is False or output.get("error"):
                    status = f"error: {output.get('error', 'failed')}"
                elif output.get("success") is True:
                    status = "success"
                else:
                    status = "completed"
                detail = _compact_text(json.dumps(output, default=str), limit=180)
            else:
                status = "completed"
                detail = _compact_text(str(output), limit=180)
            recent_actions.append(f"{tool_name}: {status}. {detail}")

    lines: List[str] = []
    if recent_requests:
        lines.append("Recent user requests:")
        for item in recent_requests[-6:]:
            lines.append(f"- {item}")
    if recent_actions:
        lines.append("Recent tool outcomes:")
        for item in recent_actions[-6:]:
            lines.append(f"- {item}")

    return "\n".join(lines)


class ChatSessionStore:
    """Persistent JSON-backed chat/project store for the AI assistant."""

    def __init__(self, path: Path):
        self.path = path
        self._lock = threading.RLock()
        self._data: Dict[str, Any] = {"sessions": []}
        self._load()

    def _load(self):
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if not self.path.exists():
                self._save_locked()
                return
            try:
                raw = json.loads(self.path.read_text(encoding="utf-8"))
                sessions = raw.get("sessions", []) if isinstance(raw, dict) else []
                projects = raw.get("projects", []) if isinstance(raw, dict) else []
                self._data = {
                    "sessions": [self._normalize_session(session) for session in sessions],
                    "projects": projects,
                }
            except Exception as exc:
                logger.warning("Could not read chat store %s: %s", self.path, exc)
                self._data = {"sessions": [], "projects": []}
                self._save_locked()

    def _save_locked(self):
        tmp_path = self.path.with_suffix(".tmp")
        tmp_path.write_text(json.dumps(self._data, indent=2), encoding="utf-8")
        attempts = 4
        for attempt in range(attempts):
            try:
                tmp_path.replace(self.path)
                return
            except PermissionError:
                if attempt >= attempts - 1:
                    raise
                time.sleep(0.05 * (attempt + 1))

    def _normalize_session(self, session: Dict[str, Any]) -> Dict[str, Any]:
        normalized = {
            "id": session.get("id") or uuid4().hex,
            "title": session.get("title") or "New Chat",
            "project_id": session.get("project_id") or "",
            "created_at": session.get("created_at") or _iso_now(),
            "updated_at": session.get("updated_at") or session.get("created_at") or _iso_now(),
            "messages": list(session.get("messages") or []),
            "memory_summary": session.get("memory_summary") or "",
            "last_provider": session.get("last_provider") or "",
            "last_model": session.get("last_model") or "",
        }
        return normalized

    def _find_session_locked(self, chat_id: str) -> Optional[Dict[str, Any]]:
        for session in self._data["sessions"]:
            if session.get("id") == chat_id:
                return session
        return None

    def _session_summary(self, session: Dict[str, Any]) -> Dict[str, Any]:
        last_message = session["messages"][-1] if session["messages"] else None
        preview = _compact_text(last_message.get("content", ""), limit=96) if last_message else ""
        return {
            "id": session["id"],
            "title": session["title"],
            "project_id": session.get("project_id", ""),
            "created_at": session["created_at"],
            "updated_at": session["updated_at"],
            "message_count": len(session["messages"]),
            "preview": preview,
            "last_provider": session.get("last_provider", ""),
            "last_model": session.get("last_model", ""),
        }

    def search_sessions(self, query: str, limit: int = 6, exclude_chat_id: str = "") -> List[Dict[str, Any]]:
        """Search prior chats/projects by title, memory summary, and recent messages."""
        query = (query or "").strip()
        if not query:
            return []

        with self._lock:
            scored: List[tuple[float, Dict[str, Any]]] = []
            for session in self._data["sessions"]:
                if exclude_chat_id and session.get("id") == exclude_chat_id:
                    continue

                recent_messages = "\n".join(
                    str(message.get("content") or "")
                    for message in session.get("messages", [])[-8:]
                )
                score = _score_text_match(
                    query,
                    str(session.get("title") or ""),
                    str(session.get("memory_summary") or ""),
                    recent_messages,
                )
                if score <= 0:
                    continue

                summary = self._session_summary(session)
                summary["memory_summary"] = session.get("memory_summary", "")
                summary["recent_excerpt"] = _compact_text(recent_messages, limit=220)
                scored.append((score, summary))

            scored.sort(key=lambda item: (item[0], item[1].get("updated_at", "")), reverse=True)
            return [item[1] for item in scored[:limit]]

    def create_session(self, title: str = "", project_id: str = "") -> Dict[str, Any]:
        with self._lock:
            now = _iso_now()
            session = {
                "id": uuid4().hex,
                "title": title.strip() or "New Chat",
                "project_id": project_id,
                "created_at": now,
                "updated_at": now,
                "messages": [],
                "memory_summary": "",
                "last_provider": "",
                "last_model": "",
            }
            self._data["sessions"].append(session)
            self._save_locked()
            return copy.deepcopy(session)

    def get_session(self, chat_id: str) -> Optional[Dict[str, Any]]:
        with self._lock:
            session = self._find_session_locked(chat_id)
            if not session:
                return None
            return copy.deepcopy(self._normalize_session(session))

    def ensure_session(self, chat_id: str = "", title: str = "") -> Dict[str, Any]:
        if chat_id:
            existing = self.get_session(chat_id)
            if existing:
                return existing
        return self.create_session(title=title)

    def append_messages(self, chat_id: str, messages: List[Dict[str, Any]], provider: str = "", model: str = "") -> Optional[Dict[str, Any]]:
        with self._lock:
            session = self._find_session_locked(chat_id)
            if not session:
                return None

            for message in messages:
                session["messages"].append(message)

            if session["title"] == "New Chat":
                for message in session["messages"]:
                    if message.get("role") == "user" and message.get("content"):
                        session["title"] = _title_from_message(message["content"])
                        break

            session["updated_at"] = _iso_now()
            session["memory_summary"] = _build_session_memory(session["messages"])
            if provider:
                session["last_provider"] = provider
            if model:
                session["last_model"] = model
            self._save_locked()
            return copy.deepcopy(self._normalize_session(session))

    def update_session(self, chat_id: str, title: Optional[str] = None, clear_messages: bool = False) -> Optional[Dict[str, Any]]:
        """Rename a session and/or clear its message history while keeping the project container."""
        with self._lock:
            session = self._find_session_locked(chat_id)
            if not session:
                return None

            if title is not None:
                session["title"] = title.strip() or "New Chat"

            if clear_messages:
                session["messages"] = []
                session["memory_summary"] = ""

            session["updated_at"] = _iso_now()
            self._save_locked()
            return copy.deepcopy(self._normalize_session(session))

    def delete_session(self, chat_id: str) -> bool:
        with self._lock:
            before = len(self._data["sessions"])
            self._data["sessions"] = [session for session in self._data["sessions"] if session.get("id") != chat_id]
            deleted = len(self._data["sessions"]) != before
            if deleted:
                self._save_locked()
            return deleted

    # ── Project methods ───────────────────────────────────────────────────

    def list_projects(self) -> List[Dict[str, Any]]:
        with self._lock:
            projects = self._data.get("projects", [])
            result = []
            for p in sorted(projects, key=lambda x: x.get("updated_at", ""), reverse=True):
                chat_count = sum(1 for s in self._data["sessions"] if s.get("project_id") == p["id"])
                result.append({**p, "chat_count": chat_count})
            return result

    def create_project(self, name: str, icon: str = "", color: str = "") -> Dict[str, Any]:
        with self._lock:
            if "projects" not in self._data:
                self._data["projects"] = []
            now = _iso_now()
            project = {
                "id": uuid4().hex,
                "name": name.strip() or "New project",
                "icon": icon or "📁",
                "color": color or "",
                "created_at": now,
                "updated_at": now,
            }
            self._data["projects"].append(project)
            self._save_locked()
            return copy.deepcopy(project)

    def update_project(self, project_id: str, name: str = "", icon: str = "", color: str = "") -> Optional[Dict[str, Any]]:
        with self._lock:
            for p in self._data.get("projects", []):
                if p["id"] == project_id:
                    if name:
                        p["name"] = name
                    if icon:
                        p["icon"] = icon
                    if color is not None and color != "":
                        p["color"] = color
                    p["updated_at"] = _iso_now()
                    self._save_locked()
                    chat_count = sum(1 for s in self._data["sessions"] if s.get("project_id") == project_id)
                    return {**copy.deepcopy(p), "chat_count": chat_count}
            return None

    def delete_project(self, project_id: str) -> bool:
        with self._lock:
            projects = self._data.get("projects", [])
            before = len(projects)
            self._data["projects"] = [p for p in projects if p["id"] != project_id]
            # Also delete all chats in this project
            self._data["sessions"] = [s for s in self._data["sessions"] if s.get("project_id") != project_id]
            deleted = len(self._data.get("projects", [])) != before
            if deleted:
                self._save_locked()
            return deleted

    def list_sessions(self, project_id: str = "") -> List[Dict[str, Any]]:
        """List sessions, optionally filtered by project_id. Empty string = all."""
        with self._lock:
            sessions = self._data["sessions"]
            if project_id:
                sessions = [s for s in sessions if s.get("project_id") == project_id]
            ordered = sorted(sessions, key=lambda session: session.get("updated_at", ""), reverse=True)
            return [self._session_summary(session) for session in ordered]


class KnowledgeStore:
    """Persistent shared knowledge base used by chat, research, and planning."""

    def __init__(self, path: Path):
        self.path = path
        self._lock = threading.RLock()
        self._data: Dict[str, Any] = {"items": []}
        self._load()

    def _load(self):
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if not self.path.exists():
                self._save_locked()
                return
            try:
                raw = json.loads(self.path.read_text(encoding="utf-8"))
                items = raw.get("items", []) if isinstance(raw, dict) else []
                self._data = {"items": [self._normalize_item(item) for item in items]}
            except Exception as exc:
                logger.warning("Could not read knowledge store %s: %s", self.path, exc)
                self._data = {"items": []}
                self._save_locked()

    def _save_locked(self):
        tmp_path = self.path.with_suffix(".tmp")
        tmp_path.write_text(json.dumps(self._data, indent=2), encoding="utf-8")
        tmp_path.replace(self.path)

    def _normalize_item(self, item: Dict[str, Any]) -> Dict[str, Any]:
        created_at = item.get("created_at") or item.get("addedAt") or _iso_now()
        updated_at = item.get("updated_at") or created_at
        normalized = {
            "id": item.get("id") or uuid4().hex,
            "type": item.get("type") or item.get("source_type") or "text",
            "source_type": item.get("source_type") or item.get("type") or "manual",
            "scope": item.get("scope") or "global",
            "title": item.get("title") or "Untitled",
            "content": item.get("content") or "",
            "quality": int(item.get("quality", 1) or 0),
            "tags": [str(tag) for tag in (item.get("tags") or [])],
            "sourceUrl": item.get("sourceUrl") or item.get("source_url") or "",
            "chat_id": item.get("chat_id") or "",
            "created_at": created_at,
            "updated_at": updated_at,
            "addedAt": item.get("addedAt") or created_at,
        }
        return normalized

    def list_items(self, include_excluded: bool = True) -> List[Dict[str, Any]]:
        with self._lock:
            items = sorted(self._data["items"], key=lambda item: item.get("updated_at", ""), reverse=True)
            if not include_excluded:
                items = [item for item in items if item.get("quality", 1) != 0]
            return [copy.deepcopy(item) for item in items]

    def get_item(self, item_id: str) -> Optional[Dict[str, Any]]:
        with self._lock:
            for item in self._data["items"]:
                if item.get("id") == item_id:
                    return copy.deepcopy(item)
        return None

    def add_item(
        self,
        *,
        item_type: str,
        title: str,
        content: str,
        source_type: str = "manual",
        scope: str = "global",
        tags: Optional[List[str]] = None,
        quality: int = 1,
        source_url: str = "",
        chat_id: str = "",
    ) -> Dict[str, Any]:
        with self._lock:
            now = _iso_now()
            compact_content = _compact_text(content, limit=1200)

            for idx, existing in enumerate(self._data["items"]):
                if (
                    existing.get("source_type") == (source_type or item_type or "manual")
                    and existing.get("title") == (title.strip() or "Untitled")
                    and existing.get("chat_id", "") == (chat_id or "")
                    and _compact_text(existing.get("content", ""), limit=1200) == compact_content
                ):
                    refreshed = dict(existing)
                    refreshed["updated_at"] = now
                    refreshed["quality"] = quality
                    refreshed["tags"] = [str(tag) for tag in (tags or refreshed.get("tags") or [])]
                    self._data["items"][idx] = self._normalize_item(refreshed)
                    self._save_locked()
                    return copy.deepcopy(self._data["items"][idx])

            item = self._normalize_item({
                "id": uuid4().hex,
                "type": item_type or "text",
                "source_type": source_type or item_type or "manual",
                "scope": scope or "global",
                "title": title.strip() or "Untitled",
                "content": content,
                "quality": quality,
                "tags": tags or [],
                "sourceUrl": source_url,
                "chat_id": chat_id,
                "created_at": now,
                "updated_at": now,
                "addedAt": now,
            })
            self._data["items"].append(item)
            self._save_locked()
            return copy.deepcopy(item)

    def update_item(self, item_id: str, **changes: Any) -> Optional[Dict[str, Any]]:
        with self._lock:
            for idx, item in enumerate(self._data["items"]):
                if item.get("id") != item_id:
                    continue

                updated = dict(item)
                if "title" in changes and changes["title"] is not None:
                    updated["title"] = str(changes["title"]).strip() or updated["title"]
                if "content" in changes and changes["content"] is not None:
                    updated["content"] = str(changes["content"])
                if "quality" in changes and changes["quality"] is not None:
                    updated["quality"] = int(changes["quality"])
                if "tags" in changes and changes["tags"] is not None:
                    updated["tags"] = [str(tag) for tag in changes["tags"]]
                if "sourceUrl" in changes and changes["sourceUrl"] is not None:
                    updated["sourceUrl"] = str(changes["sourceUrl"])
                updated["updated_at"] = _iso_now()
                self._data["items"][idx] = self._normalize_item(updated)
                self._save_locked()
                return copy.deepcopy(self._data["items"][idx])
        return None

    def delete_item(self, item_id: str) -> bool:
        with self._lock:
            before = len(self._data["items"])
            self._data["items"] = [item for item in self._data["items"] if item.get("id") != item_id]
            deleted = len(self._data["items"]) != before
            if deleted:
                self._save_locked()
            return deleted

    def search(self, query: str, limit: int = 8, include_excluded: bool = False) -> List[Dict[str, Any]]:
        query = (query or "").strip()
        if not query:
            return []

        with self._lock:
            scored: List[tuple[float, Dict[str, Any]]] = []
            for item in self._data["items"]:
                if item.get("quality", 1) == 0 and not include_excluded:
                    continue
                score = _score_text_match(
                    query,
                    str(item.get("title") or ""),
                    str(item.get("content") or ""),
                    " ".join(item.get("tags") or []),
                )
                if score <= 0:
                    continue
                summary = copy.deepcopy(item)
                summary["content"] = _compact_text(summary.get("content", ""), limit=360)
                scored.append((score, summary))

            scored.sort(key=lambda row: (row[0], row[1].get("updated_at", "")), reverse=True)
            return [row[1] for row in scored[:limit]]

    def remember_interaction(
        self,
        *,
        chat_id: str,
        chat_title: str,
        user_message: str,
        assistant_message: Dict[str, Any],
    ) -> None:
        """Persist explicit preferences and successful tool outcomes as shared memory."""
        user_text = (user_message or "").strip()
        if not user_text:
            return

        lowered = user_text.lower()
        if any(trigger in lowered for trigger in ["remember ", "remember that", "always ", "prefer ", "use this style", "don't forget"]):
            self.add_item(
                item_type="text",
                source_type="preference",
                scope="global",
                title=f"Preference from {chat_title or 'chat'}",
                content=user_text,
                tags=["preference", "memory", "chat"],
                chat_id=chat_id,
            )

        tool_result = assistant_message.get("toolResult") or {}
        tool_name = str(tool_result.get("tool") or "").strip()
        output = tool_result.get("output")
        if tool_name and output:
            success = True
            if isinstance(output, dict):
                success = output.get("success", True) is not False and not output.get("error")
            if success:
                self.add_item(
                    item_type="text",
                    source_type="tool_outcome",
                    scope="global",
                    title=f"{tool_name} outcome",
                    content=f"Chat: {chat_title or chat_id}\nUser request: {user_text}\nOutcome: {_compact_text(json.dumps(output, default=str), limit=600)}",
                    tags=["tool", tool_name, "learned"],
                    chat_id=chat_id,
                )

    def remember_attachments(
        self,
        *,
        chat_id: str,
        chat_title: str,
        attachments: List[Dict[str, Any]],
    ) -> None:
        """Persist useful attachment contents so later chats can search them."""
        context_title = chat_title or chat_id or "chat"

        for att in (attachments or [])[:MAX_ATTACHMENT_MEMORY_ITEMS]:
            name = str(att.get("name") or "attachment").strip() or "attachment"
            attachment_type = str(att.get("type") or "file").strip().lower()
            mime_type = str(att.get("mimeType") or att.get("mime_type") or "").strip()
            source_url = str(att.get("sourceUrl") or att.get("source_url") or "").strip()
            content = str(att.get("content") or "").strip()
            analysis_text = str(att.get("analysisText") or att.get("analysis_text") or "").strip()
            analysis_caption = str(att.get("analysisCaption") or att.get("analysis_caption") or "").strip()
            analysis_handwriting = str(att.get("analysisHandwriting") or att.get("analysis_handwriting") or "").strip()
            analysis_meta = att.get("analysisMeta") or att.get("analysis_meta") or {}

            if source_url and (content or analysis_text):
                host = urllib.parse.urlsplit(source_url).netloc or "web"
                web_text = content or analysis_text
                self.add_item(
                    item_type="text",
                    source_type="web_page",
                    scope="global",
                    title=f"Web resource: {name}",
                    content=f"Source URL: {source_url}\nHost: {host}\nFilename: {name}\n\n{web_text[:MAX_ATTACHMENT_MEMORY_TEXT_CHARS]}",
                    tags=["attachment", "web", host],
                    chat_id=chat_id,
                )
                continue

            if attachment_type == "file" and content:
                suffix = Path(name).suffix.lower().lstrip(".") or "text"
                self.add_item(
                    item_type="text",
                    source_type="attachment_text",
                    scope="global",
                    title=f"Attachment: {name}",
                    content=f"Chat: {context_title}\nAttachment type: {mime_type or 'text file'}\nFilename: {name}\n\n{content[:MAX_ATTACHMENT_MEMORY_TEXT_CHARS]}",
                    tags=["attachment", "file", suffix],
                    chat_id=chat_id,
                )
                continue

            if attachment_type == "image" and (analysis_text or analysis_caption):
                self.add_item(
                    item_type="text",
                    source_type="attachment_ocr" if analysis_text else "attachment_visual",
                    scope="global",
                    title=f"Image analysis: {name}",
                    content=(
                        f"Chat: {context_title}\nAttachment type: {mime_type or 'image'}\nFilename: {name}\n\n"
                        + (f"Visual description:\n{analysis_caption[:2000]}\n\n" if analysis_caption else "")
                        + (f"Handwriting guess:\n{analysis_handwriting[:2000]}\n\n" if analysis_handwriting else "")
                        + (
                            f"Visual diagnostics:\n{json.dumps(analysis_meta, default=str)[:1500]}\n\n"
                            if isinstance(analysis_meta, dict) and analysis_meta else ""
                        )
                        + (f"Visible text detected in the image:\n{analysis_text[:MAX_ATTACHMENT_MEMORY_TEXT_CHARS]}" if analysis_text else "")
                    ),
                    tags=["attachment", "image", "ocr" if analysis_text else "visual", "handwriting" if analysis_handwriting else "scene"],
                    chat_id=chat_id,
                )
                continue

            if analysis_text:
                suffix = Path(name).suffix.lower().lstrip(".")
                tags = ["attachment", attachment_type, "analysis"]
                if suffix:
                    tags.append(suffix)
                self.add_item(
                    item_type="text",
                    source_type="attachment_analysis",
                    scope="global",
                    title=f"Attachment analysis: {name}",
                    content=(
                        f"Chat: {context_title}\nAttachment type: {mime_type or attachment_type or 'unknown'}\nFilename: {name}\n\n"
                        + (f"Handwriting guess:\n{analysis_handwriting[:2000]}\n\n" if analysis_handwriting else "")
                        + (f"Visual description:\n{analysis_caption[:1200]}\n\n" if analysis_caption else "")
                        + (f"Visual diagnostics:\n{json.dumps(analysis_meta, default=str)[:1500]}\n\n" if isinstance(analysis_meta, dict) and analysis_meta else "")
                        + analysis_text[:MAX_ATTACHMENT_MEMORY_TEXT_CHARS]
                    ),
                    tags=tags,
                    chat_id=chat_id,
                )


class AttachmentAnalysisCache:
    """Persistent cache of attachment analysis so repeated files do not require full reprocessing."""

    def __init__(self, path: Path):
        self.path = path
        self._lock = threading.RLock()
        self._data: Dict[str, Any] = {"entries": []}
        self._load()

    def _load(self):
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if not self.path.exists():
                self._save_locked()
                return
            try:
                raw = json.loads(self.path.read_text(encoding="utf-8"))
                entries = raw.get("entries", []) if isinstance(raw, dict) else []
                self._data = {"entries": [self._normalize_entry(entry) for entry in entries]}
            except Exception as exc:
                logger.warning("Could not read attachment analysis cache %s: %s", self.path, exc)
                self._data = {"entries": []}
                self._save_locked()

    def _save_locked(self):
        tmp_path = self.path.with_suffix(".tmp")
        tmp_path.write_text(json.dumps(self._data, indent=2), encoding="utf-8")
        tmp_path.replace(self.path)

    def _normalize_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        now = _iso_now()
        return {
            "fingerprint": str(entry.get("fingerprint") or "").strip(),
            "payload": copy.deepcopy(entry.get("payload") or {}),
            "created_at": entry.get("created_at") or now,
            "updated_at": entry.get("updated_at") or entry.get("created_at") or now,
        }

    def _find_entry_locked(self, fingerprint: str) -> Optional[Dict[str, Any]]:
        for entry in self._data["entries"]:
            if entry.get("fingerprint") == fingerprint:
                return entry
        return None

    def get(self, fingerprint: str) -> Optional[Dict[str, Any]]:
        if not fingerprint:
            return None
        with self._lock:
            entry = self._find_entry_locked(fingerprint)
            if not entry:
                return None
            return copy.deepcopy(entry.get("payload") or {})

    def put(self, fingerprint: str, payload: Dict[str, Any]) -> None:
        if not fingerprint:
            return
        with self._lock:
            now = _iso_now()
            normalized = self._normalize_entry({
                "fingerprint": fingerprint,
                "payload": payload or {},
                "created_at": now,
                "updated_at": now,
            })
            existing = self._find_entry_locked(fingerprint)
            if existing:
                existing.update(normalized)
            else:
                self._data["entries"].append(normalized)
            if len(self._data["entries"]) > MAX_ATTACHMENT_CACHE_ITEMS:
                ordered = sorted(self._data["entries"], key=lambda item: item.get("updated_at", ""), reverse=True)
                self._data["entries"] = ordered[:MAX_ATTACHMENT_CACHE_ITEMS]
            self._save_locked()

    def stats(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "entries": len(self._data.get("entries") or []),
                "max_entries": MAX_ATTACHMENT_CACHE_ITEMS,
                "path": str(self.path),
            }


chat_store = ChatSessionStore(CHAT_STORE_FILE)
knowledge_store = KnowledgeStore(KNOWLEDGE_STORE_FILE)
attachment_analysis_cache = AttachmentAnalysisCache(ATTACHMENT_CACHE_FILE)
uefn_bridge = UEFNBridge()

# ============================================================================
# AUTO-START UEFN LISTENER
# ============================================================================

_uefn_listener_process = None

def auto_start_uefn_listener():
    """Try to auto-start the UEFN MCP listener if not already running."""
    global _uefn_listener_process

    # Check if already running
    port = discover_uefn_listener_port()
    if port:
        logger.info("UEFN listener already running on port %s — skipping auto-start", port)
        return

    listener_script = WORKSPACE_ROOT / "vendor" / "uefn-mcp-server" / "uefn_listener.py"
    if not listener_script.exists():
        logger.warning("uefn_listener.py not found at %s — cannot auto-start", listener_script)
        return

    try:
        # Try to find Python executable
        python_exe = sys.executable  # Use the same Python running this server
        logger.info("Auto-starting UEFN listener: %s %s", python_exe, listener_script)
        _uefn_listener_process = subprocess.Popen(
            [python_exe, str(listener_script)],
            cwd=str(listener_script.parent),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0),  # Windows: no console window
        )
        logger.info("UEFN listener process started (PID %s)", _uefn_listener_process.pid)

        # Wait briefly and check if it's actually running
        time.sleep(2)
        port = discover_uefn_listener_port()
        if port:
            logger.info("UEFN listener confirmed on port %s", port)
        else:
            logger.warning("UEFN listener process started but not responding yet — it may need UEFN editor open")
    except Exception as e:
        logger.warning("Could not auto-start UEFN listener: %s", e)


# Background health check
def background_health_check():
    """Check UEFN connection periodically."""
    while True:
        time.sleep(5)
        uefn_bridge.check_connection()

health_thread = threading.Thread(target=background_health_check, daemon=True)
health_thread.start()

# Try auto-start in background
threading.Thread(target=auto_start_uefn_listener, daemon=True).start()

# ============================================================================
# API ROUTES
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health():
    """Health check."""
    refresh = (request.args.get("refresh") or "").strip().lower() in {"1", "true", "yes"}
    if refresh:
        uefn_bridge.check_connection()
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "api_version": "2.4",
        "features": {
            "chat_sessions": True,
            "knowledge_store": True,
            "workspace_search": True,
            "project_overview": True,
            "attachment_previews": True,
            "structured_local_vision": True,
            "hosted_llm_retries": True,
            "local_model_prewarm": True,
            "fast_startup_ping": True,
        },
        "tools": len(tool_registry.tools),
        "uefn_connected": uefn_bridge.is_connected,
        "uefn_listener_port": uefn_bridge.discovered_port,
        "backend": {
            "host": BACKEND_HOST,
            "port": BACKEND_PORT,
        },
    })


@app.route('/api/ping', methods=['GET'])
def ping():
    """Fast readiness probe for Electron startup."""
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "api_version": "2.4",
        "features": {
            "chat_sessions": True,
            "attachment_previews": True,
            "structured_local_vision": True,
            "hosted_llm_retries": True,
            "local_model_prewarm": True,
            "fast_startup_ping": True,
        },
        "backend": {
            "host": BACKEND_HOST,
            "port": BACKEND_PORT,
        },
    })


@app.route('/api/uefn/mcp/status', methods=['GET'])
def uefn_mcp_status():
    """Detailed status for the in-editor MCP HTTP listener (Codex app + Claude MCP bridge use this)."""
    uefn_bridge.check_connection()
    return jsonify({
        "connected": uefn_bridge.is_connected,
        "port": uefn_bridge.discovered_port,
        "health": uefn_bridge.last_info or None,
        "hint": "Run uefn_listener.py in UEFN or use sync-mcp-listener so the listener auto-starts.",
    })


@app.route('/api/uefn/mcp/command', methods=['POST'])
def uefn_mcp_command():
    """Proxy a command to the UEFN listener (same JSON as vendor/uefn-mcp-server HTTP API)."""
    data = request.get_json() or {}
    command = data.get("command")
    if not command:
        return jsonify({"success": False, "error": "Missing 'command'"}), 400
    params = data.get("params") or {}
    port = discover_uefn_listener_port()
    if not port:
        uefn_bridge.check_connection()
        port = uefn_bridge.discovered_port
    if not port:
        return jsonify({
            "success": False,
            "error": "UEFN MCP listener not running. In UEFN: py path/to/uefn_listener.py (or sync-mcp-listener).",
        }), 503
    try:
        return jsonify(mcp_listener_post_command(port, str(command), params))
    except Exception as e:
        logger.error("UEFN MCP command failed: %s", e)
        return jsonify({"success": False, "error": str(e)}), 502


# ============================================================================
# UEFN ACTOR / LEVEL / VIEWPORT ENDPOINTS
# ============================================================================


@app.route('/api/uefn/actors', methods=['GET'])
def uefn_get_actors():
    """Get all actors in the current UEFN level."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        class_filter = request.args.get('class', '')
        params = {"class_filter": class_filter} if class_filter else {}
        result = mcp_listener_post_command(port, "get_all_actors", params)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/selected', methods=['GET'])
def uefn_get_selected_actors():
    """Get currently selected actors in the UEFN viewport."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "get_selected_actors", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/spawn', methods=['POST'])
def uefn_spawn_actor():
    """Spawn an actor in the UEFN level."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        result = mcp_listener_post_command(port, "spawn_actor", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/delete', methods=['POST'])
def uefn_delete_actors():
    """Delete actors from the UEFN level."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        result = mcp_listener_post_command(port, "delete_actors", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/transform', methods=['POST'])
def uefn_transform_actor():
    """Set transform (location/rotation/scale) on an actor."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        result = mcp_listener_post_command(port, "set_actor_transform", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/select', methods=['POST'])
def uefn_select_actors():
    """Select actors in the UEFN viewport."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        result = mcp_listener_post_command(port, "select_actors", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/focus', methods=['POST'])
def uefn_focus_selected():
    """Focus the viewport camera on selected actors."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "focus_selected", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/actors/properties', methods=['POST'])
def uefn_actor_properties():
    """Get or set properties on an actor."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        action = data.pop("action", "get")
        if action == "set":
            result = mcp_listener_post_command(port, "set_actor_properties", data)
        else:
            result = mcp_listener_post_command(port, "get_actor_properties", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/level', methods=['GET'])
def uefn_level_info():
    """Get current level info (world name, actor count)."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "get_level_info", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/level/save', methods=['POST'])
def uefn_save_level():
    """Save the current level."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "save_current_level", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/project', methods=['GET'])
def uefn_project_info():
    """Get UEFN project info."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "get_project_info", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/viewport', methods=['GET'])
def uefn_get_viewport():
    """Get viewport camera position."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        result = mcp_listener_post_command(port, "get_viewport_camera", {})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/viewport', methods=['POST'])
def uefn_set_viewport():
    """Set viewport camera position and rotation."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        result = mcp_listener_post_command(port, "set_viewport_camera", data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/assets', methods=['GET'])
def uefn_list_assets():
    """List assets in the UEFN project content browser."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        directory = request.args.get('directory', '/Game/')
        class_filter = request.args.get('class', '')
        params = {"directory": directory, "recursive": True}
        if class_filter:
            params["class_filter"] = class_filter
        result = mcp_listener_post_command(port, "list_assets", params)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/execute-python', methods=['POST'])
def uefn_execute_python():
    """Execute arbitrary Python code in the UEFN editor."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        data = request.get_json() or {}
        code = data.get("code", "")
        if not code:
            return jsonify({"success": False, "error": "Missing 'code'"}), 400
        result = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=45.0)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/uefn/editor-log', methods=['GET'])
def uefn_editor_log():
    """Get recent lines from the UEFN editor output log."""
    port = discover_uefn_listener_port()
    if not port:
        return jsonify({"success": False, "error": "UEFN not connected"}), 503
    try:
        last_n = int(request.args.get('last_n', 100))
        filter_str = request.args.get('filter', '')
        result = mcp_listener_post_command(port, "get_editor_log", {"last_n": last_n, "filter_str": filter_str})
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 502


@app.route('/api/tools', methods=['GET'])
def get_tools():
    """Get all tools with descriptions."""
    return jsonify({
        "tools": tool_registry.get_all_tools(),
        "count": len(tool_registry.tools),
        "categories": list(tool_registry.categories.keys())
    })


@app.route('/api/tools/categories', methods=['GET'])
def get_categories():
    """Get tool categories."""
    return jsonify({
        "categories": list(tool_registry.categories.keys()),
        "count": len(tool_registry.categories)
    })


@app.route('/api/tools/category/<category>', methods=['GET'])
def get_tools_by_category(category):
    """Get tools by category."""
    tools = tool_registry.get_tools_by_category(category)
    return jsonify({
        "category": category,
        "tools": tools,
        "count": len(tools)
    })


@app.route('/api/tools/search', methods=['GET'])
def search_tools():
    """Search tools."""
    query = request.args.get('q', '')
    results = tool_registry.search_tools(query)
    return jsonify({
        "query": query,
        "results": results,
        "count": len(results)
    })


@app.route('/api/tools/<tool_name>/execute', methods=['POST'])
def execute_tool(tool_name):
    """Execute tool."""
    try:
        parameters = request.get_json() or {}
        result = tool_registry.execute_tool(tool_name, parameters)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Execution error: {e}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/assets/shortlist', methods=['GET'])
def get_assets():
    """Get asset shortlist with 3D metadata."""
    try:
        shortlist_file = DATA_DIR / "catalog" / "shortlist.json"
        if shortlist_file.exists():
            with open(shortlist_file) as f:
                assets = json.load(f)
        else:
            assets = [
                {
                    "id": "asset_001",
                    "name": "Modern Chair",
                    "type": "furniture",
                    "category": "props",
                    "description": "Contemporary sitting furniture piece",
                    "tags": ["seating", "indoor", "modern"],
                    "dimensions": {"width": 0.8, "depth": 0.8, "height": 0.85},
                    "mesh_count": 3,
                    "material_count": 2,
                    "lod_levels": 2,
                    "polycount": 12500,
                    "trust_score": 0.95,
                    "composite_asset": True,
                    "viewer_model_url": "https://modelviewer.dev/shared-assets/models/Astronaut.glb",
                    "viewer_note": "Demo GLB; replace with your exported composite when available."
                },
                {
                    "id": "asset_002",
                    "name": "Office Desk",
                    "type": "furniture",
                    "category": "props",
                    "description": "Professional work surface with storage",
                    "tags": ["furniture", "workspace", "office"],
                    "dimensions": {"width": 1.6, "depth": 0.8, "height": 0.75},
                    "mesh_count": 5,
                    "material_count": 3,
                    "lod_levels": 2,
                    "polycount": 28000,
                    "trust_score": 0.92
                },
                {
                    "id": "asset_003",
                    "name": "Building Module",
                    "type": "building",
                    "category": "structures",
                    "description": "Pre-fab building block for construction",
                    "tags": ["structure", "building", "modular"],
                    "dimensions": {"width": 4.0, "depth": 4.0, "height": 3.5},
                    "mesh_count": 12,
                    "material_count": 6,
                    "lod_levels": 3,
                    "polycount": 125000,
                    "trust_score": 0.88
                },
            ]
        
        return jsonify({
            "assets": assets,
            "count": len(assets)
        })
    except Exception as e:
        logger.error(f"Asset error: {e}")
        return jsonify({"error": str(e), "assets": []}), 500


@app.route('/api/assets/<asset_id>/3d-view', methods=['GET'])
def get_asset_3d_view(asset_id):
    """Get 3D view data for asset."""
    return jsonify({
        "asset_id": asset_id,
        "3d_data": {
            "model_url": f"/models/{asset_id}.glb",
            "thumbnail": f"/thumbnails/{asset_id}.jpg",
            "description": "3D model available",
            "cameras": [
                {"position": [0, 2, 5], "target": [0, 0, 0], "label": "Front"},
                {"position": [5, 2, 0], "target": [0, 0, 0], "label": "Side"},
                {"position": [2, 5, 2], "target": [0, 0, 0], "label": "Top"}
            ]
        }
    })


@app.route('/api/codex/tools', methods=['GET'])
def get_codex_tools():
    """Get tools formatted for Codex AI to understand."""
    return jsonify({
        "available_tools": tool_registry.get_tools_for_ai(),
        "tool_count": len(tool_registry.tools),
        "categories": list(tool_registry.categories.keys()),
        "timestamp": datetime.now().isoformat()
    })


@app.route('/api/knowledge', methods=['GET'])
def get_knowledge_items():
    """List shared knowledge items used by planning, research, and chat memory."""
    include_excluded = request.args.get("include_excluded", "true").lower() != "false"
    items = knowledge_store.list_items(include_excluded=include_excluded)
    return jsonify({
        "items": items,
        "count": len(items),
    })


@app.route('/api/knowledge', methods=['POST'])
def create_knowledge_item():
    """Create a shared knowledge item."""
    data = request.get_json() or {}
    title = (data.get("title") or "").strip()
    content = data.get("content") or ""
    if not title or not content:
        return jsonify({"error": "Both 'title' and 'content' are required"}), 400

    item = knowledge_store.add_item(
        item_type=str(data.get("type") or "text"),
        source_type=str(data.get("source_type") or data.get("type") or "manual"),
        scope=str(data.get("scope") or "global"),
        title=title,
        content=str(content),
        tags=[str(tag) for tag in (data.get("tags") or [])],
        quality=int(data.get("quality", 1) or 0),
        source_url=str(data.get("sourceUrl") or ""),
        chat_id=str(data.get("chat_id") or ""),
    )
    return jsonify({"success": True, "item": item}), 201


@app.route('/api/knowledge/<item_id>', methods=['PATCH'])
def update_knowledge_item(item_id):
    """Update a shared knowledge item."""
    data = request.get_json() or {}
    item = knowledge_store.update_item(
        item_id,
        title=data.get("title"),
        content=data.get("content"),
        quality=data.get("quality"),
        tags=data.get("tags"),
        sourceUrl=data.get("sourceUrl"),
    )
    if not item:
        return jsonify({"error": "Knowledge item not found"}), 404
    return jsonify({"success": True, "item": item})


@app.route('/api/knowledge/<item_id>', methods=['DELETE'])
def delete_knowledge_item(item_id):
    """Delete a shared knowledge item."""
    deleted = knowledge_store.delete_item(item_id)
    if not deleted:
        return jsonify({"error": "Knowledge item not found"}), 404
    return jsonify({"success": True})


@app.route('/api/knowledge/search', methods=['POST'])
def search_knowledge_items():
    """Search shared knowledge plus related chats."""
    data = request.get_json() or {}
    query = (data.get("query") or "").strip()
    if not query:
        return jsonify({"error": "Missing query", "results": []}), 400

    results = _search_shared_context(query, current_chat_id=str(data.get("chat_id") or ""))
    return jsonify({"success": True, "results": results})


@app.route('/api/project/overview', methods=['GET'])
def get_project_overview_endpoint():
    """Return non-tool project/app capabilities plus live UEFN state."""
    query = request.args.get("q", "")
    include_assets = request.args.get("include_assets", "false").lower() == "true"
    include_workspace_hits = request.args.get("include_workspace_hits", "false").lower() == "true"
    return jsonify({
        "success": True,
        "overview": _get_project_overview(
            query=query,
            include_assets=include_assets,
            include_workspace_hits=include_workspace_hits,
        ),
    })


@app.route('/api/workspace/search', methods=['POST'])
def search_workspace_endpoint():
    """Search the local workspace for relevant implementation details."""
    data = request.get_json() or {}
    query = (data.get("query") or "").strip()
    if not query:
        return jsonify({"error": "Missing query", "results": []}), 400

    results = _search_workspace_sources(
        query,
        limit=int(data.get("limit") or 8),
        include_vendor=bool(data.get("include_vendor", True)),
    )
    return jsonify({"success": True, "results": results})


@app.route('/api/config', methods=['GET'])
def get_config():
    """Get configuration."""
    return jsonify({
        "backend": {
            "host": BACKEND_HOST,
            "port": BACKEND_PORT,
        },
        "uefn": uefn_bridge.get_status(),
        "tools": {
            "total": len(tool_registry.tools),
            "categories": len(tool_registry.categories)
        }
    })


@app.route('/api/config/reload', methods=['POST'])
def reload_config():
    """Reload configuration."""
    tool_registry.scan_tools()
    uefn_bridge.check_connection()
    
    return jsonify({
        "status": "reloaded",
        "tools": len(tool_registry.tools),
        "timestamp": datetime.now().isoformat()
    })


@app.route('/api/codex/plan', methods=['POST'])
def create_plan():
    """Create Codex plan with tool awareness — AI-powered when available."""
    try:
        data = request.get_json() or {}

        tools_context = tool_registry.get_tools_for_ai()
        desc = (data.get("description") or "").strip()
        goals = data.get("goals") or ""
        constraints = data.get("constraints") or ""
        if isinstance(goals, list):
            goals = "\n".join(str(g) for g in goals)
        if isinstance(constraints, list):
            constraints = "\n".join(str(c) for c in constraints)
        kb_parts = []
        incoming_kb = (data.get("knowledge_context") or "").strip()
        if incoming_kb:
            kb_parts.append(incoming_kb)
        shared_kb = [
            f"[{item.get('source_type', item.get('type', 'note'))}] {item.get('title', 'Untitled')}: {item.get('content', '')[:600]}"
            for item in knowledge_store.list_items(include_excluded=False)[:16]
        ]
        if shared_kb:
            kb_parts.append("\n".join(shared_kb))
        kb = "\n---\n".join(part for part in kb_parts if part)

        # Get UEFN snapshot
        uefn_snapshot = None
        uefn_bridge.check_connection()
        if uefn_bridge.discovered_port:
            try:
                snap = mcp_listener_post_command(
                    uefn_bridge.discovered_port, "get_level_info", {}, timeout=8.0
                )
                if snap.get("success"):
                    uefn_snapshot = snap.get("result")
            except Exception as ex:
                logger.warning("UEFN get_level_info for plan: %s", ex)

        # Try AI-powered plan generation
        client = _get_llm_client()
        if client and desc:
            try:
                model = _get_active_model()
                provider = _get_active_provider()
                system_prompt = f"""You are a UEFN (Unreal Editor for Fortnite) project planner. Create actionable build plans.

AVAILABLE TOOLS (can be used in steps):
{tools_context[:3000]}

{"CURRENT LEVEL STATE:" + chr(10) + json.dumps(uefn_snapshot, default=str)[:1500] if uefn_snapshot else "UEFN editor is connected but no level data fetched."}

{"KNOWLEDGE BASE CONTEXT:" + chr(10) + kb[:2000] if kb else ""}

RESPOND WITH VALID JSON ONLY — an array of step objects:
[
  {{"id": 1, "title": "Step title", "description": "Detailed description of what to do", "status": "pending", "tools": ["tool_name1", "tool_name2"]}}
]
Generate 4-8 specific, actionable steps for the user's request. Use real tool names from the list above. Be specific to UEFN/Fortnite Creative."""

                response = _chat_completion_with_retry(
                    client,
                    provider,
                    model=model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": f"Create a build plan for: {desc}" + (f"\nGoals: {goals}" if goals else "") + (f"\nConstraints: {constraints}" if constraints else "")},
                    ],
                    max_tokens=2000,
                    temperature=0.7,
                )

                tok_in = getattr(response.usage, 'prompt_tokens', 0) if response.usage else 0
                tok_out = getattr(response.usage, 'completion_tokens', 0) if response.usage else 0
                _track_usage(provider, tok_in, tok_out)

                reply = (response.choices[0].message.content or "").strip()
                # Parse JSON from response (handle markdown code blocks)
                if "```" in reply:
                    reply = reply.split("```")[1]
                    if reply.startswith("json"):
                        reply = reply[4:]
                    reply = reply.strip()

                steps = json.loads(reply)
                if isinstance(steps, list) and len(steps) > 0:
                    plan = {
                        "description": desc,
                        "goals": goals,
                        "constraints": constraints,
                        "steps": steps,
                        "uefn_editor_snapshot": uefn_snapshot,
                        "ai_generated": True,
                        "timestamp": datetime.now().isoformat(),
                    }
                    return jsonify({"success": True, "plan": plan})
            except Exception as ai_err:
                logger.warning(f"AI plan generation failed, falling back: {ai_err}")

        # Fallback: generate plan from description keywords
        steps = _generate_fallback_plan(desc, tools_context, uefn_snapshot)
        plan = {
            "description": desc,
            "goals": goals,
            "constraints": constraints,
            "steps": steps,
            "uefn_editor_snapshot": uefn_snapshot,
            "ai_generated": False,
            "timestamp": datetime.now().isoformat(),
        }
        return jsonify({"success": True, "plan": plan})
    except Exception as e:
        logger.error(f"Plan error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


def _generate_fallback_plan(desc: str, tools_context: str, snapshot: Any) -> list:
    """Generate a plan from keywords when AI is not available."""
    desc_lower = (desc or "").lower()
    steps = []
    step_id = 1

    # Always start with scope
    steps.append({"id": step_id, "title": "Analyze current level state", "description": f"Review what's already in the level and identify what needs to change for: {desc}", "status": "pending", "tools": ["scene_analyzer", "get_all_actors"]})
    step_id += 1

    if any(kw in desc_lower for kw in ['parkour', 'obstacle', 'course', 'race', 'jump']):
        steps.append({"id": step_id, "title": "Design the course layout", "description": "Plan the path, checkpoint positions, difficulty progression, and finish line placement.", "status": "pending", "tools": ["bulk_align", "text_placer"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Place platforms and obstacles", "description": "Use props and static meshes to build jump pads, moving platforms, walls, and gaps.", "status": "pending", "tools": ["asset_importer", "bulk_align"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Add devices and mechanics", "description": "Wire up Checkpoints, Timer, Speed Boost pads, and Kill Zones using Fortnite devices.", "status": "pending", "tools": ["verse_validator"]})
        step_id += 1
    elif any(kw in desc_lower for kw in ['battle', 'arena', 'pvp', 'fight', 'combat', 'deathmatch']):
        steps.append({"id": step_id, "title": "Design arena layout", "description": "Plan spawn points, cover positions, sight lines, and loot placement.", "status": "pending", "tools": ["bulk_align"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Place combat elements", "description": "Add weapon spawners, ammo crates, health items, and defensive cover.", "status": "pending", "tools": ["asset_importer"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Configure game rules", "description": "Set up team spawns, elimination scoring, round timers, and victory conditions.", "status": "pending", "tools": ["verse_validator"]})
        step_id += 1
    elif any(kw in desc_lower for kw in ['horror', 'scary', 'escape', 'puzzle']):
        steps.append({"id": step_id, "title": "Design atmosphere and layout", "description": "Plan room connections, dark corridors, jump scare positions, and escape route.", "status": "pending", "tools": ["material_master", "capture_setup"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Place environmental elements", "description": "Add lighting, fog, sound emitters, and destructible props for tension.", "status": "pending", "tools": ["asset_importer"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Wire puzzle mechanics", "description": "Set up trigger zones, locked doors, key items, and sequence puzzles.", "status": "pending", "tools": ["verse_validator"]})
        step_id += 1
    else:
        steps.append({"id": step_id, "title": "Select and place core assets", "description": f"Choose the right props, materials, and building pieces for: {desc}", "status": "pending", "tools": ["asset_importer", "asset_analyzer"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Arrange layout and flow", "description": "Position elements for good player flow, visual composition, and performance.", "status": "pending", "tools": ["bulk_align", "text_placer"]})
        step_id += 1
        steps.append({"id": step_id, "title": "Add gameplay devices", "description": "Wire up interactive elements, triggers, scoring, and win conditions.", "status": "pending", "tools": ["verse_validator"]})
        step_id += 1

    # Always end with polish and validate
    steps.append({"id": step_id, "title": "Polish visuals and materials", "description": "Refine materials, lighting, effects, and overall visual quality.", "status": "pending", "tools": ["material_master", "capture_setup"]})
    step_id += 1
    steps.append({"id": step_id, "title": "Test and validate", "description": "Playtest the experience, fix blockers, check performance, and verify all mechanics work.", "status": "pending", "tools": ["verse_validator", "memory_scan"]})

    return steps


def discover_web_candidates(query: str) -> Dict[str, Any]:
    """Lightweight web discovery via DuckDuckGo Instant Answer API (no API key)."""
    q = urllib.parse.quote(query)
    url = f"https://api.duckduckgo.com/?q={q}&format=json&no_html=1&skip_disambig=1"
    req = urllib.request.Request(url, headers={"User-Agent": "UEFN-Codex-Agent/1.0"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        data = json.loads(resp.read().decode())
    candidates: List[Dict[str, Any]] = []
    if data.get("AbstractURL"):
        candidates.append({
            "title": data.get("Heading") or "Instant answer",
            "url": data["AbstractURL"],
            "snippet": (data.get("AbstractText") or "")[:2000],
            "source": "duckduckgo",
        })
    for topic in data.get("RelatedTopics", []) or []:
        if isinstance(topic, dict):
            if topic.get("FirstURL") and topic.get("Text"):
                candidates.append({
                    "title": topic["Text"][:200],
                    "url": topic["FirstURL"],
                    "snippet": "",
                    "source": "duckduckgo",
                })
            for sub in topic.get("Topics", []) or []:
                if isinstance(sub, dict) and sub.get("FirstURL") and sub.get("Text"):
                    candidates.append({
                        "title": sub["Text"][:200],
                        "url": sub["FirstURL"],
                        "snippet": "",
                        "source": "duckduckgo",
                    })
    seen: set[str] = set()
    uniq: List[Dict[str, Any]] = []
    for c in candidates:
        u = c.get("url")
        if not u or u in seen:
            continue
        seen.add(str(u))
        uniq.append(c)
    return {
        "abstract": data.get("AbstractText", ""),
        "heading": data.get("Heading", ""),
        "candidates": uniq[:20],
    }


@app.route('/api/research/discover', methods=['POST'])
def research_discover():
    """Return candidate URLs/snippets for a query (research mode — user keeps good ones into KB)."""
    data = request.get_json() or {}
    q = (data.get("query") or "").strip()
    if not q:
        return jsonify({"success": False, "error": "Missing query", "candidates": []}), 400
    try:
        payload = discover_web_candidates(q)
        payload["success"] = True
        return jsonify(payload)
    except Exception as e:
        logger.error("Research discover failed: %s", e)
        return jsonify({"success": False, "error": str(e), "candidates": []}), 502


@app.route('/api/research', methods=['POST'])
def research():
    """Research endpoint — AI-powered analysis of query against knowledge base."""
    try:
        data = request.get_json() or {}
        query = (data.get("query") or "").strip()
        incoming_knowledge = data.get("knowledge_base") or []
        knowledge_base = list(incoming_knowledge)
        knowledge_base.extend(knowledge_store.list_items(include_excluded=False))
        context = (data.get("context") or "UEFN island design and development").strip()

        if not query:
            return jsonify({"success": False, "error": "Missing query", "findings": [], "recommendations": []}), 400

        # Gather KB snippets
        snippets: List[str] = []
        for item in knowledge_base:
            if item.get("quality") == 0:
                continue
            title = (item.get("title") or "Untitled")[:200]
            content = (item.get("content") or "")[:4000]
            itype = item.get("type") or "note"
            snippets.append(f"[{itype}] {title}: {content[:800]}")

        # Try AI-powered research
        client = _get_llm_client()
        if client and (snippets or query):
            try:
                model = _get_active_model()
                provider = _get_active_provider()
                kb_text = "\n\n".join(snippets[:15]) if snippets else "(No knowledge base items provided)"
                tools_text = tool_registry.get_tools_for_ai()[:2000]

                system_prompt = f"""You are a UEFN research assistant. Analyze the user's question using the provided knowledge base and your expertise in Unreal Editor for Fortnite, Verse programming, and Fortnite Creative.

KNOWLEDGE BASE ({len(snippets)} items):
{kb_text[:4000]}

AVAILABLE TOOLS:
{tools_text}

CONTEXT: {context}

Respond with valid JSON:
{{
  "findings": ["Finding 1", "Finding 2", ...],
  "recommendations": ["Recommendation 1", "Recommendation 2", ...]
}}

Findings should be specific, actionable insights. Recommendations should be next steps. Reference knowledge base items when relevant."""

                response = _chat_completion_with_retry(
                    client,
                    provider,
                    model=model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": query},
                    ],
                    max_tokens=1500,
                    temperature=0.7,
                )

                tok_in = getattr(response.usage, 'prompt_tokens', 0) if response.usage else 0
                tok_out = getattr(response.usage, 'completion_tokens', 0) if response.usage else 0
                _track_usage(provider, tok_in, tok_out)

                reply = (response.choices[0].message.content or "").strip()
                if "```" in reply:
                    reply = reply.split("```")[1]
                    if reply.startswith("json"):
                        reply = reply[4:]
                    reply = reply.strip()

                parsed = json.loads(reply)
                return jsonify({
                    "success": True,
                    "query": query,
                    "findings": parsed.get("findings", []),
                    "recommendations": parsed.get("recommendations", []),
                    "sources_used": len(snippets),
                    "ai_powered": True,
                    "timestamp": datetime.now().isoformat(),
                })
            except Exception as ai_err:
                logger.warning(f"AI research failed, falling back: {ai_err}")

        # Fallback: keyword-based research
        findings: List[str] = []
        recommendations: List[str] = []

        if snippets:
            findings.append(f"Scanned {len(snippets)} knowledge base item(s) for: \"{query}\"")
            joined = "\n\n".join(snippets)
            if query.lower() in joined.lower():
                findings.append("Keywords from your question were found in the knowledge base.")
            else:
                findings.append("No direct keyword match found. Try adding more specific docs or rephrasing.")
            recommendations.append("Add Epic documentation URLs or paste Verse code snippets for better results.")
            recommendations.append("Set up a free AI provider (Groq/Ollama) in Settings for AI-powered analysis.")
        else:
            findings.append("No knowledge base items provided. Add files, URLs, or notes under Research tab, then try again.")
            recommendations.append("Upload reference docs or design notes to build your knowledge base.")
            recommendations.append("Use 'Discover on Web' to find relevant UEFN documentation.")

        return jsonify({
            "success": True,
            "query": query,
            "findings": findings,
            "recommendations": recommendations,
            "sources_used": len(snippets),
            "ai_powered": False,
            "timestamp": datetime.now().isoformat(),
        })
    except Exception as e:
        logger.error(f"Research error: {e}")
        return jsonify({"success": False, "error": str(e), "findings": [], "recommendations": []}), 500


# ============================================================================
# CHAT API  — AI-powered assistant + UEFN integration
# ============================================================================

# ── Usage tracking ────────────────────────────────────────────────────────
# Tracks requests and tokens per provider for the current session + persisted daily
_usage_data = {
    "groq":     {"requests": 0, "tokens_in": 0, "tokens_out": 0, "limit_requests": 14400, "limit_tokens": 0,    "period": "day"},
    "cerebras": {"requests": 0, "tokens_in": 0, "tokens_out": 0, "limit_requests": 0,     "limit_tokens": 1000000, "period": "day"},
    "gemini":   {"requests": 0, "tokens_in": 0, "tokens_out": 0, "limit_requests": 1500,  "limit_tokens": 0,    "period": "day"},
    "ollama":   {"requests": 0, "tokens_in": 0, "tokens_out": 0, "limit_requests": 0,     "limit_tokens": 0,    "period": "unlimited"},
    "openai":   {"requests": 0, "tokens_in": 0, "tokens_out": 0, "limit_requests": 0,     "limit_tokens": 0,    "period": "pay-per-use"},
}
_usage_date = datetime.now().strftime("%Y-%m-%d")

def _load_usage():
    """Load persisted usage from disk."""
    global _usage_data, _usage_date
    usage_file = WORKSPACE_ROOT / "data" / "ai_usage.json"
    try:
        if usage_file.exists():
            raw = json.loads(usage_file.read_text(encoding="utf-8"))
            saved_date = raw.get("date", "")
            today = datetime.now().strftime("%Y-%m-%d")
            if saved_date == today:
                for prov in _usage_data:
                    if prov in raw.get("providers", {}):
                        saved = raw["providers"][prov]
                        _usage_data[prov]["requests"] = saved.get("requests", 0)
                        _usage_data[prov]["tokens_in"] = saved.get("tokens_in", 0)
                        _usage_data[prov]["tokens_out"] = saved.get("tokens_out", 0)
                _usage_date = today
            else:
                # New day — reset counters
                _usage_date = today
                for prov in _usage_data:
                    _usage_data[prov]["requests"] = 0
                    _usage_data[prov]["tokens_in"] = 0
                    _usage_data[prov]["tokens_out"] = 0
    except Exception:
        pass

def _save_usage():
    """Persist usage to disk."""
    usage_file = WORKSPACE_ROOT / "data" / "ai_usage.json"
    try:
        usage_file.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "date": _usage_date,
            "providers": {
                prov: {"requests": d["requests"], "tokens_in": d["tokens_in"], "tokens_out": d["tokens_out"]}
                for prov, d in _usage_data.items()
            }
        }
        usage_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except Exception:
        pass

def _track_usage(provider: str, tokens_in: int = 0, tokens_out: int = 0):
    """Record a request and token usage for a provider."""
    global _usage_date
    today = datetime.now().strftime("%Y-%m-%d")
    if today != _usage_date:
        # Reset on new day
        _usage_date = today
        for prov in _usage_data:
            _usage_data[prov]["requests"] = 0
            _usage_data[prov]["tokens_in"] = 0
            _usage_data[prov]["tokens_out"] = 0
    if provider in _usage_data:
        _usage_data[provider]["requests"] += 1
        _usage_data[provider]["tokens_in"] += tokens_in
        _usage_data[provider]["tokens_out"] += tokens_out
    _save_usage()

# Load saved usage on startup
_load_usage()

# LLM client (lazy-init) — auto-detects best free provider
_llm_client = None
_llm_provider = None  # tracks which provider the client is connected to
_ocr_engine = None
_local_structured_vlm_model = None
_local_structured_vlm_processor = None
_local_vlm_pipeline = None
_local_htr_pipeline = None
_local_htr_processor = None
_local_model_prewarm_thread = None
_local_model_prewarm_lock = threading.Lock()
_local_model_prewarm_status = {
    "enabled": LOCAL_MODEL_PREWARM_ENABLED,
    "running": False,
    "started_at": "",
    "completed_at": "",
    "components": {
        "ocr": {"status": "pending", "ready": False, "detail": "RapidOCR warmup pending."},
        "structured_vlm": {"status": "pending", "ready": False, "model": LOCAL_STRUCTURED_VLM_MODEL_ID, "detail": "Florence warmup pending."},
        "vlm": {"status": "pending", "ready": False, "model": LOCAL_VLM_MODEL_ID, "detail": "Caption fallback warmup pending."},
        "handwriting": {"status": "pending", "ready": False, "model": LOCAL_HTR_MODEL_ID, "detail": "Handwriting warmup pending."},
    },
}


def _copy_local_model_prewarm_status() -> Dict[str, Any]:
    with _local_model_prewarm_lock:
        return copy.deepcopy(_local_model_prewarm_status)


def _update_local_model_prewarm_component(name: str, *, status: str, ready: bool, detail: str = "", model: str = "") -> None:
    with _local_model_prewarm_lock:
        component = _local_model_prewarm_status.setdefault("components", {}).setdefault(name, {})
        component.update({
            "status": status,
            "ready": ready,
            "detail": detail,
        })
        if model:
            component["model"] = model


def _mark_local_model_prewarm_state(*, running: bool, started_at: str = "", completed_at: str = "") -> None:
    with _local_model_prewarm_lock:
        _local_model_prewarm_status["enabled"] = LOCAL_MODEL_PREWARM_ENABLED
        _local_model_prewarm_status["running"] = running
        if started_at:
            _local_model_prewarm_status["started_at"] = started_at
        if completed_at:
            _local_model_prewarm_status["completed_at"] = completed_at


def _get_exception_status_code(exc: Exception) -> Optional[int]:
    response = getattr(exc, "response", None)
    status_code = getattr(exc, "status_code", None)
    if isinstance(status_code, int):
        return status_code
    if isinstance(getattr(exc, "code", None), int):
        return int(getattr(exc, "code"))
    if response is not None and isinstance(getattr(response, "status_code", None), int):
        return int(response.status_code)
    return None


def _get_exception_headers(exc: Exception) -> Dict[str, str]:
    headers_obj = getattr(exc, "headers", None)
    if headers_obj is None:
        response = getattr(exc, "response", None)
        headers_obj = getattr(response, "headers", None) if response is not None else None
    if headers_obj is None:
        return {}
    try:
        return {str(k): str(v) for k, v in headers_obj.items()}
    except Exception:
        return {}


def _parse_retry_after_seconds(value: str) -> Optional[float]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return max(0.0, float(raw))
    except Exception:
        pass
    try:
        target = parsedate_to_datetime(raw)
        if target.tzinfo is None:
            target = target.replace(tzinfo=UTC)
        delta = (target - datetime.now(UTC)).total_seconds()
        return max(0.0, delta)
    except Exception:
        return None


def _is_retryable_llm_error(exc: Exception) -> bool:
    status_code = _get_exception_status_code(exc)
    if status_code in {408, 409, 425, 429, 500, 502, 503, 504}:
        return True

    if isinstance(exc, (TimeoutError, socket.timeout, ConnectionError, urllib.error.URLError)):
        return True

    name = exc.__class__.__name__.lower()
    if any(token in name for token in ("ratelimit", "timeout", "connection", "overloaded", "internalserver")):
        return True

    message = str(exc).lower()
    return any(token in message for token in (
        "429",
        "rate limit",
        "rate_limit",
        "quota",
        "resource_exhausted",
        "temporarily unavailable",
        "try again",
        "timed out",
        "timeout",
        "connection reset",
        "connection aborted",
        "connection refused",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
    ))


def _compute_retry_delay_seconds(exc: Exception, attempt: int) -> float:
    headers = _get_exception_headers(exc)
    retry_after = _parse_retry_after_seconds(headers.get("retry-after") or headers.get("Retry-After"))
    if retry_after is not None:
        return min(HOSTED_LLM_RETRY_MAX_SECONDS, max(0.5, retry_after))

    base = min(HOSTED_LLM_RETRY_MAX_SECONDS, HOSTED_LLM_RETRY_BASE_SECONDS * (2 ** attempt))
    jitter = min(0.5, base * 0.25) * random.random()
    return min(HOSTED_LLM_RETRY_MAX_SECONDS, base + jitter)


def _chat_completion_with_retry(client: Any, provider: str, **kwargs):
    attempts = HOSTED_LLM_MAX_RETRIES + 1
    last_error: Optional[Exception] = None

    for attempt in range(attempts):
        try:
            return client.chat.completions.create(**kwargs)
        except Exception as exc:
            last_error = exc
            if attempt >= attempts - 1 or not _is_retryable_llm_error(exc):
                raise
            delay = _compute_retry_delay_seconds(exc, attempt)
            logger.warning(
                "Transient %s chat failure on attempt %s/%s. Retrying in %.2fs: %s",
                provider,
                attempt + 1,
                attempts,
                delay,
                exc,
            )
            time.sleep(delay)

    if last_error is not None:
        raise last_error
    raise RuntimeError("LLM request failed before a response could be returned.")


def _prewarm_local_models_worker():
    if not LOCAL_MODEL_PREWARM_ENABLED:
        _update_local_model_prewarm_component("ocr", status="disabled", ready=False, detail="Local model prewarm is disabled.")
        _update_local_model_prewarm_component("structured_vlm", status="disabled", ready=False, detail="Local model prewarm is disabled.", model=LOCAL_STRUCTURED_VLM_MODEL_ID)
        _update_local_model_prewarm_component("vlm", status="disabled", ready=False, detail="Local model prewarm is disabled.", model=LOCAL_VLM_MODEL_ID)
        _update_local_model_prewarm_component("handwriting", status="disabled", ready=False, detail="Local model prewarm is disabled.", model=LOCAL_HTR_MODEL_ID)
        return

    _mark_local_model_prewarm_state(running=True, started_at=_iso_now())
    if LOCAL_MODEL_PREWARM_DELAY_SECONDS > 0:
        time.sleep(LOCAL_MODEL_PREWARM_DELAY_SECONDS)

    components = [
        (
            "ocr",
            lambda: _get_ocr_engine(),
            lambda result: result is not None,
            "RapidOCR ready.",
            "RapidOCR unavailable or failed to initialize.",
            "",
        ),
        (
            "structured_vlm",
            lambda: _get_local_structured_vlm(),
            lambda result: bool(result and result[0] and result[1]),
            "Structured local vision ready.",
            "Structured local vision unavailable or failed to initialize.",
            LOCAL_STRUCTURED_VLM_MODEL_ID,
        ),
        (
            "vlm",
            lambda: _get_local_vlm_pipeline(),
            lambda result: result is not None,
            "Caption fallback model ready.",
            "Caption fallback model unavailable or failed to initialize.",
            LOCAL_VLM_MODEL_ID,
        ),
        (
            "handwriting",
            lambda: _get_local_htr_pipeline(),
            lambda result: bool(result and result[0] and result[1]),
            "Handwriting model ready.",
            "Handwriting model unavailable or failed to initialize.",
            LOCAL_HTR_MODEL_ID,
        ),
    ]

    for name, loader, checker, ready_detail, unavailable_detail, model_id in components:
        _update_local_model_prewarm_component(name, status="loading", ready=False, detail="Loading...", model=model_id)
        try:
            loaded = loader()
            is_ready = checker(loaded)
            _update_local_model_prewarm_component(
                name,
                status="ready" if is_ready else "unavailable",
                ready=is_ready,
                detail=ready_detail if is_ready else unavailable_detail,
                model=model_id,
            )
        except Exception as exc:
            logger.warning("Local model prewarm failed for %s: %s", name, exc)
            _update_local_model_prewarm_component(
                name,
                status="failed",
                ready=False,
                detail=str(exc)[:300],
                model=model_id,
            )

    _mark_local_model_prewarm_state(running=False, completed_at=_iso_now())


def _start_local_model_prewarm() -> None:
    global _local_model_prewarm_thread
    if not LOCAL_MODEL_PREWARM_ENABLED:
        return
    if _local_model_prewarm_thread is not None and _local_model_prewarm_thread.is_alive():
        return
    if _copy_local_model_prewarm_status().get("completed_at"):
        return
    _local_model_prewarm_thread = threading.Thread(target=_prewarm_local_models_worker, daemon=True)
    _local_model_prewarm_thread.start()


def _provider_has_access(provider: str) -> bool:
    """Return whether the requested provider is currently usable."""
    if provider == "ollama":
        return _check_ollama_running()
    if provider == "openai":
        return bool(os.environ.get("OPENAI_API_KEY", "").strip())
    if provider in AI_PROVIDERS:
        key_env = AI_PROVIDERS[provider].get("key_env")
        return bool(os.environ.get(key_env, "").strip()) if key_env else False
    return False


def _get_provider_models(provider: str) -> List[str]:
    """Return the current model list for a provider."""
    if provider == "ollama":
        return _get_ollama_models()
    if provider == "openai":
        return ["gpt-4o"]
    if provider in AI_PROVIDERS:
        return list(AI_PROVIDERS[provider].get("models") or [])
    return []


def _get_provider_model_env_key(provider: str) -> str:
    """Return the env var used to store the saved default model for a provider."""
    return PROVIDER_MODEL_ENV_KEYS.get(provider, "")


def _get_provider_default_model(provider: str) -> str:
    """Return the default model for a provider."""
    model_env = _get_provider_model_env_key(provider)
    if model_env:
        persisted = os.environ.get(model_env, "").strip()
        if persisted:
            return persisted

    if provider == "ollama":
        return os.environ.get("OLLAMA_MODEL", OLLAMA_MODEL)
    if provider == "openai":
        return "gpt-4o"
    if provider in AI_PROVIDERS:
        return AI_PROVIDERS[provider]["default_model"]
    return OLLAMA_MODEL


def _coerce_model_for_provider(provider: str, requested_model: str = "") -> str:
    """Return a safe model for the provider, falling back to its default when needed."""
    candidate = (requested_model or "").strip()

    if provider == "ollama":
        if candidate:
            return candidate
        models = _get_provider_models(provider)
        return models[0] if models else _get_provider_default_model(provider)

    models = _get_provider_models(provider)
    if candidate and (not models or candidate in models):
        return candidate
    return _get_provider_default_model(provider)


def _validate_requested_model(provider: str, requested_model: str = "") -> str:
    """Validate a user-selected model for a provider and return the resolved value."""
    model = (requested_model or "").strip()

    if provider == "ollama":
        return _coerce_model_for_provider(provider, model)

    models = _get_provider_models(provider)
    if model and models and model not in models:
        allowed = ", ".join(models)
        raise ValueError(f"Model '{model}' is not available for {provider}. Available models: {allowed}")

    return _coerce_model_for_provider(provider, model)


def _build_available_providers() -> Dict[str, Dict[str, Any]]:
    """Build provider metadata for the frontend from current runtime state."""
    ollama_models = _get_ollama_models() if _check_ollama_running() else []
    available_providers: Dict[str, Dict[str, Any]] = {}

    for name, prov in AI_PROVIDERS.items():
        models = prov["models"] if name != "ollama" else ollama_models
        available_providers[name] = {
            "label": prov["label"],
            "hint": prov["hint"],
            "has_key": bool(os.environ.get(prov.get("key_env"), "").strip()) if prov.get("key_env") else bool(ollama_models),
            "is_available": _provider_has_access(name),
            "models": models,
            "default_model": _coerce_model_for_provider(name),
        }

    return available_providers


def _get_saved_provider_models() -> Dict[str, str]:
    """Return the saved default model for each provider."""
    result: Dict[str, str] = {}
    for provider in list(AI_PROVIDERS.keys()) + ["openai"]:
        result[provider] = _get_provider_default_model(provider)
    return result


def _get_llm_client():
    """Return an OpenAI-compatible client using the best available free provider."""
    global _llm_client, _llm_provider

    provider_name, api_key = _detect_provider()
    if _llm_client is not None and provider_name == _llm_provider:
        return _llm_client

    _llm_client = None
    _llm_provider = None

    if not _HAS_OPENAI_PKG:
        return None
    if not provider_name or not api_key:
        return None

    if provider_name == "openai":
        # Legacy OpenAI key support
        _llm_client = _OpenAI(api_key=api_key)
        _llm_provider = "openai"
    elif provider_name in AI_PROVIDERS:
        prov = AI_PROVIDERS[provider_name]
        _llm_client = _OpenAI(
            base_url=prov["base_url"],
            api_key=api_key,
        )
        _llm_provider = provider_name
    else:
        return None

    return _llm_client


def _get_active_provider() -> str:
    """Return the name of the currently active AI provider."""
    global _llm_provider
    name, _ = _detect_provider()
    if name:
        return name
    if _llm_provider:
        return _llm_provider
    return "none"


def _get_active_model() -> str:
    """Return the model name to use for the active provider."""
    provider = _get_active_provider()
    # Check explicit AI_MODEL env var first
    explicit = os.environ.get("AI_MODEL", "").strip()
    return _coerce_model_for_provider(provider, explicit)


def _get_runtime_model_for_provider(provider: str) -> str:
    """Return the saved model for an arbitrary provider without switching global state."""
    configured_provider = os.environ.get("AI_PROVIDER", "").strip().lower()
    if provider == configured_provider:
        explicit = os.environ.get("AI_MODEL", "").strip()
        return _coerce_model_for_provider(provider, explicit)
    return _get_provider_default_model(provider)


def _create_llm_client_for_provider(provider_name: str):
    """Create a one-off client for a specific provider."""
    if not _HAS_OPENAI_PKG or not provider_name:
        return None

    if provider_name == "openai":
        api_key = os.environ.get("OPENAI_API_KEY", "").strip()
        if not api_key:
            return None
        return _OpenAI(api_key=api_key)

    if provider_name not in AI_PROVIDERS:
        return None

    provider = AI_PROVIDERS[provider_name]
    key_env = provider.get("key_env")
    if provider_name == "ollama":
        if not _check_ollama_running():
            return None
        api_key = "ollama"
    else:
        api_key = os.environ.get(key_env, "").strip() if key_env else ""
        if not api_key:
            return None

    return _OpenAI(
        base_url=provider["base_url"],
        api_key=api_key,
    )


def _attachments_include_images(attachments: list) -> bool:
    return any(str(att.get("type") or "").strip().lower() == "image" for att in (attachments or []))


def _attachments_include_pdfs(attachments: list) -> bool:
    for att in attachments or []:
        mime_type = _infer_attachment_mime_type(
            str(att.get("name") or ""),
            str(att.get("mimeType") or att.get("mime_type") or ""),
            str(att.get("sourceUrl") or att.get("source_url") or ""),
        )
        if mime_type == "application/pdf":
            return True
    return False


def _attachments_need_native_media_reasoning(attachments: list) -> bool:
    return _attachments_include_images(attachments) or _attachments_include_pdfs(attachments)


MAX_REASONING_ATTACHMENTS = 64
MAX_REASONING_ATTACHMENTS_HISTORY = 20
MAX_REASONING_NATIVE_MEDIA_ATTACHMENTS = 10
MAX_REASONING_NATIVE_MEDIA_ATTACHMENTS_HISTORY = 3
MAX_ATTACHMENT_DOSSIER_CHARS = 48000
MAX_ATTACHMENT_DOSSIER_CHARS_HISTORY = 12000
MAX_ATTACHMENT_DOSSIER_ENTRY_CHARS = 2800
MAX_ATTACHMENT_MEMORY_ITEMS = 64
MAX_ATTACHMENT_INDEX_RESULTS = 64
MAX_ATTACHMENT_ANALYSIS_RESULTS = 32
MAX_STORED_IMAGE_ATTACHMENT_CHARS = 12_000_000
MAX_STORED_BINARY_ATTACHMENT_CHARS = 4_000_000
MAX_STORED_TEXT_ATTACHMENT_CHARS = 750_000
MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS = 64_000
MAX_STORED_ANALYSIS_TEXT_CHARS = MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS
MAX_STORED_ANALYSIS_SUMMARY_CHARS = 1000
MAX_ATTACHMENT_CACHE_ITEMS = 800
MAX_ATTACHMENT_CACHE_CHUNKS = 32
ATTACHMENT_CHUNK_CHARS = 2400
ATTACHMENT_CHUNK_OVERLAP = 240
MAX_ATTACHMENT_MEMORY_TEXT_CHARS = 12000


def _history_includes_images(history: list) -> bool:
    for message in history or []:
        if _attachments_include_images(message.get("attachments") or []):
            return True
    return False


def _history_includes_pdfs(history: list) -> bool:
    for message in history or []:
        if _attachments_include_pdfs(message.get("attachments") or []):
            return True
    return False


def _history_needs_native_media_reasoning(history: list) -> bool:
    return _history_includes_images(history) or _history_includes_pdfs(history)


def _provider_supports_vision(provider: str) -> bool:
    return provider in VISION_CAPABLE_PROVIDERS


def _resolve_llm_request_target(attachments: list, history: Optional[list] = None):
    """Pick the best client/provider/model for this request."""
    provider = _get_active_provider()
    model = _get_active_model()
    client = _get_llm_client()

    if not client:
        return None, provider, model

    needs_vision = _attachments_need_native_media_reasoning(attachments) or _history_needs_native_media_reasoning(history or [])

    if needs_vision and not _provider_supports_vision(provider):
        for candidate in ("gemini", "openai"):
            if candidate == provider:
                continue
            vision_client = _create_llm_client_for_provider(candidate)
            if vision_client:
                return vision_client, candidate, _get_runtime_model_for_provider(candidate)

    return client, provider, model


def _resolve_text_fallback_target(preferred_provider: str = ""):
    candidates: List[str] = []

    if preferred_provider and preferred_provider not in candidates:
        candidates.append(preferred_provider)

    active_provider = _get_active_provider()
    if active_provider and active_provider not in candidates:
        candidates.append(active_provider)

    for candidate in ("groq", "cerebras", "ollama", "openai", "gemini"):
        if candidate not in candidates:
            candidates.append(candidate)

    for candidate in candidates:
        client = _get_llm_client() if candidate == _get_active_provider() else _create_llm_client_for_provider(candidate)
        if client:
            return client, candidate, _get_runtime_model_for_provider(candidate)

    return None, "", ""


def _get_ocr_engine():
    global _ocr_engine

    if _ocr_engine is False:
        return None
    if _ocr_engine is not None:
        return _ocr_engine
    if not _HAS_RAPIDOCR:
        _ocr_engine = False
        return None

    try:
        _ocr_engine = _RapidOCR()
    except Exception as exc:
        logger.warning("Failed to initialize RapidOCR fallback: %s", exc)
        _ocr_engine = False
        return None

    return _ocr_engine


def _get_local_structured_vlm():
    global _local_structured_vlm_model, _local_structured_vlm_processor

    if _local_structured_vlm_model is False:
        return None, None
    if _local_structured_vlm_model is not None and _local_structured_vlm_processor is not None:
        return _local_structured_vlm_processor, _local_structured_vlm_model
    if not LOCAL_STRUCTURED_VLM_ENABLED or not _HAS_PIL or not _HAS_TRANSFORMERS or not _HAS_TORCH or not _HAS_FLORENCE2:
        _local_structured_vlm_model = False
        _local_structured_vlm_processor = False
        return None, None

    try:
        _local_structured_vlm_processor = _AutoProcessor.from_pretrained(LOCAL_STRUCTURED_VLM_MODEL_ID)
        _local_structured_vlm_model = _Florence2ForConditionalGeneration.from_pretrained(
            LOCAL_STRUCTURED_VLM_MODEL_ID,
            torch_dtype=_torch.float32 if _HAS_TORCH and _torch is not None else None,
        )
        _local_structured_vlm_model.to("cpu")
        _local_structured_vlm_model.eval()
    except Exception as exc:
        logger.warning("Failed to initialize local structured VLM '%s': %s", LOCAL_STRUCTURED_VLM_MODEL_ID, exc)
        _local_structured_vlm_model = False
        _local_structured_vlm_processor = False
        return None, None

    return _local_structured_vlm_processor, _local_structured_vlm_model


def _get_local_vlm_pipeline():
    global _local_vlm_pipeline

    if _local_vlm_pipeline is False:
        return None
    if _local_vlm_pipeline is not None:
        return _local_vlm_pipeline
    if not LOCAL_VLM_ENABLED or not _HAS_PIL or not _HAS_TRANSFORMERS or not _HAS_TORCH:
        _local_vlm_pipeline = False
        return None

    last_error: Optional[Exception] = None
    for task_name in ("image-text-to-text", "image-to-text"):
        try:
            _local_vlm_pipeline = _hf_pipeline(
                task_name,
                model=LOCAL_VLM_MODEL_ID,
                device=-1,
            )
            return _local_vlm_pipeline
        except Exception as exc:
            last_error = exc

    logger.warning("Failed to initialize local VLM '%s': %s", LOCAL_VLM_MODEL_ID, last_error)
    _local_vlm_pipeline = False
    return None


def _get_local_htr_pipeline():
    global _local_htr_pipeline, _local_htr_processor

    if _local_htr_pipeline is False:
        return None, None
    if _local_htr_pipeline is not None and _local_htr_processor is not None:
        return _local_htr_processor, _local_htr_pipeline
    if not LOCAL_HTR_ENABLED or not _HAS_PIL or not _HAS_TRANSFORMERS or not _HAS_TORCH:
        _local_htr_pipeline = False
        _local_htr_processor = False
        return None, None

    try:
        _local_htr_processor = _TrOCRProcessor.from_pretrained(LOCAL_HTR_MODEL_ID)
        _local_htr_pipeline = _VisionEncoderDecoderModel.from_pretrained(LOCAL_HTR_MODEL_ID)
        _local_htr_pipeline.to("cpu")
        _local_htr_pipeline.eval()
    except Exception as exc:
        logger.warning("Failed to initialize local handwriting model '%s': %s", LOCAL_HTR_MODEL_ID, exc)
        _local_htr_pipeline = False
        _local_htr_processor = False
        return None, None

    return _local_htr_processor, _local_htr_pipeline


def _decode_image_attachments_to_pil_images(attachments: list, limit: int = 3) -> List[Any]:
    if not _HAS_PIL or _PILImage is None:
        return []
    images: List[Any] = []
    for att in (attachments or [])[:limit]:
        if str(att.get("type") or "").strip().lower() != "image":
            continue
        _, base64_data = _split_base64_data_url(str(att.get("content") or ""))
        if not base64_data:
            continue
        try:
            image = _PILImage.open(io.BytesIO(base64.b64decode(base64_data))).convert("RGB")
        except Exception as exc:
            logger.warning("Failed to decode image attachment '%s': %s", att.get("name", "image"), exc)
            continue
        images.append(image)
    return images


def _resize_pil_image(image: Any, scale: float) -> Any:
    if not image or scale <= 1.0:
        return image
    width, height = image.size
    target = (
        max(1, int(round(width * scale))),
        max(1, int(round(height * scale))),
    )
    return image.resize(target, resample=_PILImage.Resampling.LANCZOS)


def _document_ocr_base_image(image: Any) -> Any:
    if not _HAS_PIL or image is None:
        return image
    prepared = image.convert("RGB")
    width, height = prepared.size
    longest_edge = max(width, height)
    if longest_edge and longest_edge < 1600:
        scale = min(3.0, 1600 / max(longest_edge, 1))
        prepared = _resize_pil_image(prepared, scale)
    return prepared


def _build_ocr_image_variants(image: Any) -> List[tuple[str, Any]]:
    if not _HAS_PIL or image is None:
        return []

    base = _document_ocr_base_image(image)
    gray = _ImageOps.grayscale(base)
    autocontrast = _ImageOps.autocontrast(gray)
    denoised = autocontrast.filter(_ImageFilter.MedianFilter(size=3))
    high_contrast = _ImageEnhance.Contrast(denoised).enhance(2.2)
    sharpened = high_contrast.filter(_ImageFilter.SHARPEN)
    threshold = sharpened.point(lambda value: 255 if value > 170 else 0, mode="L")

    variants: List[tuple[str, Any]] = [
        ("original_upscaled", base),
        ("grayscale_autocontrast", autocontrast),
        ("high_contrast_sharpened", sharpened),
        ("binary_threshold", threshold),
    ]
    return variants


def _build_htr_image_variants(image: Any) -> List[tuple[str, Any]]:
    if not _HAS_PIL or image is None:
        return []
    base = _document_ocr_base_image(image)
    gray = _ImageOps.grayscale(base)
    autocontrast = _ImageOps.autocontrast(gray)
    softened = autocontrast.filter(_ImageFilter.MedianFilter(size=3))
    return [
        ("handwriting_autocontrast", autocontrast.convert("RGB")),
        ("handwriting_softened", softened.convert("RGB")),
        ("handwriting_original", base),
    ]


def _extract_rapidocr_lines(result: Any) -> List[str]:
    txts = getattr(result, "txts", None)
    if txts:
        return [str(text).strip() for text in txts if str(text).strip()]

    if isinstance(result, tuple):
        for item in result:
            if isinstance(item, (list, tuple)):
                candidate_lines = [str(text).strip() for text in item if isinstance(text, (str, bytes)) and str(text).strip()]
                if candidate_lines:
                    return candidate_lines

    if isinstance(result, list):
        candidate_lines = [str(text).strip() for text in result if isinstance(text, (str, bytes)) and str(text).strip()]
        if candidate_lines:
            return candidate_lines

    text = str(result or "").strip()
    return [text] if text else []


def _score_ocr_candidate_text(text: str) -> float:
    cleaned = str(text or "").strip()
    if not cleaned:
        return -1.0
    non_space = [char for char in cleaned if not char.isspace()]
    if not non_space:
        return -1.0

    tokens = re.findall(r"[A-Za-z0-9][A-Za-z0-9'/-]{1,}", cleaned)
    mixed_alpha_digit_tokens = sum(1 for token in tokens if re.search(r"[A-Za-z]", token) and re.search(r"\d", token))
    alpha_numeric = sum(1 for char in non_space if char.isalnum())
    punctuation = sum(1 for char in non_space if not char.isalnum())
    repeated_penalty = len(re.findall(r"(.)\1{4,}", cleaned))
    line_count = max(1, len([line for line in cleaned.splitlines() if line.strip()]))

    score = 0.0
    score += min(len(cleaned), 2400) / 90.0
    score += len(tokens) * 1.25
    score += (alpha_numeric / max(len(non_space), 1)) * 10.0
    score += min(line_count, 8) * 0.5
    score -= punctuation * 0.04
    score -= mixed_alpha_digit_tokens * 2.5
    score -= repeated_penalty * 2.0
    return score


def _score_handwriting_candidate_text(text: str) -> float:
    cleaned = str(text or "").strip()
    if not cleaned:
        return -1.0
    score = _score_ocr_candidate_text(cleaned)
    if re.search(r"\b[a-zA-Z]{3,}\b", cleaned):
        score += 2.0
    return score


def _run_rapidocr_on_pil_image(engine: Any, image: Any, *, variant_label: str = "original") -> List[str]:
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            image.save(tmp, format="PNG")
            temp_path = tmp.name
        result = engine(temp_path)
        return _extract_rapidocr_lines(result)
    except Exception as exc:
        logger.warning("RapidOCR failed on %s image variant: %s", variant_label, exc)
        return []
    finally:
        if temp_path:
            try:
                os.unlink(temp_path)
            except OSError:
                pass


def _ocr_pil_image_multistage(image: Any, engine: Any) -> Dict[str, Any]:
    if not engine or image is None:
        return {}

    candidates: List[Dict[str, Any]] = []
    for variant_label, variant_image in _build_ocr_image_variants(image):
        lines = _run_rapidocr_on_pil_image(engine, variant_image, variant_label=variant_label)
        text = "\n".join(line for line in lines if line).strip()
        if not text:
            continue
        candidates.append({
            "variant": variant_label,
            "text": text,
            "score": _score_ocr_candidate_text(text),
        })

    if not candidates:
        return {}

    candidates.sort(key=lambda item: (-float(item.get("score") or 0.0), len(str(item.get("text") or "")) * -1))
    best = candidates[0]
    return {
        "text": str(best.get("text") or "").strip(),
        "variant": str(best.get("variant") or "original"),
        "score": round(float(best.get("score") or 0.0), 2),
        "candidates": [
            {
                "variant": str(item.get("variant") or ""),
                "score": round(float(item.get("score") or 0.0), 2),
            }
            for item in candidates[:4]
        ],
    }


def _basic_color_name(hex_color: str) -> str:
    palette = {
        "black": (0, 0, 0),
        "white": (255, 255, 255),
        "gray": (128, 128, 128),
        "red": (220, 20, 60),
        "orange": (255, 140, 0),
        "yellow": (240, 220, 70),
        "green": (50, 160, 90),
        "teal": (0, 140, 140),
        "blue": (65, 105, 225),
        "purple": (138, 43, 226),
        "brown": (139, 90, 43),
        "pink": (255, 105, 180),
    }
    try:
        rgb = tuple(int(hex_color[index:index + 2], 16) for index in (1, 3, 5))
    except Exception:
        return hex_color
    nearest_name = "unknown"
    nearest_distance = float("inf")
    for name, target in palette.items():
        distance = sum((rgb[channel] - target[channel]) ** 2 for channel in range(3))
        if distance < nearest_distance:
            nearest_name = name
            nearest_distance = distance
    return nearest_name


def _extract_visual_metadata_from_pil_image(image: Any) -> Dict[str, Any]:
    width, height = image.size
    orientation = "square"
    if width > height:
        orientation = "landscape"
    elif height > width:
        orientation = "portrait"

    sample = image.copy().convert("RGB")
    sample.thumbnail((80, 80))
    try:
        quantized = sample.quantize(colors=4)
        palette = quantized.getpalette() or []
        color_counts = sorted((quantized.getcolors() or []), reverse=True)
        dominant: List[str] = []
        color_names: List[str] = []
        for count, palette_index in color_counts[:3]:
            base = palette_index * 3
            rgb = tuple(int(palette[base + offset]) for offset in range(3))
            hex_color = "#{:02x}{:02x}{:02x}".format(*rgb)
            dominant.append(hex_color)
            color_names.append(_basic_color_name(hex_color))
    except Exception:
        dominant = []
        color_names = []

    return {
        "width": int(width),
        "height": int(height),
        "orientation": orientation,
        "aspectRatio": round(width / max(height, 1), 3),
        "dominantColors": dominant,
        "dominantColorNames": color_names,
    }


def _normalize_attachment_analysis_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _merge_analysis_text(existing: str, candidate: str, limit: int = 16000) -> str:
    existing = str(existing or "").strip()
    candidate = str(candidate or "").strip()
    if not existing:
        return _compact_text(candidate, limit=limit)
    if not candidate:
        return _compact_text(existing, limit=limit)
    if _normalize_attachment_analysis_text(existing) == _normalize_attachment_analysis_text(candidate):
        preferred = candidate if len(candidate) > len(existing) else existing
        return _compact_text(preferred, limit=limit)
    return _compact_text(f"{existing}\n\n{candidate}", limit=limit)


def _merge_visual_metadata(base_meta: Dict[str, Any], extra_meta: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(base_meta or {})
    if not isinstance(extra_meta, dict):
        return merged

    for key, value in extra_meta.items():
        if value in (None, "", [], {}):
            continue
        existing = merged.get(key)
        if isinstance(existing, dict) and isinstance(value, dict):
            merged[key] = _merge_visual_metadata(existing, value)
            continue
        if isinstance(existing, list) and isinstance(value, list):
            merged[key] = list(dict.fromkeys([*existing, *value]))
            continue
        if existing in (None, "", [], {}):
            merged[key] = value
            continue
        if key in {"detectedObjects", "pageVisualDiagnostics", "visionDetails"}:
            existing_list = existing if isinstance(existing, list) else [existing]
            value_list = value if isinstance(value, list) else [value]
            merged[key] = list(dict.fromkeys([*existing_list, *value_list]))
            continue
        merged[key] = existing

    return merged


def _extract_florence_task_value(parsed: Any, task_prompt: str) -> Any:
    if isinstance(parsed, dict):
        if task_prompt in parsed:
            return parsed[task_prompt]
        if len(parsed) == 1:
            return next(iter(parsed.values()))
    return parsed


def _run_local_structured_vlm_task(image: Any, task_prompt: str, max_new_tokens: int = 256) -> Any:
    processor, model = _get_local_structured_vlm()
    if not processor or not model:
        return None

    try:
        inputs = processor(text=task_prompt, images=image, return_tensors="pt")
        generated_ids = model.generate(
            **inputs,
            max_new_tokens=max_new_tokens,
            num_beams=3,
        )
        generated_text = processor.batch_decode(
            generated_ids,
            skip_special_tokens=False,
            clean_up_tokenization_spaces=False,
        )[0]
        try:
            parsed = processor.post_process_generation(generated_text, task=task_prompt, image_size=image.size)
        except Exception:
            parsed = generated_text
        return _extract_florence_task_value(parsed, task_prompt)
    except Exception as exc:
        logger.warning("Local structured VLM task %s failed: %s", task_prompt, exc)
        return None


def _analyze_pil_images_with_local_structured_vlm(images: list) -> List[Dict[str, Any]]:
    analyses: List[Dict[str, Any]] = []
    if not images:
        return analyses

    for image in images[:2]:
        caption_result = _run_local_structured_vlm_task(image, "<MORE_DETAILED_CAPTION>", max_new_tokens=160)
        ocr_image = next((variant for label, variant in _build_ocr_image_variants(image) if label == "grayscale_autocontrast"), image)
        if hasattr(ocr_image, "convert"):
            ocr_image = ocr_image.convert("RGB")
        ocr_result = _run_local_structured_vlm_task(ocr_image, "<OCR>", max_new_tokens=256)
        od_result = _run_local_structured_vlm_task(image, "<OD>", max_new_tokens=256)

        caption_text = _compact_text(str(caption_result or "").strip(), limit=600)
        ocr_text = _compact_text(str(ocr_result or "").strip(), limit=4000)

        labels: List[str] = []
        bboxes: List[Any] = []
        if isinstance(od_result, dict):
            labels = [str(label).strip() for label in (od_result.get("labels") or []) if str(label).strip()]
            bboxes = list(od_result.get("bboxes") or [])

        analyses.append({
            "caption": caption_text,
            "ocr": ocr_text,
            "detectedObjects": labels[:12],
            "detectedRegionCount": len(bboxes),
            "model": LOCAL_STRUCTURED_VLM_MODEL_ID,
            "source": "florence2",
        })

    return analyses


def _combine_structured_visual_analyses(analyses: list, label_prefix: str = "Image") -> Dict[str, Any]:
    if not analyses:
        return {}

    multi_item = len(analyses) > 1
    caption_chunks: List[str] = []
    ocr_chunks: List[str] = []
    object_labels: List[str] = []
    vision_details: List[str] = []
    region_count = 0

    for index, analysis in enumerate(analyses, start=1):
        prefix = f"{label_prefix} {index}: " if multi_item else ""
        caption = str(analysis.get("caption") or "").strip()
        if caption:
            caption_chunks.append(prefix + caption)
        ocr_text = str(analysis.get("ocr") or "").strip()
        if ocr_text:
            ocr_chunks.append(prefix + ocr_text)
        labels = [str(label).strip() for label in (analysis.get("detectedObjects") or []) if str(label).strip()]
        if labels:
            object_labels.extend(labels)
            vision_details.append(prefix + "objects: " + ", ".join(labels[:6]))
        count = int(analysis.get("detectedRegionCount") or 0)
        if count:
            region_count += count

    unique_labels = list(dict.fromkeys(object_labels))[:12]
    meta: Dict[str, Any] = {
        "visionSource": "florence2",
        "visionModel": LOCAL_STRUCTURED_VLM_MODEL_ID,
    }
    if unique_labels:
        meta["detectedObjects"] = unique_labels
        meta["detectedObjectCount"] = len(unique_labels)
    if region_count:
        meta["detectedRegionCount"] = region_count
    if vision_details:
        meta["visionDetails"] = vision_details[:4]

    result: Dict[str, Any] = {"analysisMeta": meta}
    caption_text = "\n".join(caption_chunks).strip()
    ocr_text = "\n".join(ocr_chunks).strip()
    if caption_text:
        result["analysisCaption"] = _compact_text(caption_text, limit=1200)
    if ocr_text:
        result["analysisText"] = _compact_text(ocr_text, limit=6000)
    return result


def _extract_structured_visual_analysis_from_image_attachments(attachments: list) -> Dict[str, Any]:
    images = _decode_image_attachments_to_pil_images(attachments, limit=2)
    return _combine_structured_visual_analyses(_analyze_pil_images_with_local_structured_vlm(images))


def _caption_pil_images(images: list) -> List[str]:
    captioner = _get_local_vlm_pipeline()
    if not captioner or not images:
        return []

    descriptions: List[str] = []
    for image in images[:3]:
        try:
            result = captioner(image, max_new_tokens=80)
        except Exception as exc:
            logger.warning("Local VLM captioning failed: %s", exc)
            continue

        caption = ""
        if isinstance(result, list) and result:
            item = result[0]
            if isinstance(item, dict):
                caption = str(item.get("generated_text") or item.get("caption") or "").strip()
            else:
                caption = str(item).strip()
        else:
            caption = str(result or "").strip()

        if caption:
            descriptions.append(_compact_text(caption, limit=300))
    return descriptions


def _extract_semantic_descriptions_from_image_attachments(attachments: list) -> str:
    structured = _extract_structured_visual_analysis_from_image_attachments(attachments)
    structured_caption = str(structured.get("analysisCaption") or "").strip()
    if structured_caption:
        return structured_caption
    return "\n".join(_caption_pil_images(_decode_image_attachments_to_pil_images(attachments))).strip()


def _extract_handwriting_from_pil_images(images: list) -> List[str]:
    processor, model = _get_local_htr_pipeline()
    if not processor or not model or not images:
        return []

    outputs: List[str] = []
    for image in images[:3]:
        best_text = ""
        best_score = -1.0
        for variant_label, variant_image in _build_htr_image_variants(image):
            try:
                pixel_values = processor(images=variant_image, return_tensors="pt").pixel_values
                generated_ids = model.generate(pixel_values, max_new_tokens=96)
                result = processor.batch_decode(generated_ids, skip_special_tokens=True)
            except Exception as exc:
                logger.warning("Local handwriting OCR failed on %s variant: %s", variant_label, exc)
                continue
            text = ""
            if isinstance(result, list) and result:
                text = str(result[0]).strip()
            else:
                text = str(result or "").strip()
            score = _score_handwriting_candidate_text(text)
            if text and score > best_score:
                best_text = text
                best_score = score
        if best_text:
            outputs.append(_compact_text(best_text, limit=600))
    return outputs


def _extract_handwriting_from_image_attachments(attachments: list) -> str:
    images = _decode_image_attachments_to_pil_images(attachments)
    return "\n".join(_extract_handwriting_from_pil_images(images)).strip()


def _extract_visual_metadata_from_image_attachments(attachments: list) -> Dict[str, Any]:
    images = _decode_image_attachments_to_pil_images(attachments, limit=1)
    if not images:
        return {}
    meta = _extract_visual_metadata_from_pil_image(images[0])
    structured = _extract_structured_visual_analysis_from_image_attachments(attachments)
    return _merge_visual_metadata(meta, structured.get("analysisMeta") or {})


def _render_visual_metadata_summary(meta: Dict[str, Any]) -> str:
    if not meta:
        return ""
    parts: List[str] = []
    width = meta.get("width")
    height = meta.get("height")
    if width and height:
        parts.append(f"{width}x{height}")
    orientation = str(meta.get("orientation") or "").strip()
    if orientation:
        parts.append(orientation)
    aspect_ratio = meta.get("aspectRatio")
    if aspect_ratio:
        parts.append(f"aspect {aspect_ratio}")
    color_names = [str(name) for name in (meta.get("dominantColorNames") or []) if str(name).strip()]
    if color_names:
        parts.append("dominant colors: " + ", ".join(color_names[:3]))
    detected_objects = [str(name) for name in (meta.get("detectedObjects") or []) if str(name).strip()]
    if detected_objects:
        parts.append("objects: " + ", ".join(detected_objects[:6]))
    page_count = meta.get("pageCount")
    if page_count:
        parts.append(f"pages analyzed: {page_count}")
    page_visual_diagnostics = [str(item).strip() for item in (meta.get("pageVisualDiagnostics") or []) if str(item).strip()]
    if page_visual_diagnostics:
        parts.append("page visuals: " + " | ".join(page_visual_diagnostics[:2]))
    return "; ".join(parts)


def _visual_review_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    return any(phrase in lowered for phrase in _VISUAL_REVIEW_PHRASES)


def _visual_comparison_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    return any(phrase in lowered for phrase in _VISUAL_COMPARE_PHRASES)


def _attachment_summary_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    return any(phrase in lowered for phrase in _ATTACHMENT_SUMMARY_PHRASES)


def _attachment_followup_reference_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    if any(phrase in lowered for phrase in _ATTACHMENT_FOLLOWUP_REFERENCE_PHRASES):
        return True
    tokens = _tokenize_query(lowered)
    if len(tokens) > 18:
        return False
    has_reference = any(token in _ATTACHMENT_REFERENCE_TOKENS for token in tokens)
    if not has_reference:
        return False
    return any(token in _ATTACHMENT_REFERENCE_INTENT_TOKENS for token in tokens)


def _extract_visual_request_terms(message: str) -> Dict[str, List[str]]:
    tokens = [token for token in re.findall(r"[a-z0-9]+", (message or "").lower()) if token]
    colors = [token for token in tokens if token in _VISUAL_COLOR_TERMS]
    focus_terms: List[str] = []
    for token in tokens:
        if token in _VISUAL_STOPWORDS or token in _VISUAL_COLOR_TERMS:
            continue
        if len(token) < 4:
            continue
        if token not in focus_terms:
            focus_terms.append(token)
        if len(focus_terms) >= 8:
            break
    return {
        "colors": colors,
        "focusTerms": focus_terms,
    }


def _attachment_visual_blob(attachment: Dict[str, Any]) -> str:
    analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
    text_parts = [
        str(attachment.get("name") or ""),
        str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or ""),
        str(attachment.get("analysisHandwriting") or attachment.get("analysis_handwriting") or ""),
        str(attachment.get("analysisText") or attachment.get("analysis_text") or ""),
        str(attachment.get("analysisSummary") or attachment.get("analysis_summary") or ""),
        " ".join(str(keyword) for keyword in (attachment.get("analysisKeywords") or attachment.get("analysis_keywords") or [])),
        " ".join(str(item) for item in (analysis_meta.get("detectedObjects") or [])),
        " ".join(str(item) for item in (analysis_meta.get("dominantColorNames") or [])),
        " ".join(str(item) for item in (analysis_meta.get("visionDetails") or [])),
    ]
    return " ".join(part for part in text_parts if part).strip().lower()


def _collect_present_visual_colors(attachment: Dict[str, Any]) -> List[str]:
    analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
    present: List[str] = []
    for color in analysis_meta.get("dominantColorNames") or []:
        color_name = str(color).strip().lower()
        if color_name and color_name not in present:
            present.append(color_name)
    blob = _attachment_visual_blob(attachment)
    for color in _VISUAL_COLOR_TERMS:
        if color in present:
            continue
        if re.search(rf"\b{re.escape(color)}\b", blob):
            present.append(color)
    return present


def _build_visual_alignment_report(message: str, attachments: list) -> str:
    review_requested = _visual_review_requested(message)
    comparison_requested = _visual_comparison_requested(message)
    request_terms = _extract_visual_request_terms(message)
    visual_attachments = []
    for attachment in attachments or []:
        attachment_type = str(attachment.get("type") or "").strip().lower()
        if attachment_type == "image":
            visual_attachments.append(attachment)
            continue
        analysis_caption = str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or "").strip()
        analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
        if analysis_caption or analysis_meta:
            visual_attachments.append(attachment)

    if not visual_attachments:
        return ""

    lines: List[str] = []

    if comparison_requested and len(visual_attachments) >= 2:
        first = visual_attachments[0]
        second = visual_attachments[1]
        first_colors = _collect_present_visual_colors(first)
        second_colors = _collect_present_visual_colors(second)
        first_caption = str(first.get("analysisCaption") or first.get("analysis_caption") or "").strip()
        second_caption = str(second.get("analysisCaption") or second.get("analysis_caption") or "").strip()
        first_text = str(first.get("analysisText") or first.get("analysis_text") or "").strip()
        second_text = str(second.get("analysisText") or second.get("analysis_text") or "").strip()
        first_objects = [str(item).strip() for item in ((first.get("analysisMeta") or {}).get("detectedObjects") or []) if str(item).strip()]
        second_objects = [str(item).strip() for item in ((second.get("analysisMeta") or {}).get("detectedObjects") or []) if str(item).strip()]

        lines.append("Visual comparison:")
        if first_colors or second_colors:
            lines.append(
                f"- Color shift: {', '.join(first_colors[:4]) or 'unknown'} -> {', '.join(second_colors[:4]) or 'unknown'}."
            )
        if first_text and second_text and first_text != second_text:
            lines.append(f'- Visible text changed: "{_compact_text(first_text, 80)}" -> "{_compact_text(second_text, 80)}".')
        elif second_text:
            lines.append(f'- Current visible text: "{_compact_text(second_text, 100)}".')
        if first_caption and second_caption and first_caption != second_caption:
            lines.append(f"- Scene description shift: {_compact_text(first_caption, 120)} -> {_compact_text(second_caption, 120)}.")
        elif second_caption:
            lines.append(f"- Current scene description: {_compact_text(second_caption, 140)}.")
        if first_objects or second_objects:
            lines.append(
                f"- Object labels: {', '.join(first_objects[:6]) or 'unknown'} -> {', '.join(second_objects[:6]) or 'unknown'}."
            )
        return "\n".join(lines)

    if not review_requested and not request_terms["colors"] and not request_terms["focusTerms"]:
        return ""

    attachment = visual_attachments[0]
    blob = _attachment_visual_blob(attachment)
    colors = _collect_present_visual_colors(attachment)
    analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
    detected_objects = [str(item).strip() for item in (analysis_meta.get("detectedObjects") or []) if str(item).strip()]
    caption = str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or "").strip()
    text = str(attachment.get("analysisText") or attachment.get("analysis_text") or "").strip()

    requested_colors = list(dict.fromkeys(request_terms["colors"]))
    matched_colors = [color for color in requested_colors if color in colors or re.search(rf"\b{re.escape(color)}\b", blob)]
    missing_colors = [color for color in requested_colors if color not in matched_colors]
    focus_matches = [term for term in request_terms["focusTerms"] if re.search(rf"\b{re.escape(term)}\b", blob)]

    verdict = ""
    if requested_colors:
        if matched_colors and not missing_colors:
            verdict = f"Likely matches the requested color change: {', '.join(matched_colors[:3])} is present."
        elif matched_colors and missing_colors:
            verdict = (
                f"Partial match: {', '.join(matched_colors[:3])} is present, but I do not see "
                f"{', '.join(missing_colors[:3])} clearly in the analysis."
            )
        else:
            verdict = f"Possible mismatch: I do not see the requested color {', '.join(requested_colors[:3])} in the analysis."
    elif review_requested:
        verdict = "Visual review based on the current attachment analysis."

    if not verdict:
        return ""

    lines.append(verdict)
    if colors:
        lines.append(f"- Detected colors: {', '.join(colors[:6])}.")
    if detected_objects:
        lines.append(f"- Detected objects: {', '.join(detected_objects[:6])}.")
    if focus_matches:
        lines.append(f"- Request terms seen in the analysis: {', '.join(focus_matches[:6])}.")
    elif request_terms["focusTerms"]:
        lines.append(f"- I could not clearly confirm these requested elements: {', '.join(request_terms['focusTerms'][:6])}.")
    if caption:
        lines.append(f"- Scene reading: {_compact_text(caption, 180)}.")
    if text:
        lines.append(f'- Visible text: "{_compact_text(text, 120)}".')
    return "\n".join(lines)


def _coerce_attachment_for_reasoning(att: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize attachment records without repeating heavy extraction when analysis already exists."""
    sanitized = _sanitize_chat_attachment(att)
    sanitized["mimeType"] = _infer_attachment_mime_type(
        str(sanitized.get("name") or ""),
        str(sanitized.get("mimeType") or ""),
        str(sanitized.get("sourceUrl") or ""),
    )

    already_analyzed = any([
        sanitized.get("analysisText"),
        sanitized.get("analysisCaption"),
        sanitized.get("analysisHandwriting"),
        sanitized.get("analysisMeta"),
        sanitized.get("analysisSummary"),
        sanitized.get("analysisKeywords"),
    ])
    if already_analyzed and sanitized.get("analysisSummary") and sanitized.get("analysisKeywords"):
        return sanitized

    return _prepare_chat_attachment(copy.deepcopy(sanitized))


def _prepare_attachments_for_reasoning(attachments: list, *, limit: int) -> tuple[List[Dict[str, Any]], int]:
    prepared: List[Dict[str, Any]] = []
    source = attachments or []
    clipped = max(0, len(source) - limit)
    for index, raw_attachment in enumerate(source[:limit]):
        attachment = _coerce_attachment_for_reasoning(copy.deepcopy(raw_attachment))
        attachment["_reasoningIndex"] = index
        prepared.append(attachment)
    return prepared, clipped


def _attachment_identity_key(att: Dict[str, Any]) -> str:
    sanitized = _sanitize_chat_attachment(att)
    fingerprint = str(sanitized.get("attachmentFingerprint") or "").strip()
    if not fingerprint:
        fingerprint = _build_attachment_fingerprint(sanitized)
    if fingerprint:
        return fingerprint
    return "|".join([
        str(sanitized.get("name") or "").strip().lower(),
        str(sanitized.get("type") or "").strip().lower(),
        str(sanitized.get("mimeType") or "").strip().lower(),
        str(sanitized.get("sourceUrl") or "").strip().lower(),
        str(sanitized.get("size") or 0),
    ])


def _collect_recent_attachments_from_history(history: list, *, limit: int = 12, source_turn_limit: int = 3) -> List[Dict[str, Any]]:
    collected: List[Dict[str, Any]] = []
    seen: set[str] = set()
    source_turns = 0

    for message in reversed(history or []):
        if str(message.get("role") or "") != "user":
            continue
        hist_attachments = message.get("attachments") or []
        if not hist_attachments:
            continue
        source_turns += 1
        for raw_attachment in reversed(hist_attachments):
            sanitized = _sanitize_chat_attachment(raw_attachment)
            key = _attachment_identity_key(sanitized)
            if key in seen:
                continue
            seen.add(key)
            collected.append(sanitized)
            if len(collected) >= limit:
                break
        if len(collected) >= limit or source_turns >= source_turn_limit:
            break

    collected.reverse()
    return collected


def _should_reuse_recent_attachments(message: str, attachments: list, history: list) -> bool:
    if not history:
        return False
    if not message:
        return False
    if not attachments and (
        _visual_review_requested(message)
        or _visual_comparison_requested(message)
        or _attachment_summary_requested(message)
        or _exact_attachment_text_requested(message)
        or _which_attachment_requested(message)
        or _attachment_followup_reference_requested(message)
    ):
        return True
    if _visual_comparison_requested(message) and len(attachments or []) < 2:
        return True
    return False


def _resolve_effective_turn_attachments(message: str, attachments: list, history: list) -> List[Dict[str, Any]]:
    effective = [_sanitize_chat_attachment(att) for att in (attachments or [])]
    if not _should_reuse_recent_attachments(message, effective, history):
        return effective

    recent = _collect_recent_attachments_from_history(history)
    if not recent:
        return effective

    seen = {_attachment_identity_key(att) for att in effective}
    merged = list(effective)
    for attachment in recent:
        key = _attachment_identity_key(attachment)
        if key in seen:
            continue
        merged.append(attachment)
        seen.add(key)
        if len(merged) >= MAX_REASONING_ATTACHMENTS:
            break
    return merged


def _score_attachment_relevance(query: str, attachment: Dict[str, Any]) -> float:
    query_text = (query or "").strip()
    score = _score_text_match(
        query_text,
        str(attachment.get("name") or ""),
        str(attachment.get("mimeType") or attachment.get("mime_type") or ""),
        str(attachment.get("sourceUrl") or attachment.get("source_url") or ""),
        str(attachment.get("analysisSummary") or attachment.get("analysis_summary") or ""),
        str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or ""),
        str(attachment.get("analysisHandwriting") or attachment.get("analysis_handwriting") or ""),
        str(attachment.get("analysisText") or attachment.get("analysis_text") or ""),
        str(attachment.get("content") or ""),
        json.dumps(attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}, default=str)[:2400],
        " ".join(str(keyword) for keyword in (attachment.get("analysisKeywords") or attachment.get("analysis_keywords") or [])),
    )
    attachment_type = str(attachment.get("type") or "").strip().lower()
    mime_type = str(attachment.get("mimeType") or attachment.get("mime_type") or "").strip().lower()
    lowered = query_text.lower()
    if query_text:
        if attachment_type == "image" and (_visual_review_requested(query_text) or _visual_comparison_requested(query_text)):
            score += 6.0
        if mime_type == "application/pdf" and "pdf" in lowered:
            score += 5.0
        if attachment.get("sourceUrl") and any(token in lowered for token in ("website", "web", "url", "page", "link")):
            score += 3.0
        if any(token in lowered for token in ("doc", "document", "spreadsheet", "sheet", "slide", "presentation", "csv", "tsv", "archive", "zip", "epub")):
            if attachment_type in {"file", "binary"}:
                score += 1.2
    analysis_text = str(attachment.get("analysisText") or attachment.get("analysis_text") or "")
    analysis_caption = str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or "")
    score += min(len(analysis_text) / 4000.0, 2.5)
    score += min(len(analysis_caption) / 500.0, 1.2)
    if query_text and _attachment_summary_requested(query_text) and (analysis_text or str(attachment.get("content") or "").strip()):
        score += 6.0 + min(3.0, len(_get_attachment_cached_chunks(attachment)) * 0.35)
    if query_text:
        chunk_hits = _select_attachment_evidence_snippets(attachment, query_text, limit=2)
        for chunk in chunk_hits:
            score += min(4.0, _score_text_match(query_text, chunk))
    return score


def _rank_attachments_for_query(query: str, attachments: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ranked: List[tuple[float, int, Dict[str, Any]]] = []
    for index, attachment in enumerate(attachments):
        source_index = int(attachment.get("_reasoningIndex", index))
        score = _score_attachment_relevance(query, attachment)
        attachment["_relevanceScore"] = round(score, 3)
        attachment["_relevanceConfidence"] = _score_to_confidence_label(score)
        ranked.append((score, source_index, attachment))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    ordered = [item[2] for item in ranked]
    for rank, attachment in enumerate(ordered, start=1):
        attachment["_dossierRank"] = rank
    return ordered


def _collect_repeated_attachment_terms(attachments: List[Dict[str, Any]], *, extractor, limit: int = 8) -> List[str]:
    counts: Dict[str, int] = {}
    for attachment in attachments:
        seen: List[str] = []
        for value in extractor(attachment):
            term = str(value).strip().lower()
            if not term or term in seen:
                continue
            seen.append(term)
            counts[term] = counts.get(term, 0) + 1
    repeated = [(count, term) for term, count in counts.items() if count >= 2]
    repeated.sort(key=lambda item: (-item[0], item[1]))
    return [term for _, term in repeated[:limit]]


def _render_attachment_synthesis(message: str, attachments: List[Dict[str, Any]], clipped: int = 0) -> str:
    if not attachments:
        return ""

    type_counts: Dict[str, int] = {}
    for attachment in attachments:
        label = str(attachment.get("type") or "file").strip().lower() or "file"
        type_counts[label] = type_counts.get(label, 0) + 1

    count_summary = ", ".join(
        f"{count} {label}{'s' if count != 1 else ''}"
        for label, count in sorted(type_counts.items(), key=lambda item: item[0])
    )
    lines = [f"Analyzed {len(attachments)} attachments across this request: {count_summary}."]

    if message:
        most_relevant = ", ".join(
            str(attachment.get("name") or "attachment")
            for attachment in attachments[: min(6, len(attachments))]
        )
        if most_relevant:
            lines.append(f"Most relevant files for the current request: {most_relevant}.")

    repeated_colors = _collect_repeated_attachment_terms(
        attachments,
        extractor=lambda attachment: (attachment.get("analysisMeta") or {}).get("dominantColorNames") or [],
        limit=6,
    )
    if repeated_colors:
        lines.append("Recurring colors across the file set: " + ", ".join(repeated_colors) + ".")

    repeated_objects = _collect_repeated_attachment_terms(
        attachments,
        extractor=lambda attachment: (attachment.get("analysisMeta") or {}).get("detectedObjects") or [],
        limit=8,
    )
    if repeated_objects:
        lines.append("Recurring visual entities: " + ", ".join(repeated_objects) + ".")

    repeated_keywords = _collect_repeated_attachment_terms(
        attachments,
        extractor=lambda attachment: attachment.get("analysisKeywords") or [],
        limit=10,
    )
    if repeated_keywords:
        lines.append("Cross-file themes: " + ", ".join(repeated_keywords) + ".")

    text_rich_files = [
        str(attachment.get("name") or "attachment")
        for attachment in attachments
        if str(attachment.get("analysisText") or attachment.get("content") or "").strip()
    ][:8]
    if text_rich_files:
        lines.append("Files with extracted readable content: " + ", ".join(text_rich_files) + ".")

    if clipped:
        lines.append(
            f"{clipped} additional attachments are still stored but were not expanded in this turn to keep the model stable."
        )

    return "\n".join(lines)


def _render_attachment_brief(
    attachment: Dict[str, Any],
    *,
    rank: int,
    query: str = "",
    compact: bool = False,
    detail_limit: int = MAX_ATTACHMENT_DOSSIER_ENTRY_CHARS,
) -> str:
    name = str(attachment.get("name") or "attachment")
    attachment_type = str(attachment.get("type") or "file")
    mime_type = str(attachment.get("mimeType") or attachment.get("mime_type") or "")
    summary = str(attachment.get("analysisSummary") or attachment.get("analysis_summary") or "").strip()
    analysis_text = str(attachment.get("analysisText") or attachment.get("analysis_text") or "").strip()
    analysis_caption = str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or "").strip()
    analysis_handwriting = str(attachment.get("analysisHandwriting") or attachment.get("analysis_handwriting") or "").strip()
    analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
    source_url = str(attachment.get("sourceUrl") or attachment.get("source_url") or "").strip()
    keywords = [str(keyword) for keyword in (attachment.get("analysisKeywords") or attachment.get("analysis_keywords") or []) if str(keyword).strip()]
    content = str(attachment.get("content") or "").strip()
    relevance_score = float(attachment.get("_relevanceScore") or 0.0)
    relevance_confidence = str(attachment.get("_relevanceConfidence") or _score_to_confidence_label(relevance_score))

    lines = [f"[Attachment {rank}: {name}]", f"Type: {attachment_type}"]
    if mime_type:
        lines.append(f"MIME: {mime_type}")
    lines.append(f"Reference: [{_attachment_reference_label(attachment, rank)}]")
    lines.append(f"Relevance confidence: {relevance_confidence} ({relevance_score:.1f})")
    if source_url:
        lines.append(f"Source URL: {source_url}")
    if summary:
        lines.append(f"Summary: {summary}")
    if keywords:
        lines.append(f"Keywords: {', '.join(keywords[:12])}")

    evidence_records = _select_attachment_evidence_records(attachment, query, limit=3)

    if compact:
        excerpt = (
            str(evidence_records[0].get("text") or "").strip()
            if evidence_records
            else (analysis_text or content or analysis_caption or analysis_handwriting)
        )
        if excerpt:
            prefix = f"[{evidence_records[0]['reference']}] " if evidence_records else ""
            lines.append(f"Key excerpt: {prefix}{_compact_text(excerpt, limit=max(220, detail_limit - 240))}")
        return "\n".join(lines)

    if analysis_caption:
        lines.append(f"Visual description:\n{_compact_text(analysis_caption, limit=1200)}")
    if analysis_handwriting:
        lines.append(f"Handwriting guess:\n{_compact_text(analysis_handwriting, limit=2000)}")
    meta_summary = _render_visual_metadata_summary(analysis_meta if isinstance(analysis_meta, dict) else {})
    if meta_summary:
        lines.append(f"Visual diagnostics:\n{meta_summary}")
    if evidence_records:
        evidence_lines = [
            f"- [{record['reference']}] {_compact_text(str(record.get('text') or ''), limit=max(400, detail_limit - 900))}"
            for record in evidence_records
        ]
        lines.append("Relevant evidence:\n" + "\n".join(evidence_lines))
    elif analysis_text:
        lines.append(f"Extracted content:\n{_compact_text(analysis_text, limit=max(600, detail_limit - 700))}")
    elif attachment_type == "file" and content:
        lines.append(f"Text content:\n{_compact_text(content, limit=max(600, detail_limit - 700))}")
    elif attachment_type == "binary" and content.startswith("data:"):
        lines.append("Binary file attached with no text extracted.")

    return "\n".join(lines)


def _build_attachment_dossier(message: str, attachments: list, *, history_mode: bool = False) -> Dict[str, Any]:
    if not attachments:
        return {
            "prepared": [],
            "ranked": [],
            "native_media": [],
            "dossier": "",
            "clipped": 0,
            "evidence": [],
            "conflicts": [],
            "confidence": {"label": "low", "score": 0.0, "reason": "no attachments"},
        }

    prepared, clipped = _prepare_attachments_for_reasoning(
        attachments,
        limit=MAX_REASONING_ATTACHMENTS_HISTORY if history_mode else MAX_REASONING_ATTACHMENTS,
    )
    ranked = _rank_attachments_for_query(message, prepared)

    native_media_limit = (
        MAX_REASONING_NATIVE_MEDIA_ATTACHMENTS_HISTORY
        if history_mode else MAX_REASONING_NATIVE_MEDIA_ATTACHMENTS
    )
    native_media: List[Dict[str, Any]] = []
    for attachment in ranked:
        attachment_type = str(attachment.get("type") or "").strip().lower()
        mime_type = str(attachment.get("mimeType") or attachment.get("mime_type") or "").strip().lower()
        if attachment_type == "image" or mime_type == "application/pdf":
            native_media.append(attachment)
        if len(native_media) >= native_media_limit:
            break

    summary_requested = _attachment_summary_requested(message)
    evidence_records = _collect_cross_attachment_evidence(
        message,
        ranked,
        limit=(12 if summary_requested and not history_mode else (8 if not history_mode else 4)),
    )
    conflicts = _detect_attachment_conflicts(message, ranked)
    confidence = _summarize_attachment_confidence(
        message,
        ranked,
        evidence_records,
        conflicts,
        clipped=clipped,
    )

    char_budget = MAX_ATTACHMENT_DOSSIER_CHARS_HISTORY if history_mode else MAX_ATTACHMENT_DOSSIER_CHARS
    detail_quota = 4 if history_mode else (24 if summary_requested else 18)
    multi_file_note = ""
    if len(ranked) > 1 and not history_mode:
        file_types = set(str(a.get("type") or "file") for a in ranked)
        file_names = ", ".join(str(a.get("name") or "file") for a in ranked[:8])
        multi_file_note = (
            f"MULTI-FILE ANALYSIS ({len(ranked)} files: {file_names}):\n"
            "You MUST cross-reference ALL attached files and treat them as related context. "
            "If the user asks for a summary, compile ONE unified summary across all files. "
            "If files contain complementary information (e.g. a PDF spec + an image mockup), synthesize them together. "
            "If files contain code, analyze all files as a codebase. "
            "If files conflict, note the discrepancy."
        )
    lines = [
        "ATTACHMENT DOSSIER:",
        multi_file_note,
        _render_attachment_synthesis(message, ranked, clipped=clipped),
        f"Evidence confidence for this turn: {confidence['label']} ({confidence['score']}/100) based on {confidence['reason']}.",
    ]
    if evidence_records:
        evidence_lines = [
            f"- [{record['reference']}] ({record['attachment_name']}, {record['confidence']}) "
            f"{_compact_text(str(record.get('text') or ''), limit=260)}"
            for record in evidence_records[: (4 if history_mode else 8)]
        ]
        lines.append("TOP SUPPORTING EVIDENCE:\n" + "\n".join(evidence_lines))
    if conflicts:
        lines.append("POTENTIAL CONFLICTS:\n" + "\n".join(f"- {line}" for line in conflicts))
    total_chars = len("\n\n".join(lines))
    used = 0
    for index, attachment in enumerate(ranked):
        remaining = char_budget - total_chars
        if remaining < 240:
            break
        compact = index >= detail_quota or remaining < 1800
        entry_budget = min(
            MAX_ATTACHMENT_DOSSIER_ENTRY_CHARS if not compact else 700,
            max(260, remaining - 120),
        )
        brief = _render_attachment_brief(
            attachment,
            rank=index + 1,
            query=message,
            compact=compact,
            detail_limit=entry_budget,
        )
        if len(brief) > remaining:
            brief = _compact_text(brief, limit=max(220, remaining - 20))
        lines.append(brief)
        total_chars += len(brief) + 2
        used = index + 1

    if used < len(ranked):
        remaining_names = ", ".join(
            str(attachment.get("name") or "attachment")
            for attachment in ranked[used: used + 20]
        )
        extra_count = len(ranked) - used
        lines.append(
            f"Additional analyzed attachments ({extra_count} more): {remaining_names}"
            + (" ..." if extra_count > 20 else "")
        )

    dossier = "\n\n".join(line for line in lines if line).strip()
    return {
        "prepared": prepared,
        "ranked": ranked,
        "native_media": native_media,
        "dossier": dossier,
        "clipped": clipped,
        "evidence": evidence_records,
        "conflicts": conflicts,
        "confidence": confidence,
    }


def _attachment_has_analysis(att: Dict[str, Any]) -> bool:
    return any([
        str(att.get("analysisText") or att.get("analysis_text") or "").strip(),
        str(att.get("analysisCaption") or att.get("analysis_caption") or "").strip(),
        str(att.get("analysisHandwriting") or att.get("analysis_handwriting") or "").strip(),
        bool(att.get("analysisMeta") or att.get("analysis_meta") or {}),
        str(att.get("analysisSummary") or att.get("analysis_summary") or "").strip(),
        bool(att.get("analysisKeywords") or att.get("analysis_keywords") or []),
    ])


def _attachment_cache_payload(att: Dict[str, Any]) -> Dict[str, Any]:
    chunk_source = str(att.get("analysisText") or att.get("analysis_text") or "").strip()
    if not chunk_source and str(att.get("type") or "").strip().lower() == "file":
        chunk_source = str(att.get("content") or "").strip()
    analysis_chunks = _split_attachment_text_into_chunks(chunk_source)
    return {
        "type": str(att.get("type") or "file"),
        "mimeType": str(att.get("mimeType") or att.get("mime_type") or ""),
        "analysisText": _compact_text(str(att.get("analysisText") or att.get("analysis_text") or ""), limit=MAX_STORED_ANALYSIS_TEXT_CHARS),
        "analysisCaption": _compact_text(str(att.get("analysisCaption") or att.get("analysis_caption") or ""), limit=1200),
        "analysisHandwriting": _compact_text(str(att.get("analysisHandwriting") or att.get("analysis_handwriting") or ""), limit=2000),
        "analysisMeta": copy.deepcopy(att.get("analysisMeta") or att.get("analysis_meta") or {}),
        "analysisSummary": _compact_text(str(att.get("analysisSummary") or att.get("analysis_summary") or ""), limit=MAX_STORED_ANALYSIS_SUMMARY_CHARS),
        "analysisKeywords": [str(keyword).strip() for keyword in (att.get("analysisKeywords") or att.get("analysis_keywords") or []) if str(keyword).strip()][:24],
        "analysisChunks": analysis_chunks,
        "analysisChunkCount": len(analysis_chunks),
    }


def _build_attachment_fingerprint(att: Dict[str, Any]) -> str:
    fingerprint_input = {
        "name": str(att.get("name") or "").strip(),
        "type": str(att.get("type") or "file").strip().lower(),
        "mimeType": _infer_attachment_mime_type(
            str(att.get("name") or ""),
            str(att.get("mimeType") or att.get("mime_type") or ""),
            str(att.get("sourceUrl") or att.get("source_url") or ""),
        ),
        "sourceUrl": str(att.get("sourceUrl") or att.get("source_url") or "").strip(),
        "size": int(att.get("size") or 0),
        "nameSuffix": Path(str(att.get("name") or "")).suffix.lower(),
        "content": str(att.get("content") or ""),
    }
    encoded = json.dumps(fingerprint_input, sort_keys=True, ensure_ascii=True).encode("utf-8", errors="ignore")
    return hashlib.sha256(encoded).hexdigest()


def _split_attachment_text_into_chunks(
    text: str,
    *,
    chunk_chars: int = ATTACHMENT_CHUNK_CHARS,
    overlap: int = ATTACHMENT_CHUNK_OVERLAP,
    limit: int = MAX_ATTACHMENT_CACHE_CHUNKS,
) -> List[str]:
    normalized = re.sub(r"\s+", " ", (text or "")).strip()
    if not normalized:
        return []
    if len(normalized) <= chunk_chars:
        return [normalized]

    chunks: List[str] = []
    start = 0
    step = max(200, chunk_chars - overlap)
    while start < len(normalized) and len(chunks) < limit:
        end = min(len(normalized), start + chunk_chars)
        chunk = normalized[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(normalized):
            break
        start += step
    return chunks


def _get_attachment_cached_chunks(att: Dict[str, Any]) -> List[str]:
    fingerprint = str(att.get("attachmentFingerprint") or att.get("attachment_fingerprint") or "").strip()
    if not fingerprint:
        return []
    cached_payload = attachment_analysis_cache.get(fingerprint) or {}
    return [
        str(chunk).strip()
        for chunk in (cached_payload.get("analysisChunks") or [])
        if str(chunk).strip()
    ]


def _select_attachment_evidence_snippets(attachment: Dict[str, Any], query: str, limit: int = 3) -> List[str]:
    chunks = _get_attachment_cached_chunks(attachment)
    if not chunks:
        fallback_text = str(attachment.get("analysisText") or attachment.get("content") or "").strip()
        return [fallback_text] if fallback_text else []
    if not (query or "").strip():
        return chunks[:limit]

    scored: List[tuple[float, int, str]] = []
    for index, chunk in enumerate(chunks):
        score = _score_text_match(query, chunk)
        if score <= 0:
            continue
        scored.append((score, index, chunk))
    if not scored:
        return chunks[:limit]
    scored.sort(key=lambda item: (-item[0], item[1]))
    return [item[2] for item in scored[:limit]]


def _select_attachment_summary_records(attachment: Dict[str, Any], limit: int = 5) -> List[Dict[str, Any]]:
    chunks = _get_attachment_cached_chunks(attachment)
    records: List[Dict[str, Any]] = []
    reference_prefix = _attachment_reference_label(attachment)

    if chunks:
        indexes = _sample_document_indexes(len(chunks), min(limit, len(chunks)))
        for index in indexes:
            chunk = str(chunks[index] or "").strip()
            if not chunk:
                continue
            richness = min(2.0, len(chunk) / 900.0)
            records.append({
                "text": chunk,
                "chunk_index": index + 1,
                "score": 6.5 + richness,
                "confidence": "medium",
                "reference": f"{reference_prefix} §{index + 1}",
            })
        if records:
            return records[:limit]

    fallback_text = str(
        attachment.get("analysisText")
        or attachment.get("content")
        or attachment.get("contentSnippet")
        or attachment.get("analysisCaption")
        or attachment.get("analysisHandwriting")
        or ""
    ).strip()
    if fallback_text:
        records.append({
            "text": fallback_text,
            "chunk_index": 1,
            "score": 6.0,
            "confidence": "medium",
            "reference": f"{reference_prefix} §1",
        })
    return records[:limit]


def _attachment_reference_label(attachment: Dict[str, Any], rank: Optional[int] = None) -> str:
    attachment_rank = rank or int(attachment.get("_dossierRank") or 0)
    return f"Attachment {attachment_rank}" if attachment_rank > 0 else "Attachment"


def _score_to_confidence_label(score: float) -> str:
    if score >= 14:
        return "high"
    if score >= 7:
        return "medium"
    return "low"


def _select_attachment_evidence_records(attachment: Dict[str, Any], query: str, limit: int = 3) -> List[Dict[str, Any]]:
    if _attachment_summary_requested(query):
        return _select_attachment_summary_records(attachment, limit=limit)

    chunks = _get_attachment_cached_chunks(attachment)
    records: List[Dict[str, Any]] = []
    reference_prefix = _attachment_reference_label(attachment)

    if chunks:
        scored: List[tuple[float, int, str]] = []
        for index, chunk in enumerate(chunks):
            score = _score_text_match(query, chunk) if (query or "").strip() else max(0.2, 1.0 - (index * 0.08))
            if score <= 0 and (query or "").strip():
                continue
            scored.append((score, index, chunk))
        if not scored:
            scored = [(0.1, index, chunk) for index, chunk in enumerate(chunks[:limit])]
        scored.sort(key=lambda item: (-item[0], item[1]))
        for score, index, chunk in scored[:limit]:
            records.append({
                "text": chunk,
                "chunk_index": index + 1,
                "score": score,
                "confidence": _score_to_confidence_label(score),
                "reference": f"{reference_prefix} §{index + 1}",
            })
        return records

    fallback_text = str(
        attachment.get("analysisText")
        or attachment.get("content")
        or attachment.get("contentSnippet")
        or attachment.get("analysisCaption")
        or attachment.get("analysisHandwriting")
        or ""
    ).strip()
    if fallback_text:
        score = _score_text_match(query, fallback_text) if (query or "").strip() else 0.5
        records.append({
            "text": fallback_text,
            "chunk_index": 1,
            "score": score,
            "confidence": _score_to_confidence_label(score),
            "reference": f"{reference_prefix} §1",
        })
    return records[:limit]


def _collect_cross_attachment_evidence(query: str, attachments: List[Dict[str, Any]], limit: int = 8) -> List[Dict[str, Any]]:
    collected: List[tuple[float, int, int, Dict[str, Any]]] = []
    for attachment in attachments:
        rank = int(attachment.get("_dossierRank") or 0)
        for record in _select_attachment_evidence_records(attachment, query, limit=3):
            collected.append((float(record.get("score") or 0.0), rank, int(record.get("chunk_index") or 1), {
                **record,
                "attachment_name": str(attachment.get("name") or "attachment"),
                "attachment_rank": rank,
            }))
    collected.sort(key=lambda item: (-item[0], item[1], item[2]))
    return [item[3] for item in collected[:limit]]


def _detect_attachment_conflicts(message: str, attachments: List[Dict[str, Any]], limit: int = 3) -> List[str]:
    if len(attachments) < 2 or _visual_comparison_requested(message):
        return []

    conflicts: List[str] = []
    ranked = attachments[: min(6, len(attachments))]
    exact_text_requested = any(token in (message or "").lower() for token in (
        "exact text",
        "exact sentence",
        "reply with the exact",
        "what text",
        "what does this say",
        "what does the handwriting say",
    ))

    if exact_text_requested:
        seen_snippets: Dict[str, Dict[str, Any]] = {}
        for attachment in ranked:
            record = next(iter(_select_attachment_evidence_records(attachment, message, limit=1)), None)
            if not record:
                continue
            normalized = re.sub(r"\s+", " ", str(record.get("text") or "").strip()).lower()
            if not normalized:
                continue
            if normalized not in seen_snippets:
                seen_snippets[normalized] = {
                    "reference": record.get("reference"),
                    "text": _compact_text(str(record.get("text") or ""), limit=100),
                }
            if len(seen_snippets) >= 2:
                values = list(seen_snippets.values())[:2]
                conflicts.append(
                    f'Potential text conflict: [{values[0]["reference"]}] says "{values[0]["text"]}" while '
                    f'[{values[1]["reference"]}] says "{values[1]["text"]}".'
                )
                break

    query_terms = _extract_visual_request_terms(message)
    if _visual_review_requested(message) or query_terms["colors"] or query_terms["focusTerms"]:
        color_mentions: List[tuple[str, str]] = []
        for attachment in ranked:
            colors = _collect_present_visual_colors(attachment)
            if not colors:
                continue
            color_mentions.append((_attachment_reference_label(attachment), colors[0]))
        distinct_colors = list(dict.fromkeys(color for _, color in color_mentions))
        if len(distinct_colors) >= 2:
            first_ref, first_color = color_mentions[0]
            second_ref, second_color = next(
                ((ref, color) for ref, color in color_mentions[1:] if color != first_color),
                color_mentions[1],
            )
            conflicts.append(
                f"Potential visual conflict: [{first_ref}] leans {first_color}, while [{second_ref}] leans {second_color}."
            )

    return conflicts[:limit]


def _summarize_attachment_confidence(
    message: str,
    attachments: List[Dict[str, Any]],
    evidence_records: List[Dict[str, Any]],
    conflicts: List[str],
    clipped: int = 0,
) -> Dict[str, Any]:
    top_scores = [float(attachment.get("_relevanceScore") or 0.0) for attachment in attachments[:4]]
    mean_score = (sum(top_scores) / len(top_scores)) if top_scores else 0.0
    corroborating_files = len({int(record.get("attachment_rank") or 0) for record in evidence_records if int(record.get("attachment_rank") or 0) > 0})
    evidence_strength = sum(float(record.get("score") or 0.0) for record in evidence_records[:4])
    confidence_value = min(
        100.0,
        max(
            0.0,
            mean_score * 5.0
            + corroborating_files * 10.0
            + min(len(evidence_records), 6) * 4.0
            + min(evidence_strength, 18.0) * 1.5
            - len(conflicts) * 14.0
            - (8.0 if clipped and len(evidence_records) < 3 else 0.0),
        ),
    )

    if confidence_value >= 72:
        label = "high"
    elif confidence_value >= 46:
        label = "medium"
    else:
        label = "low"

    reasons: List[str] = []
    if evidence_records:
        reasons.append(f"{len(evidence_records)} evidence snippet(s)")
    if corroborating_files:
        reasons.append(f"{corroborating_files} corroborating file(s)")
    if conflicts:
        reasons.append(f"{len(conflicts)} conflict warning(s)")
    if clipped:
        reasons.append(f"{clipped} clipped attachment(s)")

    return {
        "label": label,
        "score": round(confidence_value, 1),
        "reason": ", ".join(reasons) if reasons else "limited extracted evidence",
    }


def _sample_document_indexes(total_count: int, max_items: int) -> List[int]:
    """Sample indexes across an entire document so later sections are still considered."""
    if total_count <= 0 or max_items <= 0:
        return []
    if total_count <= max_items:
        return list(range(total_count))

    chosen: List[int] = []
    seen = set()
    denominator = max(1, max_items - 1)
    for slot in range(max_items):
        index = int(round(slot * (total_count - 1) / denominator))
        if index in seen:
            continue
        chosen.append(index)
        seen.add(index)

    candidate = 0
    while len(chosen) < max_items and candidate < total_count:
        if candidate not in seen:
            chosen.append(candidate)
            seen.add(candidate)
        candidate += 1

    return sorted(chosen)


def _sanitize_chat_attachment(att: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize and retain enough attachment data for chat history previews."""
    raw_type = str(att.get("type") or "file").strip().lower()
    attachment_type = raw_type if raw_type in {"image", "file", "binary"} else "file"
    content = str(att.get("content") or "")
    truncated = bool(att.get("truncated", False))
    source_url = str(att.get("source_url") or att.get("sourceUrl") or "")[:2000]
    analysis_text = _compact_text(str(att.get("analysis_text") or att.get("analysisText") or ""), limit=MAX_STORED_ANALYSIS_TEXT_CHARS)
    analysis_caption = _compact_text(str(att.get("analysis_caption") or att.get("analysisCaption") or ""), limit=600)
    analysis_handwriting = _compact_text(str(att.get("analysis_handwriting") or att.get("analysisHandwriting") or ""), limit=2000)
    analysis_meta = att.get("analysis_meta") or att.get("analysisMeta") or {}
    analysis_summary = _compact_text(str(att.get("analysis_summary") or att.get("analysisSummary") or ""), limit=MAX_STORED_ANALYSIS_SUMMARY_CHARS)
    raw_keywords = att.get("analysis_keywords") or att.get("analysisKeywords") or []
    analysis_keywords = [str(keyword).strip() for keyword in raw_keywords if str(keyword).strip()][:16]
    analysis_chunk_count = int(att.get("analysis_chunk_count") or att.get("analysisChunkCount") or 0)
    attachment_fingerprint = str(att.get("attachment_fingerprint") or att.get("attachmentFingerprint") or "").strip()

    if attachment_type == "image":
        if not content.startswith("data:image/"):
            content = ""
        if len(content) > MAX_STORED_IMAGE_ATTACHMENT_CHARS:
            content = content[:MAX_STORED_IMAGE_ATTACHMENT_CHARS]
            truncated = True
    elif attachment_type == "binary":
        if len(content) > MAX_STORED_BINARY_ATTACHMENT_CHARS:
            content = content[:MAX_STORED_BINARY_ATTACHMENT_CHARS]
            truncated = True
    else:
        if len(content) > MAX_STORED_TEXT_ATTACHMENT_CHARS:
            content = content[:MAX_STORED_TEXT_ATTACHMENT_CHARS]
            truncated = True

    return {
        "name": str(att.get("name") or "file")[:240],
        "type": attachment_type,
        "mimeType": str(att.get("mime_type") or att.get("mimeType") or "")[:120],
        "size": int(att.get("size") or 0),
        "truncated": truncated,
        "content": content,
        "sourceUrl": source_url,
        "analysisText": analysis_text,
        "analysisCaption": analysis_caption,
        "analysisHandwriting": analysis_handwriting,
        "analysisMeta": analysis_meta if isinstance(analysis_meta, dict) else {},
        "analysisSummary": analysis_summary,
        "analysisKeywords": analysis_keywords,
        "analysisChunkCount": analysis_chunk_count,
        "attachmentFingerprint": attachment_fingerprint,
    }


def _prepare_chat_attachment(att: Dict[str, Any]) -> Dict[str, Any]:
    """Sanitize an attachment and enrich it with OCR when helpful."""
    sanitized = _sanitize_chat_attachment(att)
    sanitized["mimeType"] = _infer_attachment_mime_type(
        str(sanitized.get("name") or ""),
        str(sanitized.get("mimeType") or ""),
        str(sanitized.get("sourceUrl") or ""),
    )
    mime_type = str(sanitized.get("mimeType") or "")
    extension = _file_extension(str(sanitized.get("name") or "")) or _file_extension(str(sanitized.get("sourceUrl") or ""))
    fingerprint = _build_attachment_fingerprint(sanitized)
    if fingerprint:
        sanitized["attachmentFingerprint"] = fingerprint

    analysis_from_cache = False
    if fingerprint and not _attachment_has_analysis(sanitized):
        cached_payload = attachment_analysis_cache.get(fingerprint)
        if cached_payload:
            analysis_from_cache = True
            for field in ("analysisText", "analysisCaption", "analysisHandwriting", "analysisMeta", "analysisSummary", "analysisKeywords"):
                if field in cached_payload and cached_payload.get(field):
                    sanitized[field] = copy.deepcopy(cached_payload[field])
            sanitized["analysisChunkCount"] = int(cached_payload.get("analysisChunkCount") or len(cached_payload.get("analysisChunks") or []))

    if sanitized.get("type") == "image" and sanitized.get("content") and not analysis_from_cache and not sanitized.get("analysisText"):
        ocr_text = _extract_text_from_image_attachments([sanitized])
        if ocr_text:
            sanitized["analysisText"] = _compact_text(ocr_text, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)

    if sanitized.get("type") == "image" and sanitized.get("content") and not analysis_from_cache:
        structured_visual = _extract_structured_visual_analysis_from_image_attachments([sanitized])
        structured_text = str(structured_visual.get("analysisText") or "").strip()
        if structured_text:
            sanitized["analysisText"] = _merge_analysis_text(sanitized.get("analysisText") or "", structured_text, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
        structured_caption = str(structured_visual.get("analysisCaption") or "").strip()
        if structured_caption:
            sanitized["analysisCaption"] = _compact_text(structured_caption, limit=1200)
        structured_meta = structured_visual.get("analysisMeta") or {}
        if structured_meta:
            sanitized["analysisMeta"] = _merge_visual_metadata(sanitized.get("analysisMeta") or {}, structured_meta)

    if sanitized.get("type") == "image" and sanitized.get("content") and not analysis_from_cache and not sanitized.get("analysisCaption"):
        semantic_caption = _extract_semantic_descriptions_from_image_attachments([sanitized])
        if semantic_caption:
            sanitized["analysisCaption"] = _compact_text(semantic_caption, limit=600)

    if sanitized.get("type") == "image" and sanitized.get("content") and not analysis_from_cache and not sanitized.get("analysisHandwriting"):
        handwriting_guess = _extract_handwriting_from_image_attachments([sanitized])
        if handwriting_guess:
            sanitized["analysisHandwriting"] = _compact_text(handwriting_guess, limit=2000)

    if sanitized.get("type") == "image" and sanitized.get("content") and not analysis_from_cache:
        visual_meta = _extract_visual_metadata_from_image_attachments([sanitized])
        if visual_meta:
            sanitized["analysisMeta"] = _merge_visual_metadata(sanitized.get("analysisMeta") or {}, visual_meta)

    if sanitized.get("type") == "binary" and not analysis_from_cache and not sanitized.get("analysisText"):
        _, base64_data = _split_base64_data_url(str(sanitized.get("content") or ""))
        if base64_data:
            try:
                binary_bytes = base64.b64decode(base64_data)
            except Exception:
                binary_bytes = b""
            if binary_bytes:
                if mime_type == "application/pdf":
                    sanitized["analysisText"] = _extract_pdf_text_from_bytes(binary_bytes)
                    if len(str(sanitized.get("analysisText") or "").strip()) < 120:
                        scanned_pdf = _extract_scanned_pdf_analysis_from_bytes(binary_bytes)
                        if scanned_pdf.get("analysisText"):
                            if sanitized.get("analysisText"):
                                sanitized["analysisText"] = _merge_analysis_text(
                                    str(sanitized.get("analysisText") or ""),
                                    str(scanned_pdf.get("analysisText") or ""),
                                    limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS,
                                )
                            else:
                                sanitized["analysisText"] = scanned_pdf.get("analysisText")
                            if scanned_pdf.get("analysisCaption"):
                                sanitized["analysisCaption"] = scanned_pdf.get("analysisCaption")
                            if scanned_pdf.get("analysisHandwriting"):
                                sanitized["analysisHandwriting"] = scanned_pdf.get("analysisHandwriting")
                            if scanned_pdf.get("analysisMeta"):
                                sanitized["analysisMeta"] = _merge_visual_metadata(
                                    sanitized.get("analysisMeta") or {},
                                    scanned_pdf.get("analysisMeta") or {},
                                )
                elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                    sanitized["analysisText"] = _extract_docx_text_from_bytes(binary_bytes)
                elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    sanitized["analysisText"] = _extract_xlsx_text_from_bytes(binary_bytes)
                elif mime_type == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
                    sanitized["analysisText"] = _extract_pptx_text_from_bytes(binary_bytes)
                elif mime_type == "application/epub+zip":
                    sanitized["analysisText"] = _extract_epub_text_from_bytes(binary_bytes)
                elif mime_type in {"application/zip", "application/x-zip-compressed"}:
                    sanitized["analysisText"] = _extract_archive_manifest_from_bytes(binary_bytes, label="ZIP archive")
                elif mime_type in {"text/csv", "application/csv", "application/vnd.ms-excel"}:
                    sanitized["analysisText"] = _extract_delimited_text_from_bytes(binary_bytes, delimiter=",", label="CSV")
                elif mime_type == "text/tab-separated-values":
                    sanitized["analysisText"] = _extract_delimited_text_from_bytes(binary_bytes, delimiter="\t", label="TSV")

    if sanitized.get("analysisText"):
        sanitized["analysisText"] = _compact_text(str(sanitized.get("analysisText") or ""), limit=MAX_STORED_ANALYSIS_TEXT_CHARS)
    if sanitized.get("analysisCaption"):
        sanitized["analysisCaption"] = _compact_text(str(sanitized.get("analysisCaption") or ""), limit=600)
    if sanitized.get("analysisHandwriting"):
        sanitized["analysisHandwriting"] = _compact_text(str(sanitized.get("analysisHandwriting") or ""), limit=2000)

    if sanitized.get("type") == "file" and not sanitized.get("analysisText"):
        file_content = str(sanitized.get("content") or "")
        if mime_type == "text/html":
            sanitized["analysisText"] = _extract_html_text(file_content)
        elif mime_type == "text/markdown" or extension in _MARKDOWN_EXTENSIONS:
                sanitized["analysisText"] = _extract_markdown_text(file_content)
        elif mime_type in {"application/json", "text/json"} or extension == "json":
            try:
                sanitized["analysisText"] = _compact_text(json.dumps(json.loads(file_content), indent=2), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
            except Exception:
                sanitized["analysisText"] = _compact_text(file_content, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
        elif mime_type in {"application/xml", "text/xml", "image/svg+xml"} or extension in _XML_EXTENSIONS:
            sanitized["analysisText"] = _extract_xml_text(file_content)
        elif mime_type in {"text/csv", "application/csv"} or extension == "csv":
            sanitized["analysisText"] = _extract_delimited_text_from_bytes(file_content.encode("utf-8"), delimiter=",", label="CSV")
        elif mime_type == "text/tab-separated-values" or extension == "tsv":
            sanitized["analysisText"] = _extract_delimited_text_from_bytes(file_content.encode("utf-8"), delimiter="\t", label="TSV")

    if not sanitized.get("analysisSummary"):
        analysis_text = str(sanitized.get("analysisText") or "").strip()
        analysis_caption = str(sanitized.get("analysisCaption") or "").strip()
        analysis_handwriting = str(sanitized.get("analysisHandwriting") or "").strip()
        visual_meta_summary = _render_visual_metadata_summary(sanitized.get("analysisMeta") or {})
        if sanitized.get("type") == "image" and analysis_text:
            if analysis_caption:
                sanitized["analysisSummary"] = (
                    f"Image description: {_compact_text(analysis_caption, limit=110)} "
                    f"Visible text: {_compact_text(analysis_text, limit=110)}"
                )
            else:
                sanitized["analysisSummary"] = f"Image with detected text: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "image" and analysis_caption:
            summary = f"Image description: {_compact_text(analysis_caption, limit=160)}"
            if visual_meta_summary:
                summary += f" [{visual_meta_summary}]"
            sanitized["analysisSummary"] = summary
        elif sanitized.get("type") == "image" and analysis_handwriting:
            sanitized["analysisSummary"] = f"Image handwriting guess: {_compact_text(analysis_handwriting, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/pdf" and analysis_text:
            if analysis_handwriting:
                sanitized["analysisSummary"] = (
                    f"PDF with extracted text and handwriting guess: {_compact_text(analysis_handwriting, limit=90)}"
                )
            else:
                sanitized["analysisSummary"] = f"PDF with extracted text: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" and analysis_text:
            sanitized["analysisSummary"] = f"DOCX with extracted text: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" and analysis_text:
            sanitized["analysisSummary"] = f"XLSX with extracted rows: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/vnd.openxmlformats-officedocument.presentationml.presentation" and analysis_text:
            sanitized["analysisSummary"] = f"PPTX with extracted slide text: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/epub+zip" and analysis_text:
            sanitized["analysisSummary"] = f"EPUB with extracted chapters: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type in {"application/zip", "application/x-zip-compressed"} and analysis_text:
            sanitized["analysisSummary"] = f"Archive with extracted manifest: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type in {"text/csv", "application/csv", "application/vnd.ms-excel"} and analysis_text:
            sanitized["analysisSummary"] = f"CSV with extracted rows: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "text/tab-separated-values" and analysis_text:
            sanitized["analysisSummary"] = f"TSV with extracted rows: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "binary" and mime_type == "application/pdf":
            sanitized["analysisSummary"] = "PDF document attached."
        elif sanitized.get("type") == "binary" and mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            sanitized["analysisSummary"] = "Spreadsheet document attached."
        elif sanitized.get("type") == "binary" and mime_type == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
            sanitized["analysisSummary"] = "Presentation document attached."
        elif sanitized.get("type") == "binary" and mime_type == "application/epub+zip":
            sanitized["analysisSummary"] = "EPUB document attached."
        elif sanitized.get("type") == "binary" and mime_type in {"application/zip", "application/x-zip-compressed"}:
            sanitized["analysisSummary"] = "Archive attached."
        elif sanitized.get("type") == "binary":
            sanitized["analysisSummary"] = f"Binary attachment: {mime_type or 'application/octet-stream'}"
        elif sanitized.get("type") == "file" and mime_type == "text/html" and analysis_text:
            title, description = _extract_html_metadata(str(sanitized.get("content") or ""))
            sanitized["analysisSummary"] = title or description or "HTML document"
        elif sanitized.get("type") == "file" and (mime_type == "text/markdown" or extension in _MARKDOWN_EXTENSIONS):
            markdown_title = _extract_markdown_title(str(sanitized.get("content") or ""))
            sanitized["analysisSummary"] = markdown_title or "Markdown document"
        elif sanitized.get("type") == "file" and (mime_type in {"application/json", "text/json"} or extension == "json"):
            sanitized["analysisSummary"] = "JSON document"
        elif sanitized.get("type") == "file" and (mime_type in {"application/xml", "text/xml", "image/svg+xml"} or extension in _XML_EXTENSIONS):
            sanitized["analysisSummary"] = "XML document"
        elif sanitized.get("type") == "file" and (mime_type in {"text/csv", "application/csv"} or extension == "csv") and analysis_text:
            sanitized["analysisSummary"] = f"CSV with extracted rows: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "file" and (mime_type == "text/tab-separated-values" or extension == "tsv") and analysis_text:
            sanitized["analysisSummary"] = f"TSV with extracted rows: {_compact_text(analysis_text, limit=180)}"
        elif sanitized.get("type") == "file" and str(sanitized.get("content") or "").strip():
            sanitized["analysisSummary"] = f"Text/code file: {_compact_text(str(sanitized.get('content') or ''), limit=180)}"

    if not sanitized.get("analysisKeywords"):
        keyword_source = " ".join([
            str(sanitized.get("name") or ""),
            str(sanitized.get("mimeType") or ""),
            str(sanitized.get("analysisCaption") or ""),
            str(sanitized.get("analysisHandwriting") or ""),
            str(sanitized.get("analysisSummary") or ""),
            str(sanitized.get("analysisText") or ""),
        ]).lower()
        tokens: List[str] = []
        for token in re.findall(r"[a-zA-Z0-9_]{3,}", keyword_source):
            if token in {"the", "and", "with", "from", "that", "this", "image", "file", "text", "application", "plain", "code", "txt", "json", "markdown", "document"}:
                continue
            if token not in tokens:
                tokens.append(token)
            if len(tokens) >= 12:
                break
        sanitized["analysisKeywords"] = tokens

    if fingerprint and _attachment_has_analysis(sanitized):
        cached_payload = _attachment_cache_payload(sanitized)
        sanitized["analysisChunkCount"] = int(cached_payload.get("analysisChunkCount") or 0)
        attachment_analysis_cache.put(fingerprint, cached_payload)

    return sanitized


def _build_user_message_content(message: str, attachments: list, provider: str, history_mode: bool = False):
    """Build a provider-appropriate user message payload including attachments."""
    message = (message or "").strip()
    text_blocks: List[str] = []
    multimodal_parts: List[Dict[str, Any]] = []
    used_multimodal = False

    if message:
        text_blocks.append(message)

    compiled = _build_attachment_dossier(message, attachments, history_mode=history_mode)
    dossier_text = compiled.get("dossier") or ""
    if dossier_text:
        text_blocks.append(dossier_text)

    ranked_attachments = compiled.get("ranked") or []
    visual_alignment = _build_visual_alignment_report(message, ranked_attachments)
    visual_alignment_block = ""
    if visual_alignment:
        visual_alignment_block = "Request alignment review:\n" + visual_alignment

    for att in compiled.get("native_media") or []:
        name = str(att.get("name") or "file")
        atype = str(att.get("type") or "file").strip().lower()
        content = str(att.get("content") or "")
        if atype == "image":
            if _provider_supports_vision(provider) and content.startswith("data:image/"):
                used_multimodal = True
                multimodal_parts.append({
                    "type": "text",
                    "text": f"Attached image: {name}",
                })
                multimodal_parts.append({
                    "type": "image_url",
                    "image_url": {"url": content},
                })
            continue

    if used_multimodal:
        intro = "\n\n".join(block for block in text_blocks if block).strip()
        if intro:
            multimodal_parts.insert(0, {"type": "text", "text": intro})
        if visual_alignment_block:
            multimodal_parts.append({"type": "text", "text": visual_alignment_block})
        return multimodal_parts

    if visual_alignment_block:
        text_blocks.append(visual_alignment_block)
    combined = "\n\n".join(block for block in text_blocks if block).strip()
    return combined or "Analyze the attached content."


def _split_base64_data_url(value: str) -> tuple[str, str]:
    match = re.match(r"^data:([^;]+);base64,(.+)$", value or "", re.DOTALL)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def _build_gemini_generate_content_parts(message: str, attachments: list, history_mode: bool = False):
    media_parts: List[Dict[str, Any]] = []
    text_parts: List[Dict[str, Any]] = []
    message = (message or "").strip()
    compiled = _build_attachment_dossier(message, attachments, history_mode=history_mode)
    dossier_text = compiled.get("dossier") or ""
    ranked_attachments = compiled.get("ranked") or []
    visual_alignment = _build_visual_alignment_report(message, ranked_attachments)

    for att in compiled.get("native_media") or []:
        name = str(att.get("name") or "file")
        atype = str(att.get("type") or "file").strip().lower()
        mime_type = str(att.get("mimeType") or att.get("mime_type") or "")
        content = str(att.get("content") or "")

        if atype == "image":
            parsed_mime, base64_data = _split_base64_data_url(content)
            if base64_data:
                media_parts.append({
                    "inline_data": {
                        "mime_type": mime_type or parsed_mime or "image/png",
                        "data": base64_data,
                    }
                })
            else:
                text_parts.append({"text": f"[Attached image: {name}]"})
            continue

        if atype == "binary":
            parsed_mime, base64_data = _split_base64_data_url(content)
            if mime_type == "application/pdf" and base64_data:
                media_parts.append({
                    "inline_data": {
                        "mime_type": mime_type,
                        "data": base64_data,
                    }
                })
                continue

    if message:
        text_parts.insert(0, {"text": message})
    if dossier_text:
        text_parts.append({"text": dossier_text})
    if visual_alignment:
        text_parts.append({"text": "Request alignment review:\n" + visual_alignment})

    parts = media_parts + text_parts
    if not parts:
        return [{"text": "Analyze the attached content."}]
    return parts


def _build_gemini_generate_content_input(message: str, attachments: list, history: list):
    conversation: List[Dict[str, Any]] = []

    for item in history[-10:]:
        role = item.get("role", "user")
        if role not in ("user", "assistant"):
            continue
        parts = _build_gemini_generate_content_parts(
            str(item.get("content") or ""),
            item.get("attachments") or [],
            history_mode=True,
        ) if role == "user" else [{"text": str(item.get("content") or "")}]
        if not parts:
            continue
        conversation.append({
            "role": "model" if role == "assistant" else "user",
            "parts": parts,
        })

    conversation.append({
        "role": "user",
        "parts": _build_gemini_generate_content_parts(message, attachments),
    })
    return conversation


def _extract_gemini_generate_content_text(payload: Dict[str, Any]) -> str:
    candidates = payload.get("candidates") or []
    text_chunks: List[str] = []
    for candidate in candidates:
        content = candidate.get("content") or {}
        for part in content.get("parts") or []:
            text = str(part.get("text") or "").strip()
            if text:
                text_chunks.append(text)
    return "\n".join(text_chunks).strip()


def _gemini_generate_content_request(model: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is not configured")

    request_obj = urllib.request.Request(
        GEMINI_GENERATE_CONTENT_URL.format(model=urllib.parse.quote(model, safe="")),
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "x-goog-api-key": api_key,
        },
        method="POST",
    )

    attempts = HOSTED_LLM_MAX_RETRIES + 1
    last_error: Optional[Exception] = None

    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(request_obj, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            wrapped = RuntimeError(detail or str(exc))
            setattr(wrapped, "code", exc.code)
            setattr(wrapped, "headers", exc.headers)
            last_error = wrapped
            if attempt >= attempts - 1 or not _is_retryable_llm_error(wrapped):
                raise wrapped from exc
            delay = _compute_retry_delay_seconds(wrapped, attempt)
            logger.warning(
                "Transient Gemini generateContent failure on attempt %s/%s. Retrying in %.2fs: %s",
                attempt + 1,
                attempts,
                delay,
                detail or exc,
            )
            time.sleep(delay)
        except Exception as exc:
            last_error = exc
            if attempt >= attempts - 1 or not _is_retryable_llm_error(exc):
                raise
            delay = _compute_retry_delay_seconds(exc, attempt)
            logger.warning(
                "Transient Gemini transport failure on attempt %s/%s. Retrying in %.2fs: %s",
                attempt + 1,
                attempts,
                delay,
                exc,
            )
            time.sleep(delay)

    if last_error is not None:
        raise last_error
    raise RuntimeError("Gemini generateContent failed before a response could be returned.")


def _gemini_generate_content_chat(message: str, attachments: list, history: list, chat_title: str, chat_memory: str, model: str) -> Dict[str, Any]:
    # ── Pre-fetch UEFN context so Gemini can act, not just describe ──
    extra_context = ""
    port = discover_uefn_listener_port()
    if port:
        try:
            r_actors = _chat_uefn_query("get_all_actors", {})
            if r_actors.get("success"):
                actors_raw = r_actors.get("result", [])
                actors_list = actors_raw
                if isinstance(actors_raw, dict) and "actors" in actors_raw:
                    actors_list = actors_raw["actors"]
                if isinstance(actors_list, list):
                    skip_classes = {"WorldDataLayers", "WorldSettings", "LevelBounds",
                                    "NavigationData", "GameMode", "GameState", "PlayerStart",
                                    "AbstractNavData", "RecastNavMesh", "PostProcessVolume"}
                    structural = [a for a in actors_list
                                  if str(a.get("class", "")) not in skip_classes]
                    actor_summaries = []
                    for a in structural[:60]:
                        actor_summaries.append({
                            "label": a.get("label", ""),
                            "class": a.get("class", ""),
                            "loc": a.get("location", {}),
                            "rot": a.get("rotation", {}),
                            "scale": a.get("scale", {}),
                        })
                    extra_context += (
                        f"\n\n[LIVE UEFN DATA - {len(structural)} actors with transforms:\n"
                        f"{json.dumps(actor_summaries, default=str)[:8000]}]"
                    )
                else:
                    extra_context += f"\n\n[LIVE UEFN DATA:\n{json.dumps(actors_raw, default=str)[:4000]}]"
        except Exception:
            pass
        try:
            r_sel = _chat_uefn_query("get_selected_actors", {})
            if r_sel.get("success"):
                sel = r_sel.get("result", [])
                if isinstance(sel, list) and sel:
                    extra_context += f"\n\n[CURRENTLY SELECTED ACTORS ({len(sel)}):\n{json.dumps(sel[:20], default=str)[:2000]}]"
        except Exception:
            pass
        try:
            r_vp = _chat_uefn_query("get_viewport_camera", {})
            if r_vp.get("success"):
                extra_context += f"\n\n[VIEWPORT CAMERA:\n{json.dumps(r_vp.get('result', {}), default=str)[:500]}]"
        except Exception:
            pass

    matching_tools = _search_tools_for_llm(message, limit=10)
    if matching_tools:
        extra_context += f"\n\n[TOOLS MOST RELEVANT TO YOUR REQUEST:\n{json.dumps(matching_tools, default=str)[:4000]}]"

    action_instruction = """

You CAN and MUST take action using action blocks. Include these in your response to execute changes:

```action
{"action": "execute_python_in_uefn", "code": "import unreal\\n...your code...\\nresult = 'done'"}
```

Or to run a registered tool:
```action
{"action": "run_uefn_tool", "tool_name": "tool_id", "parameters": {}}
```

Or to import uploaded FBX/ZIP model attachments and place them intelligently:
```action
{"action": "import_attached_models", "request": "import these models and place them along the terrain spline"}
```

RULES:
- When the user asks to FIX, BUILD, DELETE, MOVE, COLOR anything — include action blocks. Do NOT just describe steps.
- NEVER say "select the actors first" or "I need you to..." — use the LIVE UEFN DATA above to find actors yourself.
- Keep your text response to 1-3 sentences, then action blocks. No essays or tutorials.
- NEVER autocorrect or rewrite the user's message.

STRUCTURAL FIX GUIDE (for holes, clipping, misalignment):
- CRITICAL: Use the EXACT actor labels from LIVE UEFN DATA. Do NOT invent labels.
- NEVER write conditional code like "if scale.z < 0.3". ALWAYS apply direct unconditional changes.
- For roofs: left/right should have matching scales with mirrored pitch angles.
- For walls: edges must align — set matching X/Y coordinates.

3D CONSTRUCTION — HOW TO BUILD IN UEFN:
Use /Engine/BasicShapes/Cube as building blocks. 1 Unreal unit = 1cm. Scale 1.0 = 1 meter.

WALL: Calculate length from two points, angle from atan2. Position at midpoint, scale to (length/100, thickness/100, height/100), rotate yaw=angle_deg.
FLOOR: Flat cube at ground level, scale to (width/100, depth/100, thickness/100).
GABLE ROOF: Two slopes meeting at ridge. pitch_angle = degrees(atan2(roof_height, depth/2)). slope_length = sqrt((depth/2)^2 + roof_height^2). Left slope rotated +pitch, right slope rotated -pitch. Ridge cap on top.
SHED ROOF: Single slope. pitch = atan2(height, depth).
HIP ROOF: 4 slopes meeting at peak/ridge.
STAIRS: Loop of cubes, each offset by (step_depth, step_height) from previous.
DOOR OPENING: Split wall into 3 parts (left, above, right of opening).
WINDOW: Split wall into pieces around the opening.

ALWAYS: Label every actor, set collision BlockAll, use direct set_ calls (never conditional), connect wall corners precisely.

MANDATORY ROOF ACTION — ALWAYS use this for ANY roof fix/build/rebuild (do NOT write your own roof code):
```action
{"action": "rebuild_gable_roof", "roof_height_ratio": 0.3}
```

MATERIAL/TEXTURE ACTION — apply textures to actors:
```action
{"action": "apply_material", "actor_pattern": "House_*", "material": "brick"}
```
Materials: brick, wood, metal, concrete, stucco, stone, glass, water, lava, grass, terrain, moss, dirt, sand, rock, cliff, snow, mud, ground, road, sidewalk, farmfield, ocean_floor, desert_grass, sparkle, glitter

HOUSE BUILD ACTION — use this for houses, homes, cabins, cottages, villas, mansions, townhouses, apartments, and condos:
```action
{"action": "build_house", "request": "build a cozy house here", "style": "cottage", "size": "medium"}
```
Rules for house requests:
- Prefer `build_house` over generic Python or `build_structure`.
- Houses should feel complete by default: floors, walls, roof, door, and believable circulation.
- Multi-story houses must include stairs, a stairwell opening, and a reachable landing.
- Vary footprint, roof pitch, and character within sane residential ranges instead of cloning the same shell.
- Apartments and condos should still route through `build_house`, with an explicit `story_count` when the request calls for multi-story residential.

GENERATIVE BUILD ACTION — build non-residential or scenic structures with auto-texturing:
```action
{"action": "build_structure", "structure": "fountain", "position": {"x": 5200, "y": 4200, "z": 0}, "size": "medium", "material": "stone"}
```
Structures: fountain, column, arch, tower, pool, fence, platform, bridge
Water feature structures: waterfall

TERRAIN CONTROL ACTION — create, modify, delete, or list terrain patches:
```action
{"action": "terrain", "operation": "create", "terrain_type": "hill", "position": {"x": 0, "y": 0, "z": 0}, "size": {"x": 3000, "y": 3000}, "material": "grass", "elevation": 300}
```
Operations: create, modify, delete, list
Terrain types: flat, hill, valley, plateau, slope, crater, ridge
Terrain materials: grass, dirt, sand, rock, cliff, snow, mud, ground, road, sidewalk, farmfield, ocean_floor, desert_grass, moss, terrain
Extra params: height (base Z), elevation (shape height), subdivisions (tile count), label (actor name), label_pattern (for delete), decorate (auto-add biome-matched trees/rocks/shrubs when project meshes exist)
Terrain quality rules:
- Prefer broad continuous terrain surfaces first. Use stitched tiles only when the area is large or the user explicitly wants segmentation.
- Keep terrain inside a sensible local footprint. Do not generate giant terrain that overruns the whole map unless the user asks.
- Use balanced terrain materials and matching edge/cliff materials so the terrain reads naturally.
- Use layered terrain surfacing, not one flat texture everywhere. Break up the top with secondary materials that match the biome so the terrain reads smoother and more natural.
- Avoid gaps between terrain sections and avoid leaving floating edges exposed.
- For long roads, ridges, shorelines, or embankments, prefer long continuous terrain strips instead of many disconnected patches.
- Add visually pleasing environment dressing that matches the vibe of the terrain when suitable project assets exist: trees on lush ground, rocks on mountain terrain, shrubs on desert/wetland terrain, etc.
- For hills, cliffs, waterfalls, and landforms, NEVER build the shape as stair-stepped cubes unless the user explicitly wants a blockout. Use terrain for the landform first, then add continuous water sheets or pools on top if needed.
Examples:
- Flat grass field: {"action": "terrain", "operation": "create", "terrain_type": "flat", "position": {"x": 0, "y": 0, "z": 0}, "size": {"x": 5000, "y": 5000}, "material": "grass"}
- Sandy hill: {"action": "terrain", "operation": "create", "terrain_type": "hill", "position": {"x": 2000, "y": 0, "z": 0}, "size": {"x": 2000, "y": 2000}, "material": "sand", "elevation": 400}
- Long ridge strip: {"action": "terrain", "operation": "create", "terrain_type": "ridge", "position": {"x": 0, "y": 0, "z": 0}, "size": {"x": 12000, "y": 2400}, "material": "rock"}
- Waterfall on a hill: first create the hill or ridge terrain, then build a waterfall structure with a continuous cascade instead of steps.
- Delete all terrain: {"action": "terrain", "operation": "delete", "label_pattern": "Terrain_*"}

RETRY RULE: If the user repeats a request, the previous attempt FAILED. Take a DIFFERENT approach.
"""

    system_text = _build_system_prompt(chat_title=chat_title, chat_memory=chat_memory) + action_instruction

    # Build contents with extra UEFN context injected into user message
    contents = _build_gemini_generate_content_input(message, attachments, history)
    if extra_context and contents:
        last_user = contents[-1]
        if last_user.get("role") == "user":
            last_user["parts"].append({"text": extra_context})

    payload = {
        "contents": contents,
        "system_instruction": {
            "parts": [
                {"text": system_text}
            ]
        },
    }

    data = _gemini_generate_content_request(model=model, payload=payload)

    reply = _extract_gemini_generate_content_text(data)
    if not reply:
        raise RuntimeError(f"Gemini generateContent returned no text output: {json.dumps(data, default=str)[:1000]}")

    # ── Execute any action blocks Gemini included in its response ──
    action_results = _execute_action_blocks(reply)
    tool_result_for_frontend = None
    if action_results:
        tool_result_for_frontend = action_results[-1]
        for ar in action_results:
            if ar.get("success") or ar.get("result"):
                reply += f"\n\n**Action executed:** `{ar.get('action', 'unknown')}`"
                if ar.get("result"):
                    result_preview = json.dumps(ar["result"], default=str)[:500]
                    reply += f"\n```\n{result_preview}\n```"
            elif ar.get("error"):
                reply += f"\n\n**Action failed:** {ar.get('error')}"

    result = {
        "reply": reply,
        "_provider": "gemini",
        "_model": model,
    }
    if tool_result_for_frontend:
        result["tool_result"] = tool_result_for_frontend
    return result


def _chat_uefn_query(intent: str, params: dict) -> dict:
    """Query UEFN via MCP bridge based on detected intent."""
    port = discover_uefn_listener_port()
    if not port:
        return {"error": "UEFN not connected. Open your project in UEFN with the MCP listener running."}
    try:
        result = mcp_listener_post_command(int(port), intent, params, timeout=10.0)
        return result
    except Exception as e:
        return {"error": str(e)}


_EXACT_ATTACHMENT_TEXT_PHRASES = (
    "what text is in the attached file",
    "what does the attached file say",
    "what does the attached document say",
    "what text is in this pdf",
    "what does this pdf say",
    "what does this doc say",
    "what does this website say",
    "what does this spreadsheet say",
    "what does this presentation say",
    "what does this archive contain",
    "what does this epub say",
    "what does the handwriting say",
    "read the handwriting",
    "reply with the exact sentence",
    "reply with the exact text",
    "reply with the exact text only",
    "repeat the attached file",
    "quote the attached file",
    "what does the image say",
    "read the screenshot",
    "read this image",
)
_WHICH_ATTACHMENT_PHRASES = (
    "which file",
    "which attachment",
    "which document",
    "which pdf",
    "which report",
    "which screenshot",
    "which image",
    "where does it say",
    "which one says",
)


def _exact_attachment_text_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    return any(phrase in lowered for phrase in _EXACT_ATTACHMENT_TEXT_PHRASES)


def _which_attachment_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    return any(phrase in lowered for phrase in _WHICH_ATTACHMENT_PHRASES)


def _build_direct_attachment_evidence_reply(message: str, attachments: list) -> Optional[Dict[str, Any]]:
    if not attachments:
        return None

    lowered = (message or "").strip().lower()
    if any(token in lowered for token in ("handwriting", "handwritten")) and any(
        str(att.get("analysisHandwriting") or att.get("analysis_handwriting") or "").strip()
        for att in (attachments or [])
    ):
        return None

    exact_requested = _exact_attachment_text_requested(message)
    which_requested = _which_attachment_requested(message)
    if not exact_requested and not which_requested:
        return None

    compiled = _build_attachment_dossier(message, attachments, history_mode=False)
    ranked = compiled.get("ranked") or []
    evidence = compiled.get("evidence") or []
    conflicts = compiled.get("conflicts") or []
    confidence = compiled.get("confidence") or {}
    if not ranked or not evidence:
        return None

    if exact_requested:
        if conflicts:
            lines = ["I found conflicting extracted evidence, so I’m listing the strongest candidates directly:"]
            for record in evidence[: min(3, len(evidence))]:
                lines.append(
                    f"- [{record.get('reference')}] {record.get('attachment_name')}: "
                    f"{_compact_text(str(record.get('text') or ''), limit=180)}"
                )
            return {"reply": "\n".join(lines)}

        top_record = evidence[0]
        corroborating_sources = {int(record.get("attachment_rank") or 0) for record in evidence[:3] if int(record.get("attachment_rank") or 0) > 0}
        if confidence.get("label") == "low" and len(corroborating_sources) > 1:
            return {
                "reply": (
                    "The extracted evidence is weak across multiple files, so I can't give one exact answer safely.\n"
                    + "\n".join(
                        f"- [{record.get('reference')}] {record.get('attachment_name')}: "
                        f"{_compact_text(str(record.get('text') or ''), limit=160)}"
                        for record in evidence[: min(3, len(evidence))]
                    )
                )
            }
        return {"reply": str(top_record.get("text") or "").strip()}

    top_attachment = ranked[0]
    top_record = next(iter(_select_attachment_evidence_records(top_attachment, message, limit=1)), None)
    if not top_record:
        return None
    return {
        "reply": (
            f"The strongest match is [{_attachment_reference_label(top_attachment)}] "
            f"{top_attachment.get('name') or 'attachment'}.\n"
            f"Evidence: [{top_record.get('reference')}] {_compact_text(str(top_record.get('text') or ''), limit=220)}"
        )
    }


def _maybe_answer_directly_from_attachments(message: str, attachments: list) -> Optional[Dict[str, Any]]:
    """Short-circuit deterministic attachment questions before the hosted model path."""
    lowered = (message or "").strip().lower()
    visual_alignment = _build_visual_alignment_report(message, attachments)
    if visual_alignment and (_visual_review_requested(message) or _visual_comparison_requested(message)):
        return {"reply": visual_alignment}

    evidence_reply = _build_direct_attachment_evidence_reply(message, attachments)
    if evidence_reply is not None:
        return evidence_reply

    readable_attachments: List[tuple[Dict[str, Any], str]] = []
    direct_image_description_reply = ""
    for att in (attachments or []):
        attachment_type = str(att.get("type") or "").strip().lower()
        analysis_text = str(att.get("analysisText") or att.get("analysis_text") or "").strip()
        analysis_handwriting = str(att.get("analysisHandwriting") or att.get("analysis_handwriting") or "").strip()
        analysis_caption = str(att.get("analysisCaption") or att.get("analysis_caption") or "").strip()
        content_text = str(att.get("content") or "").strip()
        if attachment_type == "image":
            if any(phrase in lowered for phrase in ("describe this image", "what is in this image", "what's in this image")) and analysis_caption:
                direct_image_description_reply = analysis_caption
                attachment_text = analysis_caption
            elif any(phrase in lowered for phrase in ("what does the handwriting say", "read the handwriting", "what is the handwriting")) and analysis_handwriting:
                attachment_text = analysis_handwriting
            elif any(phrase in lowered for phrase in ("what does the image say", "read the screenshot", "read this image", "extract the text", "ocr this")) and analysis_text:
                attachment_text = analysis_text
            else:
                attachment_text = analysis_text or analysis_handwriting or analysis_caption
        elif attachment_type == "binary":
            if any(phrase in lowered for phrase in ("handwriting", "handwritten")) and analysis_handwriting:
                attachment_text = analysis_handwriting
            else:
                attachment_text = analysis_text or analysis_handwriting or analysis_caption
        else:
            attachment_text = content_text or analysis_text or analysis_handwriting or analysis_caption
        if attachment_text:
            readable_attachments.append((att, attachment_text))

    if len(readable_attachments) != 1:
        return None

    if direct_image_description_reply:
        return {"reply": direct_image_description_reply}

    attachment_text = readable_attachments[0][1]
    if _exact_attachment_text_requested(message) or any(phrase in lowered for phrase in ("extract the text", "ocr this")):
        return {"reply": attachment_text}

    return None


def _extract_text_from_image_attachments(attachments: list) -> str:
    engine = _get_ocr_engine()
    if not engine or not _HAS_PIL or _PILImage is None:
        return ""

    extracted_chunks: List[str] = []

    for att in (attachments or [])[:3]:
        if str(att.get("type") or "").strip().lower() != "image":
            continue

        mime_type, base64_data = _split_base64_data_url(str(att.get("content") or ""))
        if not base64_data:
            continue

        try:
            image_bytes = base64.b64decode(base64_data)
            image = _PILImage.open(io.BytesIO(image_bytes)).convert("RGB")
            ocr_result = _ocr_pil_image_multistage(image, engine)
            text = str(ocr_result.get("text") or "").strip()
            if text:
                extracted_chunks.append(text)
        except Exception as exc:
            logger.warning("OCR extraction failed for attachment '%s': %s", att.get("name", "image"), exc)

    return "\n\n".join(chunk for chunk in extracted_chunks if chunk).strip()


def _render_pdf_pages_to_pil_images(pdf_bytes: bytes, max_pages: int = 6, dpi: int = 300) -> List[Any]:
    if not _HAS_PYMUPDF or not _HAS_PIL:
        return []

    images: List[Any] = []
    try:
        document = _pymupdf.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        logger.warning("PDF rasterization failed: %s", exc)
        return []

    try:
        sampled_indexes = _sample_document_indexes(len(document), min(max_pages, len(document)))
        for page_index in sampled_indexes:
            try:
                page = document.load_page(page_index)
                try:
                    pixmap = page.get_pixmap(dpi=dpi, alpha=False)
                except TypeError:
                    zoom = max(1.5, dpi / 72.0)
                    pixmap = page.get_pixmap(matrix=_pymupdf.Matrix(zoom, zoom), alpha=False)
                image = _PILImage.open(io.BytesIO(pixmap.tobytes("png"))).convert("RGB")
                images.append(image)
            except Exception as exc:
                logger.warning("PDF page rasterization failed on page %s: %s", page_index + 1, exc)
    finally:
        try:
            document.close()
        except Exception:
            pass

    return images


def _extract_scanned_pdf_analysis_from_bytes(pdf_bytes: bytes) -> Dict[str, Any]:
    images = _render_pdf_pages_to_pil_images(pdf_bytes)
    if not images:
        return {}

    structured_visual = _combine_structured_visual_analyses(
        _analyze_pil_images_with_local_structured_vlm(images),
        label_prefix="Page",
    )
    visual_summaries: List[str] = []
    for page_index, image in enumerate(images[:6], start=1):
        meta = _extract_visual_metadata_from_pil_image(image)
        meta_summary = _render_visual_metadata_summary(meta)
        if meta_summary:
            visual_summaries.append(f"Page {page_index}: {meta_summary}")

    ocr_text = str(structured_visual.get("analysisText") or "").strip()
    handwriting_text = ""
    caption_text = str(structured_visual.get("analysisCaption") or "").strip()
    if not caption_text:
        captions = _caption_pil_images(images)
        caption_text = "\n".join(
            f"Page {index + 1}: {caption}"
            for index, caption in enumerate(captions)
            if caption
        ).strip()
    handwriting_outputs = _extract_handwriting_from_pil_images(images)
    handwriting_text = "\n".join(
        f"Page {index + 1}: {text}"
        for index, text in enumerate(handwriting_outputs)
        if text
    ).strip()

    engine = _get_ocr_engine()
    if engine:
        chunks: List[str] = []
        ocr_variant_notes: List[str] = []
        for page_index, image in enumerate(images[:6], start=1):
            try:
                ocr_result = _ocr_pil_image_multistage(image, engine)
                text = str(ocr_result.get("text") or "").strip()
                if text:
                    chunks.append(f"Page {page_index}: " + text)
                variant = str(ocr_result.get("variant") or "").strip()
                if variant:
                    ocr_variant_notes.append(f"Page {page_index}: OCR variant {variant}")
            except Exception as exc:
                logger.warning("Scanned PDF OCR failed on page %s: %s", page_index, exc)
        ocr_text = _merge_analysis_text(ocr_text, "\n".join(chunks).strip(), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    else:
        ocr_variant_notes = []

    combined_parts = []
    if ocr_text:
        combined_parts.append("OCR:\n" + ocr_text)
    if handwriting_text:
        combined_parts.append("Handwriting guess:\n" + handwriting_text)
    if caption_text:
        combined_parts.append("Visual description:\n" + caption_text)
    if visual_summaries:
        combined_parts.append("Page diagnostics:\n" + "\n".join(visual_summaries))

    return {
        "analysisText": _compact_text("\n\n".join(combined_parts), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS),
        "analysisHandwriting": _compact_text(handwriting_text, limit=2000),
        "analysisCaption": _compact_text(caption_text, limit=1200),
        "analysisMeta": _merge_visual_metadata(
            structured_visual.get("analysisMeta") or {},
            {
                "pageCount": len(images),
                "pageVisualDiagnostics": visual_summaries[:3],
                "ocrDiagnostics": ocr_variant_notes[:6],
            },
        ),
    }


class _VisibleTextHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self._skip_depth = 0
        self._parts: List[str] = []

    def handle_starttag(self, tag, attrs):
        tag = (tag or "").lower()
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip_depth += 1
            return
        if tag in {"br", "p", "div", "section", "article", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self._parts.append("\n")

    def handle_endtag(self, tag):
        tag = (tag or "").lower()
        if tag in {"script", "style", "noscript", "svg"} and self._skip_depth > 0:
            self._skip_depth -= 1
            return
        if tag in {"p", "div", "section", "article", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self._parts.append("\n")

    def handle_data(self, data):
        if self._skip_depth > 0:
            return
        text = html_lib.unescape(data or "")
        if text.strip():
            self._parts.append(text)

    def get_text(self) -> str:
        raw = "".join(self._parts)
        lines = [re.sub(r"\s+", " ", line).strip() for line in raw.splitlines()]
        return "\n".join(line for line in lines if line).strip()


_GENERIC_MIME_TYPES = {
    "",
    "application/octet-stream",
    "binary/octet-stream",
    "text/plain",
}
_MARKDOWN_EXTENSIONS = {"md", "markdown", "mdx"}
_XML_EXTENSIONS = {"xml", "rss", "atom", "svg"}
_TEXT_ARCHIVE_EXTENSIONS = {
    "txt", "md", "markdown", "mdx", "json", "yaml", "yml", "toml", "ini", "cfg", "conf",
    "xml", "html", "htm", "css", "js", "ts", "jsx", "tsx", "py", "lua", "verse",
    "csv", "tsv", "sql", "log", "bat", "ps1", "sh",
}


def _file_extension(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    if "://" in raw:
        raw = urllib.parse.urlsplit(raw).path or raw
    suffix = Path(raw).suffix.lower()
    return suffix[1:] if suffix.startswith(".") else suffix


def _infer_attachment_mime_type(name: str, mime_type: str, source_url: str = "") -> str:
    normalized = (mime_type or "").strip().lower()
    ext = _file_extension(name) or _file_extension(source_url)
    extension_map = {
        "md": "text/markdown",
        "markdown": "text/markdown",
        "mdx": "text/markdown",
        "csv": "text/csv",
        "tsv": "text/tab-separated-values",
        "xml": "application/xml",
        "rss": "application/xml",
        "atom": "application/xml",
        "svg": "image/svg+xml",
        "yaml": "application/yaml",
        "yml": "application/yaml",
        "json": "application/json",
        "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "epub": "application/epub+zip",
        "zip": "application/zip",
    }
    if normalized and normalized not in _GENERIC_MIME_TYPES:
        return normalized
    if ext in extension_map:
        return extension_map[ext]
    guessed, _ = mimetypes.guess_type(name or source_url or "")
    return (guessed or normalized or "").lower()


def _decode_text_bytes(raw_bytes: bytes, charset: str = "utf-8") -> str:
    for candidate in (charset or "utf-8", "utf-8-sig", "utf-16", "latin-1"):
        try:
            return raw_bytes.decode(candidate, errors="ignore")
        except Exception:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def _safe_xml_fromstring(xml_input: bytes | str):
    return _SafeElementTree.fromstring(xml_input)


def _extract_markdown_title(content: str) -> str:
    for line in (content or "").splitlines():
        match = re.match(r"^\s{0,3}#{1,6}\s+(.+?)\s*$", line)
        if match:
            return _compact_text(match.group(1).strip(), limit=240)
    return ""


def _extract_markdown_text(content: str) -> str:
    text = html_lib.unescape(content or "")
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"`{1,3}", "", text)
    text = re.sub(r"^\s{0,3}#{1,6}\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s{0,3}[-*+]\s+", "- ", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return _compact_text(text.strip(), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_xml_text(xml_input: bytes | str) -> str:
    try:
        root = _safe_xml_fromstring(xml_input)
        lines = [re.sub(r"\s+", " ", chunk).strip() for chunk in root.itertext()]
        joined = "\n".join(line for line in lines if line)
        return _compact_text(joined, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    except Exception:
        if isinstance(xml_input, bytes):
            xml_input = _decode_text_bytes(xml_input)
        parser = _VisibleTextHTMLParser()
        try:
            parser.feed(xml_input or "")
            parser.close()
        except Exception:
            return ""
        return _compact_text(parser.get_text(), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_delimited_text_from_bytes(raw_bytes: bytes, delimiter: str, label: str, charset: str = "utf-8") -> str:
    decoded = _decode_text_bytes(raw_bytes, charset=charset)
    stream = io.StringIO(decoded)
    try:
        rows = list(csv_lib.reader(stream, delimiter=delimiter))
    except Exception:
        rows = []
    lines: List[str] = [f"{label} rows:"]
    row_count = 0
    for row in rows:
        cleaned = [re.sub(r"\s+", " ", str(cell).strip()) for cell in row if str(cell).strip()]
        if not cleaned:
            continue
        lines.append(" | ".join(cleaned[:24]))
        row_count += 1
        if row_count >= 120:
            break
    if row_count == 0:
        return _compact_text(decoded, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    return _compact_text("\n".join(lines), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_archive_manifest_from_bytes(zip_bytes: bytes, *, label: str = "ZIP archive") -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
            file_infos = [info for info in archive.infolist() if not info.is_dir()]
            listing_lines = [f"{label} contents:"]
            for info in file_infos[:80]:
                listing_lines.append(f"- {info.filename} ({info.file_size} bytes)")

            previews: List[str] = []
            preview_indexes = _sample_document_indexes(len(file_infos), min(12, len(file_infos)))
            for preview_index in preview_indexes:
                info = file_infos[preview_index]
                ext = _file_extension(info.filename)
                if ext not in _TEXT_ARCHIVE_EXTENSIONS or info.file_size > 180_000:
                    continue
                try:
                    file_bytes = archive.read(info.filename)
                except Exception:
                    continue
                if ext in {"html", "htm"}:
                    preview = _extract_html_text(_decode_text_bytes(file_bytes))
                elif ext in _XML_EXTENSIONS:
                    preview = _extract_xml_text(file_bytes)
                elif ext in {"csv", "tsv"}:
                    preview = _extract_delimited_text_from_bytes(
                        file_bytes,
                        delimiter="\t" if ext == "tsv" else ",",
                        label=ext.upper(),
                    )
                elif ext in _MARKDOWN_EXTENSIONS:
                    preview = _extract_markdown_text(_decode_text_bytes(file_bytes))
                else:
                    preview = _compact_text(_decode_text_bytes(file_bytes), limit=2400)
                if preview:
                    previews.append(f"{info.filename}:\n{preview[:2400]}")
                if len(previews) >= 6:
                    break

            combined = "\n".join(listing_lines)
            if previews:
                combined += "\n\nPreviewed text files:\n" + "\n\n".join(previews)
            return _compact_text(combined, limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    except Exception as exc:
        logger.warning("%s extraction failed: %s", label, exc)
        return ""


def _extract_epub_text_from_bytes(epub_bytes: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(epub_bytes)) as archive:
            html_entries = sorted(
                name for name in archive.namelist()
                if name.lower().endswith((".xhtml", ".html", ".htm"))
            )
            if not html_entries:
                return _extract_archive_manifest_from_bytes(epub_bytes, label="EPUB archive")

            sections: List[str] = []
            total_chars = 0
            for entry_index in _sample_document_indexes(len(html_entries), min(24, len(html_entries))):
                entry = html_entries[entry_index]
                try:
                    section_html = _decode_text_bytes(archive.read(entry))
                except Exception:
                    continue
                section_text = _extract_html_text(section_html)
                if not section_text:
                    continue
                label = Path(entry).stem.replace("_", " ").replace("-", " ").strip() or entry
                chunk = f"{label}:\n{section_text}"
                sections.append(chunk)
                total_chars += len(chunk)
                if total_chars >= MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS:
                    break
            return _compact_text("\n\n".join(sections), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    except Exception as exc:
        logger.warning("EPUB text extraction failed: %s", exc)
        return ""


_URL_PATTERN = re.compile(r"https?://[^\s<>()\"']+", re.IGNORECASE)


def _extract_pdf_text_from_bytes(pdf_bytes: bytes) -> str:
    if not _HAS_PYPDF:
        return ""

    try:
        reader = _PdfReader(io.BytesIO(pdf_bytes))
        chunks: List[str] = []
        pages = list(reader.pages)
        for page_index in _sample_document_indexes(len(pages), min(24, len(pages))):
            page = pages[page_index]
            text = (page.extract_text() or "").strip()
            if text:
                chunks.append(f"Page {page_index + 1}:\n{text}")
            if sum(len(chunk) for chunk in chunks) >= MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS:
                break
        return _compact_text("\n\n".join(chunks), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
    except Exception as exc:
        logger.warning("PDF text extraction failed: %s", exc)
        return ""


def _extract_pdf_text_from_data_url(content: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type != "application/pdf" or not base64_data:
        return ""

    try:
        return _extract_pdf_text_from_bytes(base64.b64decode(base64_data))
    except Exception as exc:
        logger.warning("PDF text extraction failed: %s", exc)
        return ""


def _extract_docx_text_from_bytes(docx_bytes: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(docx_bytes)) as archive:
            xml_bytes = archive.read("word/document.xml")
    except Exception as exc:
        logger.warning("DOCX text extraction failed: %s", exc)
        return ""

    try:
        root = _safe_xml_fromstring(xml_bytes)
    except Exception as exc:
        logger.warning("DOCX XML parsing failed: %s", exc)
        return ""

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: List[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        joined = "".join(texts).strip()
        if joined:
            paragraphs.append(joined)

    return _compact_text("\n\n".join(paragraphs), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_docx_text_from_data_url(content: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type != "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or not base64_data:
        return ""

    try:
        return _extract_docx_text_from_bytes(base64.b64decode(base64_data))
    except Exception as exc:
        logger.warning("DOCX text extraction failed: %s", exc)
        return ""


def _extract_xlsx_text_from_bytes(xlsx_bytes: bytes) -> str:
    if not _HAS_OPENPYXL:
        return ""

    try:
        workbook = _openpyxl_load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("XLSX text extraction failed: %s", exc)
        return ""

    chunks: List[str] = []
    total_chars = 0
    try:
        worksheets = list(workbook.worksheets)
        for worksheet_index in _sample_document_indexes(len(worksheets), min(12, len(worksheets))):
            worksheet = worksheets[worksheet_index]
            sheet_lines: List[str] = [f"Sheet: {worksheet.title}"]
            row_count = 0
            for row in worksheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                if not cells:
                    continue
                sheet_lines.append(" | ".join(cells))
                row_count += 1
                if row_count >= 160:
                    break
            if len(sheet_lines) > 1:
                sheet_text = "\n".join(sheet_lines)
                chunks.append(sheet_text)
                total_chars += len(sheet_text)
                if total_chars >= MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS:
                    break
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return _compact_text("\n\n".join(chunks), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_xlsx_text_from_data_url(content: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or not base64_data:
        return ""

    try:
        return _extract_xlsx_text_from_bytes(base64.b64decode(base64_data))
    except Exception as exc:
        logger.warning("XLSX text extraction failed: %s", exc)
        return ""


def _extract_pptx_text_from_bytes(pptx_bytes: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as archive:
            slide_names = sorted(
                name for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            )
            chunks: List[str] = []
            total_chars = 0
            namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}

            for slide_index in _sample_document_indexes(len(slide_names), min(40, len(slide_names))):
                slide_name = slide_names[slide_index]
                slide_bytes = archive.read(slide_name)
                root = _safe_xml_fromstring(slide_bytes)
                texts = [node.text or "" for node in root.findall(".//a:t", namespace)]
                cleaned = [text.strip() for text in texts if (text or "").strip()]
                if not cleaned:
                    continue
                slide_label = Path(slide_name).stem.replace("slide", "Slide ")
                slide_text = f"{slide_label}:\n" + "\n".join(cleaned)
                chunks.append(slide_text)
                total_chars += len(slide_text)
                if total_chars >= MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS:
                    break
    except Exception as exc:
        logger.warning("PPTX text extraction failed: %s", exc)
        return ""

    return _compact_text("\n\n".join(chunks), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_pptx_text_from_data_url(content: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type != "application/vnd.openxmlformats-officedocument.presentationml.presentation" or not base64_data:
        return ""

    try:
        return _extract_pptx_text_from_bytes(base64.b64decode(base64_data))
    except Exception as exc:
        logger.warning("PPTX text extraction failed: %s", exc)
        return ""


def _extract_archive_manifest_from_data_url(content: str, expected_mime_types: set[str], *, label: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type not in expected_mime_types or not base64_data:
        return ""
    try:
        return _extract_archive_manifest_from_bytes(base64.b64decode(base64_data), label=label)
    except Exception as exc:
        logger.warning("%s extraction failed: %s", label, exc)
        return ""


def _extract_epub_text_from_data_url(content: str) -> str:
    mime_type, base64_data = _split_base64_data_url(content)
    if mime_type != "application/epub+zip" or not base64_data:
        return ""
    try:
        return _extract_epub_text_from_bytes(base64.b64decode(base64_data))
    except Exception as exc:
        logger.warning("EPUB text extraction failed: %s", exc)
        return ""


def _extract_html_text(content: str) -> str:
    if _HAS_TRAFILATURA:
        try:
            extracted = _trafilatura.extract(content or "")
            if extracted:
                return _compact_text(str(extracted), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)
        except Exception as exc:
            logger.warning("Trafilatura extraction failed: %s", exc)

    parser = _VisibleTextHTMLParser()
    try:
        parser.feed(content or "")
        parser.close()
    except Exception:
        return ""
    return _compact_text(parser.get_text(), limit=MAX_EXTRACTED_ATTACHMENT_TEXT_CHARS)


def _extract_html_metadata(content: str) -> tuple[str, str]:
    title_match = re.search(r"<title[^>]*>(.*?)</title>", content or "", re.IGNORECASE | re.DOTALL)
    title = html_lib.unescape(title_match.group(1)).strip() if title_match else ""
    title = re.sub(r"\s+", " ", title)

    desc_match = re.search(
        r'<meta[^>]+name=["\']description["\'][^>]+content=["\'](.*?)["\']',
        content or "",
        re.IGNORECASE | re.DOTALL,
    )
    description = html_lib.unescape(desc_match.group(1)).strip() if desc_match else ""
    description = re.sub(r"\s+", " ", description)
    return title[:240], description[:400]


def _attachment_name_from_url(url: str, fallback: str = "web-page") -> str:
    try:
        parsed = urllib.parse.urlsplit(url)
        name = urllib.parse.unquote(Path(parsed.path).name or "").strip()
        if name:
            return name[:240]
        if parsed.netloc:
            return parsed.netloc[:240]
    except Exception:
        pass
    return fallback


def _extract_message_urls(message: str) -> List[str]:
    urls: List[str] = []
    for match in _URL_PATTERN.findall(message or ""):
        candidate = match.rstrip(".,;:!?)]}")
        if candidate not in urls:
            urls.append(candidate)
        if len(urls) >= 3:
            break
    return urls


def _fetch_web_attachment_from_url(url: str) -> Optional[Dict[str, Any]]:
    parsed = urllib.parse.urlsplit(url or "")
    if parsed.scheme not in {"http", "https"}:
        return None

    request_obj = urllib.request.Request(
        url,
        headers={"User-Agent": "UEFN-Codex-Agent/2.1"},
        method="GET",
    )

    try:
        with urllib.request.urlopen(request_obj, timeout=12) as response:
            final_url = response.geturl()
            header_mime_type = (response.headers.get_content_type() or "text/plain").lower()
            charset = response.headers.get_content_charset() or "utf-8"
            raw_bytes = response.read(700_000)
    except Exception as exc:
        logger.warning("Failed to fetch URL attachment '%s': %s", url, exc)
        return None

    truncated = len(raw_bytes) >= 700_000
    text_content = ""
    summary = ""
    keywords: List[str] = []
    name = _attachment_name_from_url(final_url)
    mime_type = _infer_attachment_mime_type(name, header_mime_type, final_url)

    if mime_type.startswith("image/") and mime_type != "image/svg+xml":
        attachment = {
            "name": name,
            "type": "image",
            "mime_type": mime_type,
            "size": len(raw_bytes),
            "truncated": truncated,
            "content": f"data:{mime_type};base64,{base64.b64encode(raw_bytes).decode('ascii')}",
            "analysis_summary": f"Fetched image from {urllib.parse.urlsplit(final_url).netloc}",
            "source_url": final_url,
        }
        return _prepare_chat_attachment(attachment)

    if mime_type == "application/pdf":
        text_content = _extract_pdf_text_from_bytes(raw_bytes)
        summary = f"Fetched PDF from {urllib.parse.urlsplit(final_url).netloc}"
    elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        text_content = _extract_docx_text_from_bytes(raw_bytes)
        summary = f"Fetched DOCX document from {urllib.parse.urlsplit(final_url).netloc}"
    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        text_content = _extract_xlsx_text_from_bytes(raw_bytes)
        summary = f"Fetched XLSX spreadsheet from {urllib.parse.urlsplit(final_url).netloc}"
    elif mime_type == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
        text_content = _extract_pptx_text_from_bytes(raw_bytes)
        summary = f"Fetched PPTX presentation from {urllib.parse.urlsplit(final_url).netloc}"
    elif mime_type == "application/epub+zip":
        text_content = _extract_epub_text_from_bytes(raw_bytes)
        summary = f"Fetched EPUB document from {urllib.parse.urlsplit(final_url).netloc}"
    elif mime_type in {"application/zip", "application/x-zip-compressed"}:
        text_content = _extract_archive_manifest_from_bytes(raw_bytes, label="ZIP archive")
        summary = f"Fetched archive from {urllib.parse.urlsplit(final_url).netloc}"
    else:
        decoded = _decode_text_bytes(raw_bytes, charset=charset)
        if "html" in mime_type:
            title, description = _extract_html_metadata(decoded)
            visible_text = _extract_html_text(decoded)
            text_content = visible_text
            name = (title or name)[:240]
            summary = title or f"Web page from {urllib.parse.urlsplit(final_url).netloc}"
            if description:
                summary = f"{summary} — {description}"
        elif mime_type in {"text/markdown"} or _file_extension(name) in _MARKDOWN_EXTENSIONS:
            text_content = _extract_markdown_text(decoded)
            summary = _extract_markdown_title(decoded) or f"Markdown document from {urllib.parse.urlsplit(final_url).netloc}"
        elif mime_type in {"application/xml", "text/xml", "image/svg+xml"} or _file_extension(name) in _XML_EXTENSIONS:
            text_content = _extract_xml_text(decoded)
            summary = f"XML document from {urllib.parse.urlsplit(final_url).netloc}"
        elif mime_type in {"text/csv", "application/csv"} or _file_extension(name) == "csv":
            text_content = _extract_delimited_text_from_bytes(raw_bytes, delimiter=",", label="CSV", charset=charset)
            summary = f"CSV document from {urllib.parse.urlsplit(final_url).netloc}"
        elif mime_type == "text/tab-separated-values" or _file_extension(name) == "tsv":
            text_content = _extract_delimited_text_from_bytes(raw_bytes, delimiter="\t", label="TSV", charset=charset)
            summary = f"TSV document from {urllib.parse.urlsplit(final_url).netloc}"
        elif "json" in mime_type:
            try:
                parsed_json = json.loads(decoded)
                text_content = json.dumps(parsed_json, indent=2)[:16000]
            except Exception:
                text_content = decoded[:16000]
            summary = f"JSON document from {urllib.parse.urlsplit(final_url).netloc}"
        else:
            text_content = decoded[:16000]
            summary = f"Text document from {urllib.parse.urlsplit(final_url).netloc}"

    if text_content:
        keywords = []
        for token in re.findall(r"[a-zA-Z0-9_]{3,}", f"{name} {summary} {text_content}".lower()):
            if token in {"http", "https", "www", "html", "text", "page", "document", "from"}:
                continue
            if token not in keywords:
                keywords.append(token)
            if len(keywords) >= 12:
                break

    attachment = {
        "name": name,
        "type": "file" if text_content else "binary",
        "mime_type": mime_type,
        "size": len(raw_bytes),
        "truncated": truncated,
        "content": text_content if text_content else "",
        "analysis_summary": summary,
        "analysis_keywords": keywords,
        "source_url": final_url,
    }
    return _prepare_chat_attachment(attachment)


def _collect_message_url_attachments(message: str, existing_attachments: list) -> List[Dict[str, Any]]:
    existing_urls = {
        str(att.get("sourceUrl") or att.get("source_url") or "").strip()
        for att in (existing_attachments or [])
        if str(att.get("sourceUrl") or att.get("source_url") or "").strip()
    }
    collected: List[Dict[str, Any]] = []
    for url in _extract_message_urls(message):
        if url in existing_urls:
            continue
        attachment = _fetch_web_attachment_from_url(url)
        if not attachment:
            continue
        collected.append(attachment)
        existing_urls.add(str(attachment.get("sourceUrl") or url))
    return collected


def _render_attachment_reasoning_context(attachments: list, message: str = "") -> str:
    compiled = _build_attachment_dossier(message, attachments, history_mode=False)
    blocks: List[str] = []
    dossier_text = compiled.get("dossier") or ""
    if dossier_text:
        blocks.append(dossier_text)
    if _attachment_summary_requested(message):
        summary_blocks: List[str] = []
        for attachment in (compiled.get("ranked") or [])[:4]:
            name = str(attachment.get("name") or "attachment")
            summary = str(attachment.get("analysisSummary") or attachment.get("analysis_summary") or "").strip()
            records = _select_attachment_summary_records(attachment, limit=6)
            lines = [f"[{_attachment_reference_label(attachment)}: {name}]"]
            if summary:
                lines.append(f"Attachment overview: {summary}")
            if records:
                lines.append("Representative sections across the full file:")
                for record in records:
                    lines.append(
                        f"- [{record.get('reference')}] "
                        f"{_compact_text(str(record.get('text') or ''), limit=340)}"
                    )
            summary_blocks.append("\n".join(lines))
        if summary_blocks:
            blocks.append("SUMMARY COVERAGE:\n" + "\n\n".join(summary_blocks))
    confidence = compiled.get("confidence") or {}
    if confidence:
        blocks.append(
            "GROUNDING CHECK:\n"
            f"- Confidence: {confidence.get('label', 'low')} ({confidence.get('score', 0)}/100)\n"
            f"- Reason: {confidence.get('reason', 'limited extracted evidence')}"
        )
    visual_alignment = _build_visual_alignment_report(message, compiled.get("ranked") or [])
    if visual_alignment:
        blocks.append("REQUEST ALIGNMENT REVIEW:\n" + visual_alignment)
    return "\n\n".join(blocks).strip()


def _answer_with_attachment_analysis_context(message: str, attachments: list, preferred_provider: str = "") -> Dict[str, Any]:
    attachment_context = _render_attachment_reasoning_context(attachments, message=message)
    if not attachment_context:
        return {"reply": "I stored the attachment, but I could not extract enough readable content from it yet."}

    client, provider, model = _resolve_text_fallback_target(preferred_provider=preferred_provider)
    if not client:
        return {
            "reply": (
                "I couldn't use a hosted multimodal model, but I extracted this attachment context:\n\n"
                f"{attachment_context[:12000]}"
            )
        }

    summary_requested = _attachment_summary_requested(message)
    system_prompt = (
        "You are an expert AI assistant analyzing user-provided files (images, PDFs, screenshots, documents, code). "
        "You have full understanding of what was uploaded. Respond naturally and intelligently — explain what you see, "
        "identify key information, and give clear, actionable insights. "
        "DO NOT use academic citation formats like [Attachment 1 §3] or [Attachment 2]. "
        "Instead, refer to files naturally by their filename (e.g. 'In the PDF...' or 'The screenshot shows...'). "
        "If the content relates to UEFN/Unreal Engine, identify actors, materials, geometry, errors, and suggest fixes. "
        "Be confident and thorough — analyze the FULL content, not just snippets."
    )
    if summary_requested:
        system_prompt = (
            "You are an expert AI assistant summarizing user-provided files. Build a comprehensive, well-structured "
            "summary covering the main subject, key details, components, and notable points. "
            "DO NOT use academic citation formats like [Attachment 1 §3]. Refer to files by their filename naturally. "
            "Start with a clear overview, then cover details in logical order. "
            "If multiple files are attached, synthesize them into ONE unified summary, noting connections between them. "
            "Be thorough — cover the entire document, not just the beginning."
        )

    response = _chat_completion_with_retry(
        client,
        provider or "attachment_fallback",
        model=model,
        messages=[
            {
                "role": "system",
                "content": system_prompt,
            },
            {
                "role": "user",
                "content": f"User request:\n{message}\n\nAttachment context:\n{attachment_context}",
            },
        ],
        max_tokens=2600 if summary_requested else 1800,
        temperature=0.15 if summary_requested else 0.2,
    )
    tok_in = getattr(response.usage, "prompt_tokens", 0) if response.usage else 0
    tok_out = getattr(response.usage, "completion_tokens", 0) if response.usage else 0
    _track_usage(provider, tok_in, tok_out)
    reply = response.choices[0].message.content or attachment_context
    return {"reply": reply, "_provider": provider, "_model": model}


def _answer_with_ocr_context(message: str, ocr_text: str, preferred_provider: str = "") -> Dict[str, Any]:
    lowered = (message or "").strip().lower()
    if any(phrase in lowered for phrase in (
        "what text is in",
        "what does the image say",
        "reply with the exact sentence",
        "reply with the exact text",
        "read the screenshot",
        "read this image",
    )):
        return {"reply": ocr_text}

    client, provider, model = _resolve_text_fallback_target(preferred_provider=preferred_provider)
    if not client:
        return {
            "reply": f"I couldn't use a hosted vision model, but I extracted this visible text from the image:\n\n{ocr_text}"
        }

    response = _chat_completion_with_retry(
        client,
        provider or "ocr_fallback",
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are helping interpret OCR text extracted from a user-provided image. "
                    "Answer using only the OCR text below. If the OCR looks noisy or incomplete, say so plainly. "
                    "Do not call tools."
                ),
            },
            {
                "role": "user",
                "content": f"User request:\n{message}\n\nOCR text from the image:\n{ocr_text}",
            },
        ],
        max_tokens=1200,
        temperature=0.2,
    )
    tok_in = getattr(response.usage, "prompt_tokens", 0) if response.usage else 0
    tok_out = getattr(response.usage, "completion_tokens", 0) if response.usage else 0
    _track_usage(provider or "ocr_fallback", tok_in, tok_out)
    reply = response.choices[0].message.content or ocr_text

    return {
        "reply": reply,
        "_provider": provider,
        "_model": model,
    }


def _build_keyword_attachment_reply(message: str, attachments: list) -> str:
    compiled = _build_attachment_dossier(message, attachments, history_mode=False)
    ranked = compiled.get("ranked") or []
    if not ranked:
        return "I stored the attachments, but I could not extract enough readable content from them yet."

    if _attachment_summary_requested(message):
        lines: List[str] = [f"I summarized {len(ranked)} attachment(s) using representative sections across the full file set."]
        confidence = compiled.get("confidence") or {}
        if confidence:
            lines.append(
                f"Grounding confidence: {confidence.get('label', 'low')} ({confidence.get('score', 0)}/100). "
                f"Reason: {confidence.get('reason', 'limited extracted evidence')}."
            )
        for attachment in ranked[: min(4, len(ranked))]:
            name = str(attachment.get("name") or "attachment")
            summary = str(attachment.get("analysisSummary") or "").strip()
            lines.append(f"{name}: {summary or 'Attachment analyzed.'}")
            records = _select_attachment_summary_records(attachment, limit=4)
            for record in records:
                lines.append(
                    f"- [{record.get('reference')}] "
                    f"{_compact_text(str(record.get('text') or ''), limit=240)}"
                )
        return "\n".join(lines)

    lines: List[str] = []
    if message:
        lines.append(f'I analyzed {len(ranked)} attachment(s) for: "{message}".')
    else:
        lines.append(f"I analyzed {len(ranked)} attachment(s).")

    synthesis = _render_attachment_synthesis(message, ranked, clipped=int(compiled.get("clipped") or 0))
    if synthesis:
        lines.append(synthesis)
    confidence = compiled.get("confidence") or {}
    if confidence:
        lines.append(
            f"Evidence confidence: {confidence.get('label', 'low')} ({confidence.get('score', 0)}/100). "
            f"Reason: {confidence.get('reason', 'limited extracted evidence')}."
        )

    evidence = compiled.get("evidence") or []
    if evidence:
        lines.append("Best supporting evidence:")
        for record in evidence[:4]:
            lines.append(
                f"- [{record.get('reference')}] {record.get('attachment_name')}: "
                f"{_compact_text(str(record.get('text') or ''), limit=180)}"
            )

    conflicts = compiled.get("conflicts") or []
    if conflicts:
        lines.append("Potential conflicts:")
        for item in conflicts[:3]:
            lines.append(f"- {item}")

    lines.append("Top files:")
    for attachment in ranked[: min(5, len(ranked))]:
        name = str(attachment.get("name") or "attachment")
        summary = str(attachment.get("analysisSummary") or "").strip()
        excerpt = str(attachment.get("analysisText") or attachment.get("analysisCaption") or attachment.get("content") or "").strip()
        line = f"- {name}"
        if summary:
            line += f": {summary}"
        if excerpt:
            line += f" | {_compact_text(excerpt, limit=180)}"
        lines.append(line)

    if compiled.get("clipped"):
        lines.append(
            f"I kept {compiled.get('clipped')} more attachment(s) stored for later turns even though they were not expanded here."
        )

    lines.append("Ask a more specific question if you want me to focus on one file, compare files, or extract exact text.")
    return "\n".join(lines)


def _search_tools_for_llm(query: str, limit: int = 12) -> List[Dict[str, Any]]:
    """Return compact tool search results suitable for prompt injection or function results."""
    matches = tool_registry.search_tools(query)
    compact = []
    for tool in matches[:limit]:
        compact.append({
            "id": tool.get("id"),
            "name": tool.get("name"),
            "category": tool.get("category"),
            "description": tool.get("description"),
            "parameters": tool.get("parameters", [])[:6],
        })
    return compact


def _get_uefn_context(include_assets: bool = False, directory: str = "/Game/", recursive: bool = True, class_filter: str = "", asset_limit: int = 120) -> Dict[str, Any]:
    """Return a combined UEFN snapshot so the assistant can inspect the live project."""
    context: Dict[str, Any] = {
        "connected": False,
        "project": None,
        "level": None,
        "selected_actors": [],
        "viewport": None,
        "assets": [],
    }

    if not discover_uefn_listener_port():
        return context

    context["connected"] = True

    project = _chat_uefn_query("get_project_info", {})
    if project.get("success"):
        context["project"] = project.get("result")

    level = _chat_uefn_query("get_level_info", {})
    if level.get("success"):
        context["level"] = level.get("result")

    selected = _chat_uefn_query("get_selected_actors", {})
    if selected.get("success"):
        result = selected.get("result") or []
        context["selected_actors"] = result[:25] if isinstance(result, list) else result

    viewport = _chat_uefn_query("get_viewport_camera", {})
    if viewport.get("success"):
        context["viewport"] = viewport.get("result")

    if include_assets:
        assets = _chat_uefn_query("list_assets", {
            "directory": directory or "/Game/",
            "recursive": bool(recursive),
            "class_filter": class_filter or "",
        })
        if assets.get("success"):
            result = assets.get("result") or []
            if isinstance(result, list):
                context["assets"] = result[:asset_limit]
                context["asset_count"] = len(result)
            else:
                context["assets"] = result

    return context


_WORKSPACE_SEARCH_EXTENSIONS = {".py", ".ts", ".tsx", ".js", ".jsx", ".json", ".md", ".txt", ".css", ".verse"}
_WORKSPACE_SKIP_DIRS = {"node_modules", "build", "dist", "__pycache__", ".git"}
_WORKSPACE_SEARCH_ROOTS = [
    WORKSPACE_ROOT / "app" / "backend",
    WORKSPACE_ROOT / "app" / "frontend" / "src",
    VENDOR_UEFN / "Content" / "Python" / "UEFN_Toolbelt",
]


def _iter_workspace_source_files(include_vendor: bool = True) -> List[Path]:
    """Return source files that represent app and tool knowledge."""
    files: List[Path] = []
    for root in _WORKSPACE_SEARCH_ROOTS:
        if not include_vendor and "vendor" in str(root).lower():
            continue
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if any(part in _WORKSPACE_SKIP_DIRS for part in path.parts):
                continue
            if path.suffix.lower() not in _WORKSPACE_SEARCH_EXTENSIONS:
                continue
            try:
                if path.stat().st_size > 400_000:
                    continue
            except OSError:
                continue
            files.append(path)
    return files


def _extract_workspace_snippet(content: str, query: str, radius: int = 220) -> str:
    """Extract a compact snippet around the first relevant match."""
    lowered = content.lower()
    query_text = (query or "").lower()
    index = lowered.find(query_text) if query_text else -1
    if index < 0:
        for token in _tokenize_query(query):
            index = lowered.find(token)
            if index >= 0:
                break
    if index < 0:
        index = 0
    start = max(0, index - radius // 2)
    end = min(len(content), index + radius)
    return _compact_text(content[start:end], limit=radius)


def _search_workspace_sources(query: str, limit: int = 8, include_vendor: bool = True) -> List[Dict[str, Any]]:
    """Search relevant app and tool source files for a query and return snippets."""
    query = (query or "").strip()
    if not query:
        return []

    scored: List[tuple[float, Dict[str, Any]]] = []
    for path in _iter_workspace_source_files(include_vendor=include_vendor):
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue

        score = _score_text_match(query, str(path.relative_to(WORKSPACE_ROOT)), text[:60000])
        if score <= 0:
            continue

        scored.append((
            score,
            {
                "path": str(path.relative_to(WORKSPACE_ROOT)).replace("\\", "/"),
                "snippet": _extract_workspace_snippet(text[:60000], query),
            }
        ))

    scored.sort(key=lambda item: (item[0], item[1]["path"]), reverse=True)
    return [item[1] for item in scored[:limit]]


def _get_backend_route_summaries(limit: int = 18) -> List[Dict[str, Any]]:
    """Summarize backend API routes as non-tool capabilities."""
    server_file = WORKSPACE_ROOT / "app" / "backend" / "server.py"
    if not server_file.exists():
        return []

    try:
        text = server_file.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return []

    routes: List[Dict[str, Any]] = []
    pattern = re.compile(r"@app\.route\('([^']+)', methods=\[([^\]]+)\]\)\s+def\s+([a-zA-Z0-9_]+)\(")
    for match in pattern.finditer(text):
        path, methods, func_name = match.groups()
        routes.append({
            "path": path,
            "methods": [method.strip(" '\"") for method in methods.split(",") if method.strip()],
            "handler": func_name,
        })
    return routes[:limit]


def _get_frontend_component_names(limit: int = 24) -> List[str]:
    """List key frontend components so the AI can reason about non-tool UI surfaces."""
    component_dir = WORKSPACE_ROOT / "app" / "frontend" / "src" / "components"
    if not component_dir.exists():
        return []
    names = sorted(path.stem for path in component_dir.glob("*.tsx"))
    return names[:limit]


def _get_server_action_capabilities() -> List[Dict[str, str]]:
    """Describe first-class server actions the AI can call directly across providers."""
    return [
        {"id": "rebuild_gable_roof_action", "description": "Rebuild a clean gable roof from live actor bounds and replace broken roof pieces."},
        {"id": "apply_material_action", "description": "Apply a named material or texture preset to actors matched by label pattern."},
        {"id": "build_house_action", "description": "Build a functional generative residential building with managed slots, structural rules, stairs, floors, windows, roofs, balconies, and planner-backed variation. Use this for houses, cabins, cottages, villas, mansions, townhouses, apartments, and condos."},
        {"id": "build_structure_action", "description": "Build generative structures such as garages, sheds, workshops, barns, warehouses, greenhouses, studios, hangars, kiosks, pavilions, gazebos, pergolas, canopies, carports, market stalls, and other shared-planner structures."},
        {"id": "terrain_action", "description": "Create, modify, list, and delete terrain patches with layered materials, edge surfacing, and environment dressing."},
        {"id": "import_attached_models", "description": "Import attached FBX or ZIP bundles into UEFN, preserve source materials/textures, and place meshes along splines, terrain curves, or patterns."},
    ]


def _get_non_tool_capabilities() -> List[Dict[str, str]]:
    """Describe backend/editor capabilities that are not UEFN Toolbelt tools."""
    return [
        {"id": "chat_memory", "description": "Persistent chat sessions with per-project memory and cross-project recall."},
        {"id": "knowledge_store", "description": "Shared knowledge base for research notes, design references, preferences, and learned tool outcomes."},
        {"id": "workspace_search", "description": "Search app source code, tool source files, and project scripts for relevant files and snippets."},
        {"id": "project_overview", "description": "Summarize backend routes, frontend components, configured providers, and UEFN connection state."},
        {"id": "uefn_queries", "description": "Query live UEFN project info, level state, selection, viewport camera, assets, and editor log."},
        {"id": "uefn_python", "description": "Run arbitrary Python inside the connected UEFN editor when existing tools are not enough."},
        {"id": "attachment_analysis", "description": "Analyze images, screenshots, PDFs, websites, documents, spreadsheets, archives, and cross-file evidence from chat attachments."},
        {"id": "visual_alignment_review", "description": "Compare uploaded visuals against the user request, detect mismatches, and generate grounded fix briefs before editing."},
        {"id": "model_import_runtime", "description": "Materialize uploaded model files into runtime import folders, keep import artifacts, and reuse them in later tool actions."},
        {"id": "research", "description": "Discover web references and synthesize them against the stored knowledge base."},
        {"id": "planning", "description": "Generate structured build plans using tools, knowledge, and live UEFN state."},
    ]


def _get_project_overview(query: str = "", include_assets: bool = False, include_workspace_hits: bool = False) -> Dict[str, Any]:
    """Return a compact project/capabilities snapshot for prompt injection and tool calls."""
    uefn_snapshot = _get_uefn_context(include_assets=include_assets, asset_limit=40)
    overview = {
        "tools": {
            "count": len(tool_registry.tools),
            "categories": sorted(tool_registry.categories.keys()),
        },
        "server_actions": _get_server_action_capabilities(),
        "non_tool_capabilities": _get_non_tool_capabilities(),
        "backend_routes": _get_backend_route_summaries(),
        "frontend_components": _get_frontend_component_names(),
        "uefn": {
            "connected": uefn_snapshot.get("connected", False),
            "project": uefn_snapshot.get("project"),
            "level": uefn_snapshot.get("level"),
            "selected_actors": uefn_snapshot.get("selected_actors"),
        },
    }

    if include_assets:
        overview["uefn"]["assets"] = uefn_snapshot.get("assets", [])
        overview["uefn"]["asset_count"] = uefn_snapshot.get("asset_count", 0)

    if include_workspace_hits and query:
        overview["workspace_hits"] = _search_workspace_sources(query, limit=6)

    return overview


def _search_shared_context(query: str, current_chat_id: str = "") -> Dict[str, Any]:
    """Search persisted knowledge and prior chats for context relevant to the current prompt."""
    return {
        "knowledge_items": knowledge_store.search(query, limit=6),
        "related_chats": chat_store.search_sessions(query, limit=4, exclude_chat_id=current_chat_id),
    }


def _list_chat_attachments(chat_id: str, limit: int = MAX_ATTACHMENT_INDEX_RESULTS) -> List[Dict[str, Any]]:
    session = chat_store.get_session(chat_id) if chat_id else None
    if not session:
        return []

    collected: List[Dict[str, Any]] = []
    for message in session.get("messages", []):
        if message.get("role") != "user":
            continue
        for attachment in message.get("attachments") or []:
            collected.append({
                "name": attachment.get("name") or "attachment",
                "type": attachment.get("type") or "file",
                "mimeType": attachment.get("mimeType") or "",
                "sourceUrl": attachment.get("sourceUrl") or "",
                "analysisCaption": attachment.get("analysisCaption") or "",
                "analysisHandwriting": attachment.get("analysisHandwriting") or "",
                "analysisMeta": attachment.get("analysisMeta") or {},
                "analysisSummary": attachment.get("analysisSummary") or "",
                "analysisText": _compact_text(str(attachment.get("analysisText") or ""), limit=600),
                "analysisKeywords": attachment.get("analysisKeywords") or [],
                "analysisChunkCount": int(attachment.get("analysisChunkCount") or 0),
                "attachmentFingerprint": attachment.get("attachmentFingerprint") or "",
                "contentSnippet": _compact_text(str(attachment.get("content") or ""), limit=420),
                "timestamp": message.get("timestamp") or "",
            })

    return collected[-limit:]


def _analyze_chat_attachments(chat_id: str, query: str = "", limit: int = MAX_ATTACHMENT_ANALYSIS_RESULTS) -> Dict[str, Any]:
    attachments = _list_chat_attachments(chat_id, limit=max(limit * 2, MAX_ATTACHMENT_INDEX_RESULTS))
    if query:
        scored: List[tuple[float, Dict[str, Any]]] = []
        for attachment in attachments:
            score = _score_text_match(
                query,
                str(attachment.get("name") or ""),
                str(attachment.get("analysisCaption") or ""),
                str(attachment.get("analysisHandwriting") or ""),
                json.dumps(attachment.get("analysisMeta") or {}, default=str)[:600],
                str(attachment.get("analysisSummary") or ""),
                str(attachment.get("analysisText") or ""),
                " ".join(str(keyword) for keyword in (attachment.get("analysisKeywords") or [])),
                str(attachment.get("contentSnippet") or ""),
            )
            if score <= 0:
                continue
            attachment["_relevanceScore"] = round(score, 3)
            attachment["_relevanceConfidence"] = _score_to_confidence_label(score)
            scored.append((score, attachment))
        scored.sort(key=lambda item: (item[0], item[1].get("timestamp", "")), reverse=True)
        attachments = [item[1] for item in scored[:limit]]
    else:
        attachments = attachments[-limit:]

    for rank, attachment in enumerate(attachments, start=1):
        attachment["_dossierRank"] = rank
        if "_relevanceScore" not in attachment:
            score = _score_attachment_relevance(query, attachment)
            attachment["_relevanceScore"] = round(score, 3)
            attachment["_relevanceConfidence"] = _score_to_confidence_label(score)

    evidence = _collect_cross_attachment_evidence(query, attachments, limit=min(6, limit if limit > 0 else 6))
    conflicts = _detect_attachment_conflicts(query, attachments)
    return {
        "count": len(attachments),
        "attachments": attachments,
        "evidence": evidence,
        "confidence": _summarize_attachment_confidence(
            query,
            attachments,
            evidence,
            conflicts,
            clipped=0,
        ),
        "conflicts": conflicts,
    }


def _format_retrieval_block(query: str, current_chat_id: str = "") -> str:
    """Render compact retrieved knowledge/project context for the model."""
    query = (query or "").strip()
    if not query:
        return ""

    retrieval = _search_shared_context(query, current_chat_id=current_chat_id)
    lines: List[str] = []

    if retrieval["knowledge_items"]:
        lines.append("RELEVANT SHARED KNOWLEDGE:")
        for item in retrieval["knowledge_items"]:
            lines.append(
                f"- [{item.get('source_type', item.get('type', 'note'))}] {item.get('title', 'Untitled')}: {item.get('content', '')}"
            )

    if retrieval["related_chats"]:
        lines.append("RELATED PRIOR CHATS:")
        for chat in retrieval["related_chats"]:
            excerpt = chat.get("recent_excerpt") or chat.get("memory_summary") or chat.get("preview", "")
            lines.append(f"- {chat.get('title', 'Chat')}: {_compact_text(excerpt, limit=220)}")

    if any(token in query.lower() for token in ["code", "script", "backend", "frontend", "component", "tab", "route", "api", "setting", "tool", "memory"]):
        workspace_hits = _search_workspace_sources(query, limit=5)
        if workspace_hits:
            lines.append("WORKSPACE MATCHES:")
            for hit in workspace_hits:
                lines.append(f"- {hit['path']}: {hit['snippet']}")

    return "\n".join(lines)


_COLOR_PRESET_ALIASES = {
    "red": "team_red",
    "blue": "team_blue",
    "gold": "gold",
    "chrome": "chrome",
    "green": "jade",
    "cyan": "neon",
    "teal": "neon",
    "purple": "plasma",
    "orange": "lava",
    "black": "obsidian",
    "dark": "obsidian",
    "white": "ice",
}

_MODEL_IMPORT_EXTENSIONS = {"fbx"}
_MODEL_IMPORT_ARCHIVE_EXTENSIONS = {"zip"}
_MODEL_IMPORT_PHRASES = (
    "import",
    "bring in",
    "load",
    "use this model",
    "use these models",
    "put this in uefn",
    "put these in uefn",
    "add this model",
    "add these models",
)
_CURVE_PLACEMENT_KEYWORDS = {
    "curve", "curved", "curvature", "arc", "circle", "spiral", "helix",
    "spline", "road", "path", "terrain", "landscape", "follow",
}
_SPLINE_PLACEMENT_KEYWORDS = {"spline", "road", "path", "terrain", "landscape", "follow"}
_TERRAIN_PLACEMENT_KEYWORDS = {"terrain", "landscape", "ground", "surface", "road"}


def _detect_color_preset(message: str) -> Optional[str]:
    """Map simple color-language requests to known material presets."""
    msg_lower = message.lower()
    for color, preset in _COLOR_PRESET_ALIASES.items():
        if re.search(rf"\b{re.escape(color)}\b", msg_lower):
            return preset
    return None


def _safe_runtime_asset_stem(raw_name: str, fallback: str = "ImportedMesh") -> str:
    stem = Path(raw_name or fallback).stem.strip() or fallback
    stem = re.sub(r"[^A-Za-z0-9_]+", "_", stem)
    stem = re.sub(r"_+", "_", stem).strip("_")
    return (stem or fallback)[:64]


def _guess_extension_from_mime(mime_type: str) -> str:
    ext = mimetypes.guess_extension((mime_type or "").strip().lower()) or ""
    if ext.startswith("."):
        ext = ext[1:]
    return ext.lower()


def _normalize_import_attachment(att: Dict[str, Any]) -> Dict[str, Any]:
    normalized = {
        "name": str(att.get("name") or "attachment").strip() or "attachment",
        "type": str(att.get("type") or "file").strip().lower() or "file",
        "mimeType": str(att.get("mimeType") or att.get("mime_type") or "").strip(),
        "content": str(att.get("content") or ""),
        "sourceUrl": str(att.get("sourceUrl") or att.get("source_url") or "").strip(),
        "attachmentFingerprint": str(att.get("attachmentFingerprint") or att.get("attachment_fingerprint") or "").strip(),
        "truncated": bool(att.get("truncated", False)),
    }
    normalized["mimeType"] = _infer_attachment_mime_type(
        normalized["name"],
        normalized["mimeType"],
        normalized["sourceUrl"],
    )
    if not normalized["attachmentFingerprint"]:
        normalized["attachmentFingerprint"] = _build_attachment_fingerprint(_sanitize_chat_attachment(att)) or uuid4().hex
    return normalized


def _attachment_content_to_bytes(att: Dict[str, Any]) -> bytes:
    attachment = _normalize_import_attachment(att)
    content = str(attachment.get("content") or "")
    attachment_type = str(attachment.get("type") or "").lower()
    if attachment_type in {"image", "binary"}:
        _, base64_data = _split_base64_data_url(content)
        if base64_data:
            try:
                return base64.b64decode(base64_data)
            except Exception:
                return b""
    if attachment_type == "file" and content:
        return content.encode("utf-8", errors="ignore")
    return b""


def _materialize_attachment_runtime_file(att: Dict[str, Any], subfolder: str = "") -> Optional[Path]:
    attachment = _normalize_import_attachment(att)
    file_bytes = _attachment_content_to_bytes(attachment)
    if not file_bytes:
        return None

    ext = _file_extension(attachment["name"]) or _guess_extension_from_mime(attachment["mimeType"]) or "bin"
    stem = _safe_runtime_asset_stem(attachment["name"], fallback="ImportedAsset")
    fingerprint = attachment["attachmentFingerprint"] or uuid4().hex
    target_dir = RUNTIME_IMPORT_DIR / (subfolder or fingerprint[:16])
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / f"{stem}.{ext}"
    if not target_path.exists():
        target_path.write_bytes(file_bytes)
    return target_path


def _extract_fbx_files_from_archive(archive_attachment: Dict[str, Any]) -> List[Path]:
    archive_path = _materialize_attachment_runtime_file(archive_attachment, subfolder="archives")
    if archive_path is None or not archive_path.exists():
        return []

    attachment = _normalize_import_attachment(archive_attachment)
    extract_root = RUNTIME_IMPORT_DIR / "archives" / f"{attachment['attachmentFingerprint'][:16]}_fbx"
    extract_root.mkdir(parents=True, exist_ok=True)

    extracted: List[Path] = []
    try:
        with zipfile.ZipFile(archive_path) as zf:
            for member in zf.namelist():
                if Path(member).suffix.lower() != ".fbx":
                    continue
                safe_name = f"{_safe_runtime_asset_stem(Path(member).name, fallback='ImportedMesh')}.fbx"
                out_path = extract_root / safe_name
                if not out_path.exists():
                    with zf.open(member) as src, open(out_path, "wb") as dst:
                        shutil.copyfileobj(src, dst)
                extracted.append(out_path)
    except Exception as exc:
        logger.warning("Failed to extract FBX files from archive %s: %s", archive_path, exc)
        return []

    unique: List[Path] = []
    seen: set[str] = set()
    for path in extracted:
        key = str(path).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _collect_importable_model_files(attachments: List[Dict[str, Any]]) -> List[Path]:
    collected: List[Path] = []
    seen: set[str] = set()
    for raw_attachment in attachments or []:
        attachment = _normalize_import_attachment(raw_attachment)
        ext = _file_extension(attachment["name"]) or _guess_extension_from_mime(attachment["mimeType"])
        paths: List[Path] = []
        if ext in _MODEL_IMPORT_EXTENSIONS:
            materialized = _materialize_attachment_runtime_file(attachment, subfolder="models")
            if materialized is not None:
                paths = [materialized]
        elif ext in _MODEL_IMPORT_ARCHIVE_EXTENSIONS or attachment["mimeType"] in {"application/zip", "application/x-zip-compressed"}:
            paths = _extract_fbx_files_from_archive(attachment)

        for path in paths:
            key = str(path).lower()
            if key in seen:
                continue
            seen.add(key)
            collected.append(path)
    return collected


def _model_import_requested(message: str, attachments: List[Dict[str, Any]]) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    if not _collect_importable_model_files(attachments):
        return False
    return any(phrase in lowered for phrase in _MODEL_IMPORT_PHRASES) or any(
        token in lowered for token in ("model", "models", "mesh", "meshes", "fbx")
    )


def _curve_placement_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    return any(token in lowered for token in _CURVE_PLACEMENT_KEYWORDS)


def _extract_requested_count(message: str, default: int = 12, minimum: int = 1, maximum: int = 128) -> int:
    lowered = (message or "").strip().lower()
    patterns = (
        r"\b(?:count|place|spawn|scatter|instance|instances|props?)\s+(\d{1,3})\b",
        r"\b(\d{1,3})\s+(?:models?|meshes?|props?|instances?)\b",
    )
    for pattern in patterns:
        match = re.search(pattern, lowered)
        if match:
            try:
                return max(minimum, min(maximum, int(match.group(1))))
            except Exception:
                pass
    return max(minimum, min(maximum, default))


def _extract_requested_distance(message: str, default: float = 400.0) -> float:
    lowered = (message or "").strip().lower()
    match = re.search(r"\b(?:every|spacing|spaced|distance)\s+(\d+(?:\.\d+)?)\b", lowered)
    if not match:
        return default
    try:
        return max(50.0, min(5000.0, float(match.group(1))))
    except Exception:
        return default


def _extract_requested_radius(message: str, default: float = 1200.0) -> float:
    lowered = (message or "").strip().lower()
    match = re.search(r"\b(?:radius|arc|circle|helix|spiral)\s+(\d+(?:\.\d+)?)\b", lowered)
    if not match:
        return default
    try:
        return max(100.0, min(20000.0, float(match.group(1))))
    except Exception:
        return default


def _predict_imported_asset_paths(dest_path: str, model_files: List[Path]) -> List[str]:
    asset_paths: List[str] = []
    seen: set[str] = set()
    for model_file in model_files:
        asset_path = f"{dest_path.rstrip('/')}/{_safe_runtime_asset_stem(model_file.name)}"
        key = asset_path.lower()
        if key in seen:
            continue
        seen.add(key)
        asset_paths.append(asset_path)
    return asset_paths


def _build_codex_import_destination(model_files: List[Path]) -> str:
    date_part = datetime.now().strftime("%Y%m%d")
    label = _safe_runtime_asset_stem(model_files[0].name if len(model_files) == 1 else "ImportedBatch", fallback="ImportedBatch")
    return f"/Game/CodexImports/{date_part}/{label}/Meshes"


def _extract_nested_execution_payload(result: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(result, dict):
        return {}
    payload = result.get("result")
    if isinstance(payload, dict) and isinstance(payload.get("result"), dict):
        return payload.get("result") or {}
    if isinstance(payload, dict):
        return payload
    return {}


def _build_fbx_import_execution_code(model_files: List[Path], dest_path: str, message: str) -> str:
    file_paths = [str(path) for path in model_files]
    combine_meshes = "combine" in (message or "").lower()
    return f"""import os
import unreal
file_paths = {json.dumps(file_paths)}
dest_path = {dest_path!r}
combine_meshes = {combine_meshes!r}
asset_tools = unreal.AssetToolsHelpers.get_asset_tools()
eal = unreal.EditorAssetLibrary
if not eal.does_directory_exist(dest_path):
    eal.make_directory(dest_path)
records = []
for file_path in file_paths:
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    task = unreal.AssetImportTask()
    task.filename = file_path
    task.destination_path = dest_path
    task.destination_name = base_name
    task.replace_existing = True
    task.automated = True
    task.save = False
    options = unreal.FbxImportUI()
    options.import_mesh = True
    options.import_textures = True
    options.import_materials = True
    options.import_as_skeletal = False
    try:
        options.static_mesh_import_data.combine_meshes = bool(combine_meshes)
    except Exception:
        pass
    try:
        options.static_mesh_import_data.generate_lightmap_uvs = True
    except Exception:
        pass
    try:
        options.static_mesh_import_data.auto_generate_collision = True
    except Exception:
        pass
    task.options = options
    asset_tools.import_asset_tasks([task])
    imported_paths = [str(p) for p in (task.imported_object_paths or [])]
    mesh_paths = []
    material_paths = []
    texture_paths = []
    for asset_path in imported_paths:
        asset = unreal.EditorAssetLibrary.load_asset(asset_path)
        if isinstance(asset, unreal.StaticMesh):
            mesh_paths.append(asset_path)
        elif isinstance(asset, unreal.MaterialInterface):
            material_paths.append(asset_path)
        elif isinstance(asset, unreal.Texture):
            texture_paths.append(asset_path)
    records.append({{
        "source_file": file_path,
        "imported_paths": imported_paths,
        "mesh_paths": mesh_paths,
        "material_paths": material_paths,
        "texture_paths": texture_paths,
    }})
result = {{
    "status": "ok",
    "dest_path": dest_path,
    "records": records,
    "mesh_paths": [path for record in records for path in record.get("mesh_paths", [])],
    "material_paths": [path for record in records for path in record.get("material_paths", [])],
    "texture_paths": [path for record in records for path in record.get("texture_paths", [])],
}}
"""


def _build_spline_import_placement_code(asset_paths: List[str], message: str, folder_name: str) -> str:
    count = _extract_requested_count(message, default=max(len(asset_paths) * 4, 8))
    spacing_distance = _extract_requested_distance(message, default=0.0)
    random_yaw = 0.0 if any(token in (message or "").lower() for token in ("exact", "clean", "match")) else 4.0
    snap_to_floor = any(token in (message or "").lower() for token in ("terrain", "landscape", "ground", "surface", "road"))
    code = f"""import unreal
mesh_paths = {json.dumps(asset_paths)}
folder_name = {folder_name!r}
spacing_distance = {spacing_distance!r}
target_count = {count!r}
random_yaw = {random_yaw!r}
snap_to_floor = {snap_to_floor!r}
actor_sub = unreal.get_editor_subsystem(unreal.EditorActorSubsystem)
selected = actor_sub.get_selected_level_actors() or []
spline_actor = None
spline_comp = None
for actor in selected:
    comp = actor.get_component_by_class(unreal.SplineComponent)
    if comp is not None:
        spline_actor = actor
        spline_comp = comp
        break
if spline_comp is None:
    result = {{"status": "error", "message": "Select a spline or landscape spline actor before placing imported models along terrain curvature."}}
else:
    loaded_meshes = []
    for mesh_path in mesh_paths:
        asset = unreal.EditorAssetLibrary.load_asset(mesh_path)
        if isinstance(asset, unreal.StaticMesh):
            bounds = asset.get_bounds()
            extent = getattr(bounds, "box_extent", None)
            width_hint = 200.0
            height_hint = 50.0
            if extent is not None:
                width_hint = max(float(extent.x) * 2.0, float(extent.y) * 2.0, 100.0)
                height_hint = max(float(extent.z), 10.0)
            loaded_meshes.append((mesh_path, asset, width_hint, height_hint))
    if not loaded_meshes:
        result = {{"status": "error", "message": "Imported meshes could not be loaded from the Content Browser."}}
    else:
        total_length = spline_comp.get_spline_length()
        default_spacing = max(loaded_meshes[0][2] * 1.15, 220.0)
        spacing_value = spacing_distance if spacing_distance > 0 else default_spacing
        max_count_from_length = max(1, int(total_length / max(spacing_value, 1.0)) + 1)
        if spacing_value > 0:
            distances = []
            current = 0.0
            while current <= total_length and len(distances) < min(max(target_count, 1), max_count_from_length):
                distances.append(current)
                current += spacing_value
            if len(distances) < 2:
                distances = [0.0, total_length]
        else:
            bounded_count = min(max(target_count, 1), max_count_from_length)
            distances = [total_length * i / max(bounded_count - 1, 1) for i in range(max(bounded_count, 1))]
        placed = []
        cs = unreal.SplineCoordinateSpace.WORLD
        with unreal.ScopedEditorTransaction("Codex Place Imported Meshes Along Spline"):
            for idx, dist in enumerate(distances):
                mesh_path, mesh_asset, width_hint, height_hint = loaded_meshes[idx % len(loaded_meshes)]
                loc = spline_comp.get_location_at_distance_along_spline(dist, cs)
                tangent = spline_comp.get_tangent_at_distance_along_spline(dist, cs)
                rot = unreal.MathLibrary.make_rot_from_x(tangent)
                if random_yaw > 0:
                    rot = unreal.Rotator(rot.pitch, rot.yaw + ((idx % 3) - 1) * random_yaw, rot.roll)
                loc = unreal.Vector(loc.x, loc.y, loc.z + max(height_hint * 0.92, 3.0))
                actor = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, loc)
                if actor is None:
                    continue
                mesh_comp = actor.get_component_by_class(unreal.StaticMeshComponent)
                if mesh_comp is None:
                    continue
                mesh_comp.set_static_mesh(mesh_asset)
                actor.set_actor_rotation(rot, False)
                actor.set_folder_path(f"/{{folder_name}}")
                actor.set_actor_label(f"{{folder_name}}_{{idx:03d}}")
                placed.append(actor)
        if snap_to_floor and placed:
            try:
                actor_sub.set_selected_level_actors(placed)
                unreal.EditorLevelLibrary.snap_objects_to_floor()
                for actor in placed:
                    current_loc = actor.get_actor_location()
                    actor.set_actor_location(unreal.Vector(current_loc.x, current_loc.y, current_loc.z + 1.5), False, False)
            except Exception:
                pass
        instance_summaries = []
        for actor in placed[:16]:
            loc = actor.get_actor_location()
            instance_summaries.append({{"label": actor.get_actor_label(), "location": {{"x": loc.x, "y": loc.y, "z": loc.z}}}})
        result = {{
            "status": "ok",
            "placed": len(placed),
            "folder": folder_name,
            "spline_actor": spline_actor.get_actor_label(),
            "assets": mesh_paths,
            "used_spacing": spacing_value,
            "snap_to_floor": snap_to_floor,
            "instances": instance_summaries,
        }}
"""
    return code


def _get_selected_actor_origin() -> tuple[float, float, float]:
    selected = _chat_uefn_query("get_selected_actors", {})
    payload = selected.get("result") if isinstance(selected, dict) else []
    actors = payload if isinstance(payload, list) else []
    if not actors:
        return (0.0, 0.0, 0.0)
    actor = actors[0]
    if not isinstance(actor, dict):
        return (0.0, 0.0, 0.0)
    location = actor.get("location") or actor.get("loc") or {}
    if isinstance(location, dict):
        try:
            return (
                float(location.get("x", 0.0) or 0.0),
                float(location.get("y", 0.0) or 0.0),
                float(location.get("z", 0.0) or 0.0),
            )
        except Exception:
            return (0.0, 0.0, 0.0)
    return (0.0, 0.0, 0.0)


def _place_imported_assets_on_requested_curve(message: str, asset_paths: List[str]) -> Optional[Dict[str, Any]]:
    lowered = (message or "").strip().lower()
    if not asset_paths or not _curve_placement_requested(message):
        return None

    folder_name = f"CodexCurve_{datetime.now().strftime('%H%M%S')}"
    token_set = set(re.findall(r"[a-z0-9]+", lowered))

    if _SPLINE_PLACEMENT_KEYWORDS.intersection(token_set):
        if _TERRAIN_PLACEMENT_KEYWORDS.intersection(token_set) and any(token in token_set for token in {"edge", "shoulder", "roadside", "sides"}):
            road_edge_result = tool_registry.execute_tool("scatter_road_edge", {
                "mesh_path": asset_paths[0],
                "sample_spacing": _extract_requested_distance(message, default=500.0),
                "count_per_sample": 1,
                "edge_offset": 350.0,
                "both_sides": True,
                "snap_to_surface": True,
                "folder": folder_name,
            })
            if road_edge_result.get("success"):
                return {
                    "mode": "terrain_curve",
                    "tool": "scatter_road_edge",
                    "output": road_edge_result,
                }
        spline_result = _handle_tool_call("execute_python_in_uefn", {
            "code": _build_spline_import_placement_code(asset_paths, message, folder_name),
        })
        return {
            "mode": "spline_curve",
            "tool": "execute_python_in_uefn",
            "output": spline_result,
        }

    pattern_tool = "pattern_arc"
    if "circle" in lowered:
        pattern_tool = "pattern_circle"
    elif "spiral" in lowered:
        pattern_tool = "pattern_spiral"
    elif "helix" in lowered:
        pattern_tool = "pattern_helix"

    ox, oy, oz = _get_selected_actor_origin()
    params: Dict[str, Any] = {
        "mesh_path": asset_paths[0],
        "count": _extract_requested_count(message, default=12),
        "radius": _extract_requested_radius(message, default=1200.0),
        "origin": [ox, oy, oz],
        "rotation_mode": "face_tangent",
    }
    if pattern_tool == "pattern_arc":
        params["start_angle_deg"] = 0.0
        params["end_angle_deg"] = 180.0
        params["rotation_mode"] = "face_tangent"
    elif pattern_tool == "pattern_spiral":
        params.pop("radius", None)
        params["turns"] = 2.0
        params["radius_start"] = 250.0
        params["radius_end"] = _extract_requested_radius(message, default=1600.0)
    elif pattern_tool == "pattern_helix":
        params["turns"] = 2.0
        params["rise_per_turn"] = 300.0

    pattern_result = tool_registry.execute_tool(pattern_tool, params)
    return {
        "mode": "pattern_curve",
        "tool": pattern_tool,
        "output": pattern_result,
    }


def _maybe_import_and_place_model_attachments(message: str, attachments: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    import_requested = _model_import_requested(message, attachments)
    placement_requested = _curve_placement_requested(message)
    if not import_requested and not placement_requested:
        return None

    model_files = _collect_importable_model_files(attachments)
    if not model_files:
        return None

    if not discover_uefn_listener_port():
        return {
            "reply": "UEFN is not connected, so I can’t import or place the uploaded models yet. Open the project and try again.",
            "tool_result": {
                "tool": "import_models",
                "output": {"success": False, "error": "UEFN not connected"},
            },
        }

    dest_path = _build_codex_import_destination(model_files)
    import_result = _handle_tool_call("execute_python_in_uefn", {
        "code": _build_fbx_import_execution_code(model_files, dest_path, message),
    })
    import_payload = _extract_nested_execution_payload(import_result)
    imported_mesh_paths = [str(path) for path in (import_payload.get("mesh_paths") or []) if str(path).strip()]
    imported_material_paths = [str(path) for path in (import_payload.get("material_paths") or []) if str(path).strip()]
    imported_texture_paths = [str(path) for path in (import_payload.get("texture_paths") or []) if str(path).strip()]
    import_outputs: List[Dict[str, Any]] = []
    if isinstance(import_payload.get("records"), list):
        for record in import_payload.get("records") or []:
            import_outputs.append({
                "source_file": str(record.get("source_file") or ""),
                "output": {
                    "success": bool(record.get("mesh_paths") or record.get("imported_paths")),
                    "result": record,
                },
            })

    successes = [record for record in import_outputs if record["output"].get("success")]
    predicted_assets = imported_mesh_paths or _predict_imported_asset_paths(dest_path, model_files)
    placement_output = _place_imported_assets_on_requested_curve(message, predicted_assets) if successes and placement_requested else None

    if not successes:
        return {
            "reply": "I found the uploaded model files, but UEFN rejected the import. Check the import log/output in UEFN and try again with FBX assets.",
            "tool_result": {
                "tool": "import_models",
                "output": {"success": False, "imports": import_outputs, "raw_import_result": import_result},
            },
        }

    placed_count = 0
    placement_mode = ""
    placement_success = False
    if placement_output:
        output_payload = placement_output.get("output") or {}
        placement_success = bool(output_payload.get("success") or (output_payload.get("result") or {}).get("success"))
        nested_result = output_payload.get("result") if isinstance(output_payload.get("result"), dict) else output_payload
        if isinstance(nested_result, dict):
            placed_count = int(nested_result.get("placed") or nested_result.get("count") or 0)
        placement_mode = placement_output.get("mode") or ""

    imported_count = len(successes)
    file_label = "model" if imported_count == 1 else "models"
    reply = f"Imported {imported_count} {file_label} into `{dest_path}`."
    if imported_material_paths or imported_texture_paths:
        reply += " I kept the source materials and textures from the FBX import so the meshes stay visually coherent."
    if placement_output:
        if placement_success or placed_count > 0:
            if placement_mode == "terrain_curve":
                reply += " I also placed them along the selected terrain curve/road edge, conformed them to the surface, and kept the layout in a sensible local span."
            elif placement_mode == "spline_curve":
                reply += " I also placed them along the selected spline with bounds-aware spacing and a ground snap pass so they follow the curve cleanly without sinking."
            else:
                reply += " I also laid out the first imported mesh on the requested curved pattern with tangent-facing rotation."
        else:
            reply += " The import worked, but curved placement still needs a selected spline or landscape spline target in UEFN."

    return {
        "reply": reply,
        "tool_result": {
            "tool": "import_models",
            "output": {
                "success": True,
                "dest_path": dest_path,
                "asset_paths": predicted_assets,
                "material_paths": imported_material_paths,
                "texture_paths": imported_texture_paths,
                "imports": import_outputs,
                "raw_import_result": import_result,
                "placement": placement_output,
            },
        },
    }


def _uefn_edit_requested(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    return bool(re.search(
        r"\b(fix|change|edit|adjust|improve|update|delete|remove|move|rotate|scale|align|color|paint|set|replace|add|create|spawn|build|import|place|scatter)\b",
        lowered,
    ))


def _build_attachment_fix_brief(attachments: list) -> str:
    if not attachments:
        return ""
    ranked = _build_attachment_dossier("fix the issue shown here", attachments, history_mode=False).get("ranked") or []
    if not ranked:
        return ""
    attachment = ranked[0]
    analysis_meta = attachment.get("analysisMeta") or attachment.get("analysis_meta") or {}
    lines: List[str] = []
    caption = str(attachment.get("analysisCaption") or attachment.get("analysis_caption") or "").strip()
    summary = str(attachment.get("analysisSummary") or attachment.get("analysis_summary") or "").strip()
    visible_text = str(attachment.get("analysisText") or attachment.get("analysis_text") or "").strip()
    if summary:
        lines.append(f"- Attachment issue summary: {summary}")
    elif caption:
        lines.append(f"- Attachment scene summary: {_compact_text(caption, limit=220)}")
    if visible_text:
        lines.append(f'- Visible text in the attachment: "{_compact_text(visible_text, limit=160)}"')
    detected_objects = [str(item).strip() for item in (analysis_meta.get("detectedObjects") or []) if str(item).strip()]
    if detected_objects:
        lines.append(f"- Likely objects involved: {', '.join(detected_objects[:6])}")
    visual_meta = _render_visual_metadata_summary(analysis_meta if isinstance(analysis_meta, dict) else {})
    if visual_meta:
        lines.append(f"- Visual diagnostics: {visual_meta}")
    return "\n".join(lines)


def _build_execution_precheck_block(message: str, attachments: list) -> str:
    if not _uefn_edit_requested(message):
        return ""
    if not discover_uefn_listener_port():
        return ""

    try:
        context = _get_uefn_context(include_assets=False)
    except Exception as exc:
        logger.warning("Execution precheck context failed: %s", exc)
        context = {"connected": True, "selected_actors": [], "level": None}

    lines = ["LIVE EXECUTION PRECHECK:"]
    selected = context.get("selected_actors") or []
    if isinstance(selected, list) and selected:
        selected_names: List[str] = []
        for actor in selected[:8]:
            if isinstance(actor, str):
                selected_names.append(actor)
            elif isinstance(actor, dict):
                selected_names.append(str(actor.get("name") or actor.get("label") or actor))
            else:
                selected_names.append(str(actor))
        lines.append(f"- Current selection ({len(selected)}): {', '.join(name for name in selected_names if name)}")
        lines.append("- Modify the selected actors first unless a live query proves the target is elsewhere.")
    else:
        lines.append("- No actors are currently selected.")
        lines.append("- If the target is not explicitly named, identify it with query_uefn/get_uefn_context before editing anything.")

    level = context.get("level") or {}
    if isinstance(level, dict) and level:
        level_name = str(level.get("world_name") or level.get("map_name") or "").strip()
        actor_count = level.get("actor_count")
        if level_name or actor_count is not None:
            lines.append(
                f"- Level context: {level_name or 'Unknown level'}"
                + (f" with {actor_count} actor(s)." if actor_count is not None else ".")
            )

    attachment_brief = _build_attachment_fix_brief(attachments)
    if attachment_brief:
        lines.append(attachment_brief)

    tool_routing = _build_tool_routing_guidance(message, attachments)
    if tool_routing:
        lines.append(tool_routing)
    lowered = (message or "").strip().lower()
    if any(token in lowered for token in ("curve", "curved", "curvature", "spline", "terrain", "landscape", "road", "path")):
        lines.append("- For curve or terrain-following placement, prefer the currently selected spline or landscape spline and align imported meshes to its tangent.")

    lines.append("- Never modify broad actor sets blindly from get_all_level_actors(). Narrow to a small target set first.")
    lines.append("- After any world change, verify with a live query or screenshot before claiming the fix is complete.")
    lines.append("- If the target is still ambiguous after querying, ask the user to select the exact actor instead of guessing.")
    return "\n".join(lines)


def _validate_uefn_python_action(code: str) -> Optional[str]:
    lowered = (code or "").strip().lower()
    if not lowered:
        return "Blocked Python action: no code was provided."

    mutates_world = any(token in lowered for token in (
        "destroy_actor(",
        ".set_actor_location(",
        ".set_actor_rotation(",
        ".set_actor_scale3d(",
        ".set_material(",
        "spawn_actor_from_class(",
    ))
    if mutates_world and "result =" not in lowered:
        return "Blocked Python action: end the editor script with `result = ...` so the change can be verified."

    broad_actor_loop = (
        "get_all_level_actors()" in lowered
        and re.search(r"for\s+\w+\s+in\s+actors\s*:", lowered)
        and mutates_world
    )
    has_targeting = any(token in lowered for token in (
        "get_selected",
        "selected_actor",
        "selected_actors",
        "if ",
        "filter",
        "find",
        "match",
        "get_name(",
        "get_actor_label(",
        "[a for a in actors if",
    ))
    if broad_actor_loop and not has_targeting:
        return (
            "Blocked unsafe Python action: it would mutate every actor returned by "
            "`get_all_level_actors()` without narrowing the target first."
        )

    return None


def _message_requests_waterfall_scene(message: str) -> bool:
    lowered = (message or "").strip().lower()
    if not lowered:
        return False
    has_waterfall = any(token in lowered for token in ("waterfall", "cascade"))
    if not has_waterfall:
        return False
    has_build_intent = any(
        token in lowered
        for token in ("create", "make", "build", "add", "generate", "spawn", "place", "fix", "redo", "rebuild")
    )
    return has_build_intent


def _infer_waterfall_scene_material(message: str) -> str:
    lowered = (message or "").strip().lower()
    if any(token in lowered for token in ("snow", "icy", "ice", "frozen", "arctic")):
        return "snow"
    if any(token in lowered for token in ("desert", "sand", "sandy", "dune")):
        return "sand"
    if any(token in lowered for token in ("swamp", "mud", "wetland", "bog")):
        return "mud"
    if any(token in lowered for token in ("rock", "rocky", "cliff", "mountain", "stone")):
        return "rock"
    return "grass"


def _resolve_direct_build_anchor() -> tuple[Dict[str, float], str]:
    import math

    context = _get_uefn_context(include_assets=False)
    selected = context.get("selected_actors") or []
    for actor in selected:
        if not isinstance(actor, dict):
            continue
        location = actor.get("location") or actor.get("position")
        if not isinstance(location, dict):
            continue
        try:
            return (
                {
                    "x": float(location.get("x", 0.0)),
                    "y": float(location.get("y", 0.0)),
                    "z": float(location.get("z", 0.0)),
                },
                f"near the selected actor `{actor.get('name') or actor.get('label') or 'selection'}`",
            )
        except Exception:
            continue

    viewport = context.get("viewport") or {}
    location = viewport.get("location") or {}
    rotation = viewport.get("rotation") or {}
    if isinstance(location, dict) and location:
        try:
            yaw = math.radians(float(rotation.get("yaw", 0.0)))
            distance = 2200.0
            return (
                {
                    "x": round(float(location.get("x", 0.0)) + math.cos(yaw) * distance, 2),
                    "y": round(float(location.get("y", 0.0)) + math.sin(yaw) * distance, 2),
                    "z": round(max(0.0, float(location.get("z", 0.0)) - 700.0), 2),
                },
                "in front of the current viewport",
            )
        except Exception:
            pass

    return ({"x": 0.0, "y": 0.0, "z": 0.0}, "at the world origin")


def _execute_direct_waterfall_scene(message: str) -> Dict[str, Any]:
    msg_lower = (message or "").strip().lower()
    anchor, anchor_source = _resolve_direct_build_anchor()
    terrain_material = _infer_waterfall_scene_material(message)
    wants_large = any(token in msg_lower for token in ("large", "big", "huge", "massive"))
    terrain_type = "ridge" if any(token in msg_lower for token in ("hill", "hills", "ridge", "cliff", "mountain", "terrain", "landform", "slope")) else "hill"
    terrain_size = {"x": 5200.0, "y": 3200.0} if wants_large else {"x": 4200.0, "y": 2600.0}
    terrain_elevation = 1100.0 if wants_large else 820.0
    terrain_label = f"Terrain_WaterfallScene_{int(time.time())}"

    terrain_result = _execute_terrain_control({
        "operation": "create",
        "terrain_type": terrain_type,
        "position": anchor,
        "size": terrain_size,
        "material": terrain_material,
        "height": 0.0,
        "elevation": terrain_elevation,
        "label": terrain_label,
        "subdivisions": 3,
        "decorate": False,
    })
    if not (terrain_result.get("success") or terrain_result.get("result")):
        return {
            "reply": f"I tried to create the waterfall terrain, but the terrain action failed: {terrain_result.get('error', 'unknown error')}",
            "tool_result": {"tool": "terrain_action", "output": terrain_result},
        }

    waterfall_anchor = {
        "x": anchor["x"],
        "y": anchor["y"],
        "z": anchor["z"] + max(180.0, terrain_elevation * 0.24),
    }
    structure_result = _execute_generative_build({
        "structure": "waterfall",
        "position": waterfall_anchor,
        "size": "large" if wants_large else "medium",
        "material": "cliff",
        "support_material": "cliff",
    })
    if not (structure_result.get("success") or structure_result.get("result")):
        return {
            "reply": (
                "I created the landform, but the continuous waterfall build failed: "
                f"{structure_result.get('error', 'unknown error')}"
            ),
            "tool_result": {
                "tool": "build_structure_action",
                "output": {
                    "terrain": terrain_result,
                    "structure": structure_result,
                },
            },
        }

    return {
        "reply": (
            f"I created a {terrain_type} landform {anchor_source} and built a continuous waterfall on it. "
            f"The terrain uses layered `{terrain_material}` surfacing and the waterfall uses dedicated water/cliff pieces, "
            "so this should avoid the old stair-step cube result."
        ),
        "tool_result": {
            "tool": "waterfall_scene_action",
            "output": {
                "terrain": terrain_result,
                "structure": structure_result,
                "anchor": anchor,
                "anchor_source": anchor_source,
                "terrain_type": terrain_type,
                "terrain_material": terrain_material,
            },
        },
    }


def _maybe_execute_direct_action(message: str, attachments: Optional[List[Dict[str, Any]]] = None) -> Optional[Dict[str, Any]]:
    """High-confidence shortcut actions that should work even before the LLM reasons."""
    if _visual_review_requested(message) or _visual_comparison_requested(message):
        return None

    import_result = _maybe_import_and_place_model_attachments(message, attachments or [])
    if import_result is not None:
        return import_result

    msg_lower = message.lower()
    if _message_requests_waterfall_scene(message):
        return _execute_direct_waterfall_scene(message)

    direct_structure_request = _detect_direct_structure_request(message)
    if direct_structure_request is not None:
        if direct_structure_request.get("kind") == "house":
            return _execute_build_house({
                "request": message,
                "style": "",
            })
        if direct_structure_request.get("kind") == "structure":
            return _execute_build_structure_request({
                "request": message,
                "structure": str(direct_structure_request.get("structure_type") or ""),
                "style": "",
            })

    color_words = "|".join(re.escape(color) for color in _COLOR_PRESET_ALIASES.keys())
    looks_like_color_request = any(phrase in msg_lower for phrase in [
        "color this", "make this", "turn this", "paint this",
        "color selected", "make selected", "paint selected",
    ]) or bool(re.search(
        rf"\b(color|paint|turn|make|set)\b[\w\s-]{{0,60}}\b(?:to\s+)?(?:{color_words})\b",
        msg_lower,
    ))
    preset = _detect_color_preset(message)

    if looks_like_color_request and preset:
        selection = _chat_uefn_query("get_selected_actors", {})
        if not selection.get("success"):
            return {
                "reply": f"I can do that, but I couldn't inspect the current selection: {selection.get('error', 'unknown error')}",
                "tool_result": {"tool": "get_selected_actors", "output": selection},
            }

        selected = selection.get("result") or []
        if isinstance(selected, list) and not selected:
            return {
                "reply": "I can apply that material, but nothing is selected in UEFN. Select the target actors and ask again.",
                "tool_result": {"tool": "get_selected_actors", "output": selection},
            }

        tool_result = tool_registry.execute_tool("material_apply_preset", {"preset": preset})
        if tool_result.get("success"):
            return {
                "reply": f"Applied the `{preset}` material preset to the current selection.",
                "tool_result": {"tool": "material_apply_preset", "output": tool_result},
            }
        return {
            "reply": f"I tried to apply `{preset}`, but the tool failed: {tool_result.get('error', 'unknown error')}",
            "tool_result": {"tool": "material_apply_preset", "output": tool_result},
        }

    # ── Roof rebuild detection ──
    roof_keywords = ('fix the roof', 'rebuild the roof', 'fix roof', 'rebuild roof',
                     'roof is broken', 'roof looks broken', 'fix my roof',
                     'clean the roof', 'redo the roof', 'roof is messed up',
                     'roof clipping', 'roof crossing', 'roof overlapping',
                     'roof is still broken', 'roof still broken', 'roof still looks',
                     'roof doesnt look', "roof doesn't look", 'roof is wrong',
                     'roof is bad', 'roof looks bad', 'roof looks wrong')
    # Also detect if "roof" appears with fix/broken/rebuild intent words
    has_roof = 'roof' in msg_lower
    has_intent = any(w in msg_lower for w in ('fix', 'rebuild', 'broken', 'redo', 'repair', 'wrong', 'bad', 'messed'))
    if any(phrase in msg_lower for phrase in roof_keywords) or (has_roof and has_intent):
        logger.info("Roof rebuild detected in message: %s", message[:100])
        port = discover_uefn_listener_port()
        if port:
            try:
                result = _execute_rebuild_gable_roof({"roof_height_ratio": 0.3, "overhang": 30})
                logger.info("Roof rebuild result keys: %s", list(result.keys()) if isinstance(result, dict) else type(result))
                if result.get("success") or result.get("result"):
                    details = result.get("details", {})
                    return {
                        "reply": (
                            f"Rebuilt the roof cleanly. "
                            f"Pitch={details.get('pitch_deg', 0):.1f}°, "
                            f"house={details.get('house_size', {}).get('width', 0):.0f}x{details.get('house_size', {}).get('depth', 0):.0f}, "
                            f"ridge at z={details.get('ridge_z', 0):.0f}, "
                            f"overhang={details.get('overhang', 0):.0f}. "
                            f"Deleted all old pieces, created 2 slopes + ridge + 2 gable panels with validation."
                        ),
                        "tool_result": {"tool": "rebuild_gable_roof", "output": result},
                    }
                elif result.get("error"):
                    logger.warning("Roof rebuild error: %s", result["error"])
                else:
                    logger.warning("Roof rebuild returned no success: %s", json.dumps(result, default=str)[:500])
            except Exception as e:
                logger.warning("Auto roof rebuild failed: %s", e, exc_info=True)

    return None


_PROMPT_REFERENCE_CACHE: Dict[str, str] = {}


def _load_prompt_reference_text(filename: str) -> str:
    cached = _PROMPT_REFERENCE_CACHE.get(filename)
    if cached is not None:
        return cached
    path = WORKSPACE_ROOT / "config" / "codex_prompts" / filename
    try:
        text = path.read_text(encoding="utf-8-sig").strip()
    except OSError:
        text = ""
    _PROMPT_REFERENCE_CACHE[filename] = text
    return text


def _shared_house_generation_grounding() -> str:
    sections = [
        ("Planner Grounding", _load_prompt_reference_text("system_prompt.md")),
        ("UEFN Quality Grounding", _load_prompt_reference_text("uefn_quality_grounding.md")),
        ("House Generation Grounding", _load_prompt_reference_text("uefn_house_generation_grounding.md")),
        ("Structure Generation Grounding", _load_prompt_reference_text("uefn_structure_generation_grounding.md")),
        ("Tool Execution Grounding", _load_prompt_reference_text("uefn_tool_execution_grounding.md")),
    ]
    compiled: list[str] = []
    for title, body in sections:
        if not body:
            continue
        compiled.append(f"{title}:\n{body}")
    return "\n\n".join(compiled)


def _build_system_prompt(chat_title: str = "", chat_memory: str = "") -> str:
    """Build the AI system prompt with tool, memory, and UEFN context."""
    uefn_connected = uefn_bridge.is_connected
    tool_catalog = tool_registry.get_tools_for_ai()
    server_actions = "\n".join(
        f"- {item['id']}: {item['description']}"
        for item in _get_server_action_capabilities()
    )
    non_tool_capabilities = "\n".join(
        f"- {item['id']}: {item['description']}"
        for item in _get_non_tool_capabilities()
    )
    memory_block = ""
    if chat_title or chat_memory:
        memory_block = "\nSESSION MEMORY:\n"
        if chat_title:
            memory_block += f"- Chat title: {chat_title}\n"
        if chat_memory:
            memory_block += f"{chat_memory}\n"
    grounding_block = _shared_house_generation_grounding()
    if grounding_block:
        grounding_block = f"\nSHARED BUILD GROUNDING:\n{grounding_block}\n"

    return f"""You are the UEFN Codex AI Assistant — a powerful AI that can directly control and modify Unreal Editor for Fortnite (UEFN).

CURRENT STATE:
- UEFN Editor: {"CONNECTED — you have FULL CONTROL to read, modify, delete, spawn, move, and color anything" if uefn_connected else "NOT CONNECTED — tell the user to open UEFN with the listener"}
{memory_block}
{grounding_block}
REGISTERED TOOLS (use run_uefn_tool with these IDs):
{tool_catalog}

SERVER ACTIONS (first-class callable actions available to every model path):
{server_actions}

NON-TOOL CAPABILITIES:
{non_tool_capabilities}

YOUR TOOLS (use these to take action):
1. query_uefn — Get live data: actors, level info, project info, selected actors, viewport camera, assets
   Commands: get_all_actors, get_level_info, get_project_info, get_selected_actors, get_viewport_camera, get_actor_properties, list_assets
2. run_uefn_tool — Execute ANY registered tool from the catalog above. Pass the tool ID and parameters.
3. execute_python_in_uefn — Run arbitrary Python inside UEFN (import unreal). Use when no tool covers the task.
4. search_tools — Find the right tool when you're unsure which one to use
5. get_tool_details — Inspect a tool's parameters before running it
6. get_uefn_context — Fetch combined snapshot: project + level + selection + viewport + assets
7. search_knowledge — Search user-uploaded docs, notes, and knowledge base
8. save_knowledge — ONLY save when the user EXPLICITLY asks you to remember something
9. get_project_overview — Inspect app capabilities beyond tools
10. search_workspace — Search the codebase for implementation details
11. analyze_chat_attachments — Inspect uploaded files and screenshots already attached in the current chat
12. rebuild_gable_roof_action — Rebuild a clean gable roof from live actor bounds
13. apply_material_action — Apply a named material to actor labels or groups
14. build_house_action — Build a functional, generative residential building using the shared structure planner, managed slots, and house rules. Use this for houses, cabins, cottages, villas, mansions, townhouses, apartments, and condos.
15. build_structure_action — Build a generative garage, shed, workshop, barn, warehouse, greenhouse, studio, hangar, kiosk, pavilion, gazebo, pergola, canopy, carport, market stall, or scenic structure through shared geometry code
16. terrain_action — Create or modify terrain with layered materials and environment dressing
17. import_attached_models — Import attached FBX/ZIP models and place them along splines, terrain, or patterns

═══════════════════════════════════════════════════════════════
ABSOLUTE RULES — NEVER BREAK THESE:
═══════════════════════════════════════════════════════════════

RULE 1 — EXECUTE, NEVER SUGGEST ALTERNATIVES:
When the user asks you to DO something, you MUST do it. Never respond with "I can help you in a few ways" or "Here are some options". Just DO IT.
- User says "add a roof here" → execute Python to spawn roof geometry NOW
- User says "delete the blue walls" → query actors, find blue walls, delete them NOW
- User says "move this to the left" → get selected actor, move it NOW
- If no perfect tool exists, use execute_python_in_uefn to write the code that does it
- NEVER say "I don't have a direct tool for X". You have Python — you can do ANYTHING
- If the target is ambiguous, DO NOT guess. Narrow it with live queries/selection first, and only ask the user to select the exact actor if live data still leaves multiple plausible targets.

RULE 2 — YOU CAN SEE AND UNDERSTAND ALL FILES:
You have FULL vision and document understanding. When the user attaches or pastes:
- IMAGES (screenshots, photos, diagrams, mockups): You can SEE them. Describe what you see. Identify UI elements, 3D objects, layouts, colors, text in images, errors, code in screenshots. If they show a UEFN viewport, identify the actors, geometry, materials, and spatial layout.
- PDFs: You can read EVERY page. Extract text, tables, diagrams, code snippets, formulas, and structure. Understand the document as a whole — not just individual words.
- CODE FILES (.py, .js, .ts, .verse, .cpp, .h, .json, .md, .txt): Parse and understand the code. Identify functions, classes, bugs, patterns. You can execute Verse/Python/Unreal code from these files directly.
- MULTIPLE FILES: When multiple files are attached, cross-reference them. Compare, merge, compile a unified analysis. Find connections between documents. If asked to summarize, produce ONE cohesive summary across ALL files.

RULE 3 — DEEP FILE ANALYSIS:
When analyzing any attached file:
- Read the ENTIRE content, not just the first few lines
- Identify the PURPOSE and CONTEXT of the document
- Extract ALL actionable information: dimensions, coordinates, asset paths, color values, settings, parameters
- If the file contains instructions or a plan, EXECUTE those instructions using tools
- If the file shows a UEFN screenshot, map what you see to actual actors in the level using query_uefn
- If multiple files are uploaded together, treat them as related context and cross-reference

RULE 4 — ALWAYS USE TOOLS:
- ALWAYS USE TOOLS when the user asks about their level, actors, or project. Never guess — query first.
- ALWAYS EXECUTE ACTIONS when the user asks to change, delete, move, color, or spawn anything. Use tools, don't just describe steps.
- Chain multiple tools when needed: query_uefn to find actors → run_uefn_tool or execute_python_in_uefn to modify them.
- When deleting/modifying: first query to find the exact actor names, then execute the action.
- When the user says "search for X and delete it": use query_uefn(get_all_actors) to find matches, then execute_python_in_uefn to delete them.
- If a user asks you to fix something shown in an image or document, inspect the attachment context, query live UEFN state, and then execute the change.
- Prefer run_uefn_tool when a registered tool fits. Use execute_python_in_uefn for complex multi-step operations.
- For houses and architectural requests, prefer `build_house_action` or `build_structure_action` before writing ad-hoc cube-spawn Python. Those shared planners handle managed reuse, structure fit, support selection, and variation better than improvised code.
- For uploaded FBX/model imports, prefer import_fbx or import_fbx_folder first, then place imported meshes using spline_place_props, scatter_road_edge, pattern_arc, pattern_circle, pattern_spiral, or pattern_helix.
- If model attachments are present and the user wants those assets used, run `import_attached_models` first and only fall back to generated placeholder geometry if no importable assets exist.
- For terrain, ground, ridge, plateau, hill, crater, or slope requests, prefer the dedicated `terrain` action instead of ad-hoc cube spawning. Use broad continuous terrain first, pick balanced top/edge materials, and keep long roads/ridges/shorelines as long strips rather than disconnected chunks unless the user explicitly wants segmentation.
- For waterfalls, cascades, rivers, or water running down terrain, build the landform first and then use continuous water sheets/pools. Do NOT fake waterfalls with staircase cubes unless the user explicitly asked for a blockout.
- For terrain curvature or road-following placement, prefer the selected spline or landscape spline and align meshes to its tangent instead of guessing positions manually.
- Preserve source materials and textures on imported models whenever possible. Do not replace textured assets with flat placeholder materials unless the user explicitly asks.
- Respect geometry and terrain realism: use bounds-aware spacing, avoid clipping or burying meshes through the ground, and keep placements inside a sensible local build area instead of sprawling across the map.
- If a placement looks unrealistic, too dense, too large, off-terrain, or outside the intended area, adjust scale/spacing/offset before claiming it is complete.
- For visual fix requests, prefer the current selection first. If nothing is selected, identify a small candidate set before changing the world.
- After making an edit, verify it with a live query or screenshot before saying it is fixed.

RULE 5 — NATURAL INTELLIGENCE:
- You are a REAL AI. Have natural conversations. Answer any question intelligently.
- Treat recent attachments from earlier turns as active context when the user refers to them with "fix it", "summarize it", "what does it say", "based on this", etc.
- When the user pastes an image with a message like "add a roof here" — look at the image, understand the scene geometry, and write Python to create the roof at the right position/scale.
- If UEFN is offline, say so clearly. Never pretend actions were taken.
- Be concise. Use markdown for readability.
- GRAMMAR & SPELLING TOLERANCE: The user may misspell words or use informal grammar. ALWAYS understand their intent without correcting them. NEVER say "I think you meant..." or "Did you mean..." or autocorrect their message in your response. Just understand what they want and DO it. Examples:
  - "delet the blu walls" → understand "delete the blue walls" and execute it
  - "ad a rof here" → understand "add a roof here" and execute it
  - "fix these whole and clipping" → understand "fix these holes and clipping" and execute it
  - "mov it to the lft" → understand "move it to the left" and execute it

RULE 6 — RESPONSE FORMAT:
- NEVER use academic citation formats like [Attachment 1 §3] or [Attachment 2 §1] in your responses.
- Refer to files naturally by their filename: "In the PDF..." or "The screenshot shows..." or "Based on the uploaded image..."
- When analyzing images, describe what you ACTUALLY SEE — objects, colors, layout, text, UI elements, errors — not generic metadata.
- When analyzing documents, give real insights — not just "this document contains text about X". Extract the actual useful content.
- Sound like an expert who understands the content, not a search engine returning snippets.
- Keep responses CONCISE. When the user asks you to DO something, say what you're doing in 1-3 sentences then execute. Do NOT write essays, tutorials, step-by-step guides, or "recommended tools" tables.
- NEVER use ### headers, --- dividers, or long bulleted lists in action responses. Save formatting for actual explanations when the user asks "how does X work".

COMMON UNREAL PYTHON (for execute_python_in_uefn):
```python
import unreal
# Get all actors
actors = unreal.EditorLevelLibrary.get_all_level_actors()
# Find by name/label
found = [a for a in actors if 'keyword' in a.get_name().lower() or 'keyword' in a.get_actor_label().lower()]
# Delete
for a in found: unreal.EditorLevelLibrary.destroy_actor(a)
# Move
actor.set_actor_location(unreal.Vector(x, y, z), False)
# Rotate
actor.set_actor_rotation(unreal.Rotator(roll, pitch, yaw), False)  # UEFN order: roll, pitch, yaw
# Scale
actor.set_actor_scale3d(unreal.Vector(sx, sy, sz))
# Get/set properties
loc = actor.get_actor_location()  # .x .y .z
actor.set_actor_label("NewLabel")
# Spawn static mesh
new_actor = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(0,0,0))
# Set mesh (cube, sphere, cylinder, cone, plane)
mesh_comp = new_actor.get_component_by_class(unreal.StaticMeshComponent)
cube_mesh = unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube')
mesh_comp.set_static_mesh(cube_mesh)
# Set material/color
mat = unreal.EditorAssetLibrary.load_asset('/Game/Materials/M_Red')
mesh_comp.set_material(0, mat)
# Create dynamic material with color
mat_inst = unreal.MaterialInstanceDynamic.create(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/BasicShapeMaterial'), new_actor)
mat_inst.set_vector_parameter_value('Color', unreal.LinearColor(r=1.0, g=0.0, b=0.0, a=1.0))
mesh_comp.set_material(0, mat_inst)
# Build geometry (roof = scaled+rotated cube or wedge)
roof = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(x, y, z))
roof_mesh = roof.get_component_by_class(unreal.StaticMeshComponent)
roof_mesh.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
roof.set_actor_scale3d(unreal.Vector(5.0, 5.0, 0.2))  # flat slab
roof.set_actor_rotation(unreal.Rotator(0, 0, 0), False)
# ALWAYS end with: result = <data> so the output is captured
```

3D CONSTRUCTION PATTERNS (use BasicShapes/Cube, scale 1.0 = 1 meter = 100 Unreal units):
CRITICAL: UEFN Rotator order is (roll, pitch, yaw) NOT (pitch, yaw, roll)!
  Rotator(30, 0, 0) = roll=30° (tilt around X axis)
  Rotator(0, 30, 0) = pitch=30° (tilt around Y axis)
  Rotator(0, 0, 30) = yaw=30° (rotate around Z axis, horizontal turn)
- WALL between two points: length=sqrt((x2-x1)^2+(y2-y1)^2), angle=atan2(y2-y1,x2-x1). Position at midpoint, scale=(length/100, thickness/100, height/100), Rotator(0, 0, angle_deg).
- FLOOR: Flat cube at ground level, scale=(width/100, depth/100, thickness/100).
- GABLE ROOF: Two slopes meeting at ridge. pitch=degrees(atan2(roof_height, depth/2)). slope_length=sqrt((depth/2)^2+roof_height^2). Scale=(width/100, slope_length/100, thickness/100). Left=Rotator(-pitch, 0, 0), Right=Rotator(+pitch, 0, 0).
- SHED ROOF: Single slope tilted: Rotator(pitch_angle, 0, 0).
- STAIRS: Loop of cubes, each offset by (step_depth, step_height). Typical: 16 steps, 20cm rise, 28cm depth.
- DOOR OPENING: Split wall into 3 parts (left of door, above door, right of door).
- WINDOW: Split wall into pieces around opening.
- ALWAYS label actors clearly, set collision BlockAll, use direct set_ calls (NEVER conditional).
- Connect wall corners: end of one wall = start of next.
- Symmetric pairs (left/right roof): matching scales, mirrored roll angles.

ROOF FIX/BUILD PROCEDURE (CRITICAL — follow this exactly when fixing or building roofs):
When the user says "fix the roof" or "build a roof", calculate the correct geometry from the ACTUAL house dimensions in the LIVE UEFN DATA:
```python
import unreal, math

# Step 1: Find house walls to determine dimensions
actors = unreal.EditorLevelLibrary.get_all_level_actors()
house_actors = [a for a in actors if 'house' in a.get_actor_label().lower() or 'wall' in a.get_actor_label().lower()]

# Step 2: Calculate house bounds from wall positions
xs = [a.get_actor_location().x for a in house_actors]
ys = [a.get_actor_location().y for a in house_actors]
zs = [a.get_actor_location().z for a in house_actors]
house_min_x, house_max_x = min(xs), max(xs)
house_min_y, house_max_y = min(ys), max(ys)
house_width = house_max_x - house_min_x
house_depth = house_max_y - house_min_y
center_x = (house_min_x + house_max_x) / 2
center_y = (house_min_y + house_max_y) / 2

# Step 3: Find wall top (highest non-roof actor z + half its z-scale * 100)
wall_actors = [a for a in house_actors if 'story' in a.get_actor_label().lower() or 'wall' in a.get_actor_label().lower()]
wall_top = max(a.get_actor_location().z + a.get_actor_scale3d().z * 50 for a in wall_actors) if wall_actors else 620

# Step 4: Calculate roof geometry
roof_height = house_depth * 0.25  # Good proportion: 25% of depth
ridge_z = wall_top + roof_height
half_depth = house_depth / 2
pitch_deg = math.degrees(math.atan2(roof_height, half_depth))
slope_length = math.sqrt(half_depth**2 + roof_height**2)

# Step 5: Delete old roof pieces
for a in actors:
    label = a.get_actor_label().lower()
    if 'roof' in label or 'gable' in label or 'ridge' in label:
        unreal.EditorLevelLibrary.destroy_actor(a)

# Step 6: Create new roof
mesh_path = '/Engine/BasicShapes/Cube'

# Left slope — Rotator(roll, pitch, yaw): roll tilts around X axis
left = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(center_x, center_y - half_depth/2, wall_top + roof_height/2))
lm = left.get_component_by_class(unreal.StaticMeshComponent)
lm.set_static_mesh(unreal.EditorAssetLibrary.load_asset(mesh_path))
left.set_actor_scale3d(unreal.Vector((house_width+60)/100, (slope_length+30)/100, 0.12))
left.set_actor_rotation(unreal.Rotator(-pitch_deg, 0, 0), False)  # -roll = left slope tilts toward ridge
left.set_actor_label('Roof_Left')
lm.set_collision_profile_name('BlockAll')

# Right slope (mirror roll)
right = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(center_x, center_y + half_depth/2, wall_top + roof_height/2))
rm = right.get_component_by_class(unreal.StaticMeshComponent)
rm.set_static_mesh(unreal.EditorAssetLibrary.load_asset(mesh_path))
right.set_actor_scale3d(unreal.Vector((house_width+60)/100, (slope_length+30)/100, 0.12))
right.set_actor_rotation(unreal.Rotator(pitch_deg, 0, 0), False)  # +roll = right slope tilts toward ridge
right.set_actor_label('Roof_Right')
rm.set_collision_profile_name('BlockAll')

# Ridge cap
ridge = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(center_x, center_y, ridge_z))
ridm = ridge.get_component_by_class(unreal.StaticMeshComponent)
ridm.set_static_mesh(unreal.EditorAssetLibrary.load_asset(mesh_path))
ridge.set_actor_scale3d(unreal.Vector((house_width+20)/100, 0.15, 0.15))
ridge.set_actor_label('Roof_Ridge')
ridm.set_collision_profile_name('BlockAll')

# Front gable (triangular infill - use stacked rectangles)
gable_steps = 5
for i in range(gable_steps):
    frac = (i + 0.5) / gable_steps
    gable_z = wall_top + frac * roof_height
    gable_width = house_width * (1 - frac)
    g = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(center_x, house_min_y, gable_z))
    gm = g.get_component_by_class(unreal.StaticMeshComponent)
    gm.set_static_mesh(unreal.EditorAssetLibrary.load_asset(mesh_path))
    g.set_actor_scale3d(unreal.Vector(gable_width/100, 0.2, roof_height/gable_steps/100))
    g.set_actor_label(f'Gable_Front_{{i+1:02d}}')
    gm.set_collision_profile_name('BlockAll')

# Back gable
for i in range(gable_steps):
    frac = (i + 0.5) / gable_steps
    gable_z = wall_top + frac * roof_height
    gable_width = house_width * (1 - frac)
    g = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(center_x, house_max_y, gable_z))
    gm = g.get_component_by_class(unreal.StaticMeshComponent)
    gm.set_static_mesh(unreal.EditorAssetLibrary.load_asset(mesh_path))
    g.set_actor_scale3d(unreal.Vector(gable_width/100, 0.2, roof_height/gable_steps/100))
    g.set_actor_label(f'Gable_Back_{{i+1:02d}}')
    gm.set_collision_profile_name('BlockAll')

result = f'Roof rebuilt: pitch={{pitch_deg:.1f}}deg, width={{house_width}}, slope={{slope_length:.0f}}, ridge_z={{ridge_z:.0f}}'
```
IMPORTANT: When the user says "fix it again" or repeats a request, it means the PREVIOUS attempt FAILED or was WRONG.
Do NOT generate the same code again. Instead:
1. Query current actor state to see what changed (or didn't change)
2. If the previous approach didn't work, try a DIFFERENT approach
3. If actors still look wrong, DELETE the bad ones and REBUILD from scratch using the procedure above
"""


_ACTIVE_TOOL_CONTEXT = threading.local()


def _set_active_tool_context(*, message: str = "", attachments: Optional[List[Dict[str, Any]]] = None, chat_id: str = "") -> None:
    _ACTIVE_TOOL_CONTEXT.message = str(message or "")
    _ACTIVE_TOOL_CONTEXT.attachments = list(attachments or [])
    _ACTIVE_TOOL_CONTEXT.chat_id = str(chat_id or "")


def _get_active_tool_context() -> Dict[str, Any]:
    return {
        "message": getattr(_ACTIVE_TOOL_CONTEXT, "message", ""),
        "attachments": list(getattr(_ACTIVE_TOOL_CONTEXT, "attachments", []) or []),
        "chat_id": getattr(_ACTIVE_TOOL_CONTEXT, "chat_id", ""),
    }


def _clear_active_tool_context() -> None:
    _set_active_tool_context(message="", attachments=[], chat_id="")


# OpenAI function definitions for tool calling
_CHAT_FUNCTIONS = [
    {
        "type": "function",
        "function": {
            "name": "search_tools",
            "description": "Search the registered UEFN Toolbelt catalog only when the right tool is not already obvious. Query with 2-6 intent keywords, not a full sentence.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Short intent keywords like 'color selected red', 'align actors', or 'verse devices'"
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_tool_details",
            "description": "Inspect one specific registered tool by id or display name, including its parameters.",
            "parameters": {
                "type": "object",
                "properties": {
                    "tool_name": {
                        "type": "string",
                        "description": "The tool id or display name to inspect"
                    }
                },
                "required": ["tool_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_uefn_context",
            "description": "Fetch a combined snapshot of the live UEFN project, level, selection, viewport, and optionally assets.",
            "parameters": {
                "type": "object",
                "properties": {
                    "include_assets": {
                        "type": "boolean",
                        "description": "Whether to include a content-browser asset listing"
                    },
                    "directory": {
                        "type": "string",
                        "description": "Asset root to list when include_assets is true"
                    },
                    "recursive": {
                        "type": "boolean",
                        "description": "Whether to recurse when listing assets"
                    },
                    "class_filter": {
                        "type": "string",
                        "description": "Optional class filter for list_assets"
                    }
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_uefn",
            "description": "Query the live UEFN editor for information about the current level, actors, project, or viewport.",
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {
                        "type": "string",
                        "enum": [
                            "get_all_actors", "get_level_info", "get_project_info",
                            "get_selected_actors", "get_viewport_camera", "get_actor_properties",
                            "list_assets"
                        ],
                        "description": "The UEFN query command to execute"
                    },
                    "params": {
                        "type": "object",
                        "description": "Optional parameters for the command (e.g. class_filter, directory)"
                    }
                },
                "required": ["command"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_uefn_tool",
            "description": "Execute a UEFN Toolbelt tool. Available tools include material operations, bulk transforms, scatter, patterns, snapshots, LOD, screenshots, text placement, Verse generation, and more.",
            "parameters": {
                "type": "object",
                "properties": {
                    "tool_name": {
                        "type": "string",
                        "description": "The tool ID to run (e.g. 'bulk_align', 'material_apply_preset', 'screenshot_take')"
                    },
                    "parameters": {
                        "type": "object",
                        "description": "Tool-specific parameters"
                    }
                },
                "required": ["tool_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "execute_python_in_uefn",
            "description": "Execute arbitrary Python code inside the UEFN editor. Use for operations not covered by existing tools.",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Python code to execute in the editor"
                    }
                },
                "required": ["code"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "rebuild_gable_roof_action",
            "description": "Rebuild a clean gable roof from live UEFN actor bounds. Use for broken, misaligned, or repeated roof-fix requests.",
            "parameters": {
                "type": "object",
                "properties": {
                    "roof_height_ratio": {
                        "type": "number",
                        "description": "Roof steepness ratio relative to half-depth, usually between 0.2 and 0.4"
                    },
                    "overhang": {
                        "type": "number",
                        "description": "Optional roof overhang distance in Unreal units"
                    }
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "apply_material_action",
            "description": "Apply a named material or texture preset to actors matched by label pattern.",
            "parameters": {
                "type": "object",
                "properties": {
                    "actor_pattern": {
                        "type": "string",
                        "description": "Actor label pattern such as 'House_*' or 'Terrain_*'"
                    },
                    "material": {
                        "type": "string",
                        "description": "Material name or alias such as brick, wood, grass, dirt, rock, cliff, concrete, lava, sparkle, or water"
                    }
                },
                "required": ["actor_pattern", "material"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "build_house_action",
            "description": "Build a functional residential building using the shared structure planner. Prefer this for houses, homes, cabins, cottages, villas, mansions, townhouses, apartments, condos, and similar requests instead of generic cube code.",
            "parameters": {
                "type": "object",
                "properties": {
                    "request": {
                        "type": "string",
                        "description": "Short house intent such as 'build a modern house here' or 'make a cozy cabin with stairs'"
                    },
                    "position": {
                        "type": "object",
                        "description": "Optional world position with x, y, z coordinates. If omitted, use the selected support surface."
                    },
                    "style": {
                        "type": "string",
                        "description": "Optional style hint such as suburban, modern, cabin, cottage, villa, mansion, townhouse, or apartment"
                    },
                    "size": {
                        "type": "string",
                        "description": "Approximate size label such as small, medium, large, or huge"
                    },
                    "story_count": {
                        "type": "integer",
                        "description": "Optional story count for multi-story residential builds such as a 4-story apartment"
                    },
                    "material": {
                        "type": "string",
                        "description": "Optional main wall/body material"
                    },
                    "roof_material": {
                        "type": "string",
                        "description": "Optional roof material"
                    },
                    "zone_id": {
                        "type": "string",
                        "description": "Optional managed zone id. If omitted, one is derived from the chat and request."
                    },
                    "label_prefix": {
                        "type": "string",
                        "description": "Optional label prefix for generated house pieces"
                    }
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "build_structure_action",
            "description": "Generate a non-house structure with the shared planner when possible. Use this for garages, sheds, workshops, barns, warehouses, greenhouses, studios, hangars, kiosks, pavilions, gazebos, pergolas, canopies, carports, market stalls, and also scenic structures like fountains or bridges.",
            "parameters": {
                "type": "object",
                "properties": {
                    "structure": {
                        "type": "string",
                        "description": "Structure type to generate, such as garage, shed, workshop, barn, pavilion, gazebo, pergola, fountain, arch, tower, bridge, platform, or waterfall"
                    },
                    "request": {
                        "type": "string",
                        "description": "Short build intent such as 'build a rustic gazebo here' or 'make a wide garage with a metal roof'"
                    },
                    "position": {
                        "type": "object",
                        "description": "Optional world position with x, y, z coordinates. If omitted, use the selected support surface."
                    },
                    "size": {
                        "type": "string",
                        "description": "Approximate size label such as small, medium, large, or huge"
                    },
                    "style": {
                        "type": "string",
                        "description": "Optional style hint such as rustic, modern, utility, agricultural, or garden"
                    },
                    "material": {
                        "type": "string",
                        "description": "Primary body material name or alias to apply"
                    },
                    "roof_material": {
                        "type": "string",
                        "description": "Optional roof material for structures that have a roof"
                    },
                    "support_material": {
                        "type": "string",
                        "description": "Optional support material for structures like waterfalls, cliffs, or rock-backed features"
                    },
                    "zone_id": {
                        "type": "string",
                        "description": "Optional managed zone id. If omitted, one is derived from the chat and request."
                    },
                    "label": {
                        "type": "string",
                        "description": "Optional actor label prefix"
                    }
                },
                "required": ["structure"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "terrain_action",
            "description": "Create, modify, list, or delete terrain with layered surfaces, edge materials, and scene-aware biome dressing.",
            "parameters": {
                "type": "object",
                "properties": {
                    "operation": {
                        "type": "string",
                        "enum": ["create", "modify", "delete", "list"],
                        "description": "Terrain operation to perform"
                    },
                    "terrain_type": {
                        "type": "string",
                        "enum": ["flat", "hill", "valley", "plateau", "slope", "crater", "ridge"],
                        "description": "Terrain shape to create or modify"
                    },
                    "position": {
                        "type": "object",
                        "description": "World position with x, y, z coordinates"
                    },
                    "size": {
                        "type": "object",
                        "description": "Terrain size with x and y dimensions in Unreal units"
                    },
                    "material": {
                        "type": "string",
                        "description": "Top material name such as grass, dirt, sand, rock, cliff, mud, moss, ground, road, or terrain"
                    },
                    "height": {
                        "type": "number",
                        "description": "Base terrain height in Unreal units"
                    },
                    "elevation": {
                        "type": "number",
                        "description": "Additional elevation or depth used to shape the terrain"
                    },
                    "subdivisions": {
                        "type": "integer",
                        "description": "Optional tile count for larger terrain patches"
                    },
                    "label": {
                        "type": "string",
                        "description": "Optional terrain label prefix"
                    },
                    "label_pattern": {
                        "type": "string",
                        "description": "Pattern used when deleting terrain patches"
                    },
                    "decorate": {
                        "type": "boolean",
                        "description": "Whether to add biome-matched trees, rocks, and shrubs when assets exist"
                    }
                },
                "required": ["operation"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "import_attached_models",
            "description": "Import FBX or ZIP model attachments from the current chat turn into UEFN and optionally place them along splines, terrain curves, or patterns.",
            "parameters": {
                "type": "object",
                "properties": {
                    "request": {
                        "type": "string",
                        "description": "Short intent like 'import these models', 'import and place along the terrain spline', or 'import and arrange in a circle'"
                    }
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_knowledge",
            "description": "Search the shared knowledge base, learned notes, and prior chat/project memory across all models.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "What to search for in the global knowledge and prior chats"
                    },
                    "chat_id": {
                        "type": "string",
                        "description": "Optional current chat id so it can be excluded from related-chat results"
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_project_overview",
            "description": "Inspect backend/frontend capabilities, live UEFN state, and other non-tool features available in this app.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Optional topic used to refine workspace matches"
                    },
                    "include_assets": {
                        "type": "boolean",
                        "description": "Whether to include a compact asset listing from the live UEFN project"
                    },
                    "include_workspace_hits": {
                        "type": "boolean",
                        "description": "Whether to search the workspace codebase for relevant file snippets"
                    }
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_workspace",
            "description": "Search the app codebase and tool source files for files and snippets relevant to the user's request.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Free-form code or feature search query"
                    },
                    "include_vendor": {
                        "type": "boolean",
                        "description": "Whether to include vendor tool source files in the search"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of file hits to return"
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "analyze_chat_attachments",
            "description": "Inspect and summarize uploaded files, screenshots, OCR text, and extracted attachment metadata that already exist in a chat.",
            "parameters": {
                "type": "object",
                "properties": {
                    "chat_id": {
                        "type": "string",
                        "description": "The chat id whose attachments should be inspected"
                    },
                    "query": {
                        "type": "string",
                        "description": "Optional filter such as 'castle door', 'ocr', 'pdf', or 'screenshots'"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Maximum number of attachment records to return"
                    }
                },
                "required": ["chat_id"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "save_knowledge",
            "description": "ONLY use when the user EXPLICITLY asks you to remember/save something. Do NOT auto-save chat responses or summaries.",
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {
                        "type": "string",
                        "description": "Short title for the memory"
                    },
                    "content": {
                        "type": "string",
                        "description": "The actual note to remember"
                    },
                    "tags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Optional tags for retrieval"
                    },
                    "chat_id": {
                        "type": "string",
                        "description": "Optional originating chat id"
                    }
                },
                "required": ["title", "content"]
            }
        }
    }
]


def _handle_tool_call(name: str, arguments: dict) -> dict:
    """Execute a tool call from the AI and return the result."""
    if name == "search_tools":
        return {"success": True, "results": _search_tools_for_llm(arguments.get("query", ""))}

    elif name == "get_tool_details":
        return tool_registry.describe_tool(arguments.get("tool_name", ""))

    elif name == "get_uefn_context":
        return {
            "success": True,
            "context": _get_uefn_context(
                include_assets=bool(arguments.get("include_assets", False)),
                directory=str(arguments.get("directory") or "/Game/"),
                recursive=bool(arguments.get("recursive", True)),
                class_filter=str(arguments.get("class_filter") or ""),
            )
        }

    elif name == "query_uefn":
        cmd = arguments.get("command", "")
        params = arguments.get("params", {})
        return _chat_uefn_query(cmd, params)

    elif name == "run_uefn_tool":
        tool_name = arguments.get("tool_name", "")
        params = arguments.get("parameters", {})
        return tool_registry.execute_tool(tool_name, params)

    elif name == "execute_python_in_uefn":
        code = arguments.get("code", "")
        validation_error = _validate_uefn_python_action(code)
        if validation_error:
            return {"error": validation_error}
        port = discover_uefn_listener_port()
        if not port:
            return {"error": "UEFN not connected"}
        try:
            return mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)
        except Exception as e:
            return {"error": str(e)}

    elif name == "rebuild_gable_roof_action":
        return _execute_rebuild_gable_roof(arguments)

    elif name == "apply_material_action":
        return _execute_apply_material(arguments)

    elif name == "build_house_action":
        return _execute_build_house(arguments)

    elif name == "build_structure_action":
        return _execute_structure_action_with_shared_planner(arguments)

    elif name == "terrain_action":
        return _execute_terrain_control(arguments)

    elif name == "import_attached_models":
        context = _get_active_tool_context()
        request_text = str(arguments.get("request") or context.get("message") or "import attached models into uefn")
        attachments = context.get("attachments") or []
        if not attachments:
            return {"error": "No attachments are available in the current chat turn for model import."}
        result = _maybe_import_and_place_model_attachments(request_text, attachments)
        if not result:
            return {"error": "No importable FBX or ZIP model attachments were found in the current chat turn."}
        tool_output = (result.get("tool_result") or {}).get("output") or {}
        return {
            "success": bool(tool_output.get("success", True)),
            "reply": result.get("reply") or "",
            "result": tool_output,
        }

    elif name == "search_knowledge":
        return {
            "success": True,
            "results": _search_shared_context(
                str(arguments.get("query") or ""),
                current_chat_id=str(arguments.get("chat_id") or ""),
            )
        }

    elif name == "get_project_overview":
        return {
            "success": True,
            "overview": _get_project_overview(
                query=str(arguments.get("query") or ""),
                include_assets=bool(arguments.get("include_assets", False)),
                include_workspace_hits=bool(arguments.get("include_workspace_hits", False)),
            )
        }

    elif name == "search_workspace":
        return {
            "success": True,
            "results": _search_workspace_sources(
                str(arguments.get("query") or ""),
                limit=int(arguments.get("limit") or 8),
                include_vendor=bool(arguments.get("include_vendor", True)),
            )
        }

    elif name == "analyze_chat_attachments":
        return {
            "success": True,
            "results": _analyze_chat_attachments(
                str(arguments.get("chat_id") or ""),
                query=str(arguments.get("query") or ""),
                limit=int(arguments.get("limit") or MAX_ATTACHMENT_ANALYSIS_RESULTS),
            )
        }

    elif name == "save_knowledge":
        item = knowledge_store.add_item(
            item_type="text",
            source_type="ai_saved",
            scope="global",
            title=str(arguments.get("title") or "Memory"),
            content=str(arguments.get("content") or ""),
            tags=[str(tag) for tag in (arguments.get("tags") or [])],
            chat_id=str(arguments.get("chat_id") or ""),
        )
        return {"success": True, "item": item}

    return {"error": f"Unknown function: {name}"}


def _parse_inline_function_calls(raw_text: str) -> List[Dict[str, Any]]:
    """Parse provider-emitted function markup when tool calling fails."""
    if not raw_text:
        return []

    matches: List[Dict[str, Any]] = []
    patterns = [
        re.compile(r"<function=([a-zA-Z0-9_]+)\s*>(\{.*?\})</function>", re.DOTALL),
        re.compile(r"<function=([a-zA-Z0-9_]+)\s+(\{.*?\})\s*</function>", re.DOTALL),
        re.compile(r"<function=([a-zA-Z0-9_]+)(\{.*?\})</function>", re.DOTALL),
    ]

    for pattern in patterns:
        for name, args_blob in pattern.findall(raw_text):
            arguments: Dict[str, Any]
            try:
                arguments = json.loads(args_blob)
            except json.JSONDecodeError:
                continue
            matches.append({
                "name": name.strip(),
                "arguments": arguments,
            })
        if matches:
            break

    return matches


def _format_recovered_tool_reply(tool_name: str, tool_output: Dict[str, Any]) -> str:
    """Create a deterministic reply when we cannot get a clean post-tool model response."""
    if tool_name == "search_tools":
        results = tool_output.get("results") or []
        if not results:
            return "I searched the tool catalog but did not find a strong match for that request."
        top = results[:3]
        lines = [
            "I searched the tool catalog. Best matches:",
            *[
                f"- `{tool.get('id', 'unknown')}`: {tool.get('short') or tool.get('description') or tool.get('name', 'Tool')}"
                for tool in top
            ],
        ]
        return "\n".join(lines)

    if tool_name == "run_uefn_tool":
        if tool_output.get("success"):
            return "I executed the requested UEFN tool successfully."
        return f"I tried to run the requested UEFN tool, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "query_uefn":
        if tool_output.get("success"):
            result = tool_output.get("result")
            return f"I pulled live data from UEFN: `{json.dumps(result, default=str)[:600]}`"
        return f"I tried to query UEFN, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "get_tool_details":
        tool = tool_output.get("tool") or {}
        if tool:
            return f"`{tool.get('id', 'tool')}` looks like the relevant tool. {tool.get('description', '')}".strip()

    if tool_name == "rebuild_gable_roof_action":
        if tool_output.get("success") or tool_output.get("result"):
            return "I rebuilt the roof using the dedicated roof action."
        return f"I tried to rebuild the roof, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "apply_material_action":
        if tool_output.get("success") or tool_output.get("result"):
            return "I applied the requested material successfully."
        return f"I tried to apply the material, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "build_house_action":
        if tool_output.get("success") or tool_output.get("result"):
            return "I built the requested house using the shared structure planner."
        return f"I tried to build the house, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "build_structure_action":
        if tool_output.get("success") or tool_output.get("result"):
            return "I built the requested structure in UEFN."
        return f"I tried to build the structure, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "terrain_action":
        if tool_output.get("success") or tool_output.get("result"):
            return "I executed the terrain action successfully."
        return f"I tried to run the terrain action, but it failed: {tool_output.get('error', 'unknown error')}"

    if tool_name == "import_attached_models":
        if tool_output.get("success") or tool_output.get("result"):
            return str(tool_output.get("reply") or "I imported the attached models successfully.")
        return f"I tried to import the attached models, but it failed: {tool_output.get('error', 'unknown error')}"

    return "I recovered the intended tool call and processed it."


_ACTION_ALIASES = {
    "rebuild_gable_roof_action": "rebuild_gable_roof",
    "apply_material_action": "apply_material",
    "build_house_action": "build_house",
    "build_structure_action": "build_structure",
    "terrain_action": "terrain",
}
_SUPPORTED_ACTION_TYPES = {
    "execute_python_in_uefn",
    "run_uefn_tool",
    "query_uefn",
    "rebuild_gable_roof",
    "apply_material",
    "build_house",
    "build_structure",
    "terrain",
    "import_attached_models",
}


def _normalize_action_type(action_type: Any) -> str:
    normalized = str(action_type or "").strip()
    return _ACTION_ALIASES.get(normalized, normalized)


def _recover_failed_tool_call(
    error_msg: str,
    messages: List[Dict[str, Any]],
    client: Any,
    provider: str,
    model: str,
) -> Optional[Dict[str, Any]]:
    """Recover from provider-side malformed tool-call syntax."""
    recovered_calls = _parse_inline_function_calls(error_msg)
    if not recovered_calls:
        return None

    tool_results_for_frontend = []
    recovery_messages = list(messages)
    recovery_messages.append({
        "role": "assistant",
        "content": f"[Recovered malformed tool call markup from provider]\n{json.dumps(recovered_calls, default=str)}",
    })

    for recovered in recovered_calls[:3]:
        tool_name = recovered["name"]
        args = recovered["arguments"]
        result = _handle_tool_call(tool_name, args)
        tool_results_for_frontend.append({
            "tool": tool_name,
            "output": result,
        })
        recovery_messages.append({
            "role": "system",
            "content": (
                f"Recovered tool result for `{tool_name}`:\n"
                f"{json.dumps(result, default=str)[:5000]}"
            ),
        })

    fallback_reply = _format_recovered_tool_reply(
        tool_results_for_frontend[-1]["tool"],
        tool_results_for_frontend[-1]["output"],
    )

    try:
        response = _chat_completion_with_retry(
            client,
            provider,
            model=model,
            messages=recovery_messages + [{
                "role": "system",
                "content": (
                    "The tool result above is already final. Reply to the user normally using that result. "
                    "Do not call tools again and do not emit any function markup."
                ),
            }],
            max_tokens=2000,
            temperature=0.3,
        )
        tok_in = getattr(response.usage, 'prompt_tokens', 0) if response.usage else 0
        tok_out = getattr(response.usage, 'completion_tokens', 0) if response.usage else 0
        _track_usage(provider, tok_in, tok_out)
        reply = response.choices[0].message.content or fallback_reply
    except Exception:
        reply = fallback_reply

    result = {"reply": reply}
    if tool_results_for_frontend:
        result["tool_result"] = tool_results_for_frontend[-1]
    return result


def _clean_prefetch_reply(reply: str) -> str:
    """Post-process replies from pre-fetch models (Cerebras/Ollama) to remove verbose junk.

    Strategy: If the reply contains action blocks, extract them and keep only
    1-2 intro sentences. If no action blocks, aggressively strip markdown formatting.
    """
    import re as _re

    # ── Step 1: Extract all action blocks ──
    action_pattern = r'```(?:action|json)\s*\n.*?\n```'
    action_blocks = _re.findall(action_pattern, reply, _re.DOTALL)

    if action_blocks:
        # ── Has action blocks: keep only intro + action blocks ──
        # Get the text BEFORE the first action block
        first_action_start = _re.search(action_pattern, reply, _re.DOTALL)
        intro_text = reply[:first_action_start.start()].strip() if first_action_start else ""

        # Clean intro text: remove autocorrections, headers, tables, dividers
        intro_text = _re.sub(r'#{1,6}\s+[^\n]*', '', intro_text)
        intro_text = _re.sub(r'-{3,}', '', intro_text)
        intro_text = _re.sub(r'\|[^\n]*\|', '', intro_text)
        intro_text = _re.sub(r'[\U0001F000-\U0001FFFF]+\s*', '', intro_text)
        intro_text = _re.sub(r'\*{1,2}[^*]+\*{1,2}', lambda m: m.group().strip('*'), intro_text)

        # Remove autocorrection phrases
        for pattern in (
            r'This appears to be a typo[^.]*\..*?(?:you (?:likely|probably) meant)[^\n]*',
            r'Based on context,?\s*you (?:likely|probably) meant[^\n]*',
            r'You said:\s*>?\s*\*?"[^"]*"\*?\s*',
            r'I think you meant[^\n]*',
            r'you likely meant[^\n]*',
            r'>.*?"[^"]*"',
        ):
            intro_text = _re.sub(pattern, '', intro_text, flags=_re.IGNORECASE | _re.DOTALL)

        # Collapse to just 1-2 meaningful sentences
        sentences = [s.strip() for s in _re.split(r'(?<=[.!])\s+', intro_text) if s.strip() and len(s.strip()) > 10]
        intro = ' '.join(sentences[:2]).strip()

        # Rebuild: intro + action blocks
        result = intro
        for block in action_blocks:
            result += '\n\n' + block
        return result.strip()

    # ── Step 2: No action blocks — strip verbose formatting ──
    # Remove autocorrection blocks
    for pattern in (
        r'This appears to be a typo[^.]*\..*?(?:you (?:likely|probably) meant)[^\n]*\n*(?:\s*>.*?\n)*',
        r'Based on context,?\s*you (?:likely|probably) meant[^\n]*\n*(?:\s*>.*?\n)*',
        r'You said:\s*>?\s*\*?"[^"]*"\*?\s*\n?',
    ):
        reply = _re.sub(pattern, '', reply, flags=_re.IGNORECASE | _re.DOTALL)

    lines = reply.split('\n')
    cleaned = []
    for line in lines:
        stripped = line.strip()
        lower = stripped.lower()

        # Skip markdown headers, dividers, tables, emoji lines
        if _re.match(r'^#{1,6}\s', stripped):
            continue
        if _re.match(r'^-{3,}$', stripped):
            continue
        if stripped.startswith('|') and stripped.endswith('|'):
            continue
        if _re.match(r'^[\s]*\|[-:| ]+\|$', stripped):
            continue
        if _re.match(r'^[\s]*[\U0001F000-\U0001FFFF]', stripped):
            continue

        # Skip filler phrases
        if any(phrase in lower for phrase in (
            'let me know if you', 'alternatively,', 'step-by-step',
            'pro tip', 'final notes', 'recommended fixes', 'next steps',
            'here\'s a clear breakdown', 'if you can share', 'if you can reupload',
            'here is a summary', 'summary of changes', 'clean white architectural',
            'you\'re working in unreal editor for fortnite',
        )):
            continue

        cleaned.append(line)

    result = '\n'.join(cleaned).strip()
    result = _re.sub(r'\n{3,}', '\n\n', result)
    result = _re.sub(r'\s*-{3,}\s*$', '', result)
    return result if result else reply


def _execute_rebuild_gable_roof(action: dict) -> dict:
    """Server-side gable roof rebuild with validation.

    Process:
    1. Query all actors and identify house structure
    2. Compute precise bounds from wall positions + scales
    3. Delete ALL existing roof/gable/ridge geometry
    4. Build clean roof: 2 slopes + ridge cap + 2 gable end panels
    5. Validate: check for overlaps, floating parts, alignment
    6. Auto-fix any validation issues
    """
    import math

    port = discover_uefn_listener_port()
    if not port:
        return {"error": "UEFN not connected"}

    # ── Step 1: Get all actors ──
    r = _chat_uefn_query("get_all_actors", {})
    if not r.get("success"):
        return {"error": "Could not query actors"}

    actors_raw = r.get("result", {})
    actors_list = actors_raw.get("actors", actors_raw) if isinstance(actors_raw, dict) else actors_raw
    if not isinstance(actors_list, list):
        return {"error": "Invalid actor data"}

    # ── Step 2: Identify house structure ──
    # Find wall actors by label keywords
    house_keywords = ['house', 'story']
    exclude_keywords = ['stair', 'floor', 'guard', 'landing', 'header', 'beacon', 'door']
    house_actors = [
        a for a in actors_list
        if isinstance(a, dict)
        and any(kw in str(a.get("label", "")).lower() for kw in house_keywords)
        and not any(ex in str(a.get("label", "")).lower() for ex in exclude_keywords)
        and "location" in a and "scale" in a
        and a.get("scale", {}).get("z", 0) >= 1.0  # structural walls only
    ]

    if not house_actors:
        # Fallback: try any actor with 'wall' or 'house' in label
        house_actors = [
            a for a in actors_list
            if isinstance(a, dict)
            and any(kw in str(a.get("label", "")).lower() for kw in ['house', 'story', 'wall'])
            and "location" in a and "scale" in a
        ]

    if not house_actors:
        logger.warning("Roof rebuild: no house actors found. Labels: %s",
                       [a.get("label", "?") for a in actors_list[:30]])
        return {"error": "No house/wall actors found in level"}

    logger.info("Roof rebuild: found %d structural actors", len(house_actors))

    # ── Step 3: Compute precise house bounds ──
    # Account for actor scale to find true outer edges
    outer_min_x = float('inf')
    outer_max_x = float('-inf')
    outer_min_y = float('inf')
    outer_max_y = float('-inf')
    wall_top = float('-inf')

    for a in house_actors:
        loc = a["location"]
        sc = a["scale"]
        x, y, z = loc["x"], loc["y"], loc["z"]
        sx, sy, sz = sc["x"], sc["y"], sc["z"]
        # Half-extents in world space (scale * 50 for a default 100x100x100 cube)
        hx, hy, hz = sx * 50, sy * 50, sz * 50
        outer_min_x = min(outer_min_x, x - hx)
        outer_max_x = max(outer_max_x, x + hx)
        outer_min_y = min(outer_min_y, y - hy)
        outer_max_y = max(outer_max_y, y + hy)
        top = z + hz
        if top > wall_top:
            wall_top = top

    center_x = (outer_min_x + outer_max_x) / 2
    center_y = (outer_min_y + outer_max_y) / 2
    house_width = outer_max_x - outer_min_x
    house_depth = outer_max_y - outer_min_y

    logger.info("Roof rebuild: bounds x=[%.0f..%.0f] y=[%.0f..%.0f] width=%.0f depth=%.0f wall_top=%.0f",
                outer_min_x, outer_max_x, outer_min_y, outer_max_y,
                house_width, house_depth, wall_top)

    # Sanity check: house must have reasonable dimensions
    if house_width < 50 or house_depth < 50:
        return {"error": f"House too small: {house_width:.0f}x{house_depth:.0f}"}
    if house_width > 5000 or house_depth > 5000:
        return {"error": f"House too large: {house_width:.0f}x{house_depth:.0f} — check actor filtering"}

    # ── Step 4: Calculate roof geometry ──
    overhang = action.get("overhang", 30)  # overhang past walls in UU
    roof_height_ratio = action.get("roof_height_ratio", 0.3)
    roof_thickness = action.get("thickness", 12)  # thickness in UU
    roof_height = house_depth * roof_height_ratio
    half_depth = house_depth / 2
    pitch_deg = math.degrees(math.atan2(roof_height, half_depth))
    slope_length = math.sqrt(half_depth ** 2 + roof_height ** 2)
    ridge_z = wall_top + roof_height

    # Slope dimensions with overhang
    slope_width = house_width + overhang * 2   # overhang on both sides
    slope_depth = slope_length + overhang       # overhang at eave

    # Position: each slope centered between eave and ridge
    # Left slope: from outer_min_y (eave) to center_y (ridge)
    left_center_y = (outer_min_y - overhang / 2 + center_y) / 2
    right_center_y = (center_y + outer_max_y + overhang / 2) / 2
    slope_center_z = wall_top + roof_height / 2

    # Gable end panels: single rectangle per end, covering the full triangle
    # Use height = roof_height, width = house_width, thin in Y
    gable_z = wall_top + roof_height / 2
    gable_thickness = 15  # UU

    logger.info("Roof rebuild: pitch=%.1f° slope=%.0f ridge_z=%.0f overhang=%.0f",
                pitch_deg, slope_length, ridge_z, overhang)

    # ── Step 5: Build UEFN Python code ──
    code = f"""import unreal

# ── DELETE all existing roof geometry ──
# Delete by label AND by detecting orphaned flat panels above wall_top
actors = unreal.EditorLevelLibrary.get_all_level_actors()
deleted = []
for a in actors:
    label = a.get_actor_label()
    label_lower = label.lower()
    should_delete = False
    # Known roof labels
    if any(kw in label_lower for kw in ['roof', 'gable', 'ridge']):
        should_delete = True
    # Detect orphaned flat panels above walls (old roof pieces with generic labels)
    elif 'staticmeshactor' in label_lower or label_lower.startswith('staticmesh'):
        loc = a.get_actor_location()
        sc = a.get_actor_scale3d()
        # If it's above wall_top and is flat (z-scale < 0.2), it's an old roof piece
        if loc.z > {wall_top - 100} and sc.z < 0.25 and (sc.x > 2.0 or sc.y > 2.0):
            should_delete = True
    if should_delete:
        deleted.append(label)
        unreal.EditorLevelLibrary.destroy_actor(a)
print(f'Deleted {{len(deleted)}} roof pieces: {{deleted}}')

mesh = unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube')

# ── LEFT SLOPE ──
left = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({center_x}, {left_center_y}, {slope_center_z}))
lm = left.get_component_by_class(unreal.StaticMeshComponent)
lm.set_static_mesh(mesh)
left.set_actor_scale3d(unreal.Vector({slope_width / 100}, {slope_depth / 100}, {roof_thickness / 100}))
left.set_actor_rotation(unreal.Rotator({-pitch_deg}, 0, 0), False)
left.set_actor_label('Roof_Left')
lm.set_collision_profile_name('BlockAll')

# ── RIGHT SLOPE ──
right = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({center_x}, {right_center_y}, {slope_center_z}))
rm = right.get_component_by_class(unreal.StaticMeshComponent)
rm.set_static_mesh(mesh)
right.set_actor_scale3d(unreal.Vector({slope_width / 100}, {slope_depth / 100}, {roof_thickness / 100}))
right.set_actor_rotation(unreal.Rotator({pitch_deg}, 0, 0), False)
right.set_actor_label('Roof_Right')
rm.set_collision_profile_name('BlockAll')

# ── RIDGE CAP ──
ridge = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({center_x}, {center_y}, {ridge_z}))
ridm = ridge.get_component_by_class(unreal.StaticMeshComponent)
ridm.set_static_mesh(mesh)
ridge.set_actor_scale3d(unreal.Vector({(slope_width + 20) / 100}, {roof_thickness * 2 / 100}, {roof_thickness / 100}))
ridge.set_actor_label('Roof_Ridge')
ridm.set_collision_profile_name('BlockAll')

# ── FRONT GABLE PANEL ── (single clean rectangle, not stair-steps)
gf = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({center_x}, {outer_min_y}, {gable_z}))
gfm = gf.get_component_by_class(unreal.StaticMeshComponent)
gfm.set_static_mesh(mesh)
gf.set_actor_scale3d(unreal.Vector({house_width / 100}, {gable_thickness / 100}, {roof_height / 100}))
gf.set_actor_label('Gable_Front')
gfm.set_collision_profile_name('BlockAll')

# ── BACK GABLE PANEL ──
gb = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({center_x}, {outer_max_y}, {gable_z}))
gbm = gb.get_component_by_class(unreal.StaticMeshComponent)
gbm.set_static_mesh(mesh)
gb.set_actor_scale3d(unreal.Vector({house_width / 100}, {gable_thickness / 100}, {roof_height / 100}))
gb.set_actor_label('Gable_Back')
gbm.set_collision_profile_name('BlockAll')

# ── VALIDATION ──
# Check all new roof pieces exist and have correct properties
roof_labels = ['Roof_Left', 'Roof_Right', 'Roof_Ridge', 'Gable_Front', 'Gable_Back']
all_actors = unreal.EditorLevelLibrary.get_all_level_actors()
found_roof = {{}}
for a in all_actors:
    lbl = a.get_actor_label()
    if lbl in roof_labels:
        loc = a.get_actor_location()
        sc = a.get_actor_scale3d()
        rot = a.get_actor_rotation()
        found_roof[lbl] = {{'loc': (loc.x, loc.y, loc.z), 'scale': (sc.x, sc.y, sc.z), 'rot': (rot.roll, rot.pitch, rot.yaw)}}

# Check for duplicates
label_counts = {{}}
for a in all_actors:
    lbl = a.get_actor_label()
    if any(kw in lbl.lower() for kw in ['roof', 'gable', 'ridge']):
        label_counts[lbl] = label_counts.get(lbl, 0) + 1

duplicates = {{k: v for k, v in label_counts.items() if v > 1}}
if duplicates:
    print(f'WARNING: Duplicate roof actors found: {{duplicates}}')
    # Auto-fix: delete extras
    seen = set()
    for a in all_actors:
        lbl = a.get_actor_label()
        if lbl in duplicates:
            if lbl in seen:
                unreal.EditorLevelLibrary.destroy_actor(a)
                print(f'  Removed duplicate: {{lbl}}')
            else:
                seen.add(lbl)

missing = [lbl for lbl in roof_labels if lbl not in found_roof]
if missing:
    print(f'WARNING: Missing roof pieces: {{missing}}')

print(f'Validation: {{len(found_roof)}}/{{len(roof_labels)}} pieces placed. Duplicates removed: {{len(duplicates)}}')
for lbl, info in found_roof.items():
    print(f'  {{lbl}}: pos={{info["loc"]}}, scale={{info["scale"]}}, rot={{info["rot"]}}')
"""

    # ── Step 6: Execute in UEFN ──
    exec_result = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)

    details = {
        "house_bounds": {
            "outer_min_x": outer_min_x, "outer_max_x": outer_max_x,
            "outer_min_y": outer_min_y, "outer_max_y": outer_max_y,
        },
        "house_size": {"width": house_width, "depth": house_depth},
        "wall_top": wall_top,
        "roof_height": roof_height,
        "pitch_deg": pitch_deg,
        "slope_length": slope_length,
        "ridge_z": ridge_z,
        "overhang": overhang,
        "actors_used": len(house_actors),
    }
    exec_result["details"] = details
    logger.info("Roof rebuild complete: %s", details)
    return exec_result


# ════════════════════════════════════════════════════════════════
# MATERIAL CATALOG — verified Fortnite/UEFN material paths
# ════════════════════════════════════════════════════════════════
MATERIAL_CATALOG = {
    # ── Building surfaces ──
    "brick": "/Game/Packages/PBW/Brick/Materials/Brick_lvl1_Wall",
    "brick_floor": "/Game/Packages/PBW/Brick/Materials/Brick_lv3_Floor_INST",
    "wood": "/Game/Packages/PBW/Wood/Materials/Wood_lvl1_Wall",
    "wood_floor": "/Game/Packages/PBW/Wood/Materials/Wood_lvl1_Floor",
    "wood_static": "/Game/Packages/PBW/Wood/Materials/Wood_lvl1_Static",
    "metal": "/Game/Packages/PBW/Metal/Materials/Metal_lvl1_Wall",
    "metal_floor": "/Game/Packages/PBW/Metal/Materials/Metal_lv1_Floor_INST",
    "concrete": "/Game/Environments/World/Utility/MI_Concrete_WhiteGrid512",
    "stucco": "/EpicBaseMaterials/EpicBase/facades/MI_Stucco_Facade",
    "stone": "/Game/Athena/Helios/Maps/Test/Dev/ChrisK/Materials/Helios_StoneTemp",
    # ── Natural/Terrain ──
    "grass": "/BlastBerryMapContent/Game/Environments/Materials/Grass/M_Athena_Grass_Master",
    "terrain": "/Game/Packages/Fortnite_UniqueMaterials_1/DS_Fortnight/Terrain_Mat",
    "moss": "/Game/Environments/Asteria/Foliage/ForestFloor/M_GroundMoss",
    "dirt": "/Game/Environments/FN_Biomes/Desert/Materials/M_Athena_Terrain_DirtRoad",
    "road": "/Game/Environments/FN_Biomes/Urban/Materials/M_Athena_Road_Asphalt",
    "sidewalk": "/Game/Environments/FN_Biomes/Urban/Materials/M_Athena_Sidewalk_Concrete",
    "sand": "/Game/Environments/FN_Biomes/Desert/Materials/M_Athena_Terrain_Sand",
    "farmfield": "/Game/Environments/FN_Biomes/Rural/Materials/M_Athena_Terrain_FarmField",
    "rock": "/Game/Environments/FN_Biomes/Mountains/Materials/M_Athena_Terrain_Rock",
    "cliff": "/Game/Environments/FN_Biomes/Mountains/Materials/M_Athena_Terrain_Cliff",
    "ground": "/Game/Environments/FN_Biomes/Grasslands/Materials/M_Athena_Terrain_Ground",
    "ocean_floor": "/Game/Environments/FN_Biomes/Coastal/Materials/M_Athena_Terrain_OceanFloor",
    "desert_grass": "/Game/Environments/FN_Biomes/Desert/Materials/M_Athena_Terrain_DesertGrass",
    "snow": "/Game/Environments/FN_Biomes/Arctic/Materials/M_Athena_Terrain_Snow",
    "mud": "/Game/Environments/FN_Biomes/Swamp/Materials/M_Athena_Terrain_Mud",
    # ── Water/Liquid ──
    "water": "/BlastBerryMapContent/Game/Packages/Fortnite_UniqueMaterials_1/DS_Fortnight/Rufus_Water_Shallow_New",
    "lava": "/DelMarGame/Environments/S2_Volcanic/Materials/M_RR_Lava",
    # ── Effects ──
    "glass": "/BRCosmetics/Accessories/FORT_Backpacks/Backpack_M_MED_DomeRoof/Materials/M_MED_Glass_DomeRoof",
    "glitter": "/Game/Effects/Fort_Effects/Materials/Walls/M_Wall_Glitter",
    "sparkle": "/Game/Effects/Fort_Effects/Materials/Walls/M_Wall_Sparkle2",
    # ── Default ──
    "default": "/Engine/BasicShapes/BasicShapeMaterial",
}

# Aliases for natural language matching
MATERIAL_ALIASES = {
    "bricks": "brick", "wooden": "wood", "metallic": "metal", "steel": "metal",
    "cement": "concrete", "rock": "stone", "rocks": "stone", "marble": "stone",
    "earth": "dirt", "soil": "dirt", "muddy": "mud",
    "asphalt": "road", "pavement": "road", "street": "road",
    "path": "sidewalk", "walkway": "sidewalk", "paving": "sidewalk",
    "sandy": "sand", "beach": "sand", "dune": "sand",
    "farm": "farmfield", "field": "farmfield", "crop": "farmfield",
    "rocky": "rock", "boulder": "rock", "mountain": "cliff",
    "snowy": "snow", "frozen": "snow", "arctic": "snow", "winter": "snow",
    "swamp": "mud", "wetland": "mud", "marsh": "mud",
    "seabed": "ocean_floor", "underwater": "ocean_floor",
    "arid": "desert_grass", "savanna": "desert_grass", "steppe": "desert_grass",
    "grassy": "grass", "lawn": "grass", "turf": "grass",
    "ocean": "water", "pool": "water", "pond": "water", "river": "water",
    "ice": "glass", "crystal": "glass", "transparent": "glass",
    "fire": "lava", "magma": "lava", "volcanic": "lava",
    "shiny": "sparkle", "glowing": "glitter",
    "plaster": "stucco", "white": "stucco", "smooth": "stucco",
    "timber": "wood", "lumber": "wood", "plank": "wood",
    "iron": "metal", "aluminum": "metal", "chrome": "metal",
    "cobblestone": "stone", "flagstone": "stone", "slate": "stone",
}


def _resolve_material_name(name: str) -> str:
    """Resolve a material name/alias to a catalog key."""
    name = name.lower().strip()
    if name in MATERIAL_CATALOG:
        return name
    return MATERIAL_ALIASES.get(name, name)


def _get_material_path(name: str) -> Optional[str]:
    """Get the UEFN asset path for a material name."""
    key = _resolve_material_name(name)
    return MATERIAL_CATALOG.get(key)


def _execute_apply_material(action: dict) -> dict:
    """Apply a material to actors by label pattern.

    action = {"actor_pattern": "Fountain_*", "material": "stone"}
    """
    port = discover_uefn_listener_port()
    if not port:
        return {"error": "UEFN not connected"}

    pattern = action.get("actor_pattern", action.get("actor", ""))
    mat_name = action.get("material", "default")
    mat_path = _get_material_path(mat_name)
    if not mat_path:
        return {"error": f"Unknown material: {mat_name}. Available: {', '.join(sorted(MATERIAL_CATALOG.keys()))}"}

    # Convert glob pattern to Python matching
    import fnmatch
    py_pattern = pattern.replace("*", ".*")

    code = f"""import unreal, re
mat = unreal.EditorAssetLibrary.load_asset('{mat_path}')
if not mat:
    result = 'Failed to load material: {mat_path}'
else:
    applied = 0
    pattern = '{pattern}'
    for a in unreal.EditorLevelLibrary.get_all_level_actors():
        label = a.get_actor_label()
        import fnmatch
        if fnmatch.fnmatch(label, pattern) or fnmatch.fnmatch(label.lower(), pattern.lower()):
            mc = a.get_component_by_class(unreal.StaticMeshComponent)
            if mc:
                mc.set_material(0, mat)
                applied += 1
    result = f'Applied {{mat.get_name()}} to {{applied}} actors matching {pattern}'
"""
    return mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=15.0)


def _execute_generative_build(action: dict) -> dict:
    """Generative construction: build a structure from a description.

    action = {
        "structure": "fountain|tower|bridge|arch|column|platform|fence|pool|waterfall",
        "position": {"x": 5200, "y": 4200, "z": 0},
        "size": "small|medium|large",
        "material": "stone|brick|wood|metal|concrete",
    }
    """
    import math

    port = discover_uefn_listener_port()
    if not port:
        return {"error": "UEFN not connected"}

    structure = str(action.get("structure", "")).strip().lower()
    if structure in {"house", "home", "cabin", "cottage", "townhouse", "villa", "mansion", "estate", "manor", "modern house", "suburban house", "apartment", "apartment building", "condo", "condominium"}:
        return _execute_build_house({
            "request": str(action.get("request") or structure),
            "style": (_canonical_residential_style(structure) or structure) if structure not in {"house", "home"} else "",
            "story_count": action.get("story_count")
            or (_parse_requested_story_count(str(action.get("request") or structure), default=4) if structure in {"apartment", "apartment building", "condo", "condominium"} else (_parse_requested_story_count(str(action.get("request") or structure), default=3) if structure in {"mansion", "estate", "manor", "villa"} else None)),
            "position": action.get("position") or {},
            "size": action.get("size") or "medium",
            "material": action.get("material") or "",
            "roof_material": action.get("roof_material") or "",
            "label_prefix": action.get("label") or "",
        })
    if structure in _SHARED_GENERATIVE_STRUCTURE_TYPES or structure in _DIRECT_STRUCTURE_KEYWORDS:
        return _execute_build_structure_request({
            "request": str(action.get("request") or structure),
            "structure": structure,
            "style": action.get("style") or "",
            "position": action.get("position") or {},
            "size": action.get("size") or "medium",
            "material": action.get("material") or "",
            "roof_material": action.get("roof_material") or "",
            "label_prefix": action.get("label") or "",
            "zone_id": action.get("zone_id") or "",
        })

    pos = action.get("position", {"x": 5200, "y": 4200, "z": 0})
    px, py, pz = pos.get("x", 5200), pos.get("y", 4200), pos.get("z", 0)
    size = action.get("size", "medium")
    mat_name = _resolve_material_name(action.get("material", "stone"))
    mat_path = _get_material_path(mat_name) or MATERIAL_CATALOG["stone"]

    # Size multipliers
    scale_mult = {"small": 0.5, "medium": 1.0, "large": 1.5, "huge": 2.5}.get(size, 1.0)

    # Label prefix
    prefix = structure.replace(" ", "_").title()

    # Build the spawn code for the chosen structure
    pieces = []

    if structure in ("fountain", "water fountain"):
        s = scale_mult
        water_path = MATERIAL_CATALOG["water"]
        pieces = [
            # Base
            (0, 0, 5*s, 4*s, 4*s, 0.10*s, f"{prefix}_Base", mat_path),
            # Basin walls
            (0, -190*s, 30*s, 3.8*s, 0.12, 0.5*s, f"{prefix}_Basin_F", mat_path),
            (0, 190*s, 30*s, 3.8*s, 0.12, 0.5*s, f"{prefix}_Basin_B", mat_path),
            (-190*s, 0, 30*s, 0.12, 3.8*s, 0.5*s, f"{prefix}_Basin_L", mat_path),
            (190*s, 0, 30*s, 0.12, 3.8*s, 0.5*s, f"{prefix}_Basin_R", mat_path),
            # Water
            (0, 0, 20*s, 3.6*s, 3.6*s, 0.04, f"{prefix}_Water", water_path),
            # Pedestal
            (0, 0, 15*s, 1.2*s, 1.2*s, 0.10*s, f"{prefix}_Ped_Base", mat_path),
            (0, 0, 40*s, 0.8*s, 0.8*s, 0.40*s, f"{prefix}_Ped_Mid", mat_path),
            (0, 0, 75*s, 0.5*s, 0.5*s, 0.30*s, f"{prefix}_Ped_Top", mat_path),
            # Upper bowl
            (0, -55*s, 95*s, 1.4*s, 0.08, 0.20*s, f"{prefix}_Bowl_F", mat_path),
            (0, 55*s, 95*s, 1.4*s, 0.08, 0.20*s, f"{prefix}_Bowl_B", mat_path),
            (-55*s, 0, 95*s, 0.08, 1.4*s, 0.20*s, f"{prefix}_Bowl_L", mat_path),
            (55*s, 0, 95*s, 0.08, 1.4*s, 0.20*s, f"{prefix}_Bowl_R", mat_path),
            (0, 0, 88*s, 1.2*s, 1.2*s, 0.04, f"{prefix}_UpperWater", water_path),
            # Spout
            (0, 0, 105*s, 0.3*s, 0.3*s, 0.15*s, f"{prefix}_Spout_Base", mat_path),
            (0, 0, 118*s, 0.15*s, 0.15*s, 0.12*s, f"{prefix}_Spout_Top", mat_path),
            # Corner posts
            (-185*s, -185*s, 35*s, 0.15, 0.15, 0.6*s, f"{prefix}_Post_FL", mat_path),
            (185*s, -185*s, 35*s, 0.15, 0.15, 0.6*s, f"{prefix}_Post_FR", mat_path),
            (-185*s, 185*s, 35*s, 0.15, 0.15, 0.6*s, f"{prefix}_Post_BL", mat_path),
            (185*s, 185*s, 35*s, 0.15, 0.15, 0.6*s, f"{prefix}_Post_BR", mat_path),
        ]

    elif structure in ("column", "pillar"):
        s = scale_mult
        pieces = [
            (0, 0, 5, 1.0*s, 1.0*s, 0.10, f"{prefix}_Plinth", mat_path),
            (0, 0, 150*s, 0.6*s, 0.6*s, 2.8*s, f"{prefix}_Shaft", mat_path),
            (0, 0, 300*s+5, 1.0*s, 1.0*s, 0.10, f"{prefix}_Capital", mat_path),
        ]

    elif structure in ("arch", "archway"):
        s = scale_mult
        pieces = [
            (-120*s, 0, 150*s, 0.4*s, 0.4*s, 3.0*s, f"{prefix}_Left", mat_path),
            (120*s, 0, 150*s, 0.4*s, 0.4*s, 3.0*s, f"{prefix}_Right", mat_path),
            (0, 0, 305*s, 2.8*s, 0.5*s, 0.3*s, f"{prefix}_Top", mat_path),
        ]

    elif structure in ("tower", "watchtower"):
        s = scale_mult
        h = 500 * s
        pieces = [
            (0, 0, h/2, 0.2*s, 2.0*s, h/100, f"{prefix}_Wall_F", mat_path),
            (0, 0, h/2, 2.0*s, 0.2*s, h/100, f"{prefix}_Wall_L", mat_path),
            (-100*s, 0, h/2, 0.2*s, 2.0*s, h/100, f"{prefix}_Wall_B", mat_path),
            (100*s, 0, h/2, 0.2*s, 2.0*s, h/100, f"{prefix}_Wall_R", mat_path),
            (0, 0, h+10, 2.2*s, 2.2*s, 0.10, f"{prefix}_Floor_Top", mat_path),
            (0, 0, 5, 2.2*s, 2.2*s, 0.10, f"{prefix}_Floor_Bot", mat_path),
        ]

    elif structure in ("pool", "swimming pool"):
        s = scale_mult
        water_path = MATERIAL_CATALOG["water"]
        pieces = [
            (0, 0, -20*s, 6.0*s, 3.0*s, 0.10, f"{prefix}_Floor", mat_path),
            (0, -145*s, 0, 6.0*s, 0.10, 0.4*s, f"{prefix}_Wall_F", mat_path),
            (0, 145*s, 0, 6.0*s, 0.10, 0.4*s, f"{prefix}_Wall_B", mat_path),
            (-295*s, 0, 0, 0.10, 3.0*s, 0.4*s, f"{prefix}_Wall_L", mat_path),
            (295*s, 0, 0, 0.10, 3.0*s, 0.4*s, f"{prefix}_Wall_R", mat_path),
            (0, 0, 10*s, 5.8*s, 2.8*s, 0.04, f"{prefix}_Water", water_path),
        ]

    elif structure in ("fence", "wall fence"):
        s = scale_mult
        pieces = []
        for i in range(6):
            xo = (i - 2.5) * 120 * s
            pieces.append((xo, 0, 50*s, 0.1*s, 0.1*s, 1.0*s, f"{prefix}_Post_{i}", mat_path))
            if i < 5:
                pieces.append((xo + 60*s, 0, 70*s, 1.1*s, 0.05, 0.05, f"{prefix}_Rail_T_{i}", mat_path))
                pieces.append((xo + 60*s, 0, 30*s, 1.1*s, 0.05, 0.05, f"{prefix}_Rail_B_{i}", mat_path))

    elif structure in ("platform", "stage"):
        s = scale_mult
        pieces = [
            (0, 0, 25*s, 5.0*s, 3.0*s, 0.5*s, f"{prefix}_Top", mat_path),
            (0, 0, 5, 5.4*s, 3.4*s, 0.10, f"{prefix}_Base", mat_path),
        ]

    elif structure in ("bridge",):
        s = scale_mult
        pieces = [
            (0, 0, 100*s, 1.5*s, 8.0*s, 0.15, f"{prefix}_Deck", mat_path),
            (0, -350*s, 50*s, 1.5*s, 0.4*s, 1.0*s, f"{prefix}_Support_L", mat_path),
            (0, 350*s, 50*s, 1.5*s, 0.4*s, 1.0*s, f"{prefix}_Support_R", mat_path),
            (-70*s, 0, 115*s, 0.08, 8.0*s, 0.15*s, f"{prefix}_Rail_L", mat_path),
            (70*s, 0, 115*s, 0.08, 8.0*s, 0.15*s, f"{prefix}_Rail_R", mat_path),
        ]

    elif structure in ("waterfall", "cascade"):
        s = scale_mult
        water_path = MATERIAL_CATALOG["water"]
        support_material_name = _resolve_material_name(action.get("support_material", action.get("material", "cliff")))
        support_path = _get_material_path(support_material_name) or _get_material_path("cliff") or mat_path
        pieces = [
            (0, -240*s, 35*s, 4.8*s, 2.4*s, 0.35*s, f"{prefix}_TopLedge", support_path, 0.0, 0.0, 0.0),
            (0, -220*s, 65*s, 4.2*s, 1.9*s, 0.05*s, f"{prefix}_TopPool", water_path, 0.0, 0.0, 0.0),
            (0, 10*s, 250*s, 5.6*s, 0.35*s, 5.0*s, f"{prefix}_CliffBack", support_path, 0.0, 0.0, 0.0),
            (-225*s, 15*s, 185*s, 0.42*s, 2.2*s, 3.8*s, f"{prefix}_CliffSide_L", support_path, 0.0, 0.0, 0.0),
            (225*s, 15*s, 185*s, 0.42*s, 2.2*s, 3.8*s, f"{prefix}_CliffSide_R", support_path, 0.0, 0.0, 0.0),
            (0, 30*s, 130*s, 2.9*s, 6.0*s, 0.06*s, f"{prefix}_Cascade_Main", water_path, -32.0, 0.0, 0.0),
            (0, 270*s, -5*s, 3.8*s, 2.4*s, 0.05*s, f"{prefix}_SplashSheet", water_path, -10.0, 0.0, 0.0),
            (0, 360*s, -70*s, 6.8*s, 3.8*s, 0.05*s, f"{prefix}_BottomPool", water_path, 0.0, 0.0, 0.0),
            (-315*s, 360*s, -10*s, 0.14*s, 1.5*s, 0.75*s, f"{prefix}_Bank_L", support_path, 0.0, 0.0, 0.0),
            (315*s, 360*s, -10*s, 0.14*s, 1.5*s, 0.75*s, f"{prefix}_Bank_R", support_path, 0.0, 0.0, 0.0),
        ]

    else:
        return {"error": f"Unknown structure: {structure}. Available scenic structures: fountain, column, arch, tower, pool, fence, platform, bridge, waterfall. Residential builds should use build_house/build_house_action, and planner-backed shared structures include garage, shed, workshop, barn, warehouse, greenhouse, studio, hangar, kiosk, pavilion, gazebo, pergola, canopy, carport, and market stall."}

    # Generate spawn code
    spawn_lines = []
    for piece in pieces:
        if len(piece) >= 11:
            dx, dy, dz, sx, sy, sz, label, mp, pitch, yaw, roll = piece
        else:
            dx, dy, dz, sx, sy, sz, label, mp = piece
            pitch = yaw = roll = 0.0
        x, y, z = px + dx, py + dy, pz + dz
        spawn_lines.append(
            f"a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({x},{y},{z}))\n"
            f"m = a.get_component_by_class(unreal.StaticMeshComponent)\n"
            f"m.set_static_mesh(mesh)\n"
            f"a.set_actor_scale3d(unreal.Vector({sx},{sy},{sz}))\n"
            f"a.set_actor_rotation(unreal.Rotator({roll},{pitch},{yaw}), False)\n"
            f"a.set_actor_label('{label}')\n"
            f"m.set_collision_profile_name('BlockAll')\n"
            f"mat = unreal.EditorAssetLibrary.load_asset('{mp}')\n"
            f"if mat: m.set_material(0, mat)\n"
        )

    code = "import unreal\nmesh = unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube')\n\n"
    code += "\n".join(spawn_lines)
    code += f"\nresult = 'Built {structure} ({len(pieces)} pieces) at ({px},{py},{pz})'\nprint(result)"

    exec_result = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)
    exec_result["details"] = {
        "structure": structure,
        "position": {"x": px, "y": py, "z": pz},
        "size": size,
        "material": mat_name,
        "pieces": len(pieces),
    }
    return exec_result


def _safe_triplet(value: Any, default: list[float] | None = None) -> list[float]:
    fallback = list(default or [0.0, 0.0, 0.0])
    if isinstance(value, dict):
        return [
            _safe_number(value.get("x"), fallback[0]),
            _safe_number(value.get("y"), fallback[1]),
            _safe_number(value.get("z"), fallback[2]),
        ]
    if isinstance(value, (list, tuple)):
        padded = list(value[:3]) + fallback[len(value[:3]):]
        return [
            _safe_number(padded[0], fallback[0]),
            _safe_number(padded[1], fallback[1]),
            _safe_number(padded[2], fallback[2]),
        ]
    return fallback


def _safe_number(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _clean_identifier_fragment(value: str, fallback: str = "default") -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return cleaned or fallback


def _stable_house_seed(message: str, chat_id: str, explicit_seed: Any = None) -> int:
    digest = hashlib.sha256(f"{chat_id}|{message}|{explicit_seed or ''}".encode("utf-8")).hexdigest()
    return int(digest[:8], 16)


def _house_style_from_request(message: str, explicit_style: str = "") -> str:
    preferred = str(explicit_style or "").strip().lower()
    if preferred:
        return preferred
    lowered = str(message or "").lower()
    style_keywords = (
        ("mansion", ("mansion", "estate", "manor", "grand house", "grand mansion", "luxury home")),
        ("villa", ("villa", "mediterranean", "courtyard villa", "estate villa")),
        ("modern", ("modern", "sleek", "minimal", "contemporary")),
        ("cabin", ("cabin", "lodge", "rustic", "log house")),
        ("cottage", ("cottage", "cozy", "storybook", "country house")),
        ("townhouse", ("townhouse", "town home", "row house", "rowhome")),
        ("apartment", ("apartment", "apartment building", "condo", "condominium", "residential block", "low rise")),
        ("suburban", ("suburban", "family house", "starter home", "residential")),
    )
    for style_name, keywords in style_keywords:
        if any(keyword in lowered for keyword in keywords):
            return style_name
    return "suburban"


def _size_multiplier_from_label(size_label: str) -> float:
    normalized = str(size_label or "medium").strip().lower()
    return {
        "tiny": 0.8,
        "small": 0.9,
        "medium": 1.0,
        "large": 1.15,
        "huge": 1.3,
    }.get(normalized, 1.0)


def _apply_house_request_modifiers(
    message: str,
    *,
    width_cm: float,
    depth_cm: float,
    story_height_cm: float,
    roof_pitch_deg: float,
    roof_rise_cm: float,
    balcony_depth_cm: float,
    entry_canopy_depth_cm: float,
    corner_column_diameter_cm: float,
    window_columns_per_wall: int,
    site_clearance_cm: float,
    style: str,
) -> dict[str, Any]:
    lowered = str(message or "").lower()
    width_scale = 1.0
    depth_scale = 1.0
    story_height_scale = 1.0
    roof_pitch_scale = 1.0
    roof_rise_scale = 1.0
    balcony_depth = float(balcony_depth_cm)
    canopy_depth = float(entry_canopy_depth_cm)
    corner_columns = float(corner_column_diameter_cm)
    window_columns = int(window_columns_per_wall)
    site_clearance = float(site_clearance_cm)

    if style in {"mansion", "villa"}:
        width_scale *= 1.16
        depth_scale *= 1.1
        site_clearance = max(site_clearance, 120.0)
    elif style == "apartment":
        width_scale *= 1.08
        depth_scale *= 1.08
        site_clearance = max(site_clearance, 100.0)

    if any(token in lowered for token in ("grand", "luxury", "estate", "expansive", "sprawling")):
        width_scale *= 1.12
        depth_scale *= 1.08
        site_clearance = max(site_clearance, 130.0)
    if any(token in lowered for token in ("compact", "small footprint", "tight lot")):
        width_scale *= 0.88
        depth_scale *= 0.9
    if any(token in lowered for token in ("wide", "broad", "wider")):
        width_scale *= 1.12
    if any(token in lowered for token in ("deep", "long", "extended")):
        depth_scale *= 1.12
    if any(token in lowered for token in ("tall ceiling", "tall ceilings", "lofty", "grand foyer")):
        story_height_scale *= 1.08
    if any(token in lowered for token in ("steep roof", "pitched roof", "dramatic roof", "proper roof")):
        roof_pitch_scale *= 1.08
        roof_rise_scale *= 1.1
    if any(token in lowered for token in ("flat roof", "roof deck", "parapet")):
        roof_pitch_scale *= 0.82
        roof_rise_scale *= 0.82
    if any(token in lowered for token in ("balcony", "terrace", "veranda")):
        balcony_depth = max(balcony_depth, 120.0 if style in {"mansion", "villa"} else 95.0)
    if any(token in lowered for token in ("porch", "portico", "covered entry", "canopy")):
        canopy_depth = max(canopy_depth, 110.0 if style in {"mansion", "villa"} else 85.0)
    if any(token in lowered for token in ("arched", "curved", "column", "columns")):
        corner_columns = max(corner_columns, 42.0 if style in {"mansion", "villa"} else 28.0)
    if any(token in lowered for token in ("window", "windows", "windowed", "many windows", "lots of windows")):
        window_columns += 1

    return {
        "width_cm": round(width_cm * width_scale, 3),
        "depth_cm": round(depth_cm * depth_scale, 3),
        "story_height_cm": round(story_height_cm * story_height_scale, 3),
        "roof_pitch_deg": round(roof_pitch_deg * roof_pitch_scale, 3),
        "roof_rise_cm": round(roof_rise_cm * roof_rise_scale, 3),
        "balcony_depth_cm": round(balcony_depth, 3),
        "entry_canopy_depth_cm": round(canopy_depth, 3),
        "corner_column_diameter_cm": round(corner_columns, 3),
        "window_columns_per_wall": max(1, min(5, window_columns)),
        "site_clearance_cm": round(site_clearance, 3),
    }


_NUMBER_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
}


def _parse_requested_story_count(message: str, explicit_story_count: Any = None, default: int = 2) -> int:
    try:
        explicit_value = int(explicit_story_count)
        return max(1, min(8, explicit_value))
    except (TypeError, ValueError):
        pass

    lowered = str(message or "").strip().lower()
    if not lowered:
        return max(1, min(8, default))

    digit_match = re.search(r"\b(\d+)\s*[- ]?\s*(?:story|stories|floor|floors|level|levels)\b", lowered)
    if digit_match:
        return max(1, min(8, int(digit_match.group(1))))

    word_pattern = "|".join(sorted(_NUMBER_WORDS, key=len, reverse=True))
    word_match = re.search(rf"\b({word_pattern})\s*[- ]?\s*(?:story|stories|floor|floors|level|levels)\b", lowered)
    if word_match:
        return max(1, min(8, int(_NUMBER_WORDS.get(word_match.group(1), default))))

    return max(1, min(8, default))


_DIRECT_BUILD_VERBS = ("build", "make", "create", "generate", "spawn", "place", "add", "construct")
_DIRECT_BUILD_INTENT_PHRASES = ("i want", "i need", "give me", "put a", "put an", "drop a", "drop an")
_DIRECT_BUILD_BLOCKERS = (
    "how do",
    "how to",
    "why is",
    "what is",
    "what are",
    "explain",
    "prompt",
    "code review",
    "analyze",
    "analyse",
    "look at the code",
    "plan out",
    "make sure",
    "if i say",
    "when i say",
)
_DIRECT_HOUSE_KEYWORDS = (
    "house",
    "home",
    "cabin",
    "cottage",
    "mansion",
    "estate",
    "manor",
    "townhouse",
    "apartment building",
    "apartment",
    "condominium",
    "condo",
    "villa",
    "suburban house",
    "modern house",
)
_DIRECT_STRUCTURE_KEYWORDS = {
    "garage": "garage",
    "shed": "shed",
    "tool shed": "shed",
    "storage shed": "shed",
    "workshop": "workshop",
    "barn": "barn",
    "warehouse": "warehouse",
    "storage building": "warehouse",
    "greenhouse": "greenhouse",
    "glasshouse": "greenhouse",
    "studio": "studio",
    "art studio": "studio",
    "hangar": "hangar",
    "aircraft hangar": "hangar",
    "kiosk": "kiosk",
    "booth": "kiosk",
    "pavilion": "pavilion",
    "gazebo": "gazebo",
    "pergola": "pergola",
    "canopy": "canopy",
    "carport": "carport",
    "market stall": "market_stall",
    "stall": "market_stall",
}
_SHARED_GENERATIVE_STRUCTURE_TYPES = set(SUPPORTED_GENERATIVE_STRUCTURE_TYPES)
_RESIDENTIAL_STRUCTURE_ALIASES = {
    "house": "suburban",
    "home": "suburban",
    "suburban house": "suburban",
    "modern house": "modern",
    "cabin": "cabin",
    "cottage": "cottage",
    "mansion": "mansion",
    "estate": "mansion",
    "manor": "mansion",
    "townhouse": "townhouse",
    "villa": "villa",
    "apartment": "apartment",
    "apartment building": "apartment",
    "condo": "apartment",
    "condominium": "apartment",
    "duplex": "townhouse",
}
_TOOL_FORCED_TERRAIN_KEYWORDS = (
    "terrain",
    "landscape",
    "ground",
    "hill",
    "ridge",
    "plateau",
    "slope",
    "crater",
    "shore",
    "shoreline",
    "road",
    "path",
    "river",
    "waterfall",
)
_TOOL_FORCED_IMPORT_KEYWORDS = ("import", "fbx", "mesh", "model", "zip", "glb", "gltf")


def _direct_build_intent_requested(message: str) -> bool:
    lowered = str(message or "").strip().lower()
    if not lowered:
        return False
    if any(phrase in lowered for phrase in _DIRECT_BUILD_BLOCKERS):
        return False
    if re.search(r"\b(build|create|generate|spawn|place|add|construct)\b", lowered):
        return True
    if re.search(r"\bmake\b(?!\s+sure)", lowered):
        return True
    return any(phrase in lowered for phrase in _DIRECT_BUILD_INTENT_PHRASES)


def _message_contains_keyword_phrase(message: str, keyword: str) -> bool:
    lowered = str(message or "").strip().lower()
    normalized_keyword = str(keyword or "").strip().lower()
    if not lowered or not normalized_keyword:
        return False
    parts = [re.escape(part) for part in normalized_keyword.split() if part]
    if not parts:
        return False
    pattern = r"(?<!\w)" + r"\s+".join(parts) + r"(?!\w)"
    return re.search(pattern, lowered) is not None


def _detect_direct_structure_request(message: str) -> dict[str, Any] | None:
    lowered = str(message or "").strip().lower()
    if not _direct_build_intent_requested(lowered):
        return None
    for keyword in sorted(_DIRECT_HOUSE_KEYWORDS, key=len, reverse=True):
        if _message_contains_keyword_phrase(lowered, keyword):
            return {"kind": "house", "matched_keyword": keyword}
    for keyword, structure_type in sorted(_DIRECT_STRUCTURE_KEYWORDS.items(), key=lambda item: len(item[0]), reverse=True):
        if _message_contains_keyword_phrase(lowered, keyword):
            return {"kind": "structure", "matched_keyword": keyword, "structure_type": structure_type}
    return None


def _looks_like_terrain_request(message: str) -> bool:
    lowered = str(message or "").strip().lower()
    if not lowered:
        return False
    return any(token in lowered for token in _TOOL_FORCED_TERRAIN_KEYWORDS)


def _build_tool_routing_guidance(message: str, attachments: list[dict[str, Any]] | None = None) -> str:
    lowered = str(message or "").strip().lower()
    direct_structure_request = _detect_direct_structure_request(message)
    importable_attachments = _collect_importable_model_files(attachments or [])
    matching_tools = _search_tools_for_llm(message, limit=6)

    lines: list[str] = []
    if direct_structure_request:
        if direct_structure_request.get("kind") == "house":
            lines.append("- Use `build_house_action` for this request instead of ad-hoc Python or generic prose.")
            story_count = _parse_requested_story_count(message, default=2)
            if story_count > 2:
                lines.append(f"- This is multi-story residential. Set `story_count={story_count}` and keep stairs, landings, and the roof fully functional.")
        elif direct_structure_request.get("kind") == "structure":
            lines.append(
                f"- Use `build_structure_action` for this request and set `structure=\"{direct_structure_request.get('structure_type')}\"` so shared geometry code handles the build."
            )
    if _looks_like_terrain_request(lowered):
        lines.append("- This looks like a terrain/landform request. Prefer `terrain_action` before any manual mesh placement.")
    if importable_attachments:
        lines.append("- Importable model attachments are present. Prefer `import_attached_models` before inventing placeholder geometry.")
    if any(token in lowered for token in _TOOL_FORCED_IMPORT_KEYWORDS):
        lines.append("- If the request references uploaded meshes or FBX files, use import actions/tools first and preserve their materials.")
    if matching_tools:
        tool_ids = ", ".join(f"`{tool.get('id', 'unknown')}`" for tool in matching_tools[:4] if tool.get("id"))
        if tool_ids:
            lines.append(f"- Matching registered tools for this request: {tool_ids}. Use `run_uefn_tool` when one of these fits precisely.")

    if not lines:
        return ""
    return "TOOL ROUTING GUIDANCE:\n" + "\n".join(lines)


def _should_force_tool_execution(
    message: str,
    attachments: list[dict[str, Any]] | None,
    *,
    use_tools: bool,
    user_wants_action: bool,
    requires_reasoned_targeting: bool,
) -> bool:
    if not use_tools:
        return False
    if _detect_direct_structure_request(message):
        return True
    if _looks_like_terrain_request(message):
        return True
    if _collect_importable_model_files(attachments or []):
        return True
    return user_wants_action and not requires_reasoned_targeting


def _infer_body_and_roof_materials(message: str, explicit_material: str = "", explicit_roof_material: str = "", style: str = "") -> tuple[str, str]:
    body = _resolve_material_name(explicit_material or "")
    roof = _resolve_material_name(explicit_roof_material or "")
    if not body or body not in MATERIAL_CATALOG:
        defaults = {
            "mansion": "stone",
            "villa": "stucco",
            "modern": "concrete",
            "cabin": "wood",
            "cottage": "stucco",
            "townhouse": "brick",
            "apartment": "concrete",
            "suburban": "stucco",
        }
        body = defaults.get(style or "suburban", "stucco")
        lowered = str(message or "").lower()
        for token in sorted(MATERIAL_CATALOG, key=len, reverse=True):
            if token in lowered:
                body = _resolve_material_name(token)
                break
    if not roof or roof not in MATERIAL_CATALOG:
        roof_defaults = {
            "mansion": "stone",
            "villa": "stucco",
            "modern": "metal",
            "cabin": "wood",
            "cottage": "wood",
            "townhouse": "metal",
            "apartment": "metal",
            "suburban": "metal",
        }
        roof = roof_defaults.get(style or "suburban", "metal")
    return body, roof


def _select_material_for_structure_role(
    *,
    material_role: str,
    body_material_path: str | None,
    roof_material_path: str | None,
    trim_material_path: str | None,
    floor_material_path: str | None,
    glass_material_path: str | None,
) -> str | None:
    role = str(material_role or "body").strip().lower()
    if role == "roof":
        return roof_material_path or body_material_path
    if role == "glass":
        return glass_material_path or trim_material_path or body_material_path
    if role == "trim":
        return trim_material_path or body_material_path or roof_material_path
    if role == "floor":
        return floor_material_path or body_material_path
    return body_material_path or trim_material_path or roof_material_path


def _canonical_residential_style(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    return _RESIDENTIAL_STRUCTURE_ALIASES.get(normalized, "")


def _execute_structure_action_with_shared_planner(action: dict[str, Any]) -> dict[str, Any]:
    context = _get_active_tool_context()
    explicit_structure_value = str(
        action.get("structure")
        or action.get("structure_type")
        or ""
    ).strip()
    request_text = str(
        action.get("request")
        or context.get("message")
        or explicit_structure_value
        or ""
    ).strip()
    direct_request = _detect_direct_structure_request(request_text)

    if direct_request and direct_request.get("kind") == "house":
        forwarded_action = dict(action)
        forwarded_action["request"] = request_text
        matched_keyword = str(direct_request.get("matched_keyword") or "").strip()
        residential_style = (
            _canonical_residential_style(matched_keyword)
            or _canonical_residential_style(explicit_structure_value)
        )
        if residential_style and not str(forwarded_action.get("style") or "").strip():
            forwarded_action["style"] = residential_style
        if forwarded_action.get("story_count") in (None, ""):
            story_default = 4 if residential_style == "apartment" else 2
            forwarded_action["story_count"] = _parse_requested_story_count(
                request_text,
                default=story_default,
            )
        return _execute_build_house(forwarded_action)

    if direct_request and direct_request.get("kind") == "structure":
        canonical_requested_structure = canonical_structure_type(
            str(direct_request.get("structure_type") or ""),
            fallback="",
        )
        if canonical_requested_structure:
            forwarded_action = dict(action)
            forwarded_action["request"] = request_text
            forwarded_action["structure"] = canonical_requested_structure
            return _execute_build_structure_request(forwarded_action)

    structure_value = explicit_structure_value or request_text
    residential_style = _canonical_residential_style(structure_value)
    if residential_style:
        forwarded_action = dict(action)
        forwarded_action["request"] = request_text
        if residential_style != "suburban" and not str(forwarded_action.get("style") or "").strip():
            forwarded_action["style"] = residential_style
        if residential_style == "apartment" and forwarded_action.get("story_count") in (None, ""):
            forwarded_action["story_count"] = _parse_requested_story_count(
                request_text,
                default=4,
            )
        return _execute_build_house(forwarded_action)

    canonical_structure = canonical_structure_type(structure_value, fallback="")
    if canonical_structure:
        forwarded_action = dict(action)
        forwarded_action["request"] = request_text
        forwarded_action["structure"] = canonical_structure
        return _execute_build_structure_request(forwarded_action)
    return _execute_generative_build(action)


def _support_anchor_for_actor_payload(actor: dict[str, Any]) -> list[float]:
    support_anchor = actor.get("support_anchor")
    if isinstance(support_anchor, list) and len(support_anchor) >= 3:
        return _safe_triplet(support_anchor)
    bounds = dict(actor.get("bounds_cm") or {})
    if bounds:
        origin = _safe_triplet(bounds.get("origin"), _safe_triplet(actor.get("location")))
        extent = _safe_triplet(bounds.get("box_extent"))
        return [origin[0], origin[1], origin[2] + extent[2]]
    return _safe_triplet(actor.get("location"))


def _resolve_house_support_context(action: dict[str, Any], scene_state: dict[str, Any]) -> dict[str, Any]:
    actors = [dict(actor) for actor in list(scene_state.get("actors") or []) if isinstance(actor, dict)]
    explicit_support = str(action.get("support_actor") or action.get("support_actor_label") or "").strip()
    selected = [actor for actor in actors if bool(actor.get("selected"))]
    support_actor: dict[str, Any] | None = None
    if explicit_support:
        lowered = explicit_support.lower()
        support_actor = next(
            (
                actor for actor in actors
                if str(actor.get("label") or "").strip().lower() == lowered
                or str(actor.get("actor_path") or "").strip().lower() == lowered
            ),
            None,
        )
    if support_actor is None and selected:
        support_actor = dict(selected[0])
    placement_targets = dict(scene_state.get("placement_targets") or {})
    fallback_anchor = (
        placement_targets.get("surface_anchor")
        or placement_targets.get("ground_anchor")
        or placement_targets.get("anchor_point")
        or [0.0, 0.0, 0.0]
    )
    support_anchor = _safe_triplet(_support_anchor_for_actor_payload(support_actor) if support_actor else fallback_anchor)
    requested_position = dict(action.get("position") or {})
    center_x = _safe_number(requested_position.get("x"), support_anchor[0])
    center_y = _safe_number(requested_position.get("y"), support_anchor[1])
    support_z = _safe_number(requested_position.get("z"), support_anchor[2])
    support_surface_kind = str(
        action.get("support_surface_kind")
        or (support_actor or {}).get("support_surface_kind")
        or placement_targets.get("support_surface_kind")
        or "support_surface"
    )
    support_level = int(
        action.get("support_level")
        or (support_actor or {}).get("support_level")
        or placement_targets.get("support_level")
        or 0
    )
    support_actor_label = str(
        action.get("support_actor_label")
        or (support_actor or {}).get("label")
        or placement_targets.get("support_actor_label")
        or explicit_support
    ).strip()
    support_actor_path = str(
        (support_actor or {}).get("actor_path")
        or placement_targets.get("support_actor_path")
        or ""
    ).strip()
    return {
        "center_x": center_x,
        "center_y": center_y,
        "support_z": support_z,
        "support_anchor": support_anchor,
        "support_surface_kind": support_surface_kind,
        "support_level": support_level,
        "support_actor_label": support_actor_label,
        "support_actor_path": support_actor_path,
        "support_actor": support_actor or {},
    }


def _generate_house_spec_from_request(action: dict[str, Any], *, message: str, chat_id: str, support_context: dict[str, Any]) -> tuple[HouseSpec, dict[str, Any]]:
    style = _house_style_from_request(message, str(action.get("style") or ""))
    seed = _stable_house_seed(message, chat_id, explicit_seed=action.get("variation_seed"))
    rng = random.Random(seed)
    requested_story_count = _parse_requested_story_count(
        message,
        explicit_story_count=action.get("story_count"),
        default=4 if style == "apartment" else (3 if style in {"mansion", "villa"} else 2),
    )
    style_ranges = {
        "mansion": {
            "width": (1100.0, 1520.0),
            "depth": (900.0, 1320.0),
            "story_height": (315.0, 350.0),
            "roof_pitch": (28.0, 40.0),
            "roof_overhang": (26.0, 40.0),
            "roof_rise": (135.0, 200.0),
        },
        "villa": {
            "width": (980.0, 1320.0),
            "depth": (820.0, 1120.0),
            "story_height": (305.0, 340.0),
            "roof_pitch": (24.0, 34.0),
            "roof_overhang": (24.0, 34.0),
            "roof_rise": (120.0, 175.0),
        },
        "modern": {
            "width": (820.0, 980.0),
            "depth": (620.0, 780.0),
            "story_height": (300.0, 330.0),
            "roof_pitch": (18.0, 26.0),
            "roof_overhang": (12.0, 20.0),
            "roof_rise": (90.0, 125.0),
        },
        "cabin": {
            "width": (640.0, 780.0),
            "depth": (560.0, 720.0),
            "story_height": (290.0, 315.0),
            "roof_pitch": (34.0, 42.0),
            "roof_overhang": (28.0, 42.0),
            "roof_rise": (130.0, 170.0),
        },
        "cottage": {
            "width": (660.0, 800.0),
            "depth": (560.0, 720.0),
            "story_height": (290.0, 320.0),
            "roof_pitch": (30.0, 40.0),
            "roof_overhang": (22.0, 34.0),
            "roof_rise": (120.0, 155.0),
        },
        "townhouse": {
            "width": (560.0, 680.0),
            "depth": (760.0, 920.0),
            "story_height": (295.0, 320.0),
            "roof_pitch": (24.0, 32.0),
            "roof_overhang": (12.0, 22.0),
            "roof_rise": (95.0, 135.0),
        },
        "apartment": {
            "width": (900.0, 1240.0),
            "depth": (760.0, 1080.0),
            "story_height": (285.0, 310.0),
            "roof_pitch": (18.0, 26.0),
            "roof_overhang": (12.0, 18.0),
            "roof_rise": (90.0, 125.0),
        },
        "suburban": {
            "width": (760.0, 940.0),
            "depth": (620.0, 820.0),
            "story_height": (295.0, 325.0),
            "roof_pitch": (26.0, 34.0),
            "roof_overhang": (18.0, 28.0),
            "roof_rise": (105.0, 145.0),
        },
    }
    profile = style_ranges.get(style, style_ranges["suburban"])
    size_multiplier = _size_multiplier_from_label(str(action.get("size") or "medium"))
    width_cm = _safe_number(action.get("inner_width_cm"), rng.uniform(*profile["width"]) * size_multiplier)
    depth_cm = _safe_number(action.get("inner_depth_cm"), rng.uniform(*profile["depth"]) * size_multiplier)
    story_height_cm = _safe_number(action.get("story_height_cm"), rng.uniform(*profile["story_height"]))
    roof_pitch_deg = _safe_number(action.get("roof_pitch_deg"), rng.uniform(*profile["roof_pitch"]))
    roof_overhang_cm = _safe_number(action.get("roof_overhang_cm"), rng.uniform(*profile["roof_overhang"]))
    roof_rise_cm = _safe_number(action.get("roof_rise_cm"), rng.uniform(*profile["roof_rise"]))
    stair_width_cm = _safe_number(action.get("stair_width_cm"), max(120.0, min(width_cm - 120.0, rng.uniform(120.0, 150.0))))
    stair_step_rise_cm = _safe_number(action.get("stair_step_rise_cm"), rng.uniform(18.0, 20.5))
    stair_step_run_cm = _safe_number(action.get("stair_step_run_cm"), rng.uniform(27.0, 31.0))
    stair_step_count = int(action.get("stair_step_count") or max(10, min(14, int(round((story_height_cm + 20.0) / max(stair_step_rise_cm, 1.0))))))
    roof_style = str(action.get("roof_style") or ("parapet" if style in {"apartment", "townhouse", "modern"} and requested_story_count >= 4 else "gable"))
    window_columns = int(
        action.get("window_columns_per_wall")
        or (
            4 if style == "apartment" else
            3 if style in {"mansion", "villa"} else
            2
        )
    )
    window_width_cm = _safe_number(
        action.get("window_width_cm"),
        rng.uniform(120.0, 150.0) if style == "apartment" else rng.uniform(110.0, 140.0),
    )
    window_height_cm = _safe_number(
        action.get("window_height_cm"),
        rng.uniform(115.0, 145.0) if style == "apartment" else rng.uniform(105.0, 130.0),
    )
    window_sill_height_cm = _safe_number(
        action.get("window_sill_height_cm"),
        rng.uniform(85.0, 100.0) if style == "apartment" else rng.uniform(90.0, 110.0),
    )
    entry_canopy_depth_cm = _safe_number(
        action.get("entry_canopy_depth_cm"),
        150.0 if style == "mansion" else
        125.0 if style == "villa" else
        110.0 if style == "apartment" else
        (80.0 if style in {"townhouse", "modern"} else 45.0),
    )
    balcony_depth_cm = _safe_number(
        action.get("balcony_depth_cm"),
        135.0 if style == "mansion" else
        120.0 if style == "villa" else
        100.0 if style == "apartment" else
        80.0 if style == "townhouse" else 0.0,
    )
    corner_column_diameter_cm = _safe_number(
        action.get("corner_column_diameter_cm"),
        48.0 if style == "mansion" else
        42.0 if style == "villa" else
        34.0 if style in {"apartment", "townhouse", "modern"} else 0.0,
    )
    site_clearance_cm = _safe_number(
        action.get("site_clearance_cm"),
        140.0 if style == "mansion" else
        120.0 if style == "villa" else
        100.0 if style == "apartment" else
        80.0 if requested_story_count >= 3 else 60.0,
    )
    modifiers = _apply_house_request_modifiers(
        message,
        width_cm=width_cm,
        depth_cm=depth_cm,
        story_height_cm=story_height_cm,
        roof_pitch_deg=roof_pitch_deg,
        roof_rise_cm=roof_rise_cm,
        balcony_depth_cm=balcony_depth_cm,
        entry_canopy_depth_cm=entry_canopy_depth_cm,
        corner_column_diameter_cm=corner_column_diameter_cm,
        window_columns_per_wall=window_columns,
        site_clearance_cm=site_clearance_cm,
        style=style,
    )
    width_cm = modifiers["width_cm"]
    depth_cm = modifiers["depth_cm"]
    story_height_cm = modifiers["story_height_cm"]
    roof_pitch_deg = modifiers["roof_pitch_deg"]
    roof_rise_cm = modifiers["roof_rise_cm"]
    balcony_depth_cm = modifiers["balcony_depth_cm"]
    entry_canopy_depth_cm = modifiers["entry_canopy_depth_cm"]
    corner_column_diameter_cm = modifiers["corner_column_diameter_cm"]
    window_columns = int(modifiers["window_columns_per_wall"])
    site_clearance_cm = modifiers["site_clearance_cm"]
    label_prefix = str(action.get("label_prefix") or f"UCA_{style.title()}House").strip()
    zone_id = str(action.get("zone_id") or f"zone_house_{_clean_identifier_fragment(style)}_{seed % 100000:05d}").strip()
    spec = HouseSpec(
        zone_id=zone_id,
        center_x=float(support_context["center_x"]),
        center_y=float(support_context["center_y"]),
        support_z=float(support_context["support_z"]),
        variation_seed=seed,
        story_count=requested_story_count,
        inner_width_cm=width_cm,
        inner_depth_cm=depth_cm,
        story_height_cm=story_height_cm,
        wall_thickness_cm=_safe_number(action.get("wall_thickness_cm"), 20.0),
        floor_thickness_cm=_safe_number(action.get("floor_thickness_cm"), 20.0),
        door_width_cm=_safe_number(action.get("door_width_cm"), max(150.0, min(width_cm - 140.0, 170.0))),
        door_height_cm=_safe_number(action.get("door_height_cm"), min(story_height_cm - 35.0, 230.0)),
        roof_style=roof_style,
        roof_pitch_deg=roof_pitch_deg,
        roof_thickness_cm=_safe_number(action.get("roof_thickness_cm"), 18.0),
        roof_overhang_cm=roof_overhang_cm,
        roof_rise_cm=roof_rise_cm,
        window_width_cm=window_width_cm,
        window_height_cm=window_height_cm,
        window_sill_height_cm=window_sill_height_cm,
        window_columns_per_wall=window_columns,
        entry_canopy_depth_cm=entry_canopy_depth_cm,
        balcony_depth_cm=balcony_depth_cm,
        corner_column_diameter_cm=corner_column_diameter_cm,
        site_clearance_cm=site_clearance_cm,
        stair_width_cm=stair_width_cm,
        stair_step_rise_cm=stair_step_rise_cm,
        stair_step_run_cm=stair_step_run_cm,
        stair_step_count=stair_step_count,
        stair_opening_margin_cm=_safe_number(action.get("stair_opening_margin_cm"), 8.0),
        landing_depth_cm=_safe_number(action.get("landing_depth_cm"), max(100.0, stair_step_run_cm * 3.0)),
        stair_guard_height_cm=_safe_number(action.get("stair_guard_height_cm"), 95.0),
        stair_guard_thickness_cm=_safe_number(action.get("stair_guard_thickness_cm"), 8.0),
        roof_ridge_thickness_cm=_safe_number(action.get("roof_ridge_thickness_cm"), 8.0),
        gable_infill_step_count=int(action.get("gable_infill_step_count") or 4),
        grid_snap_cm=_safe_number(action.get("grid_snap_cm"), 10.0),
        label_prefix=label_prefix,
        residential_profile=style,
        support_surface_kind=str(support_context["support_surface_kind"] or "support_surface"),
        support_level=int(support_context["support_level"] or 0),
        support_actor_label=str(support_context["support_actor_label"] or ""),
        parent_support_actor=str(support_context["support_actor_label"] or ""),
        support_reference_policy=str(action.get("support_reference_policy") or "selected_first"),
    )
    body_material, roof_material = _infer_body_and_roof_materials(
        message,
        explicit_material=str(action.get("material") or ""),
        explicit_roof_material=str(action.get("roof_material") or ""),
        style=style,
    )
    return spec, {
        "seed": seed,
        "style": style,
        "story_count": requested_story_count,
        "body_material": body_material,
        "roof_material": roof_material,
        "trim_material": "concrete" if style in {"apartment", "modern"} else ("stone" if style in {"mansion", "villa"} else body_material),
        "glass_material": "glass",
        "size_label": str(action.get("size") or "medium"),
    }


def _structure_style_from_request(message: str, structure_type: str, explicit_style: str = "") -> str:
    preferred = str(explicit_style or "").strip().lower()
    if preferred:
        return preferred
    lowered = str(message or "").lower()
    if structure_type in {"pavilion", "gazebo", "pergola"}:
        if any(token in lowered for token in ("garden", "courtyard", "park", "gazebo")):
            return "garden"
        if any(token in lowered for token in ("modern", "sleek", "minimal")):
            return "modern"
        return "open_air"
    if structure_type in {"canopy", "carport", "market_stall", "kiosk"}:
        if any(token in lowered for token in ("modern", "sleek", "minimal")):
            return "modern"
        if any(token in lowered for token in ("market", "vendor", "festival", "street")):
            return "street"
        return "lightweight"
    if structure_type in {"garage", "workshop"}:
        if any(token in lowered for token in ("modern", "clean", "sleek")):
            return "modern"
        return "utility"
    if structure_type in {"warehouse", "hangar"}:
        if any(token in lowered for token in ("industrial", "shipping", "logistics", "factory")):
            return "industrial"
        return "utility"
    if structure_type == "greenhouse":
        return "garden"
    if structure_type == "studio":
        if any(token in lowered for token in ("art", "creative", "gallery")):
            return "creative"
        return "modern"
    if structure_type == "barn":
        return "agricultural"
    if structure_type == "shed":
        if any(token in lowered for token in ("garden", "backyard", "storage")):
            return "garden"
        return "utility"
    return "utility"


def _apply_structure_request_modifiers(
    message: str,
    *,
    width_cm: float,
    depth_cm: float,
    wall_height_cm: float,
    roof_pitch_deg: float,
    roof_rise_cm: float,
) -> dict[str, float]:
    lowered = str(message or "").lower()
    width_scale = 1.0
    depth_scale = 1.0
    height_scale = 1.0
    roof_pitch_scale = 1.0
    roof_rise_scale = 1.0

    if any(token in lowered for token in ("wide", "broad", "spacious", "double-wide")):
        width_scale *= 1.14
    if any(token in lowered for token in ("narrow", "skinny", "compact width")):
        width_scale *= 0.88
    if any(token in lowered for token in ("deep", "long", "extended", "long-form")):
        depth_scale *= 1.14
    if any(token in lowered for token in ("shallow", "compact", "tight")):
        depth_scale *= 0.9
    if any(token in lowered for token in ("tall", "high ceiling", "high-ceiling", "lofty")):
        height_scale *= 1.12
        roof_rise_scale *= 1.08
    if any(token in lowered for token in ("low profile", "flat roof", "shallow roof", "minimal roof")):
        roof_pitch_scale *= 0.8
        roof_rise_scale *= 0.8
    if any(token in lowered for token in ("steep roof", "pitched roof", "dramatic roof")):
        roof_pitch_scale *= 1.12
        roof_rise_scale *= 1.1

    return {
        "width_cm": round(width_cm * width_scale, 3),
        "depth_cm": round(depth_cm * depth_scale, 3),
        "wall_height_cm": round(wall_height_cm * height_scale, 3),
        "roof_pitch_deg": round(roof_pitch_deg * roof_pitch_scale, 3),
        "roof_rise_cm": round(roof_rise_cm * roof_rise_scale, 3),
    }


def _infer_structure_body_and_roof_materials(
    structure_type: str,
    message: str,
    *,
    explicit_material: str = "",
    explicit_roof_material: str = "",
    style: str = "",
) -> tuple[str, str]:
    body = _resolve_material_name(explicit_material or "")
    roof = _resolve_material_name(explicit_roof_material or "")
    if not body or body not in MATERIAL_CATALOG:
        defaults = {
            "garage": "concrete",
            "shed": "wood",
            "workshop": "metal",
            "barn": "wood",
            "warehouse": "metal",
            "greenhouse": "glass",
            "studio": "stucco",
            "hangar": "metal",
            "kiosk": "wood",
            "pavilion": "wood",
            "gazebo": "wood",
            "pergola": "wood",
            "canopy": "metal",
            "carport": "metal",
            "market_stall": "wood",
        }
        style_overrides = {
            "modern": "concrete",
            "garden": "wood",
            "utility": "metal" if structure_type in {"workshop"} else defaults.get(structure_type, "wood"),
            "agricultural": "wood",
            "open_air": "wood",
            "street": "wood",
            "lightweight": "metal",
            "industrial": "metal",
            "creative": "stucco",
        }
        body = style_overrides.get(style, defaults.get(structure_type, "wood"))
        lowered = str(message or "").lower()
        for token in sorted(MATERIAL_CATALOG, key=len, reverse=True):
            if token in lowered:
                body = _resolve_material_name(token)
                break
    if not roof or roof not in MATERIAL_CATALOG:
        roof_defaults = {
            "garage": "metal",
            "shed": "metal",
            "workshop": "metal",
            "barn": "wood",
            "warehouse": "metal",
            "greenhouse": "glass",
            "studio": "metal",
            "hangar": "metal",
            "kiosk": "wood",
            "pavilion": "wood",
            "gazebo": "wood",
            "pergola": body,
            "canopy": "metal",
            "carport": "metal",
            "market_stall": "wood",
        }
        roof = roof_defaults.get(structure_type, "metal")
        if roof not in MATERIAL_CATALOG:
            roof = body if body in MATERIAL_CATALOG else "metal"
    return body, roof


def _generate_structure_spec_from_request(
    action: dict[str, Any],
    *,
    message: str,
    chat_id: str,
    support_context: dict[str, Any],
) -> tuple[StructureSpec, dict[str, Any]]:
    structure_type = canonical_structure_type(
        action.get("structure")
        or action.get("structure_type")
        or (action.get("request") or ""),
        fallback="shed",
    )
    style = _structure_style_from_request(message, structure_type, str(action.get("style") or ""))
    seed = _stable_house_seed(message, chat_id, explicit_seed=action.get("variation_seed"))
    rng = random.Random(seed)
    size_multiplier = _size_multiplier_from_label(str(action.get("size") or "medium"))
    profiles: dict[str, dict[str, Any]] = {
        "garage": {
            "width": (760.0, 960.0),
            "depth": (700.0, 920.0),
            "wall_height": (260.0, 300.0),
            "roof_pitch": (18.0, 28.0),
            "roof_overhang": (14.0, 24.0),
            "roof_rise": (80.0, 120.0),
            "opening_width": (250.0, 330.0),
            "opening_height": (220.0, 250.0),
            "label_prefix": "UCA_Garage",
        },
        "shed": {
            "width": (380.0, 560.0),
            "depth": (420.0, 640.0),
            "wall_height": (220.0, 265.0),
            "roof_pitch": (24.0, 34.0),
            "roof_overhang": (14.0, 24.0),
            "roof_rise": (70.0, 120.0),
            "door_width": (90.0, 120.0),
            "door_height": (190.0, 220.0),
            "label_prefix": "UCA_Shed",
        },
        "workshop": {
            "width": (680.0, 860.0),
            "depth": (620.0, 820.0),
            "wall_height": (250.0, 300.0),
            "roof_pitch": (18.0, 28.0),
            "roof_overhang": (14.0, 22.0),
            "roof_rise": (78.0, 118.0),
            "opening_width": (200.0, 280.0),
            "opening_height": (220.0, 245.0),
            "label_prefix": "UCA_Workshop",
        },
        "barn": {
            "width": (860.0, 1100.0),
            "depth": (760.0, 980.0),
            "wall_height": (280.0, 340.0),
            "roof_pitch": (28.0, 38.0),
            "roof_overhang": (18.0, 30.0),
            "roof_rise": (130.0, 190.0),
            "opening_width": (240.0, 340.0),
            "opening_height": (230.0, 270.0),
            "label_prefix": "UCA_Barn",
        },
        "warehouse": {
            "width": (1040.0, 1480.0),
            "depth": (860.0, 1220.0),
            "wall_height": (300.0, 380.0),
            "roof_pitch": (14.0, 22.0),
            "roof_overhang": (10.0, 18.0),
            "roof_rise": (90.0, 135.0),
            "opening_width": (320.0, 460.0),
            "opening_height": (250.0, 320.0),
            "label_prefix": "UCA_Warehouse",
        },
        "greenhouse": {
            "width": (620.0, 960.0),
            "depth": (680.0, 1080.0),
            "wall_height": (250.0, 320.0),
            "roof_pitch": (24.0, 34.0),
            "roof_overhang": (10.0, 18.0),
            "roof_rise": (110.0, 160.0),
            "door_width": (110.0, 150.0),
            "door_height": (210.0, 235.0),
            "label_prefix": "UCA_Greenhouse",
        },
        "studio": {
            "width": (620.0, 880.0),
            "depth": (620.0, 940.0),
            "wall_height": (270.0, 330.0),
            "roof_pitch": (18.0, 30.0),
            "roof_overhang": (14.0, 24.0),
            "roof_rise": (90.0, 130.0),
            "door_width": (105.0, 135.0),
            "door_height": (210.0, 235.0),
            "label_prefix": "UCA_Studio",
        },
        "hangar": {
            "width": (1280.0, 1900.0),
            "depth": (960.0, 1520.0),
            "wall_height": (340.0, 420.0),
            "roof_pitch": (12.0, 20.0),
            "roof_overhang": (10.0, 18.0),
            "roof_rise": (120.0, 180.0),
            "opening_width": (420.0, 620.0),
            "opening_height": (290.0, 360.0),
            "label_prefix": "UCA_Hangar",
        },
        "kiosk": {
            "width": (320.0, 520.0),
            "depth": (320.0, 520.0),
            "wall_height": (220.0, 280.0),
            "roof_pitch": (18.0, 30.0),
            "roof_overhang": (14.0, 28.0),
            "roof_rise": (60.0, 110.0),
            "door_width": (110.0, 160.0),
            "door_height": (190.0, 220.0),
            "label_prefix": "UCA_Kiosk",
        },
        "pavilion": {
            "width": (760.0, 1040.0),
            "depth": (640.0, 920.0),
            "wall_height": (250.0, 310.0),
            "roof_pitch": (18.0, 28.0),
            "roof_overhang": (28.0, 48.0),
            "roof_rise": (90.0, 145.0),
            "post_thickness": (18.0, 26.0),
            "beam_thickness": (12.0, 18.0),
            "label_prefix": "UCA_Pavilion",
        },
        "gazebo": {
            "width": (560.0, 760.0),
            "depth": (560.0, 760.0),
            "wall_height": (240.0, 300.0),
            "roof_pitch": (24.0, 36.0),
            "roof_overhang": (30.0, 48.0),
            "roof_rise": (110.0, 170.0),
            "post_thickness": (18.0, 26.0),
            "beam_thickness": (12.0, 18.0),
            "label_prefix": "UCA_Gazebo",
        },
        "pergola": {
            "width": (700.0, 980.0),
            "depth": (640.0, 920.0),
            "wall_height": (240.0, 290.0),
            "roof_overhang": (12.0, 22.0),
            "post_thickness": (18.0, 26.0),
            "beam_thickness": (12.0, 18.0),
            "label_prefix": "UCA_Pergola",
        },
        "canopy": {
            "width": (680.0, 980.0),
            "depth": (520.0, 780.0),
            "wall_height": (240.0, 300.0),
            "roof_pitch": (12.0, 22.0),
            "roof_overhang": (20.0, 36.0),
            "roof_rise": (70.0, 110.0),
            "post_thickness": (16.0, 24.0),
            "beam_thickness": (10.0, 16.0),
            "label_prefix": "UCA_Canopy",
        },
        "carport": {
            "width": (760.0, 1020.0),
            "depth": (760.0, 980.0),
            "wall_height": (250.0, 300.0),
            "roof_pitch": (14.0, 24.0),
            "roof_overhang": (18.0, 30.0),
            "roof_rise": (80.0, 120.0),
            "post_thickness": (18.0, 24.0),
            "beam_thickness": (12.0, 18.0),
            "label_prefix": "UCA_Carport",
        },
        "market_stall": {
            "width": (420.0, 660.0),
            "depth": (360.0, 560.0),
            "wall_height": (220.0, 275.0),
            "roof_pitch": (20.0, 32.0),
            "roof_overhang": (16.0, 28.0),
            "roof_rise": (70.0, 115.0),
            "post_thickness": (14.0, 20.0),
            "beam_thickness": (10.0, 15.0),
            "label_prefix": "UCA_MarketStall",
        },
    }
    profile = profiles.get(structure_type, profiles["shed"])
    width_cm = _safe_number(action.get("width_cm"), rng.uniform(*profile["width"]) * size_multiplier)
    depth_cm = _safe_number(action.get("depth_cm"), rng.uniform(*profile["depth"]) * size_multiplier)
    width_explicit = action.get("width_cm") not in (None, "")
    depth_explicit = action.get("depth_cm") not in (None, "")
    if structure_type == "gazebo":
        average_span = (width_cm + depth_cm) / 2.0
        width_cm = average_span
        depth_cm = average_span
    wall_height_cm = _safe_number(action.get("wall_height_cm"), rng.uniform(*profile["wall_height"]))
    roof_pitch_deg = _safe_number(action.get("roof_pitch_deg"), rng.uniform(*profile.get("roof_pitch", (0.0, 0.0)))) if "roof_pitch" in profile else 0.0
    roof_overhang_cm = _safe_number(action.get("roof_overhang_cm"), rng.uniform(*profile.get("roof_overhang", (14.0, 24.0))))
    roof_rise_cm = _safe_number(action.get("roof_rise_cm"), rng.uniform(*profile.get("roof_rise", (90.0, 130.0))))
    modifiers = _apply_structure_request_modifiers(
        message,
        width_cm=width_cm,
        depth_cm=depth_cm,
        wall_height_cm=wall_height_cm,
        roof_pitch_deg=roof_pitch_deg,
        roof_rise_cm=roof_rise_cm,
    )
    if not width_explicit:
        width_cm = modifiers["width_cm"]
    if not depth_explicit:
        depth_cm = modifiers["depth_cm"]
    if action.get("wall_height_cm") in (None, ""):
        wall_height_cm = modifiers["wall_height_cm"]
    if action.get("roof_pitch_deg") in (None, ""):
        roof_pitch_deg = modifiers["roof_pitch_deg"]
    if action.get("roof_rise_cm") in (None, ""):
        roof_rise_cm = modifiers["roof_rise_cm"]
    if structure_type == "gazebo":
        average_span = (width_cm + depth_cm) / 2.0
        width_cm = average_span
        depth_cm = average_span
    label_prefix = str(action.get("label_prefix") or profile["label_prefix"]).strip()
    zone_id = str(action.get("zone_id") or f"zone_{_clean_identifier_fragment(structure_type)}_{seed % 100000:05d}").strip()
    roof_style = str(action.get("roof_style") or ("beam" if structure_type == "pergola" else "gable")).strip().lower()
    post_thickness_cm = _safe_number(action.get("post_thickness_cm"), rng.uniform(*profile.get("post_thickness", (20.0, 20.0))))
    beam_thickness_cm = _safe_number(action.get("beam_thickness_cm"), rng.uniform(*profile.get("beam_thickness", (14.0, 14.0))))
    door_width_cm = _safe_number(action.get("door_width_cm"), rng.uniform(*profile.get("door_width", (100.0, 130.0))))
    door_height_cm = _safe_number(action.get("door_height_cm"), rng.uniform(*profile.get("door_height", (200.0, 220.0))))
    opening_width_cm = _safe_number(action.get("opening_width_cm"), rng.uniform(*profile.get("opening_width", (door_width_cm, max(door_width_cm, 180.0)))))
    opening_height_cm = _safe_number(action.get("opening_height_cm"), rng.uniform(*profile.get("opening_height", (door_height_cm, max(door_height_cm, 220.0)))))
    body_material, roof_material = _infer_structure_body_and_roof_materials(
        structure_type,
        message,
        explicit_material=str(action.get("material") or ""),
        explicit_roof_material=str(action.get("roof_material") or ""),
        style=style,
    )
    spec = StructureSpec(
        zone_id=zone_id,
        structure_type=structure_type,
        center_x=float(support_context["center_x"]),
        center_y=float(support_context["center_y"]),
        support_z=float(support_context["support_z"]),
        width_cm=width_cm,
        depth_cm=depth_cm,
        wall_height_cm=wall_height_cm,
        wall_thickness_cm=_safe_number(action.get("wall_thickness_cm"), 20.0),
        floor_thickness_cm=_safe_number(action.get("floor_thickness_cm"), 20.0),
        door_width_cm=door_width_cm,
        door_height_cm=door_height_cm,
        opening_width_cm=opening_width_cm,
        opening_height_cm=opening_height_cm,
        roof_style=roof_style,
        roof_pitch_deg=roof_pitch_deg,
        roof_thickness_cm=_safe_number(action.get("roof_thickness_cm"), 18.0),
        roof_overhang_cm=roof_overhang_cm,
        roof_rise_cm=roof_rise_cm,
        roof_ridge_thickness_cm=_safe_number(action.get("roof_ridge_thickness_cm"), 8.0),
        gable_infill_step_count=int(action.get("gable_infill_step_count") or 3),
        post_thickness_cm=post_thickness_cm,
        beam_thickness_cm=beam_thickness_cm,
        railing_height_cm=_safe_number(action.get("railing_height_cm"), 90.0),
        grid_snap_cm=_safe_number(action.get("grid_snap_cm"), 10.0),
        label_prefix=label_prefix,
        support_surface_kind=str(support_context["support_surface_kind"] or "support_surface"),
        support_level=int(support_context["support_level"] or 0),
        support_actor_label=str(support_context["support_actor_label"] or ""),
        parent_support_actor=str(support_context["support_actor_label"] or ""),
        support_reference_policy=str(action.get("support_reference_policy") or "selected_first"),
    )
    return spec, {
        "seed": seed,
        "style": style,
        "structure_type": structure_type,
        "body_material": body_material,
        "roof_material": roof_material,
        "size_label": str(action.get("size") or "medium"),
    }


def _execute_build_house(action: dict) -> dict:
    context = _get_active_tool_context()
    message = str(action.get("request") or context.get("message") or "build a house")
    chat_id = str(action.get("chat_id") or context.get("chat_id") or "chat")
    repo_root = WORKSPACE_ROOT
    session_slug = _clean_identifier_fragment(chat_id, fallback="chat")
    session_path = DATA_DIR / "sessions" / f"chat_houses_{session_slug}"
    session_path.mkdir(parents=True, exist_ok=True)

    scene_state = enrich_scene_state(collect_scene_state(repo_root), repo_root)
    support_context = _resolve_house_support_context(action, scene_state)
    requested_spec, variation = _generate_house_spec_from_request(
        action,
        message=message,
        chat_id=chat_id,
        support_context=support_context,
    )
    zone_records = managed_records_for_zone(session_path, requested_spec.zone_id)
    ignore_actor_paths = {
        str(record.get("actor_path") or "").strip()
        for record in zone_records
        if str(record.get("actor_path") or "").strip()
    }
    ignore_actor_labels = {
        str(record.get("actor_label") or "").strip()
        for record in zone_records
        if str(record.get("actor_label") or "").strip()
    }
    if support_context.get("support_actor_path"):
        ignore_actor_paths.add(str(support_context["support_actor_path"]))
    if support_context.get("support_actor_label"):
        ignore_actor_labels.add(str(support_context["support_actor_label"]))

    plan = plan_house_spec(
        requested_spec,
        [dict(actor) for actor in list(scene_state.get("actors") or []) if isinstance(actor, dict)],
        ignore_actor_paths=ignore_actor_paths,
        ignore_actor_labels=ignore_actor_labels,
    )
    spec = plan["spec"]
    structure_plan = build_house_structure_plan(spec)
    actions = build_house_actions(spec)
    desired_slots = {
        str(action_payload.get("managed_slot") or "").strip()
        for action_payload in actions
        if str(action_payload.get("managed_slot") or "").strip()
    }

    cleanup_paths: list[str] = []
    released_slots: list[str] = []
    for record in zone_records:
        managed_slot = str(record.get("managed_slot") or "").strip()
        actor_path = str(record.get("actor_path") or "").strip()
        ownership = dict(record.get("ownership") or {})
        if not managed_slot or managed_slot in desired_slots:
            continue
        released = release_slot(
            session_path,
            zone_id=spec.zone_id,
            managed_slot=managed_slot,
            reason="obsolete_house_slot",
        )
        if released:
            released_slots.append(managed_slot)
        if actor_path and bool(ownership.get("allow_cleanup", True)):
            cleanup_paths.append(actor_path)

    action_backend = choose_action_backend(repo_root)
    cleanup_result: dict[str, Any] = {"success": True, "deleted_count": 0, "deleted_paths": []}
    if action_backend == "uefn_mcp_apply":
        cleanup_paths.extend(_stray_tool_generated_paths_for_zone(repo_root, zone_records))
        deduped_paths = sorted({path for path in cleanup_paths if str(path).strip()})
        if deduped_paths:
            cleanup_result = _delete_actors_by_paths(repo_root, deduped_paths) or cleanup_result

    results: list[dict[str, Any]] = []
    material_results: list[dict[str, Any]] = []
    live_actors_by_slot: dict[str, dict[str, Any]] = {}
    body_material_path = _get_material_path(variation["body_material"])
    roof_material_path = _get_material_path(variation["roof_material"])
    trim_material_path = _get_material_path(variation.get("trim_material") or variation["body_material"])
    floor_material_path = _get_material_path(f"{variation['body_material']}_floor") or body_material_path
    glass_material_path = _get_material_path(variation.get("glass_material") or "glass")

    if action_backend == "uefn_mcp_apply":
        for action_payload in actions:
            result = apply_action_via_mcp(
                repo_root,
                action_payload,
                session_path=session_path,
                auto_save=False,
            )
            results.append(result)
            actor_payload = dict(result.get("actor") or {})
            managed_slot = str(action_payload.get("managed_slot") or "").strip()
            if managed_slot and actor_payload:
                actor_payload["managed_slot"] = managed_slot
                live_actors_by_slot[managed_slot] = actor_payload
            actor_path = str(actor_payload.get("actor_path") or "").strip()
            placement_hint = dict(action_payload.get("placement_hint") or {})
            material_role = str(placement_hint.get("material_role") or "").strip().lower()
            selected_material = _select_material_for_structure_role(
                material_role=material_role,
                body_material_path=body_material_path,
                roof_material_path=roof_material_path,
                trim_material_path=trim_material_path,
                floor_material_path=floor_material_path,
                glass_material_path=glass_material_path,
            )
            if actor_path and selected_material:
                material_results.append(
                    set_actor_material(
                        repo_root,
                        actor_identifier=actor_path,
                        material_path=selected_material,
                    )
                )
    elif action_backend == "uefn_verse_apply":
        dirty_zone = {
            "zone_id": spec.zone_id,
            "room_type": str(scene_state.get("room_type") or "living_room"),
            "dirty_bounds": dict(scene_state.get("dirty_bounds") or {}),
            "support_surface_kind": str(spec.support_surface_kind),
        }
        for index, action_payload in enumerate(actions, start=1):
            results.append(
                apply_action_via_verse_export(
                    repo_root=repo_root,
                    session_path=session_path,
                    cycle_number=index,
                    scene_state=scene_state,
                    dirty_zone=dirty_zone,
                    action_payload=action_payload,
                )
            )
    else:
        results.append(
            {
                "status": "planned",
                "backend": "plan_only",
                "applied_mode": "plan_only",
                "degraded_to_fallback": True,
                "applied": False,
                "reason": "No live MCP bridge or Verse export backend was available, so the house was planned but not applied.",
                "segment_count": len(actions),
            }
        )

    structure_validation = validate_structure_plan(structure_plan, live_actors_by_slot=live_actors_by_slot)
    structure_plan_payload = dict(structure_plan)
    structure_spec = structure_plan_payload.get("spec")
    if hasattr(structure_spec, "__dict__"):
        structure_plan_payload["spec"] = dict(structure_spec.__dict__)
    structure_plan_path = session_path / "structure_plans" / f"{spec.zone_id}.json"
    structure_plan_path.parent.mkdir(parents=True, exist_ok=True)
    structure_plan_path.write_text(
        json.dumps(
            {
                "structure_plan": structure_plan_payload,
                "structure_validation": structure_validation,
                "placement_plan": {
                    "requested_center": [requested_spec.center_x, requested_spec.center_y],
                    "final_center": [spec.center_x, spec.center_y],
                    "relocated": bool(plan.get("relocated", False)),
                    "offset_cm": list(plan.get("offset_cm") or [0.0, 0.0]),
                    "conflict_count": int(plan.get("conflict_count") or 0),
                    "blocking_conflicts": list(plan.get("blocking_conflicts") or []),
                },
                "variation": variation,
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    residential_label = "apartment building" if variation["style"] == "apartment" else "house"
    summary = (
        f"Built a {variation['style']} {spec.story_count}-story {residential_label} with the shared structure planner. "
        f"Used {len(actions)} managed pieces, support={spec.support_surface_kind}, "
        f"relocated={bool(plan.get('relocated', False))}, validation={'passed' if structure_validation.get('passed') else 'needs review'}."
    )
    return {
        "success": True,
        "status": "ok",
        "summary": summary,
        "backend": action_backend,
        "result": {
            "zone_id": spec.zone_id,
            "session_id": session_path.name,
            "style": variation["style"],
            "story_count": spec.story_count,
            "variation_seed": variation["seed"],
            "materials": {
                "body": variation["body_material"],
                "roof": variation["roof_material"],
            },
            "structure_plan_path": str(structure_plan_path.relative_to(WORKSPACE_ROOT)).replace("\\", "/"),
            "placement_plan": {
                "requested_center": [requested_spec.center_x, requested_spec.center_y],
                "final_center": [spec.center_x, spec.center_y],
                "relocated": bool(plan.get("relocated", False)),
                "offset_cm": list(plan.get("offset_cm") or [0.0, 0.0]),
                "conflict_count": int(plan.get("conflict_count") or 0),
                "blocking_conflicts": list(plan.get("blocking_conflicts") or []),
            },
            "cleanup": {
                "released_slots": released_slots,
                "cleanup_result": cleanup_result,
            },
            "structure_validation": structure_validation,
            "results": results,
            "material_results": material_results,
        },
    }


def _execute_build_structure_request(action: dict) -> dict:
    context = _get_active_tool_context()
    message = str(action.get("request") or context.get("message") or "build a structure")
    chat_id = str(action.get("chat_id") or context.get("chat_id") or "chat")
    repo_root = WORKSPACE_ROOT
    session_slug = _clean_identifier_fragment(chat_id, fallback="chat")
    session_path = DATA_DIR / "sessions" / f"chat_structures_{session_slug}"
    session_path.mkdir(parents=True, exist_ok=True)

    scene_state = enrich_scene_state(collect_scene_state(repo_root), repo_root)
    support_context = _resolve_house_support_context(action, scene_state)
    requested_spec, variation = _generate_structure_spec_from_request(
        action,
        message=message,
        chat_id=chat_id,
        support_context=support_context,
    )
    zone_records = managed_records_for_zone(session_path, requested_spec.zone_id)
    ignore_actor_paths = {
        str(record.get("actor_path") or "").strip()
        for record in zone_records
        if str(record.get("actor_path") or "").strip()
    }
    ignore_actor_labels = {
        str(record.get("actor_label") or "").strip()
        for record in zone_records
        if str(record.get("actor_label") or "").strip()
    }
    if support_context.get("support_actor_path"):
        ignore_actor_paths.add(str(support_context["support_actor_path"]))
    if support_context.get("support_actor_label"):
        ignore_actor_labels.add(str(support_context["support_actor_label"]))

    plan = plan_structure_spec(
        requested_spec,
        [dict(actor) for actor in list(scene_state.get("actors") or []) if isinstance(actor, dict)],
        ignore_actor_paths=ignore_actor_paths,
        ignore_actor_labels=ignore_actor_labels,
    )
    spec = plan["spec"]
    structure_plan = build_structure_plan(spec)
    actions = build_structure_actions(spec)
    desired_slots = {
        str(action_payload.get("managed_slot") or "").strip()
        for action_payload in actions
        if str(action_payload.get("managed_slot") or "").strip()
    }

    cleanup_paths: list[str] = []
    released_slots: list[str] = []
    for record in zone_records:
        managed_slot = str(record.get("managed_slot") or "").strip()
        actor_path = str(record.get("actor_path") or "").strip()
        ownership = dict(record.get("ownership") or {})
        if not managed_slot or managed_slot in desired_slots:
            continue
        released = release_slot(
            session_path,
            zone_id=spec.zone_id,
            managed_slot=managed_slot,
            reason="obsolete_structure_slot",
        )
        if released:
            released_slots.append(managed_slot)
        if actor_path and bool(ownership.get("allow_cleanup", True)):
            cleanup_paths.append(actor_path)

    action_backend = choose_action_backend(repo_root)
    cleanup_result: dict[str, Any] = {"success": True, "deleted_count": 0, "deleted_paths": []}
    if action_backend == "uefn_mcp_apply":
        cleanup_paths.extend(_stray_tool_generated_paths_for_zone(repo_root, zone_records))
        deduped_paths = sorted({path for path in cleanup_paths if str(path).strip()})
        if deduped_paths:
            cleanup_result = _delete_actors_by_paths(repo_root, deduped_paths) or cleanup_result

    results: list[dict[str, Any]] = []
    material_results: list[dict[str, Any]] = []
    live_actors_by_slot: dict[str, dict[str, Any]] = {}
    body_material_path = _get_material_path(variation["body_material"])
    roof_material_path = _get_material_path(variation["roof_material"])
    trim_material_path = _get_material_path(variation.get("trim_material") or variation["body_material"])
    floor_material_path = _get_material_path(f"{variation['body_material']}_floor") or body_material_path
    glass_material_path = _get_material_path(variation.get("glass_material") or "glass")

    if action_backend == "uefn_mcp_apply":
        for action_payload in actions:
            result = apply_action_via_mcp(
                repo_root,
                action_payload,
                session_path=session_path,
                auto_save=False,
            )
            results.append(result)
            actor_payload = dict(result.get("actor") or {})
            managed_slot = str(action_payload.get("managed_slot") or "").strip()
            if managed_slot and actor_payload:
                actor_payload["managed_slot"] = managed_slot
                live_actors_by_slot[managed_slot] = actor_payload
            actor_path = str(actor_payload.get("actor_path") or "").strip()
            placement_hint = dict(action_payload.get("placement_hint") or {})
            material_role = str(placement_hint.get("material_role") or "").strip().lower()
            selected_material = _select_material_for_structure_role(
                material_role=material_role,
                body_material_path=body_material_path,
                roof_material_path=roof_material_path,
                trim_material_path=trim_material_path,
                floor_material_path=floor_material_path,
                glass_material_path=glass_material_path,
            )
            if actor_path and selected_material:
                material_results.append(
                    set_actor_material(
                        repo_root,
                        actor_identifier=actor_path,
                        material_path=selected_material,
                    )
                )
    elif action_backend == "uefn_verse_apply":
        dirty_zone = {
            "zone_id": spec.zone_id,
            "room_type": str(scene_state.get("room_type") or "structure"),
            "dirty_bounds": dict(scene_state.get("dirty_bounds") or {}),
            "support_surface_kind": str(spec.support_surface_kind),
        }
        for index, action_payload in enumerate(actions, start=1):
            results.append(
                apply_action_via_verse_export(
                    repo_root=repo_root,
                    session_path=session_path,
                    cycle_number=index,
                    scene_state=scene_state,
                    dirty_zone=dirty_zone,
                    action_payload=action_payload,
                )
            )
    else:
        results.append(
            {
                "status": "planned",
                "backend": "plan_only",
                "applied_mode": "plan_only",
                "degraded_to_fallback": True,
                "applied": False,
                "reason": "No live MCP bridge or Verse export backend was available, so the structure was planned but not applied.",
                "segment_count": len(actions),
            }
        )

    structure_validation = validate_structure_plan(structure_plan, live_actors_by_slot=live_actors_by_slot)
    structure_plan_payload = dict(structure_plan)
    structure_spec = structure_plan_payload.get("spec")
    if hasattr(structure_spec, "__dict__"):
        structure_plan_payload["spec"] = dict(structure_spec.__dict__)
    structure_plan_path = session_path / "structure_plans" / f"{spec.zone_id}.json"
    structure_plan_path.parent.mkdir(parents=True, exist_ok=True)
    structure_plan_path.write_text(
        json.dumps(
            {
                "structure_plan": structure_plan_payload,
                "structure_validation": structure_validation,
                "placement_plan": {
                    "requested_center": [requested_spec.center_x, requested_spec.center_y],
                    "final_center": [spec.center_x, spec.center_y],
                    "relocated": bool(plan.get("relocated", False)),
                    "offset_cm": list(plan.get("offset_cm") or [0.0, 0.0]),
                    "conflict_count": int(plan.get("conflict_count") or 0),
                    "blocking_conflicts": list(plan.get("blocking_conflicts") or []),
                },
                "variation": variation,
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    summary = (
        f"Built a {variation['style']} {variation['structure_type']} through the shared structure planner. "
        f"Used {len(actions)} managed pieces, support={spec.support_surface_kind}, "
        f"relocated={bool(plan.get('relocated', False))}, validation={'passed' if structure_validation.get('passed') else 'needs review'}."
    )
    return {
        "success": True,
        "status": "ok",
        "summary": summary,
        "backend": action_backend,
        "result": {
            "zone_id": spec.zone_id,
            "session_id": session_path.name,
            "structure_type": variation["structure_type"],
            "style": variation["style"],
            "variation_seed": variation["seed"],
            "materials": {
                "body": variation["body_material"],
                "roof": variation["roof_material"],
            },
            "structure_plan_path": str(structure_plan_path.relative_to(WORKSPACE_ROOT)).replace("\\", "/"),
            "placement_plan": {
                "requested_center": [requested_spec.center_x, requested_spec.center_y],
                "final_center": [spec.center_x, spec.center_y],
                "relocated": bool(plan.get("relocated", False)),
                "offset_cm": list(plan.get("offset_cm") or [0.0, 0.0]),
                "conflict_count": int(plan.get("conflict_count") or 0),
                "blocking_conflicts": list(plan.get("blocking_conflicts") or []),
            },
            "cleanup": {
                "released_slots": released_slots,
                "cleanup_result": cleanup_result,
            },
            "structure_validation": structure_validation,
            "results": results,
            "material_results": material_results,
        },
    }


def _clamp_terrain_dimension(value: Any, default: float = 2000.0) -> float:
    try:
        numeric = float(value)
    except Exception:
        numeric = default
    return max(600.0, min(24000.0, numeric))


def _clamp_terrain_elevation(value: Any, default: float = 200.0) -> float:
    try:
        numeric = float(value)
    except Exception:
        numeric = default
    return max(40.0, min(2400.0, numeric))


def _resolve_terrain_tile_grid(size_x: float, size_y: float, subdivisions: Any) -> tuple[int, int]:
    try:
        tiles = int(subdivisions or 1)
    except Exception:
        tiles = 1
    tiles = max(1, min(36, tiles))
    if tiles == 1:
        return 1, 1

    short_side = max(1.0, min(size_x, size_y))
    long_side = max(size_x, size_y)
    aspect = long_side / short_side
    if aspect >= 2.25:
        long_tiles = max(2, min(8, tiles))
        return (long_tiles, 1) if size_x >= size_y else (1, long_tiles)

    grid_x = int(round(math.sqrt(tiles)))
    grid_y = max(1, int(math.ceil(tiles / max(grid_x, 1))))
    while grid_x * grid_y < tiles:
        grid_y += 1
    return max(1, grid_x), max(1, grid_y)


def _terrain_edge_material_name(top_material_name: str) -> str:
    material = _resolve_material_name(top_material_name or "grass")
    edge_map = {
        "grass": "dirt",
        "ground": "dirt",
        "moss": "dirt",
        "terrain": "cliff",
        "sand": "rock",
        "snow": "cliff",
        "road": "dirt",
        "sidewalk": "concrete",
        "farmfield": "dirt",
        "desert_grass": "sand",
        "ocean_floor": "rock",
        "mud": "dirt",
    }
    return edge_map.get(material, "cliff")


_TERRAIN_ENVIRONMENT_ASSET_CACHE: Dict[str, Any] = {"assets": [], "expires_at": 0.0}


def _terrain_biome_profile(material_name: str, terrain_type: str) -> Dict[str, Any]:
    material = _resolve_material_name(material_name or "grass")
    terrain_type = (terrain_type or "flat").lower()

    if material in {"sand", "desert_grass"}:
        biome = "desert"
    elif material in {"snow"}:
        biome = "arctic"
    elif material in {"rock", "cliff", "terrain"}:
        biome = "mountain"
    elif material in {"mud", "ocean_floor"}:
        biome = "wetland"
    elif material in {"farmfield"}:
        biome = "rural"
    elif material in {"road", "sidewalk"}:
        biome = "roadside"
    else:
        biome = "lush"

    layer_map = {
        "lush": ["moss", "ground", "dirt"],
        "desert": ["desert_grass", "rock", "dirt"],
        "arctic": ["rock", "cliff", "ground"],
        "mountain": ["cliff", "dirt", "moss"],
        "wetland": ["moss", "mud", "ground"],
        "rural": ["ground", "dirt", "moss"],
        "roadside": ["dirt", "ground", "moss"],
    }
    decorations = {
        "lush": [
            {"category": "trees", "keywords": ["tree", "oak", "pine", "birch", "fir"], "count": 10, "radius_scale": 0.46, "min_separation": 850.0, "scale": (0.9, 1.18)},
            {"category": "shrubs", "keywords": ["bush", "shrub", "fern", "plant", "ivy"], "count": 16, "radius_scale": 0.48, "min_separation": 340.0, "scale": (0.75, 1.1)},
            {"category": "rocks", "keywords": ["rock", "boulder", "stone", "cliff"], "count": 9, "radius_scale": 0.44, "min_separation": 420.0, "scale": (0.8, 1.2)},
        ],
        "desert": [
            {"category": "rocks", "keywords": ["rock", "boulder", "stone", "cliff"], "count": 15, "radius_scale": 0.5, "min_separation": 360.0, "scale": (0.8, 1.25)},
            {"category": "shrubs", "keywords": ["cactus", "shrub", "brush", "plant"], "count": 8, "radius_scale": 0.45, "min_separation": 520.0, "scale": (0.8, 1.15)},
        ],
        "arctic": [
            {"category": "trees", "keywords": ["pine", "fir", "tree"], "count": 7, "radius_scale": 0.42, "min_separation": 950.0, "scale": (0.9, 1.2)},
            {"category": "rocks", "keywords": ["rock", "boulder", "stone", "cliff"], "count": 12, "radius_scale": 0.47, "min_separation": 420.0, "scale": (0.85, 1.22)},
        ],
        "mountain": [
            {"category": "rocks", "keywords": ["rock", "boulder", "stone", "cliff"], "count": 18, "radius_scale": 0.5, "min_separation": 380.0, "scale": (0.82, 1.28)},
            {"category": "trees", "keywords": ["pine", "fir", "tree"], "count": 6, "radius_scale": 0.4, "min_separation": 980.0, "scale": (0.85, 1.15)},
        ],
        "wetland": [
            {"category": "trees", "keywords": ["tree", "willow", "mangrove", "swamp"], "count": 8, "radius_scale": 0.42, "min_separation": 880.0, "scale": (0.9, 1.18)},
            {"category": "shrubs", "keywords": ["reed", "grass", "fern", "bush", "plant"], "count": 18, "radius_scale": 0.48, "min_separation": 300.0, "scale": (0.72, 1.08)},
            {"category": "rocks", "keywords": ["rock", "stone", "boulder"], "count": 7, "radius_scale": 0.4, "min_separation": 420.0, "scale": (0.8, 1.14)},
        ],
        "rural": [
            {"category": "trees", "keywords": ["tree", "oak", "pine", "birch"], "count": 7, "radius_scale": 0.44, "min_separation": 980.0, "scale": (0.9, 1.16)},
            {"category": "rocks", "keywords": ["rock", "stone", "boulder"], "count": 8, "radius_scale": 0.42, "min_separation": 420.0, "scale": (0.82, 1.18)},
            {"category": "shrubs", "keywords": ["bush", "shrub", "plant", "grass"], "count": 12, "radius_scale": 0.46, "min_separation": 320.0, "scale": (0.78, 1.08)},
        ],
        "roadside": [
            {"category": "rocks", "keywords": ["rock", "stone", "boulder"], "count": 8, "radius_scale": 0.42, "min_separation": 460.0, "scale": (0.84, 1.16)},
            {"category": "shrubs", "keywords": ["bush", "shrub", "grass", "fern"], "count": 10, "radius_scale": 0.45, "min_separation": 300.0, "scale": (0.78, 1.05)},
        ],
    }

    layer_names = [name for name in layer_map.get(biome, ["dirt", "ground"]) if name != material]
    if terrain_type in {"ridge", "slope"} and biome in {"lush", "rural"} and "rock" not in layer_names:
        layer_names.append("rock")
    if terrain_type in {"crater", "valley"} and "dirt" not in layer_names:
        layer_names.insert(0, "dirt")

    return {
        "biome": biome,
        "layer_material_names": [material] + [name for name in layer_names if name != material],
        "decoration_presets": decorations.get(biome, []),
    }


def _terrain_piece(
    *,
    mesh_path: str,
    dx: float,
    dy: float,
    dz: float,
    sx: float,
    sy: float,
    sz: float,
    label: str,
    material_path: str,
    pitch: float = 0.0,
    yaw: float = 0.0,
    roll: float = 0.0,
) -> Dict[str, Any]:
    return {
        "mesh_path": mesh_path,
        "dx": round(float(dx), 3),
        "dy": round(float(dy), 3),
        "dz": round(float(dz), 3),
        "sx": round(float(sx), 4),
        "sy": round(float(sy), 4),
        "sz": round(float(sz), 4),
        "label": str(label),
        "material_path": str(material_path),
        "pitch": round(float(pitch), 3),
        "yaw": round(float(yaw), 3),
        "roll": round(float(roll), 3),
    }


def _terrain_piece_size_uu(piece: Dict[str, Any]) -> tuple[float, float]:
    return max(0.0, float(piece.get("sx", 0.0)) * 100.0), max(0.0, float(piece.get("sy", 0.0)) * 100.0)


def _terrain_surface_candidates(pieces: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for piece in pieces:
        if piece.get("mesh_path") != "/Engine/BasicShapes/Plane":
            continue
        label = str(piece.get("label") or "")
        if "_Layer_" in label:
            continue
        width_uu, depth_uu = _terrain_piece_size_uu(piece)
        if width_uu < 220.0 or depth_uu < 220.0:
            continue
        if abs(_float_from_value(piece.get("pitch"), 0.0)) > 8.0 or abs(_float_from_value(piece.get("roll"), 0.0)) > 8.0:
            continue
        if any(token in label for token in ("Skirt", "Wall", "Foundation")):
            continue
        candidates.append(piece)
    return candidates


def _make_terrain_overlay_from_surface(
    surface_piece: Dict[str, Any],
    *,
    label_suffix: str,
    material_path: str,
    width_scale: float,
    depth_scale: float,
    dz: float,
    dx: float = 0.0,
    dy: float = 0.0,
) -> Dict[str, Any]:
    width_uu, depth_uu = _terrain_piece_size_uu(surface_piece)
    return _terrain_piece(
        mesh_path="/Engine/BasicShapes/Plane",
        dx=float(surface_piece.get("dx", 0.0)) + dx,
        dy=float(surface_piece.get("dy", 0.0)) + dy,
        dz=float(surface_piece.get("dz", 0.0)) + dz,
        sx=max(120.0, width_uu * width_scale) / 100.0,
        sy=max(120.0, depth_uu * depth_scale) / 100.0,
        sz=1.0,
        label=f"{surface_piece.get('label')}_{label_suffix}",
        material_path=material_path,
        pitch=float(surface_piece.get("pitch", 0.0)),
        yaw=float(surface_piece.get("yaw", 0.0)),
        roll=float(surface_piece.get("roll", 0.0)),
    )


def _append_terrain_material_layers(
    pieces: List[Dict[str, Any]],
    *,
    layer_material_names: List[str],
    long_axis: str,
) -> List[str]:
    layer_material_names = [name for name in layer_material_names if _get_material_path(name)]
    if len(layer_material_names) < 2:
        return layer_material_names[:1]

    base_surfaces = sorted(
        _terrain_surface_candidates(pieces),
        key=lambda piece: _terrain_piece_size_uu(piece)[0] * _terrain_piece_size_uu(piece)[1],
        reverse=True,
    )
    if not base_surfaces:
        return layer_material_names[:1]

    applied_layers: List[str] = [layer_material_names[0]]
    accent_names = layer_material_names[1:]
    overlay_count = 0

    for idx, surface in enumerate(base_surfaces[:5]):
        width_uu, depth_uu = _terrain_piece_size_uu(surface)
        primary_name = accent_names[idx % len(accent_names)]
        primary_path = _get_material_path(primary_name)
        if not primary_path or primary_path == surface.get("material_path"):
            continue

        lateral_offset = (min(width_uu, depth_uu) * 0.08) * (1 if idx % 2 == 0 else -1)
        dx = lateral_offset if long_axis == "x" else 0.0
        dy = lateral_offset if long_axis == "y" else 0.0
        pieces.append(_make_terrain_overlay_from_surface(
            surface,
            label_suffix=f"Layer_{primary_name}_{idx}",
            material_path=primary_path,
            width_scale=0.72 if long_axis == "x" else 0.58,
            depth_scale=0.58 if long_axis == "x" else 0.72,
            dz=1.2 + idx * 0.15,
            dx=dx,
            dy=dy,
        ))
        overlay_count += 1
        if primary_name not in applied_layers:
            applied_layers.append(primary_name)

        if idx == 0 or max(width_uu, depth_uu) >= 2600.0:
            ribbon_name = accent_names[(idx + 1) % len(accent_names)]
            ribbon_path = _get_material_path(ribbon_name)
            if ribbon_path:
                ribbon_dx = 0.0
                ribbon_dy = 0.0
                if long_axis == "x":
                    ribbon_dy = depth_uu * (0.16 if idx % 2 == 0 else -0.16)
                else:
                    ribbon_dx = width_uu * (0.16 if idx % 2 == 0 else -0.16)
                pieces.append(_make_terrain_overlay_from_surface(
                    surface,
                    label_suffix=f"LayerRibbon_{ribbon_name}_{idx}",
                    material_path=ribbon_path,
                    width_scale=0.9 if long_axis == "x" else 0.24,
                    depth_scale=0.24 if long_axis == "x" else 0.9,
                    dz=2.1 + idx * 0.1,
                    dx=ribbon_dx,
                    dy=ribbon_dy,
                ))
                overlay_count += 1
                if ribbon_name not in applied_layers:
                    applied_layers.append(ribbon_name)

        if overlay_count >= 8:
            break

    return applied_layers


def _append_terrain_surface_tiles(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    z: float,
    material_path: str,
    grid_x: int = 1,
    grid_y: int = 1,
) -> None:
    grid_x = max(1, int(grid_x))
    grid_y = max(1, int(grid_y))
    tile_w = width_uu / grid_x
    tile_d = depth_uu / grid_y
    overlap_x = 40.0 if grid_x > 1 else 0.0
    overlap_y = 40.0 if grid_y > 1 else 0.0
    for gx in range(grid_x):
        for gy in range(grid_y):
            dx = -width_uu / 2.0 + tile_w * (gx + 0.5)
            dy = -depth_uu / 2.0 + tile_d * (gy + 0.5)
            piece_w = tile_w + overlap_x
            piece_d = tile_d + overlap_y
            suffix = f"_{gx}_{gy}" if (grid_x > 1 or grid_y > 1) else ""
            pieces.append(_terrain_piece(
                mesh_path="/Engine/BasicShapes/Plane",
                dx=dx,
                dy=dy,
                dz=z,
                sx=piece_w / 100.0,
                sy=piece_d / 100.0,
                sz=1.0,
                label=f"{label_prefix}{suffix}",
                material_path=material_path,
            ))


def _append_terrain_foundation_fill(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    top_z: float,
    thickness_uu: float,
    material_path: str,
    inset_uu: float = 40.0,
    dx: float = 0.0,
    dy: float = 0.0,
) -> None:
    thickness_uu = max(80.0, float(thickness_uu))
    inset_uu = max(0.0, min(float(inset_uu), min(width_uu, depth_uu) * 0.18))
    fill_w = max(120.0, width_uu - inset_uu * 2.0)
    fill_d = max(120.0, depth_uu - inset_uu * 2.0)
    top_gap = max(6.0, min(18.0, thickness_uu * 0.06))
    center_z = top_z - top_gap - thickness_uu / 2.0
    pieces.append(_terrain_piece(
        mesh_path="/Engine/BasicShapes/Cube",
        dx=dx,
        dy=dy,
        dz=center_z,
        sx=fill_w / 100.0,
        sy=fill_d / 100.0,
        sz=thickness_uu / 100.0,
        label=f"{label_prefix}_Foundation",
        material_path=material_path,
    ))


def _append_terrain_ramped_surface(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    base_z: float,
    rise_uu: float,
    material_path: str,
    rise_axis: str = "x",
    rise_direction: float = 1.0,
    dx: float = 0.0,
    dy: float = 0.0,
) -> None:
    rise_axis = "x" if str(rise_axis).lower() == "x" else "y"
    span = max(1.0, width_uu if rise_axis == "x" else depth_uu)
    angle_deg = math.degrees(math.atan2(float(rise_uu), span))
    pitch = angle_deg * float(rise_direction) if rise_axis == "x" else 0.0
    roll = -angle_deg * float(rise_direction) if rise_axis == "y" else 0.0
    pieces.append(_terrain_piece(
        mesh_path="/Engine/BasicShapes/Plane",
        dx=dx,
        dy=dy,
        dz=base_z + float(rise_uu) / 2.0,
        sx=(width_uu + 80.0) / 100.0,
        sy=(depth_uu + 80.0) / 100.0,
        sz=1.0,
        label=label_prefix,
        material_path=material_path,
        pitch=pitch,
        roll=roll,
    ))


def _append_terrain_progressive_fill(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    base_z: float,
    rise_uu: float,
    material_path: str,
    rise_axis: str = "x",
    rise_direction: float = 1.0,
    steps: int = 3,
) -> None:
    rise_axis = "x" if str(rise_axis).lower() == "x" else "y"
    rise_direction = 1.0 if float(rise_direction) >= 0 else -1.0
    rise_uu = max(40.0, float(rise_uu))
    steps = max(2, min(5, int(steps)))
    bottom_z = base_z - max(100.0, rise_uu * 0.28)
    inset = max(30.0, min(width_uu, depth_uu) * 0.04)
    for idx in range(steps):
        coverage = (idx + 1) / steps
        top_z = base_z + rise_uu * coverage - 8.0
        thickness = max(80.0, top_z - bottom_z)
        if rise_axis == "x":
            seg_w = max(160.0, width_uu * coverage)
            center_x = rise_direction * (-width_uu / 2.0 + seg_w / 2.0)
            pieces.append(_terrain_piece(
                mesh_path="/Engine/BasicShapes/Cube",
                dx=center_x,
                dy=0.0,
                dz=bottom_z + thickness / 2.0,
                sx=max(120.0, seg_w - inset * 2.0) / 100.0,
                sy=max(120.0, depth_uu - inset * 2.0) / 100.0,
                sz=thickness / 100.0,
                label=f"{label_prefix}_Fill_{idx + 1}",
                material_path=material_path,
            ))
        else:
            seg_d = max(160.0, depth_uu * coverage)
            center_y = rise_direction * (-depth_uu / 2.0 + seg_d / 2.0)
            pieces.append(_terrain_piece(
                mesh_path="/Engine/BasicShapes/Cube",
                dx=0.0,
                dy=center_y,
                dz=bottom_z + thickness / 2.0,
                sx=max(120.0, width_uu - inset * 2.0) / 100.0,
                sy=max(120.0, seg_d - inset * 2.0) / 100.0,
                sz=thickness / 100.0,
                label=f"{label_prefix}_Fill_{idx + 1}",
                material_path=edge_material_path,
            ))


def _append_terrain_perimeter_skirts(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    top_z: float,
    drop_uu: float,
    material_path: str,
) -> None:
    drop_uu = max(80.0, float(drop_uu))
    lip = max(80.0, min(width_uu, depth_uu) * 0.06)
    center_z = top_z - drop_uu / 2.0
    pieces.extend([
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=0.0,
            dy=-(depth_uu / 2.0) - (lip / 2.0),
            dz=center_z,
            sx=(width_uu + lip * 2.0) / 100.0,
            sy=lip / 100.0,
            sz=drop_uu / 100.0,
            label=f"{label_prefix}_Skirt_N",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=0.0,
            dy=(depth_uu / 2.0) + (lip / 2.0),
            dz=center_z,
            sx=(width_uu + lip * 2.0) / 100.0,
            sy=lip / 100.0,
            sz=drop_uu / 100.0,
            label=f"{label_prefix}_Skirt_S",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=-(width_uu / 2.0) - (lip / 2.0),
            dy=0.0,
            dz=center_z,
            sx=lip / 100.0,
            sy=depth_uu / 100.0,
            sz=drop_uu / 100.0,
            label=f"{label_prefix}_Skirt_W",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=(width_uu / 2.0) + (lip / 2.0),
            dy=0.0,
            dz=center_z,
            sx=lip / 100.0,
            sy=depth_uu / 100.0,
            sz=drop_uu / 100.0,
            label=f"{label_prefix}_Skirt_E",
            material_path=material_path,
        ),
    ])


def _append_terrain_inner_walls(
    pieces: List[Dict[str, Any]],
    *,
    label_prefix: str,
    width_uu: float,
    depth_uu: float,
    low_z: float,
    high_z: float,
    material_path: str,
) -> None:
    wall_height = max(60.0, high_z - low_z)
    wall_center_z = low_z + wall_height / 2.0
    lip = max(60.0, min(width_uu, depth_uu) * 0.06)
    pieces.extend([
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=0.0,
            dy=-(depth_uu / 2.0),
            dz=wall_center_z,
            sx=width_uu / 100.0,
            sy=lip / 100.0,
            sz=wall_height / 100.0,
            label=f"{label_prefix}_Wall_N",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=0.0,
            dy=(depth_uu / 2.0),
            dz=wall_center_z,
            sx=width_uu / 100.0,
            sy=lip / 100.0,
            sz=wall_height / 100.0,
            label=f"{label_prefix}_Wall_S",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=-(width_uu / 2.0),
            dy=0.0,
            dz=wall_center_z,
            sx=lip / 100.0,
            sy=depth_uu / 100.0,
            sz=wall_height / 100.0,
            label=f"{label_prefix}_Wall_W",
            material_path=material_path,
        ),
        _terrain_piece(
            mesh_path="/Engine/BasicShapes/Cube",
            dx=(width_uu / 2.0),
            dy=0.0,
            dz=wall_center_z,
            sx=lip / 100.0,
            sy=depth_uu / 100.0,
            sz=wall_height / 100.0,
            label=f"{label_prefix}_Wall_E",
            material_path=material_path,
        ),
    ])


def _build_terrain_piece_specs(
    *,
    terrain_type: str,
    size_x: float,
    size_y: float,
    height: float,
    elevation: float,
    label: str,
    material_name: str,
    material_path: str,
    subdivisions: Any,
) -> Dict[str, Any]:
    size_x = _clamp_terrain_dimension(size_x)
    size_y = _clamp_terrain_dimension(size_y)
    height = float(height or 0.0)
    elevation = _clamp_terrain_elevation(elevation)
    grid_x, grid_y = _resolve_terrain_tile_grid(size_x, size_y, subdivisions)
    edge_material_name = _terrain_edge_material_name(material_name)
    edge_material_path = _get_material_path(edge_material_name) or material_path
    biome_profile = _terrain_biome_profile(material_name, terrain_type)
    terrain_type = (terrain_type or "flat").lower()
    pieces: List[Dict[str, Any]] = []
    outer_drop = max(140.0, min(size_x, size_y) * 0.14, elevation * 0.65)
    long_axis = "x" if size_x >= size_y else "y"

    if terrain_type == "flat":
        _append_terrain_surface_tiles(
            pieces,
            label_prefix=f"{label}_Top",
            width_uu=size_x,
            depth_uu=size_y,
            z=height,
            material_path=material_path,
            grid_x=grid_x,
            grid_y=grid_y,
        )
        _append_terrain_foundation_fill(
            pieces,
            label_prefix=f"{label}_Top",
            width_uu=size_x,
            depth_uu=size_y,
            top_z=height,
            thickness_uu=max(120.0, outer_drop),
            material_path=edge_material_path,
        )

    elif terrain_type == "hill":
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Base", width_uu=size_x, depth_uu=size_y, z=height, material_path=material_path, grid_x=grid_x, grid_y=grid_y)
        _append_terrain_foundation_fill(pieces, label_prefix=f"{label}_Base", width_uu=size_x, depth_uu=size_y, top_z=height, thickness_uu=max(140.0, outer_drop), material_path=edge_material_path)
        layer_width = size_x
        layer_depth = size_y
        last_z = height
        for index, scale in enumerate((0.82, 0.62, 0.45, 0.28), start=1):
            layer_width *= scale
            layer_depth *= scale
            layer_z = height + elevation * (index / 4.0)
            _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Hill_{index}", width_uu=layer_width, depth_uu=layer_depth, z=layer_z, material_path=material_path)
            _append_terrain_foundation_fill(
                pieces,
                label_prefix=f"{label}_Hill_{index}",
                width_uu=layer_width,
                depth_uu=layer_depth,
                top_z=layer_z,
                thickness_uu=max(80.0, layer_z - last_z),
                material_path=edge_material_path,
            )
            last_z = layer_z

    elif terrain_type == "valley":
        outer_z = height + elevation * 0.35
        inner_z = height - elevation * 0.25
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Outer", width_uu=size_x, depth_uu=size_y, z=outer_z, material_path=material_path, grid_x=grid_x, grid_y=grid_y)
        _append_terrain_foundation_fill(pieces, label_prefix=f"{label}_Outer", width_uu=size_x, depth_uu=size_y, top_z=outer_z, thickness_uu=max(140.0, outer_drop), material_path=edge_material_path)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_ValleyFloor", width_uu=size_x * 0.42, depth_uu=size_y * 0.42, z=inner_z, material_path=material_path)
        _append_terrain_inner_walls(pieces, label_prefix=f"{label}_Valley", width_uu=size_x * 0.42, depth_uu=size_y * 0.42, low_z=inner_z, high_z=outer_z, material_path=edge_material_path)

    elif terrain_type == "plateau":
        plateau_z = height + elevation
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Ground", width_uu=size_x, depth_uu=size_y, z=height, material_path=material_path, grid_x=grid_x, grid_y=grid_y)
        _append_terrain_foundation_fill(pieces, label_prefix=f"{label}_Ground", width_uu=size_x, depth_uu=size_y, top_z=height, thickness_uu=max(140.0, outer_drop), material_path=edge_material_path)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_PlateauTop", width_uu=size_x * 0.62, depth_uu=size_y * 0.62, z=plateau_z, material_path=material_path)
        _append_terrain_foundation_fill(
            pieces,
            label_prefix=f"{label}_PlateauTop",
            width_uu=size_x * 0.62,
            depth_uu=size_y * 0.62,
            top_z=plateau_z,
            thickness_uu=max(100.0, plateau_z - height),
            material_path=edge_material_path,
        )
        _append_terrain_inner_walls(pieces, label_prefix=f"{label}_Plateau", width_uu=size_x * 0.62, depth_uu=size_y * 0.62, low_z=height, high_z=plateau_z, material_path=edge_material_path)

    elif terrain_type == "slope":
        if long_axis == "x":
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_SlopeSurface",
                width_uu=size_x,
                depth_uu=size_y,
                base_z=height,
                rise_uu=elevation,
                material_path=material_path,
                rise_axis="x",
                rise_direction=1.0,
            )
            _append_terrain_progressive_fill(
                pieces,
                label_prefix=f"{label}_Slope",
                width_uu=size_x,
                depth_uu=size_y,
                base_z=height,
                rise_uu=elevation,
                material_path=edge_material_path,
                rise_axis="x",
                rise_direction=1.0,
                steps=max(2, min(4, grid_x)),
            )
        else:
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_SlopeSurface",
                width_uu=size_x,
                depth_uu=size_y,
                base_z=height,
                rise_uu=elevation,
                material_path=material_path,
                rise_axis="y",
                rise_direction=1.0,
            )
            _append_terrain_progressive_fill(
                pieces,
                label_prefix=f"{label}_Slope",
                width_uu=size_x,
                depth_uu=size_y,
                base_z=height,
                rise_uu=elevation,
                material_path=edge_material_path,
                rise_axis="y",
                rise_direction=1.0,
                steps=max(2, min(4, grid_y)),
            )
        _append_terrain_perimeter_skirts(pieces, label_prefix=label, width_uu=size_x, depth_uu=size_y, top_z=height + elevation * 0.5, drop_uu=outer_drop + elevation * 0.3, material_path=edge_material_path)

    elif terrain_type == "crater":
        crater_floor_z = height - elevation * 0.35
        rim_z = height + elevation * 0.18
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Outer", width_uu=size_x, depth_uu=size_y, z=height, material_path=material_path, grid_x=grid_x, grid_y=grid_y)
        _append_terrain_foundation_fill(pieces, label_prefix=f"{label}_Outer", width_uu=size_x, depth_uu=size_y, top_z=height, thickness_uu=max(140.0, outer_drop), material_path=edge_material_path)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_CraterFloor", width_uu=size_x * 0.36, depth_uu=size_y * 0.36, z=crater_floor_z, material_path=material_path)
        _append_terrain_inner_walls(pieces, label_prefix=f"{label}_Crater", width_uu=size_x * 0.36, depth_uu=size_y * 0.36, low_z=crater_floor_z, high_z=height, material_path=edge_material_path)
        ring_w = size_x * 0.18
        ring_d = size_y * 0.18
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Rim_N", width_uu=size_x * 0.7, depth_uu=ring_d, z=rim_z, material_path=material_path)
        for piece in pieces[-1:]:
            piece["dy"] = round(piece["dy"] - size_y * 0.24, 3)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Rim_S", width_uu=size_x * 0.7, depth_uu=ring_d, z=rim_z, material_path=material_path)
        for piece in pieces[-1:]:
            piece["dy"] = round(piece["dy"] + size_y * 0.24, 3)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Rim_W", width_uu=ring_w, depth_uu=size_y * 0.7, z=rim_z, material_path=material_path)
        for piece in pieces[-1:]:
            piece["dx"] = round(piece["dx"] - size_x * 0.24, 3)
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Rim_E", width_uu=ring_w, depth_uu=size_y * 0.7, z=rim_z, material_path=material_path)
        for piece in pieces[-1:]:
            piece["dx"] = round(piece["dx"] + size_x * 0.24, 3)
    elif terrain_type == "ridge":
        ridge_top_z = height + elevation
        shoulder_z = height + elevation * 0.45
        _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_Ground", width_uu=size_x, depth_uu=size_y, z=height, material_path=material_path, grid_x=grid_x, grid_y=grid_y)
        _append_terrain_foundation_fill(pieces, label_prefix=f"{label}_Ground", width_uu=size_x, depth_uu=size_y, top_z=height, thickness_uu=max(140.0, outer_drop), material_path=edge_material_path)
        if long_axis == "x":
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_Shoulder_N",
                width_uu=size_x * 0.92,
                depth_uu=size_y * 0.26,
                base_z=shoulder_z,
                rise_uu=ridge_top_z - shoulder_z,
                material_path=edge_material_path,
                rise_axis="y",
                rise_direction=1.0,
                dy=-size_y * 0.14,
            )
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_Shoulder_S",
                width_uu=size_x * 0.92,
                depth_uu=size_y * 0.26,
                base_z=shoulder_z,
                rise_uu=ridge_top_z - shoulder_z,
                material_path=edge_material_path,
                rise_axis="y",
                rise_direction=-1.0,
                dy=size_y * 0.14,
            )
            _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_RidgeTop", width_uu=size_x * 0.9, depth_uu=size_y * 0.18, z=ridge_top_z, material_path=material_path)
            _append_terrain_foundation_fill(
                pieces,
                label_prefix=f"{label}_RidgeTop",
                width_uu=size_x * 0.9,
                depth_uu=size_y * 0.18,
                top_z=ridge_top_z,
                thickness_uu=max(100.0, ridge_top_z - shoulder_z),
                material_path=edge_material_path,
            )
            _append_terrain_inner_walls(pieces, label_prefix=f"{label}_Ridge", width_uu=size_x * 0.9, depth_uu=size_y * 0.18, low_z=shoulder_z, high_z=ridge_top_z, material_path=edge_material_path)
        else:
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_Shoulder_W",
                width_uu=size_x * 0.26,
                depth_uu=size_y * 0.92,
                base_z=shoulder_z,
                rise_uu=ridge_top_z - shoulder_z,
                material_path=edge_material_path,
                rise_axis="x",
                rise_direction=1.0,
                dx=-size_x * 0.14,
            )
            _append_terrain_ramped_surface(
                pieces,
                label_prefix=f"{label}_Shoulder_E",
                width_uu=size_x * 0.26,
                depth_uu=size_y * 0.92,
                base_z=shoulder_z,
                rise_uu=ridge_top_z - shoulder_z,
                material_path=edge_material_path,
                rise_axis="x",
                rise_direction=-1.0,
                dx=size_x * 0.14,
            )
            _append_terrain_surface_tiles(pieces, label_prefix=f"{label}_RidgeTop", width_uu=size_x * 0.18, depth_uu=size_y * 0.9, z=ridge_top_z, material_path=material_path)
            _append_terrain_foundation_fill(
                pieces,
                label_prefix=f"{label}_RidgeTop",
                width_uu=size_x * 0.18,
                depth_uu=size_y * 0.9,
                top_z=ridge_top_z,
                thickness_uu=max(100.0, ridge_top_z - shoulder_z),
                material_path=edge_material_path,
            )
            _append_terrain_inner_walls(pieces, label_prefix=f"{label}_Ridge", width_uu=size_x * 0.18, depth_uu=size_y * 0.9, low_z=shoulder_z, high_z=ridge_top_z, material_path=edge_material_path)
        _append_terrain_perimeter_skirts(pieces, label_prefix=label, width_uu=size_x, depth_uu=size_y, top_z=height, drop_uu=outer_drop, material_path=edge_material_path)

    else:
        _append_terrain_surface_tiles(
            pieces,
            label_prefix=f"{label}_Top",
            width_uu=size_x,
            depth_uu=size_y,
            z=height,
            material_path=material_path,
            grid_x=grid_x,
            grid_y=grid_y,
        )
        _append_terrain_perimeter_skirts(pieces, label_prefix=label, width_uu=size_x, depth_uu=size_y, top_z=height, drop_uu=outer_drop, material_path=edge_material_path)

    material_layers = _append_terrain_material_layers(
        pieces,
        layer_material_names=biome_profile["layer_material_names"],
        long_axis=long_axis,
    )

    return {
        "pieces": pieces,
        "size_x": size_x,
        "size_y": size_y,
        "height": height,
        "elevation": elevation,
        "biome": biome_profile["biome"],
        "material_layers": material_layers,
        "decoration_presets": biome_profile["decoration_presets"],
        "edge_material_name": edge_material_name,
        "edge_material_path": edge_material_path,
        "grid": {"x": grid_x, "y": grid_y},
        "continuous": grid_x == 1 and grid_y == 1,
    }


def _list_project_static_mesh_assets(force_refresh: bool = False) -> List[str]:
    now = time.time()
    if not force_refresh and _TERRAIN_ENVIRONMENT_ASSET_CACHE["assets"] and now < float(_TERRAIN_ENVIRONMENT_ASSET_CACHE["expires_at"] or 0.0):
        return list(_TERRAIN_ENVIRONMENT_ASSET_CACHE["assets"])

    result = _chat_uefn_query("list_assets", {
        "directory": "/Game/",
        "recursive": True,
        "class_filter": "StaticMesh",
    })
    payload = result.get("result") if isinstance(result, dict) else None
    assets: List[str] = []
    if isinstance(payload, dict):
        raw_assets = payload.get("assets") or []
    elif isinstance(payload, list):
        raw_assets = payload
    else:
        raw_assets = []

    for asset in raw_assets:
        path = str(asset).strip()
        if path.startswith("/Game/"):
            assets.append(path)

    _TERRAIN_ENVIRONMENT_ASSET_CACHE["assets"] = assets
    _TERRAIN_ENVIRONMENT_ASSET_CACHE["expires_at"] = now + 120.0
    return list(assets)


def _score_environment_asset(asset_path: str, keywords: List[str]) -> int:
    lowered = asset_path.lower()
    name = lowered.rsplit("/", 1)[-1]
    reject_tokens = ("material", "texture", "fx", "vfx", "niagara", "audio", "sound", "skeletal")
    if any(token in lowered for token in reject_tokens):
        return -999
    score = 0
    if "/foliage/" in lowered or "/nature/" in lowered:
        score += 8
    if name.startswith("sm_"):
        score += 4
    for keyword in keywords:
        if keyword in name:
            score += 14
        elif keyword in lowered:
            score += 6
    return score


def _select_environment_assets(asset_paths: List[str], decoration_presets: List[Dict[str, Any]]) -> Dict[str, str]:
    selected: Dict[str, str] = {}
    used: set[str] = set()

    for preset in decoration_presets:
        category = str(preset.get("category") or "").strip().lower()
        keywords = [str(keyword).lower() for keyword in (preset.get("keywords") or []) if str(keyword).strip()]
        ranked: List[tuple[int, str]] = []
        for path in asset_paths:
            score = _score_environment_asset(path, keywords)
            if score > 0:
                ranked.append((score, path))
        ranked.sort(key=lambda item: item[0], reverse=True)
        for _, path in ranked:
            if path not in used:
                selected[category] = path
                used.add(path)
                break

    return selected


def _float_from_value(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _actor_scale_average(actor: Dict[str, Any]) -> float:
    scale = actor.get("scale") or {}
    if isinstance(scale, dict):
        values = [
            max(0.01, _float_from_value(scale.get("x"), 1.0)),
            max(0.01, _float_from_value(scale.get("y"), 1.0)),
            max(0.01, _float_from_value(scale.get("z"), 1.0)),
        ]
        return sum(values) / len(values)
    return 1.0


def _actor_location_xy(actor: Dict[str, Any]) -> tuple[float, float]:
    location = actor.get("location") or actor.get("loc") or {}
    if isinstance(location, dict):
        return _float_from_value(location.get("x"), 0.0), _float_from_value(location.get("y"), 0.0)
    return 0.0, 0.0


def _match_environment_category_for_actor(actor: Dict[str, Any]) -> Optional[str]:
    label = str(actor.get("label") or actor.get("name") or "").lower()
    actor_class = str(actor.get("class") or actor.get("actor_class") or actor.get("type") or "").lower()
    blob = f"{label} {actor_class}"

    if any(token in blob for token in ("tree", "pine", "oak", "birch", "fir", "willow", "mangrove")):
        return "trees"
    if any(token in blob for token in ("bush", "shrub", "fern", "grass", "plant", "ivy", "reed", "cactus")):
        return "shrubs"
    if any(token in blob for token in ("rock", "stone", "boulder", "cliff", "pebble")):
        return "rocks"
    if any(token in blob for token in ("road", "street", "path", "sidewalk", "track")):
        return "roads"
    if any(token in blob for token in ("wall", "house", "building", "roof", "floor", "fence", "bridge", "tower", "platform")):
        return "structures"
    return None


def _collect_nearby_environment_context(px: float, py: float, radius: float) -> Dict[str, Any]:
    response = _chat_uefn_query("get_all_actors", {})
    payload = response.get("result") if isinstance(response, dict) else None
    actors = payload if isinstance(payload, list) else []
    if not actors:
        return {"counts": {}, "scale_means": {}, "total_nearby": 0}

    radius = max(600.0, float(radius))
    radius_sq = radius * radius
    counts: Dict[str, int] = {}
    scale_buckets: Dict[str, List[float]] = {}
    nearby_total = 0

    for actor in actors:
        if not isinstance(actor, dict):
            continue
        ax, ay = _actor_location_xy(actor)
        dx = ax - px
        dy = ay - py
        if dx * dx + dy * dy > radius_sq:
            continue
        nearby_total += 1
        category = _match_environment_category_for_actor(actor)
        if not category:
            continue
        counts[category] = counts.get(category, 0) + 1
        scale_buckets.setdefault(category, []).append(_actor_scale_average(actor))

    scale_means = {
        category: round(sum(values) / max(len(values), 1), 3)
        for category, values in scale_buckets.items()
        if values
    }
    return {
        "counts": counts,
        "scale_means": scale_means,
        "total_nearby": nearby_total,
    }


def _apply_scene_context_to_decoration_presets(
    decoration_presets: List[Dict[str, Any]],
    scene_context: Dict[str, Any],
    *,
    size_x: float,
    size_y: float,
    terrain_type: str,
) -> List[Dict[str, Any]]:
    counts = scene_context.get("counts") or {}
    scale_means = scene_context.get("scale_means") or {}
    short_side = min(size_x, size_y)
    local_area = max(1.0, size_x * size_y)
    density_area = max(local_area / 1_000_000.0, 1.0)
    structure_pressure = _float_from_value(counts.get("structures"), 0.0) / density_area
    road_presence = _float_from_value(counts.get("roads"), 0.0)
    terrain_type = (terrain_type or "flat").lower()
    aspect = max(size_x, size_y) / max(1.0, short_side)
    narrow_landform = terrain_type in {"ridge", "slope"} and (short_side < 3200.0 or aspect >= 3.0)

    adjusted: List[Dict[str, Any]] = []
    for preset in decoration_presets:
        updated = dict(preset)
        category = str(updated.get("category") or "").strip().lower()
        base_count = max(0.0, _float_from_value(updated.get("count"), 0.0))
        scale_min_value, scale_max_value = updated.get("scale") or (0.85, 1.15)
        avg_scale = _float_from_value(scale_means.get(category), 0.0)

        if avg_scale > 0:
            scale_center = max(0.55, min(2.4, avg_scale))
            spread = max(0.1, min(0.4, (float(scale_max_value) - float(scale_min_value)) / 2.0))
            scale_min_value = max(0.45, round(scale_center - spread, 3))
            scale_max_value = min(2.8, round(scale_center + spread, 3))
            updated["min_separation"] = round(max(
                220.0,
                _float_from_value(updated.get("min_separation"), 0.0) * max(0.75, min(1.7, scale_center))
            ), 3)
        if category in counts:
            local_density = _float_from_value(counts.get(category), 0.0) / density_area
            if local_density > 0:
                count_scale = max(0.45, min(1.9, 0.85 + local_density * 0.22))
                updated["count"] = int(round(base_count * count_scale))
        if structure_pressure > 1.4:
            if category == "trees":
                updated["count"] = int(round(_float_from_value(updated.get("count"), base_count) * 0.62))
                updated["radius_scale"] = round(max(0.22, _float_from_value(updated.get("radius_scale"), 0.4) * 0.84), 3)
            elif category == "shrubs":
                updated["count"] = int(round(_float_from_value(updated.get("count"), base_count) * 0.82))
            elif category == "rocks":
                updated["count"] = int(round(_float_from_value(updated.get("count"), base_count) * 0.9))
        if road_presence > 0 and category in {"trees", "shrubs"}:
            updated["radius_scale"] = round(max(0.2, _float_from_value(updated.get("radius_scale"), 0.4) * 0.9), 3)
        if short_side < 1800.0 and category == "trees":
            updated["count"] = min(int(updated.get("count") or 0), 6)
        if narrow_landform:
            if category == "trees":
                updated["count"] = 0
                updated["radius_scale"] = round(min(_float_from_value(updated.get("radius_scale"), 0.4), 0.2), 3)
            elif category == "rocks":
                updated["count"] = min(max(1, int(updated.get("count") or 0)), 4)
                updated["radius_scale"] = round(min(_float_from_value(updated.get("radius_scale"), 0.4), 0.14), 3)
                updated["min_separation"] = round(max(_float_from_value(updated.get("min_separation"), 0.0), short_side * 0.22), 3)
                scale_min_value = max(0.35, min(float(scale_min_value), 0.66))
                scale_max_value = max(scale_min_value + 0.08, min(float(scale_max_value), 0.84))
            elif category == "shrubs":
                updated["count"] = min(max(2, int(updated.get("count") or 0)), 5)
                updated["radius_scale"] = round(min(_float_from_value(updated.get("radius_scale"), 0.4), 0.12), 3)
                updated["min_separation"] = round(max(_float_from_value(updated.get("min_separation"), 0.0), short_side * 0.18), 3)
                scale_min_value = max(0.28, min(float(scale_min_value), 0.58))
                scale_max_value = max(scale_min_value + 0.08, min(float(scale_max_value), 0.74))
        updated["count"] = max(0, int(updated.get("count") or 0))
        updated["scale"] = (float(scale_min_value), float(scale_max_value))
        adjusted.append(updated)
    return adjusted


def _build_terrain_spine_points(px: float, py: float, pz: float, size_x: float, size_y: float, top_z: float, long_axis: str) -> List[List[float]]:
    points: List[List[float]] = []
    samples = 6
    if long_axis == "x":
        start_x = px - size_x * 0.42
        end_x = px + size_x * 0.42
        for idx in range(samples):
            t = idx / max(samples - 1, 1)
            points.append([round(start_x + (end_x - start_x) * t, 3), round(py, 3), round(pz + top_z + 24.0, 3)])
    else:
        start_y = py - size_y * 0.42
        end_y = py + size_y * 0.42
        for idx in range(samples):
            t = idx / max(samples - 1, 1)
            points.append([round(px, 3), round(start_y + (end_y - start_y) * t, 3), round(pz + top_z + 24.0, 3)])
    return points


def _build_terrain_environment_plan(
    *,
    label: str,
    material_name: str,
    terrain_type: str,
    px: float,
    py: float,
    pz: float,
    size_x: float,
    size_y: float,
    top_z: float,
    decoration_presets: List[Dict[str, Any]],
    available_assets: Optional[List[str]] = None,
) -> Dict[str, Any]:
    available_assets = list(available_assets if available_assets is not None else _list_project_static_mesh_assets())
    scene_context = _collect_nearby_environment_context(px, py, max(size_x, size_y) * 0.75)
    adapted_presets = _apply_scene_context_to_decoration_presets(
        decoration_presets,
        scene_context,
        size_x=size_x,
        size_y=size_y,
        terrain_type=terrain_type,
    )
    selected_assets = _select_environment_assets(available_assets, adapted_presets)
    if not selected_assets:
        return {
            "selected_assets": {},
            "operations": [],
            "notes": "No matching project environment meshes were found for this terrain biome.",
            "scene_context": scene_context,
        }

    long_axis = "x" if size_x >= size_y else "y"
    short_side = min(size_x, size_y)
    aspect = max(size_x, size_y) / max(1.0, short_side)
    narrow_landform = terrain_type in {"ridge", "slope"} and (short_side < 3200.0 or aspect >= 3.0)
    area_factor = max(0.75, min(2.25, (size_x * size_y) / (3200.0 * 3200.0)))
    radius = max(900.0, short_side * 0.44)
    operations: List[Dict[str, Any]] = []

    for preset in adapted_presets:
        category = str(preset.get("category") or "").strip().lower()
        asset_path = selected_assets.get(category)
        if not asset_path:
            continue
        base_count = int(round(float(preset.get("count") or 0) * area_factor))
        if base_count <= 0:
            continue
        count = max(1, min(42, base_count))
        scale_min, scale_max = preset.get("scale") or (0.85, 1.15)
        folder = f"TerrainDecor_{label}_{category.title()}"
        if aspect >= 2.3 and category in {"rocks", "shrubs"}:
            path_points = _build_terrain_spine_points(px, py, pz, size_x, size_y, top_z, long_axis)
            if narrow_landform and len(path_points) > 4:
                path_points = [path_points[index] for index in (0, 2, 3, 5)]
            operations.append({
                "tool": "scatter_along_path",
                "category": category,
                "mesh_path": asset_path,
                "path_points": path_points,
                "spread": max(120.0, short_side * (0.12 if category == "rocks" else 0.16)) if narrow_landform else max(180.0, short_side * (0.22 if category == "rocks" else 0.28)),
                "count_per_point": 1 if narrow_landform else (1 if category == "rocks" else 2),
                "scale_min": float(scale_min),
                "scale_max": float(scale_max),
                "folder": folder,
                "seed": abs(hash((label, category, asset_path))) % 100000,
            })
        else:
            operations.append({
                "tool": "scatter_props",
                "category": category,
                "mesh_path": asset_path,
                "count": count,
                "radius": round(radius * float(preset.get("radius_scale") or 0.44), 3),
                "center": [round(px, 3), round(py, 3), round(pz + top_z + 24.0, 3)],
                "min_separation": float(preset.get("min_separation") or 0.0),
                "scale_min": float(scale_min),
                "scale_max": float(scale_max),
                "rot_yaw_range": 360.0,
                "snap_to_surface": True,
                "folder": folder,
                "seed": abs(hash((label, category, asset_path))) % 100000,
            })

    return {
        "selected_assets": selected_assets,
        "operations": operations,
        "notes": "" if operations else "Matching assets were found, but no environment operations were generated.",
        "scene_context": scene_context,
    }


def _execute_terrain_environment_plan(plan: Dict[str, Any]) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    for operation in plan.get("operations") or []:
        tool_name = str(operation.get("tool") or "").strip()
        if not tool_name:
            continue
        params = {key: value for key, value in operation.items() if key not in {"tool", "category"}}
        output = tool_registry.execute_tool(tool_name, params)
        results.append({
            "tool": tool_name,
            "category": operation.get("category"),
            "mesh_path": operation.get("mesh_path"),
            "output": output,
        })
    return results


def _execute_terrain_control(action: dict) -> dict:
    """Terrain control: create, modify, or texture terrain patches.

    action = {
        "action": "terrain",
        "operation": "create|modify|delete|list",
        "terrain_type": "flat|hill|valley|plateau|slope|crater|ridge",
        "position": {"x": 0, "y": 0, "z": 0},
        "size": {"x": 2000, "y": 2000},  # ground plane dimensions in UU
        "material": "grass",
        "height": 0,  # base height for terrain
        "elevation": 0,  # extra height shaping (for hills/valleys)
        "label": "Terrain_Grass_01",
        "subdivisions": 1,  # how many tiles to break into (1=single, 4=2x2, 9=3x3)
        "decorate": True,   # auto-scatter biome-matched environment props when possible
    }
    """
    port = discover_uefn_listener_port()
    if not port:
        return {"error": "UEFN not connected"}

    operation = action.get("operation", "create").lower()
    terrain_type = action.get("terrain_type", "flat").lower()
    pos = action.get("position", {"x": 0, "y": 0, "z": 0})
    px, py, pz = pos.get("x", 0), pos.get("y", 0), pos.get("z", 0)
    size = action.get("size", {"x": 2000, "y": 2000})
    size_x, size_y = size.get("x", 2000), size.get("y", 2000)
    mat_name = _resolve_material_name(action.get("material", "grass"))
    mat_path = _get_material_path(mat_name) or MATERIAL_CATALOG["grass"]
    height = action.get("height", 0)
    elevation = action.get("elevation", 200)
    label = action.get("label", f"Terrain_{mat_name.title()}")
    subdivisions = action.get("subdivisions", 1)
    decorate = bool(action.get("decorate", True))

    if operation == "delete":
        # Delete terrain actors matching label pattern
        pattern = action.get("label_pattern", "Terrain_*")
        code = f"""import unreal
actors = unreal.EditorLevelLibrary.get_all_level_actors()
deleted = 0
for a in actors:
    lab = a.get_actor_label()
    if lab and (lab.startswith('{pattern.replace("*", "")}') if '*' in '{pattern}' else lab == '{pattern}'):
        unreal.EditorLevelLibrary.destroy_actor(a)
        deleted += 1
result = f'Deleted {{deleted}} terrain actors matching {pattern}'
print(result)
"""
        return mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=15.0)

    if operation == "list":
        code = """import unreal
actors = unreal.EditorLevelLibrary.get_all_level_actors()
terrain_actors = []
for a in actors:
    lab = a.get_actor_label()
    if lab and lab.startswith('Terrain_'):
        loc = a.get_actor_location()
        sc = a.get_actor_scale3d()
        terrain_actors.append(f'{lab}: pos=({loc.x:.0f},{loc.y:.0f},{loc.z:.0f}) scale=({sc.x:.1f},{sc.y:.1f},{sc.z:.1f})')
result = f'Found {len(terrain_actors)} terrain actors:\\n' + '\\n'.join(terrain_actors) if terrain_actors else 'No terrain actors found'
print(result)
"""
        return mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=15.0)

    terrain_spec = _build_terrain_piece_specs(
        terrain_type=terrain_type,
        size_x=size_x,
        size_y=size_y,
        height=height,
        elevation=elevation,
        label=label,
        material_name=mat_name,
        material_path=mat_path,
        subdivisions=subdivisions,
    )
    pieces = terrain_spec["pieces"]
    size_x = terrain_spec["size_x"]
    size_y = terrain_spec["size_y"]
    height = terrain_spec["height"]
    elevation = terrain_spec["elevation"]
    material_layers = terrain_spec.get("material_layers") or [mat_name]
    biome = str(terrain_spec.get("biome") or "mixed")

    spawn_lines = [
        "mesh_cache = {}",
        "mat_cache = {}",
        "spawned = []",
    ]
    for piece in pieces:
        x = px + float(piece["dx"])
        y = py + float(piece["dy"])
        z = pz + float(piece["dz"])
        mesh_path = str(piece["mesh_path"])
        mat_piece_path = str(piece["material_path"])
        spawn_lines.append(
            f"mesh = mesh_cache.get({mesh_path!r}) or unreal.EditorAssetLibrary.load_asset({mesh_path!r})\n"
            f"mesh_cache[{mesh_path!r}] = mesh\n"
            f"a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector({x},{y},{z}))\n"
            f"m = a.get_component_by_class(unreal.StaticMeshComponent)\n"
            f"if mesh: m.set_static_mesh(mesh)\n"
            f"a.set_actor_scale3d(unreal.Vector({piece['sx']},{piece['sy']},{piece['sz']}))\n"
            f"a.set_actor_rotation(unreal.Rotator({piece['roll']},{piece['pitch']},{piece['yaw']}), False)\n"
            f"a.set_actor_label({piece['label']!r})\n"
            f"m.set_collision_profile_name('BlockAll')\n"
            f"mat = mat_cache.get({mat_piece_path!r})\n"
            f"if mat is None:\n"
            f"    mat = unreal.EditorAssetLibrary.load_asset({mat_piece_path!r})\n"
            f"    mat_cache[{mat_piece_path!r}] = mat\n"
            f"if mat:\n"
            f"    m.set_material(0, mat)\n"
            f"spawned.append(a.get_actor_label())\n"
        )

    code = "import unreal\n" + "\n".join(spawn_lines)
    code += (
        f"\nresult = 'Created {terrain_type} terrain ({len(pieces)} pieces, grid {terrain_spec['grid']['x']}x{terrain_spec['grid']['y']}) "
        f"at ({px},{py},{pz}) with {mat_name} top and {terrain_spec['edge_material_name']} edges'\nprint(result)"
    )

    exec_result = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)
    top_surface_z = height
    top_planes = [piece for piece in pieces if piece.get("mesh_path") == "/Engine/BasicShapes/Plane"]
    if top_planes:
        top_surface_z = max(float(piece.get("dz", height)) for piece in top_planes)

    environment_plan = {"selected_assets": {}, "operations": [], "notes": ""}
    environment_results: List[Dict[str, Any]] = []
    if operation == "create" and decorate:
        environment_plan = _build_terrain_environment_plan(
            label=label,
            material_name=mat_name,
            terrain_type=terrain_type,
            px=px,
            py=py,
            pz=pz,
            size_x=size_x,
            size_y=size_y,
            top_z=top_surface_z,
            decoration_presets=terrain_spec.get("decoration_presets") or [],
        )
        if environment_plan.get("operations"):
            environment_results = _execute_terrain_environment_plan(environment_plan)

    decor_summary = ""
    successful_decor = [entry for entry in environment_results if (entry.get("output") or {}).get("success")]
    if successful_decor:
        parts: List[str] = []
        for entry in successful_decor:
            category = str(entry.get("category") or "details")
            output = entry.get("output") or {}
            nested = output.get("result") if isinstance(output.get("result"), dict) else output
            placed = 0
            if isinstance(nested, dict):
                placed = int(nested.get("placed") or nested.get("count") or 0)
            parts.append(f"{placed} {category}")
        decor_summary = ", ".join(parts[:3])

    exec_result["details"] = {
        "operation": operation,
        "terrain_type": terrain_type,
        "position": {"x": px, "y": py, "z": pz},
        "size": {"x": size_x, "y": size_y},
        "material": mat_name,
        "material_layers": material_layers,
        "biome": biome,
        "edge_material": terrain_spec["edge_material_name"],
        "grid": terrain_spec["grid"],
        "continuous": terrain_spec["continuous"],
        "pieces": len(pieces),
        "environment_assets": environment_plan.get("selected_assets") or {},
        "environment_operations": environment_plan.get("operations") or [],
        "environment_results": environment_results,
        "environment_notes": environment_plan.get("notes") or "",
        "environment_scene_context": environment_plan.get("scene_context") or {},
    }
    layer_summary = ", ".join(material_layers[:4])
    summary = (
        f"Created {terrain_type} terrain with layered materials ({layer_summary})"
        f" and {terrain_spec['edge_material_name']} edges."
    )
    if decor_summary:
        summary += f" Added environment dressing: {decor_summary}."
    elif decorate and environment_plan.get("notes"):
        summary += f" {environment_plan['notes']}"
    exec_result["summary"] = summary
    return exec_result


def _execute_action_blocks(reply: str) -> list:
    """Extract action blocks from the AI's reply and execute them.

    Supports multiple formats:
      ```action\n{...}\n```
      ```json\n{"action": ...}\n```
      ```\n{"action": ...}\n```
    """
    import re as _re
    results = []
    # Match ```action\n{...}\n``` or ```json\n{"action":...}\n``` or bare ```\n{"action":...}\n```
    pattern = _re.compile(r'```(?:action|json)?\s*\n(\{.*?\})\s*\n```', _re.DOTALL)
    matches = list(pattern.finditer(reply))

    # Fallback: look for inline JSON with "action" key if no fenced blocks found
    if not matches:
        inline_pattern = _re.compile(r'\{[^{}]*"action"\s*:\s*"[^"]+?"[^{}]*\}')
        for m in inline_pattern.finditer(reply):
            try:
                candidate = json.loads(m.group(0))
                action_name = _normalize_action_type(candidate.get("action"))
                if action_name in _SUPPORTED_ACTION_TYPES:
                    matches.append(m)
            except json.JSONDecodeError:
                pass

    for match in matches:
        raw = match.group(1).strip() if match.lastindex else match.group(0).strip()
        try:
            action = json.loads(raw)
        except json.JSONDecodeError:
            results.append({"error": f"Invalid action JSON: {raw[:200]}"})
            continue

        action_type = _normalize_action_type(action.get("action", ""))
        if not action_type:
            continue
        if action_type not in _SUPPORTED_ACTION_TYPES:
            results.append({"action": action_type, "error": f"Unknown action type: {action_type}"})
            continue
        action["action"] = action_type
        try:
            if action_type == "execute_python_in_uefn":
                code = action.get("code", "")
                validation_error = _validate_uefn_python_action(code)
                if validation_error:
                    results.append({"action": action_type, "error": validation_error})
                    continue
                port = discover_uefn_listener_port()
                if not port:
                    results.append({"action": action_type, "error": "UEFN not connected"})
                    continue
                r = mcp_listener_post_command(port, "execute_python", {"code": code}, timeout=30.0)
                r["action"] = action_type
                results.append(r)

            elif action_type == "run_uefn_tool":
                tool_name = action.get("tool_name", "")
                params = action.get("parameters", {})
                r = tool_registry.execute_tool(tool_name, params)
                r["action"] = action_type
                results.append(r)

            elif action_type == "query_uefn":
                cmd = action.get("command", "")
                params = action.get("params", {})
                r = _chat_uefn_query(cmd, params)
                r["action"] = action_type
                results.append(r)

            elif action_type == "rebuild_gable_roof":
                # Server-side roof builder — does all math correctly
                r = _execute_rebuild_gable_roof(action)
                r["action"] = action_type
                results.append(r)

            elif action_type == "apply_material":
                # Apply a material to actors by label pattern
                r = _execute_apply_material(action)
                r["action"] = action_type
                results.append(r)

            elif action_type == "build_house":
                # Functional house generation through the shared structure planner
                r = _execute_build_house(action)
                r["action"] = action_type
                results.append(r)

            elif action_type == "build_structure":
                # Backward-compatible structure routing through the shared planner
                r = _execute_structure_action_with_shared_planner(action)
                r["action"] = action_type
                results.append(r)

            elif action_type == "terrain":
                # Terrain control — create/modify/delete/list terrain
                r = _execute_terrain_control(action)
                r["action"] = action_type
                results.append(r)

            elif action_type == "import_attached_models":
                context = _get_active_tool_context()
                request_text = str(action.get("request") or context.get("message") or "import attached models into uefn")
                attachments = context.get("attachments") or []
                payload = _maybe_import_and_place_model_attachments(request_text, attachments)
                if not payload:
                    results.append({
                        "action": action_type,
                        "error": "No importable FBX or ZIP model attachments were available in the current chat turn.",
                    })
                    continue
                tool_output = (payload.get("tool_result") or {}).get("output") or {}
                results.append({
                    "action": action_type,
                    "success": bool(tool_output.get("success", True)),
                    "reply": payload.get("reply") or "",
                    "result": tool_output,
                })

            else:
                results.append({"action": action_type, "error": f"Unknown action type: {action_type}"})

        except Exception as e:
            results.append({"action": action_type, "error": str(e)})

    return results


def _ai_chat(
    message: str,
    attachments: list,
    history: list,
    chat_title: str = "",
    chat_memory: str = "",
    chat_id: str = "",
    _tried_providers: Optional[List[str]] = None,
) -> dict:
    """AI-powered chat using the best available free provider with tool calling."""
    global _llm_client, _llm_provider
    effective_attachments = _resolve_effective_turn_attachments(message, attachments, history)
    _set_active_tool_context(message=message, attachments=effective_attachments, chat_id=chat_id)
    direct_result = _maybe_execute_direct_action(message, attachments or effective_attachments)
    if direct_result is not None:
        return direct_result

    attachment_result = _maybe_answer_directly_from_attachments(message, effective_attachments)
    if attachment_result is not None:
        return attachment_result

    # Remember the original provider/client before any vision swaps
    original_provider = _get_active_provider()
    original_model = _get_active_model()
    original_client = _get_llm_client()

    client, provider, model = _resolve_llm_request_target(effective_attachments, history)
    if not client:
        return None  # Signal to fall back to keyword matching

    needs_native_media = _attachments_need_native_media_reasoning(effective_attachments) or _history_needs_native_media_reasoning(history)

    if provider == "gemini" and needs_native_media:
        try:
            return _gemini_generate_content_chat(
                message=message,
                attachments=effective_attachments,
                history=history,
                chat_title=chat_title,
                chat_memory=chat_memory,
                model=model,
            )
        except Exception as gemini_error:
            error_text = str(gemini_error)
            logger.warning("Gemini multimodal generateContent fallback failed: %s", error_text)

            try:
                attachment_fallback = _answer_with_attachment_analysis_context(
                    message=message,
                    attachments=effective_attachments,
                    preferred_provider=original_provider or _get_active_provider(),
                )
                if attachment_fallback.get("reply"):
                    return attachment_fallback
            except Exception as fallback_error:
                logger.warning("Attachment analysis fallback also failed: %s", fallback_error)

            # ── Fall back to original provider with tool-calling path (text-only dossier) ──
            if original_client and original_provider != "gemini":
                logger.info("Gemini vision failed, falling back to %s with tool-calling path", original_provider)
                client = original_client
                provider = original_provider
                model = original_model
                needs_native_media = False
            else:
                # Gemini was the user's chosen provider and it failed — try any other provider
                fallback_client, fallback_provider, fallback_model = _resolve_text_fallback_target(preferred_provider="groq")
                if fallback_client:
                    client = fallback_client
                    provider = fallback_provider
                    model = fallback_model
                    needs_native_media = False
                else:
                    return {"reply": f"Gemini vision failed: {error_text[:200]}. Try switching to a different provider.", "_provider": "gemini", "_model": model}

    # Build messages
    messages = [{"role": "system", "content": _build_system_prompt(chat_title=chat_title, chat_memory=chat_memory)}]
    retrieval_block = _format_retrieval_block(message, current_chat_id=chat_id)
    if retrieval_block:
        messages.append({"role": "system", "content": retrieval_block})
    execution_precheck = _build_execution_precheck_block(message, effective_attachments)
    if execution_precheck:
        messages.append({"role": "system", "content": execution_precheck})
    tool_routing_guidance = _build_tool_routing_guidance(message, effective_attachments)
    if tool_routing_guidance:
        messages.append({"role": "system", "content": tool_routing_guidance})

    # Add conversation history (last 10 messages)
    for h in history[-10:]:
        role = h.get("role", "user")
        content = h.get("content", "")
        hist_attachments = h.get("attachments") or []
        if role == "user":
            rendered_content = _build_user_message_content(content, hist_attachments, provider, history_mode=True)
        else:
            rendered_content = content
        if role in ("user", "assistant") and rendered_content:
            messages.append({"role": role, "content": rendered_content})

    # Build current user message with attachment context
    user_content = _build_user_message_content(message, effective_attachments, provider)

    # ── When images are attached, pre-fetch UEFN context so the AI has real data ──
    # This prevents the AI from saying "please select actors" — it already has actor data.
    has_images = any(
        str(a.get("type", "")).lower() == "image"
        for a in effective_attachments
    )
    if has_images or needs_native_media:
        prefetch_context = ""
        port = discover_uefn_listener_port()
        if port:
            try:
                r_actors = _chat_uefn_query("get_all_actors", {})
                if r_actors.get("success"):
                    actors_raw = r_actors.get("result", [])
                    # Extract the actors list (may be nested under 'actors' key)
                    actors_list = actors_raw
                    if isinstance(actors_raw, dict) and "actors" in actors_raw:
                        actors_list = actors_raw["actors"]
                    if isinstance(actors_list, list):
                        # Filter to structural/visible actors (skip system actors)
                        skip_classes = {"WorldDataLayers", "WorldSettings", "LevelBounds",
                                        "NavigationData", "GameMode", "GameState", "PlayerStart",
                                        "AbstractNavData", "RecastNavMesh", "PostProcessVolume"}
                        structural = [a for a in actors_list
                                      if str(a.get("class", "")) not in skip_classes
                                      and not str(a.get("label", "")).startswith("_")]
                        # Build compact actor summary with position/rotation/scale
                        actor_summaries = []
                        for a in structural[:60]:
                            summary = {
                                "label": a.get("label", ""),
                                "class": a.get("class", ""),
                                "loc": a.get("location", {}),
                                "rot": a.get("rotation", {}),
                                "scale": a.get("scale", {}),
                            }
                            actor_summaries.append(summary)
                        prefetch_context += (
                            f"\n\n[LIVE UEFN DATA - {len(structural)} structural actors in level. "
                            f"Each has label, class, location (x,y,z), rotation (pitch,yaw,roll), scale (x,y,z):\n"
                            f"{json.dumps(actor_summaries, default=str)[:8000]}]"
                        )
                    else:
                        prefetch_context += f"\n\n[LIVE UEFN DATA:\n{json.dumps(actors_raw, default=str)[:4000]}]"
            except Exception:
                pass
            try:
                r_sel = _chat_uefn_query("get_selected_actors", {})
                if r_sel.get("success"):
                    sel = r_sel.get("result", [])
                    if isinstance(sel, list) and sel:
                        prefetch_context += f"\n\n[CURRENTLY SELECTED ACTORS ({len(sel)}):\n{json.dumps(sel[:20], default=str)[:2000]}]"
            except Exception:
                pass
        matching_tools = _search_tools_for_llm(message, limit=8)
        if matching_tools:
            prefetch_context += f"\n\n[RELEVANT TOOLS:\n{json.dumps(matching_tools, default=str)[:3000]}]"
        if prefetch_context:
            if isinstance(user_content, str):
                user_content += prefetch_context
            elif isinstance(user_content, list):
                user_content.append({"type": "text", "text": prefetch_context})
        # Add safer visual execution guidance for image-driven edit requests
        messages[0]["content"] += (
            "\n\nIMAGE ATTACHED — CRITICAL INSTRUCTIONS:"
            "\n- You have LIVE UEFN DATA with EXACT labels, positions, rotations, and scales for every actor."
            "\n- Use the EXACT actor labels from LIVE UEFN DATA. Do NOT invent or modify labels."
            "\n- Analyze the transform data to identify the problem (wrong rotation, position, or asymmetric scale)."
            "\n- NEVER write conditional code like 'if scale.z < 0.3'. ALWAYS apply direct unconditional changes."
            "\n- For paired actors (left/right roof, etc.), make scales match and rotations symmetric."
            "\n- Use execute_python_in_uefn to directly fix the actors. DO NOT just describe the problem."
            "\n- NEVER ask the user to select actors — you already have the full actor list."
        )

    messages.append({"role": "user", "content": user_content})

    # Determine if this provider supports tool calling
    # Cerebras does NOT reliably support OpenAI-style tool calling — use pre-fetch path
    use_tools = provider in ("groq", "gemini", "openai")

    # Force tool calling when user has action intent (fix, delete, move, etc.)
    msg_lower = message.lower()
    action_keywords = ('fix', 'delete', 'remove', 'add', 'move', 'build', 'create', 'spawn',
                       'color', 'change', 'rotate', 'scale', 'align', 'place', 'set', 'make',
                       'clean', 'organize', 'rename', 'destroy', 'replace')
    user_wants_action = any(kw in msg_lower for kw in action_keywords)
    requires_reasoned_targeting = bool(effective_attachments) or _attachment_followup_reference_requested(message)
    effective_tool_choice = (
        "required"
        if _should_force_tool_execution(
            message,
            effective_attachments,
            use_tools=use_tools,
            user_wants_action=user_wants_action,
            requires_reasoned_targeting=requires_reasoned_targeting,
        )
        else "auto"
    )

    # ── Retry detection: if user is repeating a request, inject stronger directive ──
    retry_phrases = ('again', 'still', 'didnt work', "didn't work", 'not working',
                     'same thing', 'nothing changed', 'try again', 'do it again',
                     'fix it', 'still the same', 'still broken', 'still bad')
    is_retry = any(phrase in msg_lower for phrase in retry_phrases)
    if is_retry and user_wants_action:
        retry_msg = (
            "\n\n[RETRY: The user's previous request FAILED or produced no visible change. "
            "Take a COMPLETELY DIFFERENT approach: DELETE the problematic actors and REBUILD from scratch. "
            "Do NOT repeat the same code. Use execute_python_in_uefn with the ROOF FIX PROCEDURE if it involves a roof.]"
        )
        if isinstance(user_content, str):
            user_content += retry_msg
        elif isinstance(user_content, list):
            user_content.append({"type": "text", "text": retry_msg})
        # Update the message in messages list
        messages[-1]["content"] = user_content

    try:
        if use_tools:
            # ── OpenAI path: real tool calling ──
            response = _chat_completion_with_retry(
                client,
                provider,
                model=model,
                messages=messages,
                tools=_CHAT_FUNCTIONS,
                tool_choice=effective_tool_choice,
                max_tokens=2000,
                temperature=0.7
            )

            assistant_msg = response.choices[0].message
            tool_results_for_frontend = []
            total_in = getattr(response.usage, 'prompt_tokens', 0) if response.usage else 0
            total_out = getattr(response.usage, 'completion_tokens', 0) if response.usage else 0

            rounds = 0
            while assistant_msg.tool_calls and rounds < 3:
                rounds += 1
                messages.append(assistant_msg)

                for tc in assistant_msg.tool_calls:
                    try:
                        args = json.loads(tc.function.arguments) if isinstance(tc.function.arguments, str) else tc.function.arguments
                    except json.JSONDecodeError:
                        args = {}

                    result = _handle_tool_call(tc.function.name, args)
                    result_str = json.dumps(result, default=str)[:4000]

                    tool_results_for_frontend.append({
                        "tool": tc.function.name,
                        "output": result
                    })

                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc.id,
                        "content": result_str
                    })

                response = _chat_completion_with_retry(
                    client,
                    provider,
                    model=model,
                    messages=messages,
                    tools=_CHAT_FUNCTIONS,
                    tool_choice="auto",  # After first round, switch to auto for follow-ups
                    max_tokens=2000,
                    temperature=0.7
                )
                assistant_msg = response.choices[0].message
                if response.usage:
                    total_in += getattr(response.usage, 'prompt_tokens', 0)
                    total_out += getattr(response.usage, 'completion_tokens', 0)

            _track_usage(provider, total_in, total_out)

            reply = assistant_msg.content or "I executed your request."
            result = {"reply": reply, "_provider": provider, "_model": model}
            if tool_results_for_frontend:
                result["tool_result"] = tool_results_for_frontend[-1]
            return result

        else:
            # ── Pre-fetch path (Cerebras, Ollama, etc.) ──
            # These providers don't support OpenAI tool calling reliably.
            # Strategy: ALWAYS fetch live UEFN context so the AI knows what's in the level,
            # then ask the AI to respond AND emit structured action blocks we can execute.
            tool_result_for_frontend = None
            msg_lower = message.lower()
            port = discover_uefn_listener_port()

            # ── ALWAYS pre-fetch level actors + selection + level info when UEFN is connected ──
            extra_context = ""
            if port:
                try:
                    # Actors list with transforms
                    r_actors = _chat_uefn_query("get_all_actors", {})
                    if r_actors.get("success"):
                        actors_raw = r_actors.get("result", [])
                        actors_list = actors_raw
                        if isinstance(actors_raw, dict) and "actors" in actors_raw:
                            actors_list = actors_raw["actors"]
                        if isinstance(actors_list, list):
                            skip_classes = {"WorldDataLayers", "WorldSettings", "LevelBounds",
                                            "NavigationData", "GameMode", "GameState", "PlayerStart",
                                            "AbstractNavData", "RecastNavMesh", "PostProcessVolume"}
                            structural = [a for a in actors_list
                                          if str(a.get("class", "")) not in skip_classes]
                            actor_summaries = []
                            for a in structural[:60]:
                                actor_summaries.append({
                                    "label": a.get("label", ""),
                                    "class": a.get("class", ""),
                                    "loc": a.get("location", {}),
                                    "rot": a.get("rotation", {}),
                                    "scale": a.get("scale", {}),
                                })
                            extra_context += (
                                f"\n\n[LIVE UEFN DATA - {len(structural)} actors with location/rotation/scale:\n"
                                f"{json.dumps(actor_summaries, default=str)[:8000]}]"
                            )
                        else:
                            extra_context += f"\n\n[LIVE UEFN DATA:\n{json.dumps(actors_raw, default=str)[:4000]}]"
                except Exception as e:
                    logger.warning("Pre-fetch actors failed: %s", e)

                try:
                    # Selected actors
                    r_sel = _chat_uefn_query("get_selected_actors", {})
                    if r_sel.get("success"):
                        sel = r_sel.get("result", [])
                        if isinstance(sel, list) and sel:
                            extra_context += f"\n\n[CURRENTLY SELECTED ACTORS ({len(sel)}):\n{json.dumps(sel[:20], default=str)[:2000]}]"
                        elif sel:
                            extra_context += f"\n\n[SELECTED:\n{json.dumps(sel, default=str)[:2000]}]"
                except Exception as e:
                    logger.warning("Pre-fetch selection failed: %s", e)

                try:
                    # Level info
                    r_lvl = _chat_uefn_query("get_level_info", {})
                    if r_lvl.get("success"):
                        extra_context += f"\n\n[LEVEL INFO:\n{json.dumps(r_lvl.get('result', {}), default=str)[:1500]}]"
                except Exception as e:
                    logger.warning("Pre-fetch level info failed: %s", e)

                try:
                    # Viewport
                    r_vp = _chat_uefn_query("get_viewport_camera", {})
                    if r_vp.get("success"):
                        extra_context += f"\n\n[VIEWPORT CAMERA:\n{json.dumps(r_vp.get('result', {}), default=str)[:500]}]"
                except Exception as e:
                    logger.warning("Pre-fetch viewport failed: %s", e)

            # Additional context for specific requests
            if any(kw in msg_lower for kw in ['project tree', 'asset tree', 'list assets', 'content browser']):
                r = _chat_uefn_query("list_assets", {"directory": "/Game/", "recursive": True, "class_filter": ""})
                if r.get("success"):
                    assets = r.get("result", [])
                    sample = assets[:80] if isinstance(assets, list) else assets
                    extra_context += f"\n\n[ASSET TREE ({len(assets) if isinstance(assets, list) else '?'} assets):\n{json.dumps(sample, default=str)[:5000]}]"

            if any(kw in msg_lower for kw in ['list tools', 'available tools', 'what tools', 'show tools', 'what can you do']):
                extra_context += f"\n\n[FULL TOOL LIST:\n{tool_registry.get_tools_for_ai()[:5000]}]"

            # Always search for relevant tools so the AI knows what's available
            matching_tools = _search_tools_for_llm(message, limit=10)
            if matching_tools:
                extra_context += f"\n\n[TOOLS MOST RELEVANT TO YOUR REQUEST (use run_uefn_tool with these):\n{json.dumps(matching_tools, default=str)[:4000]}]"

            execution_precheck = _build_execution_precheck_block(message, effective_attachments)
            if execution_precheck:
                extra_context += f"\n\n[{execution_precheck}]"

            # Inject action instruction into the system prompt for pre-fetch mode
            action_instruction = """

═══════════════════════════════════════════════════════════════
CRITICAL RESPONSE RULES — FOLLOW THESE EXACTLY:
═══════════════════════════════════════════════════════════════

1. BE CONCISE. Your response should be SHORT — a few sentences max, then action blocks. NO essays, NO tutorials, NO step-by-step guides.
2. NEVER AUTOCORRECT the user's message. If they misspell something, just understand it and act. NEVER write "I think you meant..." or "Based on context, you likely meant..." or rewrite their words.
3. NEVER use markdown headers (###), horizontal rules (---), or emoji bullet points in your response.
4. NEVER list "Next Steps", "Recommended Tools", "Final Cleanup Tips", or "What you can do" sections.
5. NEVER say "I can help in a few ways", "Here are some options", "Let me know if you'd like", or "If you can reupload..."
6. NEVER analyze or describe what the user's images show in long paragraphs. Just briefly acknowledge what you see and EXECUTE the fix.
7. Your response format should be: 1-3 short sentences saying what you're doing → action block(s). That's it.

BAD RESPONSE (NEVER do this):
"Based on your request, you likely meant 'fix these holes'. Here's a step-by-step guide: ### Step 1: Check geometry ### Step 2: Add collision..."

GOOD RESPONSE (ALWAYS do this):
"I can see the holes in your walls and the clipping issues. Fixing them now by sealing the geometry and adding proper collision."
```action
{"action": "execute_python_in_uefn", "code": "import unreal\\n..."}
```

═══════════════════════════════════════════════════════════════

ACTION EXECUTION (MANDATORY):
You CANNOT call tools directly. Instead, include ACTION BLOCKS in your response. The system executes them automatically.

FORMAT — always use this exact format:
```action
{"action": "execute_python_in_uefn", "code": "import unreal\\nactors = unreal.EditorLevelLibrary.get_all_level_actors()\\nresult = [a.get_name() for a in actors]"}
```

Or to run a registered tool:
```action
{"action": "run_uefn_tool", "tool_name": "screenshot_take", "parameters": {}}
```

Or to rebuild a roof (ALWAYS use this for roof fixes — do NOT write your own roof code):
```action
{"action": "rebuild_gable_roof", "roof_height_ratio": 0.3}
```
This auto-calculates dimensions from LIVE UEFN DATA, deletes old pieces, and builds a correct gable roof.

To apply a material/texture to actors:
```action
{"action": "apply_material", "actor_pattern": "Fountain_*", "material": "stone"}
```
Available materials: brick, wood, metal, concrete, stucco, stone, glass, water, lava, grass, terrain, moss, sparkle, glitter, wood_floor, brick_floor, metal_floor, default
Aliases: rocks→stone, cement→concrete, shiny→sparkle, pool→water, etc.

To build a functional residential building with the shared structure planner:
```action
{"action": "build_house", "request": "build a modern house here", "style": "modern", "size": "large"}
```
Use this for house/home/cabin/cottage/townhouse/apartment/condo requests.
- Prefer this over `build_structure` for anything residential.
- Houses must include coherent walls, floors, roof, door, and sane circulation.
- If the build is multi-story, the stairs must arrive into a real landing and open floor area.
- Vary the house within sane residential ranges instead of cloning the same shape every time.
- For apartments and condos, include `story_count` when the request clearly calls for 3+ stories.

To build a non-residential or scenic structure generatively (with auto-texturing):
```action
{"action": "build_structure", "structure": "fountain", "position": {"x": 5200, "y": 4200, "z": 0}, "size": "medium", "material": "stone"}
```
Available structures: garage, shed, workshop, barn, warehouse, greenhouse, studio, hangar, kiosk, pavilion, gazebo, pergola, canopy, carport, market stall, fountain, column, arch, tower, pool, fence, platform, bridge, waterfall
Sizes: small, medium, large, huge

To create layered terrain with biome-aware decoration:
```action
{"action": "terrain", "operation": "create", "terrain_type": "ridge", "position": {"x": 0, "y": 0, "z": 0}, "size": {"x": 12000, "y": 2400}, "material": "grass", "elevation": 450, "decorate": true}
```

To import uploaded model files from this chat turn:
```action
{"action": "import_attached_models", "request": "import these models and place them along the terrain spline"}
```
This imports attached FBX/ZIP files, keeps source materials and textures when available, and can place the result along splines, terrain curves, or curved patterns.

RULES:
- DO NOT use query_uefn — you ALREADY have all the actor data in the LIVE UEFN DATA above. Go straight to execute.
1. You MUST include an action block whenever the user asks to DO something (delete, move, color, find, count, etc.)
2. For execute_python_in_uefn: code runs via exec(). ALWAYS end with `result = <value>` so output is captured.
3. Use \\n for newlines inside the JSON code string.
4. Always `import unreal` first.
5. You can include MULTIPLE action blocks in one response if needed.
6. Say what you're doing in 1-2 sentences BEFORE the action block, then include it. Nothing after.
7. NEVER just describe steps without including the action block. ALWAYS execute.
8. NEVER say "I can help in a few ways" or list alternatives. Just DO the action.
9. If no perfect tool exists, use execute_python_in_uefn — you can do ANYTHING with Python.
10. NEVER mutate broad actor sets blindly. Query first, narrow to the selection or a small candidate set, then edit only those actors.
11. After any world-changing edit, include a verification query or screenshot action block before saying the fix is done.
12. For waterfalls, hills, cliffs, and landforms, do NOT fake them with stair-step cubes unless the user explicitly asked for a blockout. Use terrain for the landform and continuous water geometry for the cascade.

FILE & IMAGE ANALYSIS:
- The ATTACHMENT DOSSIER below contains extracted content from user uploads.
- When you see a UEFN screenshot, identify the scene and EXECUTE fixes — don't write essays about what you see.
- When multiple files are attached, cross-reference them and act on the combined context.
- ALWAYS act on what you see in attachments. If the user pastes a screenshot and says "fix this", look at the spatial data and write Python to fix it.

STRUCTURAL FIX GUIDE (for holes, clipping, misalignment):
- You have FULL transform data for every actor (location, rotation, scale) in the LIVE UEFN DATA.
- CRITICAL: Use the EXACT actor labels from the LIVE UEFN DATA. Do NOT invent labels. Copy them exactly as shown.
- Look at ROTATION values — misaligned actors need rotation correction via set_actor_rotation.
- Look at LOCATION values — gapped or overlapping actors need position adjustment via set_actor_location.
- Look at SCALE values — asymmetric paired pieces (left/right roof, etc.) should have matching scales.
- Use explicit target selection or exact actor labels before changing transforms.
- It is OK to inspect current values before changing them. Do that when it helps prevent overcorrecting the scene.
- For roofs: left/right should be symmetric (same scale, mirrored pitch). Fix any asymmetry.
- For walls: edges must align — check X/Y coordinates to close gaps.

BAD CODE (never do this — it changes NOTHING when values are already at the threshold):
```python
if new_scale.z < 0.3:
    new_scale.z = 0.3  # USELESS if z is already 0.3!
```

GOOD CODE (direct fix — always changes something):
```python
# Make roof sides symmetric by matching scales and positions
actor.set_actor_scale3d(unreal.Vector(6.9, 6.9, 0.3))
actor.set_actor_location(unreal.Vector(4200, 4200, 680), False)
```

PYTHON PATTERNS:
- Find actor by EXACT label: `[a for a in unreal.EditorLevelLibrary.get_all_level_actors() if a.get_actor_label() == 'ExactLabel']`
- Find by partial name: `[a for a in actors if 'keyword' in a.get_actor_label().lower()]`
- Move: `actor.set_actor_location(unreal.Vector(x,y,z), False)`
- Rotate: `actor.set_actor_rotation(unreal.Rotator(roll, pitch, yaw), False)` — NOTE: UEFN order is (roll, pitch, yaw) NOT (pitch, yaw, roll)
- Scale: `actor.set_actor_scale3d(unreal.Vector(sx, sy, sz))`
- Delete: `unreal.EditorLevelLibrary.destroy_actor(actor)`
- Spawn: `a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(x,y,z))`
- Set mesh: `comp.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))`
- Collision: `comp.set_collision_profile_name("BlockAll")`
- Label: `actor.set_actor_label('MyLabel')`
- Get mesh component: `comp = actor.get_component_by_class(unreal.StaticMeshComponent)`

3D CONSTRUCTION GUIDE — HOW TO BUILD STRUCTURES IN UEFN:
Use BasicShapes/Cube as building blocks. Each piece is a cube actor positioned, rotated, and scaled.
One Unreal unit = 1 cm. Standard wall height = 300 units (3m). Standard wall thickness = 20 units.

WALL CONSTRUCTION (two-point walls):
```python
import unreal, math
def build_wall(start_x, start_y, end_x, end_y, z_base, height=300, thickness=20, label='Wall'):
    length = math.sqrt((end_x - start_x)**2 + (end_y - start_y)**2)
    angle_rad = math.atan2(end_y - start_y, end_x - start_x)
    angle_deg = math.degrees(angle_rad)
    cx = (start_x + end_x) / 2
    cy = (start_y + end_y) / 2
    cz = z_base + height / 2
    a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(cx, cy, cz))
    m = a.get_component_by_class(unreal.StaticMeshComponent)
    m.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
    a.set_actor_scale3d(unreal.Vector(length/100, thickness/100, height/100))
    a.set_actor_rotation(unreal.Rotator(0, 0, angle_deg), False)  # UEFN: (roll, pitch, yaw) — yaw for horizontal rotation
    a.set_actor_label(label)
    m.set_collision_profile_name('BlockAll')
    return a
```

FLOOR/SLAB (flat rectangle):
```python
def build_floor(cx, cy, cz, width, depth, thickness=10, label='Floor'):
    a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(cx, cy, cz))
    m = a.get_component_by_class(unreal.StaticMeshComponent)
    m.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
    a.set_actor_scale3d(unreal.Vector(width/100, depth/100, thickness/100))
    a.set_actor_label(label)
    m.set_collision_profile_name('BlockAll')
    return a
```

ROOF TYPES — use pitched cubes:
GABLE ROOF (two symmetric slopes meeting at ridge):
```python
import math
def build_gable_roof(cx, cy, cz_base, width, depth, roof_height, thickness=10):
    pitch_angle = math.degrees(math.atan2(roof_height, depth/2))
    slope_length = math.sqrt((depth/2)**2 + roof_height**2)
    ridge_z = cz_base + roof_height
    # Left slope
    left = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(cx, cy - depth/4, cz_base + roof_height/2))
    lm = left.get_component_by_class(unreal.StaticMeshComponent)
    lm.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
    left.set_actor_scale3d(unreal.Vector(width/100, slope_length/100, thickness/100))
    left.set_actor_rotation(unreal.Rotator(-pitch_angle, 0, 0), False)  # -roll = left slope tilts toward ridge
    left.set_actor_label('Roof_Left')
    # Right slope (mirror pitch)
    right = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(cx, cy + depth/4, cz_base + roof_height/2))
    rm = right.get_component_by_class(unreal.StaticMeshComponent)
    rm.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
    right.set_actor_scale3d(unreal.Vector(width/100, slope_length/100, thickness/100))
    right.set_actor_rotation(unreal.Rotator(pitch_angle, 0, 0), False)  # +roll = right slope tilts toward ridge
    right.set_actor_label('Roof_Right')
    # Ridge cap
    ridge = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(cx, cy, ridge_z))
    ridm = ridge.get_component_by_class(unreal.StaticMeshComponent)
    ridm.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
    ridge.set_actor_scale3d(unreal.Vector(width/100, 0.08, 0.08))
    ridge.set_actor_label('Roof_Ridge')
    return [left, right, ridge]
```

SHED ROOF (single slope):
  pitch = atan2(roof_height, depth), single slab tilted at that angle

HIP ROOF (4 slopes meeting at peak):
  Same as gable but add front/back triangular slopes

BUILDING A COMPLETE HOUSE:
1. Build floor slab at ground level
2. Build 4 walls (front, back, left, right) using build_wall with connected corners
3. Build second floor slab if multi-story
4. Build roof using gable/hip/shed pattern
5. Add gable end walls (triangular infill between roof and walls)
6. Set collision on ALL pieces

DOOR OPENING (subtract from wall):
  Split wall into 3 pieces: left of door, above door, right of door
  Or: spawn a slightly smaller wall piece with a gap

WINDOW OPENING:
  Split wall into 5 pieces: left, right, above, below, and two side columns

STAIRCASE:
```python
def build_stairs(start_x, start_y, start_z, num_steps, step_width, step_depth, step_height):
    actors = []
    for i in range(num_steps):
        x = start_x
        y = start_y + i * step_depth
        z = start_z + i * step_height
        a = unreal.EditorLevelLibrary.spawn_actor_from_class(unreal.StaticMeshActor, unreal.Vector(x, y, z + step_height/2))
        m = a.get_component_by_class(unreal.StaticMeshComponent)
        m.set_static_mesh(unreal.EditorAssetLibrary.load_asset('/Engine/BasicShapes/Cube'))
        a.set_actor_scale3d(unreal.Vector(step_width/100, step_depth/100, step_height/100))
        a.set_actor_label(f'Stair_{i+1:02d}')
        m.set_collision_profile_name('BlockAll')
        actors.append(a)
    return actors
```

MATERIAL APPLICATION:
```python
# Set material on a mesh component
mat = unreal.EditorAssetLibrary.load_asset('/Game/Materials/M_White')
comp.set_material(0, mat)
# For basic colors, use engine materials:
# /Engine/BasicShapes/BasicShapeMaterial (white)
# Or create dynamic material instances for custom colors
```

IMPORTANT CONSTRUCTION RULES:
- ALWAYS use direct set calls (set_actor_location, set_actor_rotation, set_actor_scale3d) — NEVER conditional
- ALWAYS label every actor clearly (Wall_Front, Roof_Left, Floor_Ground, etc.)
- ALWAYS set collision on structural pieces
- Connect wall corners precisely — end of one wall = start of next
- For symmetry: left/right pieces must have matching scales with mirrored rotation
- The BasicShapes/Cube default size is 100x100x100 (1m), so scale 1.0 = 1 meter

ROOF FIX (CRITICAL): When fixing a roof, ALWAYS:
1. Calculate house bounds from wall actor positions in the LIVE UEFN DATA
2. half_depth = (max_y - min_y) / 2
3. roof_height = half_depth * 0.25 (or specified by user)
4. pitch = degrees(atan2(roof_height, half_depth))
5. slope_length = sqrt(half_depth^2 + roof_height^2)
6. DELETE old roof/gable/ridge actors
7. Create left slope: scale=(house_width/100, slope_length/100, 0.15), pitch=+pitch_deg, at center_y - half_depth/2
8. Create right slope: same scale, pitch=-pitch_deg, at center_y + half_depth/2
9. Create ridge cap at ridge height

MANDATORY ROOF ACTION — ALWAYS use this for ANY roof fix/build/rebuild (do NOT write your own roof code):
```action
{"action": "rebuild_gable_roof", "roof_height_ratio": 0.25}
```
This automatically calculates house dimensions from LIVE UEFN DATA, deletes all old roof/gable/ridge pieces, and rebuilds a clean gable roof with correct geometry. The roof_height_ratio (default 0.25) controls steepness.

RETRY RULE: If the user says "fix it again" or repeats a request, the previous attempt FAILED.
Do NOT generate the same code. DELETE the broken actors and REBUILD from scratch.
Use the rebuild_gable_roof action for quick, correct results.
"""
            messages[0]["content"] += action_instruction

            if extra_context:
                messages[-1]["content"] += extra_context

            # ── Force directive suffix to override verbose models ──
            # This goes at the END of the user message so the model sees it last
            has_action_intent = any(kw in msg_lower for kw in [
                'fix', 'delete', 'remove', 'add', 'move', 'build', 'create', 'spawn',
                'color', 'change', 'rotate', 'scale', 'align', 'place', 'set', 'make',
                'clean', 'organize', 'rename', 'duplicate', 'copy', 'replace',
            ])
            if has_action_intent:
                # Detect retry/repeat intent — user is asking again because last attempt failed
                retry_phrases = ('again', 'still', 'didnt work', "didn't work", 'not working',
                                 'same thing', 'nothing changed', 'try again', 'do it again',
                                 'fix it', 'still the same', 'still broken', 'still bad')
                is_retry = any(phrase in msg_lower for phrase in retry_phrases)

                retry_directive = ""
                if is_retry:
                    retry_directive = (
                        "\n\nRETRY DETECTED: The user's PREVIOUS request FAILED or produced no visible change. "
                        "You MUST take a DIFFERENT approach this time: "
                        "1. DELETE the problematic actors first. "
                        "2. REBUILD from scratch using the construction procedures above. "
                        "3. Do NOT repeat the same code that failed before. "
                        "4. Use the ROOF FIX PROCEDURE if this involves a roof."
                    )

                messages[-1]["content"] += (
                    "\n\n[SYSTEM: Respond with 1-2 sentences MAX, then include ```action blocks. "
                    "Do NOT write tutorials, step-by-step guides, analysis sections, or tables. "
                    "Do NOT rewrite or correct the user's message. Just act.]"
                    + retry_directive
                )

            response = _chat_completion_with_retry(
                client,
                provider,
                model=model,
                messages=messages,
                max_tokens=3000,
                temperature=0.5
            )

            tok_in = getattr(response.usage, 'prompt_tokens', 0) if response.usage else 0
            tok_out = getattr(response.usage, 'completion_tokens', 0) if response.usage else 0
            _track_usage(provider, tok_in, tok_out)

            reply = response.choices[0].message.content or "I processed your request."

            # ── Post-process: clean up verbose/messy responses from weaker models ──
            reply = _clean_prefetch_reply(reply)

            # ── Post-process: extract and execute action blocks from the AI's response ──
            action_results = _execute_action_blocks(reply)
            if action_results:
                tool_result_for_frontend = action_results[-1]  # show last action result
                # Append execution results to the reply
                for ar in action_results:
                    if ar.get("success") or ar.get("result"):
                        reply += f"\n\n✅ **Action executed:** `{ar.get('action', 'unknown')}`"
                        if ar.get("result"):
                            result_preview = json.dumps(ar["result"], default=str)[:500]
                            reply += f"\n```\n{result_preview}\n```"
                    elif ar.get("error"):
                        reply += f"\n\n❌ **Action failed:** {ar.get('error')}"

            result = {"reply": reply, "_provider": provider, "_model": model}
            if tool_result_for_frontend:
                result["tool_result"] = tool_result_for_frontend
            return result

    except Exception as e:
        logger.error(f"LLM chat error: {e}")
        error_msg = str(e)
        lowered = error_msg.lower()

        if "tool_use_failed" in lowered or "<function=" in error_msg:
            recovered = _recover_failed_tool_call(
                error_msg=error_msg,
                messages=messages,
                client=client,
                provider=provider,
                model=model,
            )
            if recovered is not None:
                recovered["_provider"] = provider
                recovered["_model"] = model
                return recovered

        if provider in AI_PROVIDERS and any(token in lowered for token in ("model_not_found", "model_decommissioned", "no longer supported", "decommissioned")):
            fallback_model = AI_PROVIDERS[provider].get("default_model", "").strip()
            if fallback_model and fallback_model != model:
                env_updates = {"AI_MODEL": fallback_model}
                os.environ["AI_MODEL"] = fallback_model
                model_env = _get_provider_model_env_key(provider)
                if model_env:
                    os.environ[model_env] = fallback_model
                    env_updates[model_env] = fallback_model
                _persist_env(env_updates)
                _llm_client = None
                _llm_provider = None
                logger.warning("Retrying AI chat with fallback model '%s' for provider '%s'", fallback_model, provider)
                return _ai_chat(message, attachments, history, chat_title=chat_title, chat_memory=chat_memory, chat_id=chat_id)

        # ── Auto-fallback to another provider on rate limit / 429 ──
        if _is_retryable_llm_error(e) or "rate_limit" in lowered or "429" in error_msg or "quota" in lowered or "resource_exhausted" in lowered:
            attempted = set(_tried_providers or [])
            attempted.add(provider)
            fallback_order = (
                ["openai", "gemini", "groq", "cerebras", "ollama"]
                if needs_native_media
                else ["groq", "cerebras", "gemini", "openai", "ollama"]
            )
            original_provider_env = os.environ.get("AI_PROVIDER", "")
            original_model_env = os.environ.get("AI_MODEL", "")
            for fb_prov in fallback_order:
                if fb_prov in attempted:
                    continue
                if fb_prov == "openai":
                    fb_key = os.environ.get("OPENAI_API_KEY", "").strip()
                    fb_model = _get_provider_default_model("openai")
                else:
                    fb_info = AI_PROVIDERS.get(fb_prov, {})
                    fb_key_env = fb_info.get("key_env")
                    fb_key = os.environ.get(fb_key_env, "").strip() if fb_key_env else ""
                    fb_model = fb_info.get("default_model", "")
                if fb_prov == "ollama":
                    if not _check_ollama_running():
                        continue
                    fb_key = "ollama"
                elif not fb_key:
                    continue

                try:
                    logger.warning("Provider '%s' had a transient failure. Retrying this request with '%s'.", provider, fb_prov)
                    _llm_client = None
                    _llm_provider = None
                    os.environ["AI_PROVIDER"] = fb_prov
                    os.environ["AI_MODEL"] = fb_model
                    return _ai_chat(
                        message,
                        effective_attachments,
                        history,
                        chat_title=chat_title,
                        chat_memory=chat_memory,
                        chat_id=chat_id,
                        _tried_providers=sorted(attempted | {fb_prov}),
                    )
                except Exception:
                    continue
                finally:
                    if original_provider_env:
                        os.environ["AI_PROVIDER"] = original_provider_env
                    else:
                        os.environ.pop("AI_PROVIDER", None)
                    if original_model_env:
                        os.environ["AI_MODEL"] = original_model_env
                    else:
                        os.environ.pop("AI_MODEL", None)
                    _llm_client = None
                    _llm_provider = None

            return {"reply": f"All AI providers are rate-limited. Please wait a few minutes and try again.\n\nOriginal error: {error_msg[:200]}"}

        if "Connection refused" in error_msg or "connection" in error_msg.lower():
            prov = _get_active_provider()
            if prov == "ollama":
                return {"reply": "Could not connect to Ollama. Make sure Ollama is running (`ollama serve`) and you have a model pulled (`ollama pull llama3.1`)."}
            return {"reply": f"Could not connect to AI provider ({prov}). Check your API key in Settings."}
        return {"reply": f"AI error: {error_msg}"}


def _keyword_chat(message: str, attachments: list, history: list) -> dict:
    """Improved keyword-based chat fallback when no AI API is configured."""
    msg_lower = message.lower()
    effective_attachments = _resolve_effective_turn_attachments(message, attachments, history)
    tool_result = None
    reply = ""

    # ── Intent: List actors
    if any(kw in msg_lower for kw in ['actors', 'what is in my level', 'list actors', "what's in my level"]):
        result = _chat_uefn_query("get_all_actors", {})
        if result.get("success"):
            actors = result.get("result", [])
            if isinstance(actors, list):
                count = len(actors)
                sample = actors[:15]
                actor_list = "\n".join(
                    f"  - {a}" if isinstance(a, str) else f"  - {a.get('name', a.get('label', str(a)))}"
                    for a in sample
                )
                reply = f"Your level has **{count} actors**. Here are the first {min(15, count)}:\n{actor_list}"
                if count > 15:
                    reply += f"\n\n...and {count - 15} more."
            else:
                reply = f"Actor query returned: {str(actors)[:500]}"
            tool_result = {"tool": "get_all_actors", "output": actors}
        else:
            reply = f"Couldn't query actors: {result.get('error', 'Unknown error')}"

    elif any(kw in msg_lower for kw in ['level info', 'level name', 'current level', 'map info']):
        result = _chat_uefn_query("get_level_info", {})
        if result.get("success"):
            info = result.get("result", {})
            reply = f"**Level Info:**\n- World: {info.get('world_name', 'N/A')}\n- Map: {info.get('map_name', 'N/A')}\n- Actors: {info.get('actor_count', '?')}"
            tool_result = {"tool": "get_level_info", "output": info}
        else:
            reply = f"Couldn't get level info: {result.get('error', 'Unknown')}"

    elif any(kw in msg_lower for kw in ['project info', 'project name', 'show project']):
        result = _chat_uefn_query("get_project_info", {})
        if result.get("success"):
            info = result.get("result", {})
            reply = f"**Project:** {info.get('project_name', 'N/A')}\n- Engine: {info.get('engine_version', 'N/A')}\n- Content Dir: {info.get('content_dir', 'N/A')}"
            tool_result = {"tool": "get_project_info", "output": info}
        else:
            reply = f"Couldn't get project info: {result.get('error', 'Unknown')}"

    elif any(kw in msg_lower for kw in ['screenshot', 'capture', 'take a pic']):
        port = discover_uefn_listener_port()
        if port:
            code = "import UEFN_Toolbelt as tb; result = tb.run('screenshot_take', {})"
            result = mcp_listener_post_command(int(port), "execute_python", {"code": code}, timeout=15.0)
            if result.get("success"):
                reply = "Screenshot captured! Check the UEFN screenshots folder."
                tool_result = {"tool": "screenshot_take", "output": result.get("result", {})}
            else:
                reply = f"Screenshot failed: {result.get('error', 'Unknown')}"
        else:
            reply = "Can't take screenshots — UEFN is not connected."

    elif any(kw in msg_lower for kw in ['memory scan', 'memory usage', 'optimization', 'performance']):
        port = discover_uefn_listener_port()
        if port:
            code = "import UEFN_Toolbelt as tb; result = tb.run('memory_scan', {})"
            result = mcp_listener_post_command(int(port), "execute_python", {"code": code}, timeout=30.0)
            if result.get("success"):
                reply = "Memory scan complete! See the results below."
                tool_result = {"tool": "memory_scan", "output": result.get("result", {})}
            else:
                reply = f"Memory scan failed: {result.get('error', 'Unknown')}"
        else:
            reply = "Can't run memory scan — UEFN is not connected."

    elif any(kw in msg_lower for kw in ['list tools', 'available tools', 'what tools', 'show tools', 'what can you do']):
        tools = list(tool_registry.tools.values())
        cats = {}
        for t in tools:
            cat = t.get('category', 'Other')
            if cat not in cats:
                cats[cat] = []
            cats[cat].append(t['name'])
        lines = [f"I have access to **{len(tools)} tools** across {len(cats)} categories:\n"]
        for cat, names in sorted(cats.items()):
            lines.append(f"**{cat}** ({len(names)}): {', '.join(names[:5])}" + (f" +{len(names)-5} more" if len(names) > 5 else ""))
        server_actions = _get_server_action_capabilities()
        if server_actions:
            lines.append("")
            lines.append("**Server Actions**: " + ", ".join(item["id"] for item in server_actions))
        reply = "\n".join(lines)
        reply += "\n\nAsk me to run any tool or describe what you want to build!"

    elif any(kw in msg_lower for kw in ['selected', 'selection', 'what did i select']):
        result = _chat_uefn_query("get_selected_actors", {})
        if result.get("success"):
            actors = result.get("result", [])
            if actors:
                actor_list = "\n".join(
                    f"  - {a}" if isinstance(a, str) else f"  - {a.get('name', str(a))}"
                    for a in actors[:20]
                )
                reply = f"You have {len(actors)} actor(s) selected:\n{actor_list}"
            else:
                reply = "No actors are selected in the viewport."
            tool_result = {"tool": "get_selected_actors", "output": actors}
        else:
            reply = f"Couldn't check selection: {result.get('error', 'Unknown')}"

    elif any(kw in msg_lower for kw in ['viewport', 'camera position', 'where am i looking']):
        result = _chat_uefn_query("get_viewport_camera", {})
        if result.get("success"):
            cam = result.get("result", {})
            loc = cam.get("location", {})
            rot = cam.get("rotation", {})
            reply = f"**Viewport Camera:**\n- Position: X={loc.get('x', '?'):.0f}, Y={loc.get('y', '?'):.0f}, Z={loc.get('z', '?'):.0f}\n- Rotation: P={rot.get('pitch', '?'):.1f}, Y={rot.get('yaw', '?'):.1f}"
            tool_result = {"tool": "get_viewport_camera", "output": cam}
        else:
            reply = f"Couldn't get viewport: {result.get('error', 'Unknown')}"

    elif effective_attachments:
        reply = _build_keyword_attachment_reply(message, effective_attachments)

    else:
        port = discover_uefn_listener_port()
        status = "connected" if port else "not connected"

        # Handle greetings naturally
        greetings = ['hello', 'hi', 'hey', 'sup', 'what up', 'yo', 'howdy', 'whats up', "what's up", 'good morning', 'good evening', 'hola']
        if any(g in msg_lower for g in greetings):
            reply = f"Hey! I'm your UEFN AI assistant. UEFN is **{status}**.\n\n"
            reply += "What are you working on? I can:\n"
            reply += "- Query your level (actors, project info, viewport)\n"
            reply += "- Run any of 90+ UEFN tools\n"
            reply += "- Take screenshots and analyze your scene\n"
            reply += "- Help plan and build new features\n"
            reply += "\nJust tell me what you need!"
        elif any(kw in msg_lower for kw in ['help', 'what can you do', 'how do']):
            reply = "Here's what I can do:\n\n"
            reply += '- **"What actors are in my level?"** — query live editor state\n'
            reply += '- **"List all tools"** — see the 90+ tools available\n'
            reply += '- **"Take a screenshot"** — capture the viewport\n'
            reply += '- **"Show project info"** — project details\n'
            reply += '- **"Run a memory scan"** — check optimization\n'
            reply += "- **Attach files** with the + button for context\n"
            reply += "\nFor full AI conversations, set up a free provider in **Settings**."
        elif any(kw in msg_lower for kw in ['thank', 'thanks', 'thx', 'ty']):
            reply = "You're welcome! Let me know if you need anything else."
        else:
            reply = f"I got your message: *\"{message}\"*\n\n"
            reply += f"UEFN is **{status}**. "
            reply += "I'm currently in **basic mode** (keyword matching). I can respond to specific commands like:\n"
            reply += "- \"What actors are in my level?\"\n"
            reply += "- \"Take a screenshot\"\n"
            reply += "- \"Show project info\"\n\n"
            reply += "For free-form AI conversations, set up **Groq** (free, fastest) or **Ollama** (free, local) in **Settings**. It takes 30 seconds!"

    response = {"reply": reply}
    if tool_result:
        response["tool_result"] = tool_result
    return response


def _make_chat_message(role: str, content: str, attachments: Optional[List[Dict[str, Any]]] = None, tool_result: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Create a frontend-compatible chat message record."""
    return {
        "id": uuid4().hex,
        "role": role,
        "content": content,
        "timestamp": _iso_now(),
        "attachments": attachments or [],
        "toolResult": tool_result,
    }


# ── Project endpoints ─────────────────────────────────────────────────────

@app.route('/api/projects', methods=['GET'])
def list_projects():
    """List all projects."""
    projects = chat_store.list_projects()
    return jsonify({"projects": projects, "count": len(projects)})


@app.route('/api/projects', methods=['POST'])
def create_project():
    """Create a new project."""
    data = request.get_json() or {}
    name = (data.get("name") or "").strip() or "New project"
    icon = (data.get("icon") or "").strip()
    color = (data.get("color") or "").strip()
    project = chat_store.create_project(name, icon=icon, color=color)
    return jsonify({"success": True, "project": project}), 201


@app.route('/api/projects/<project_id>', methods=['PATCH'])
def update_project(project_id):
    """Update a project (name, icon, color)."""
    data = request.get_json() or {}
    name = data.get("name", "")
    icon = data.get("icon", "")
    color = data.get("color", "")
    if not name and not icon and color == "":
        return jsonify({"error": "Nothing to update"}), 400
    project = chat_store.update_project(project_id, name=str(name).strip() if name else "", icon=str(icon).strip() if icon else "", color=str(color).strip() if color is not None else "")
    if not project:
        return jsonify({"error": "Project not found"}), 404
    return jsonify({"success": True, "project": project})


@app.route('/api/projects/<project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project and all its chats."""
    deleted = chat_store.delete_project(project_id)
    if not deleted:
        return jsonify({"error": "Project not found"}), 404
    return jsonify({"success": True})


# ── Chat endpoints ────────────────────────────────────────────────────────

@app.route('/api/chats', methods=['GET'])
def list_chats():
    """List all persisted chat sessions. Optional ?project_id= filter."""
    project_id = request.args.get("project_id", "")
    sessions = chat_store.list_sessions(project_id=project_id)
    return jsonify({
        "chats": sessions,
        "count": len(sessions),
    })


@app.route('/api/chats', methods=['POST'])
def create_chat():
    """Create a new persisted chat session."""
    data = request.get_json() or {}
    session = chat_store.create_session(
        title=str(data.get("title") or ""),
        project_id=str(data.get("project_id") or ""),
    )
    return jsonify({"success": True, "chat": session}), 201


@app.route('/api/chats/<chat_id>', methods=['GET'])
def get_chat(chat_id):
    """Get one persisted chat/project session."""
    session = chat_store.get_session(chat_id)
    if not session:
        return jsonify({"error": "Chat not found"}), 404
    return jsonify({"chat": session})


@app.route('/api/chats/<chat_id>', methods=['PATCH'])
def update_chat(chat_id):
    """Rename a stored chat/project or clear its memory while keeping the project shell."""
    data = request.get_json() or {}
    raw_title = data.get("title")
    clear_messages = bool(data.get("clear_messages"))

    if raw_title is None and not clear_messages:
        return jsonify({"error": "No chat update provided"}), 400

    session = chat_store.update_session(
        chat_id,
        title=str(raw_title) if raw_title is not None else None,
        clear_messages=clear_messages,
    )
    if not session:
        return jsonify({"error": "Chat not found"}), 404
    return jsonify({"success": True, "chat": session})


@app.route('/api/chats/<chat_id>', methods=['DELETE'])
def delete_chat(chat_id):
    """Delete one persisted chat/project session."""
    deleted = chat_store.delete_session(chat_id)
    if not deleted:
        return jsonify({"error": "Chat not found"}), 404
    return jsonify({"success": True})


@app.route('/api/chats/<chat_id>/attachments', methods=['GET'])
def chat_attachment_index(chat_id):
    """Return analyzed attachment records for a chat, optionally filtered by a query."""
    session = chat_store.get_session(chat_id)
    if not session:
        return jsonify({"error": "Chat not found"}), 404

    query = (request.args.get("query", "") or "").strip()
    try:
        limit = int(request.args.get("limit", str(MAX_ATTACHMENT_ANALYSIS_RESULTS)) or MAX_ATTACHMENT_ANALYSIS_RESULTS)
    except ValueError:
        limit = MAX_ATTACHMENT_ANALYSIS_RESULTS

    limit = max(1, min(limit, MAX_ATTACHMENT_INDEX_RESULTS))
    result = _analyze_chat_attachments(chat_id, query=query, limit=limit)
    return jsonify({"success": True, **result})


@app.route('/api/chat', methods=['POST'])
def chat_endpoint():
    """AI chat endpoint — uses OpenAI when available, falls back to keyword matching."""
    try:
        data = request.get_json() or {}
        chat_id = str(data.get('chat_id') or '').strip()
        message = (data.get('message', '') or '').strip()
        attachments = [_prepare_chat_attachment(att) for att in (data.get('attachments', []) or [])]
        attachments.extend(_collect_message_url_attachments(message, attachments))
        session = chat_store.ensure_session(chat_id=chat_id, title=str(data.get("title") or ""))
        chat_id = session["id"]
        history = session.get("messages", [])

        if not message and not attachments:
            return jsonify({
                "reply": "Send me a message or attach a file to get started.",
                "chat_id": chat_id,
                "chat": session,
            }), 200

        attachment_meta = attachments
        user_content = message or ("Shared attachments: " + ", ".join(att["name"] for att in attachment_meta))
        user_message = _make_chat_message("user", user_content, attachments=attachment_meta)

        # Try AI-powered chat first
        ai_result = _ai_chat(
            message,
            attachments,
            history,
            chat_title=session.get("title", ""),
            chat_memory=session.get("memory_summary", ""),
            chat_id=chat_id,
        )
        if ai_result is not None:
            response_provider = ai_result.get("_provider") or _get_active_provider()
            response_model = ai_result.get("_model") or _get_active_model()
            assistant_message = _make_chat_message(
                "assistant",
                ai_result.get("reply") or ai_result.get("message") or "No response",
                tool_result=ai_result.get("tool_result"),
            )
            updated_session = chat_store.append_messages(
                chat_id,
                [user_message, assistant_message],
                provider=response_provider,
                model=response_model,
            )
            if updated_session:
                knowledge_store.remember_attachments(
                    chat_id=chat_id,
                    chat_title=updated_session.get("title", ""),
                    attachments=attachment_meta,
                )
                knowledge_store.remember_interaction(
                    chat_id=chat_id,
                    chat_title=updated_session.get("title", ""),
                    user_message=user_content,
                    assistant_message=assistant_message,
                )
            result = dict(ai_result)
            result.pop("_provider", None)
            result.pop("_model", None)
            result.update({
                "chat_id": chat_id,
                "provider": response_provider,
                "model": response_model,
                "chat": updated_session,
            })
            return jsonify(result)

        # Fall back to keyword matching
        keyword_result = _keyword_chat(message, attachments, history)
        assistant_message = _make_chat_message(
            "assistant",
            keyword_result.get("reply") or keyword_result.get("message") or "No response",
            tool_result=keyword_result.get("tool_result"),
        )
        updated_session = chat_store.append_messages(
            chat_id,
            [user_message, assistant_message],
            provider=_get_active_provider(),
            model=_get_active_model(),
        )
        if updated_session:
            knowledge_store.remember_attachments(
                chat_id=chat_id,
                chat_title=updated_session.get("title", ""),
                attachments=attachment_meta,
            )
            knowledge_store.remember_interaction(
                chat_id=chat_id,
                chat_title=updated_session.get("title", ""),
                user_message=user_content,
                assistant_message=assistant_message,
            )
        keyword_result.update({
            "chat_id": chat_id,
            "provider": _get_active_provider(),
            "model": _get_active_model(),
            "chat": updated_session,
        })
        return jsonify(keyword_result)

    except Exception as e:
        logger.error(f"Chat error: {e}")
        error_text = str(e).lower()
        # Give user-friendly messages instead of raw 500
        if "resource_exhausted" in error_text or "quota" in error_text or "rate_limit" in error_text or "429" in str(e):
            reply = "The AI provider is temporarily rate-limited. Please wait about 30-60 seconds and try again, or switch to a different provider in Settings."
        elif "api_key" in error_text or "authentication" in error_text or "unauthorized" in error_text or "401" in str(e):
            reply = "API key issue — please check your API key in Settings."
        elif "timeout" in error_text or "timed out" in error_text:
            reply = "The request timed out. Please try again."
        else:
            reply = f"Something went wrong: {str(e)[:200]}"
        # Return 200 so the frontend shows the message in chat instead of an error toast
        return jsonify({
            "reply": reply,
            "chat_id": chat_id,
            "provider": _get_active_provider(),
            "model": _get_active_model(),
        }), 200


@app.route('/api/chat/status', methods=['GET'])
def chat_status():
    """Check AI chat configuration status."""
    ollama_running = _check_ollama_running()
    ollama_models = _get_ollama_models() if ollama_running else []
    prewarm_status = _copy_local_model_prewarm_status()

    provider_name, api_key = _detect_provider()

    if provider_name and _HAS_OPENAI_PKG:
        mode = provider_name
        ai_enabled = True
        model = _get_active_model()
    else:
        mode = "keyword"
        ai_enabled = False
        model = None

    return jsonify({
        "ai_enabled": ai_enabled,
        "mode": mode,
        "provider": provider_name,
        "model": model,
        "provider_models": _get_saved_provider_models(),
        "ollama_running": ollama_running,
        "ollama_models": ollama_models,
        "ollama_base_url": OLLAMA_BASE_URL,
        "has_openai_pkg": _HAS_OPENAI_PKG,
        "available_providers": _build_available_providers(),
        "local_structured_vlm": {
            "enabled": LOCAL_STRUCTURED_VLM_ENABLED,
            "model": LOCAL_STRUCTURED_VLM_MODEL_ID,
            "has_transformers": _HAS_TRANSFORMERS,
            "has_florence2": _HAS_FLORENCE2,
            "has_torch": _HAS_TORCH,
            "has_pil": _HAS_PIL,
            "ready": bool(_local_structured_vlm_model not in (None, False) and _local_structured_vlm_processor not in (None, False)),
            "available": bool(LOCAL_STRUCTURED_VLM_ENABLED and _HAS_TRANSFORMERS and _HAS_FLORENCE2 and _HAS_TORCH and _HAS_PIL),
            "prewarm": prewarm_status.get("components", {}).get("structured_vlm", {}),
        },
        "local_vlm": {
            "enabled": LOCAL_VLM_ENABLED,
            "model": LOCAL_VLM_MODEL_ID,
            "has_transformers": _HAS_TRANSFORMERS,
            "has_torch": _HAS_TORCH,
            "has_pil": _HAS_PIL,
            "ready": bool(_local_vlm_pipeline not in (None, False)),
            "available": bool(LOCAL_VLM_ENABLED and _HAS_TRANSFORMERS and _HAS_TORCH and _HAS_PIL),
            "prewarm": prewarm_status.get("components", {}).get("vlm", {}),
        },
        "local_htr": {
            "enabled": LOCAL_HTR_ENABLED,
            "model": LOCAL_HTR_MODEL_ID,
            "has_transformers": _HAS_TRANSFORMERS,
            "has_torch": _HAS_TORCH,
            "has_pil": _HAS_PIL,
            "ready": bool(_local_htr_pipeline not in (None, False) and _local_htr_processor not in (None, False)),
            "available": bool(LOCAL_HTR_ENABLED and _HAS_TRANSFORMERS and _HAS_TORCH and _HAS_PIL),
            "prewarm": prewarm_status.get("components", {}).get("handwriting", {}),
        },
        "local_pdf_vision": {
            "has_pymupdf": _HAS_PYMUPDF,
            "ready": bool(_HAS_PYMUPDF and _HAS_PIL),
        },
        "local_model_prewarm": prewarm_status,
        "hosted_llm_retries": {
            "max_retries": HOSTED_LLM_MAX_RETRIES,
            "base_seconds": HOSTED_LLM_RETRY_BASE_SECONDS,
            "max_seconds": HOSTED_LLM_RETRY_MAX_SECONDS,
        },
        "attachment_analysis_cache": attachment_analysis_cache.stats(),
    })


@app.route('/api/settings/ai', methods=['POST'])
def set_ai_settings():
    """Update AI chat settings — provider, model, API keys (supports saving all keys at once)."""
    global _llm_client, _llm_provider
    data = request.get_json() or {}
    provider = (data.get("provider") or "").strip().lower()
    model = (data.get("model") or "").strip()
    api_key = (data.get("api_key") or "").strip()  # single key (legacy)
    keys = data.get("keys") or {}  # dict of {provider_name: key_value} for batch save
    provider_models = data.get("provider_models") or {}

    env_updates = {}

    # Handle batch keys: {"groq": "gsk_...", "cerebras": "csk_...", "gemini": "AIza..."}
    for prov_name, key_val in keys.items():
        prov_name = prov_name.strip().lower()
        key_val = (key_val or "").strip()
        if prov_name in AI_PROVIDERS and key_val:
            key_env = AI_PROVIDERS[prov_name].get("key_env")
            if key_env:
                os.environ[key_env] = key_val
                env_updates[key_env] = key_val

    # Single key for the selected provider (legacy path)
    if provider and provider in AI_PROVIDERS and api_key:
        key_env = AI_PROVIDERS[provider].get("key_env")
        if key_env:
            os.environ[key_env] = api_key
            env_updates[key_env] = api_key

    if provider and provider not in AI_PROVIDERS:
        return jsonify({"error": f"Unknown provider: {provider}"}), 400

    for prov_name, model_name in provider_models.items():
        prov_name = str(prov_name).strip().lower()
        model_name = str(model_name or "").strip()
        if prov_name not in AI_PROVIDERS or not model_name:
            continue
        try:
            resolved = _validate_requested_model(prov_name, model_name)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        model_env = _get_provider_model_env_key(prov_name)
        if model_env:
            os.environ[model_env] = resolved
            env_updates[model_env] = resolved

    if provider:
        if not _provider_has_access(provider):
            return jsonify({"error": f"Provider '{provider}' is not ready. Add a key in Settings or start Ollama first."}), 400
        try:
            resolved_model = _validate_requested_model(provider, model or _get_provider_default_model(provider))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        os.environ["AI_PROVIDER"] = provider
        os.environ["AI_MODEL"] = resolved_model
        env_updates["AI_PROVIDER"] = provider
        env_updates["AI_MODEL"] = resolved_model
        model_env = _get_provider_model_env_key(provider)
        if model_env:
            os.environ[model_env] = resolved_model
            env_updates[model_env] = resolved_model
    elif model:
        active_provider = _get_active_provider()
        try:
            resolved_model = _validate_requested_model(active_provider, model)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        os.environ["AI_MODEL"] = resolved_model
        env_updates["AI_MODEL"] = resolved_model
        model_env = _get_provider_model_env_key(active_provider)
        if model_env:
            os.environ[model_env] = resolved_model
            env_updates[model_env] = resolved_model

    # Force re-init on next chat
    _llm_client = None
    _llm_provider = None

    # Persist to .env file
    _persist_env(env_updates)

    # Return updated status
    active_provider = _get_active_provider()
    active_model = _get_active_model()

    return jsonify({
        "success": True,
        "provider": active_provider,
        "model": active_model,
        "provider_models": _get_saved_provider_models(),
    })


def _persist_env(env_updates: dict):
    """Write key=value pairs to .env file (upsert)."""
    env_path = WORKSPACE_ROOT / ".env"
    try:
        lines = []
        if env_path.exists():
            lines = env_path.read_text(encoding="utf-8").splitlines()

        def _upsert(key: str, val: str):
            for i, line in enumerate(lines):
                if line.startswith(f"{key}="):
                    lines[i] = f"{key}={val}"
                    return
            lines.append(f"{key}={val}")

        for key, val in env_updates.items():
            _upsert(key, val)

        env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except Exception as e:
        logger.warning(f"Could not write .env: {e}")


@app.route('/api/chat/usage', methods=['GET'])
def chat_usage():
    """Return usage stats for all AI providers."""
    # Check for day rollover
    today = datetime.now().strftime("%Y-%m-%d")
    if today != _usage_date:
        _track_usage("_rollover_check")  # triggers reset

    result = {}
    for prov, data in _usage_data.items():
        limit_req = data["limit_requests"]
        limit_tok = data["limit_tokens"]
        total_tokens = data["tokens_in"] + data["tokens_out"]

        # Calculate percentage used
        if limit_req > 0:
            pct = round(data["requests"] / limit_req * 100, 1)
        elif limit_tok > 0:
            pct = round(total_tokens / limit_tok * 100, 1)
        else:
            pct = 0  # unlimited or pay-per-use

        result[prov] = {
            "requests": data["requests"],
            "tokens_in": data["tokens_in"],
            "tokens_out": data["tokens_out"],
            "tokens_total": total_tokens,
            "limit_requests": limit_req,
            "limit_tokens": limit_tok,
            "period": data["period"],
            "percent_used": min(pct, 100),
            "label": AI_PROVIDERS.get(prov, {}).get("label", prov),
        }

    return jsonify({
        "date": _usage_date,
        "providers": result,
        "active_provider": _get_active_provider(),
        "active_model": _get_active_model(),
    })


@app.route('/api/chat/model', methods=['POST'])
def quick_switch_model():
    """Quick model switch from the chat panel — no API key change."""
    global _llm_client, _llm_provider
    data = request.get_json() or {}
    model = (data.get("model") or "").strip()
    provider = (data.get("provider") or "").strip().lower()

    if not provider and not model:
        return jsonify({"error": "No provider or model specified"}), 400

    target_provider = provider or _get_active_provider()
    if target_provider not in AI_PROVIDERS and target_provider != "openai":
        return jsonify({"error": f"Unknown provider: {target_provider}"}), 400
    if target_provider in AI_PROVIDERS and not _provider_has_access(target_provider):
        return jsonify({"error": f"Provider '{target_provider}' is not ready."}), 400

    try:
        resolved_model = _validate_requested_model(target_provider, model or _get_provider_default_model(target_provider))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    env_updates = {
        "AI_PROVIDER": target_provider,
        "AI_MODEL": resolved_model,
    }
    os.environ["AI_PROVIDER"] = target_provider
    os.environ["AI_MODEL"] = resolved_model

    model_env = _get_provider_model_env_key(target_provider)
    if model_env:
        os.environ[model_env] = resolved_model
        env_updates[model_env] = resolved_model

    _llm_client = None
    _llm_provider = None

    _persist_env(env_updates)

    return jsonify({
        "success": True,
        "provider": _get_active_provider(),
        "model": _get_active_model(),
        "provider_models": _get_saved_provider_models(),
    })


# ============================================================================
# ASSET GENERATION PIPELINE
# ============================================================================

_pipeline_instance = None
_job_manager_instance = None

def _get_pipeline():
    global _pipeline_instance
    if _pipeline_instance is None:
        import sys
        apps_root = str(WORKSPACE_ROOT / "apps")
        if apps_root not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.pipeline import AssetPipeline
        _pipeline_instance = AssetPipeline()
    return _pipeline_instance


def _get_job_manager():
    global _job_manager_instance
    if _job_manager_instance is None:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.job_manager import JobManager
        _job_manager_instance = JobManager(pipeline=_get_pipeline())
    return _job_manager_instance


@app.route('/api/pipeline/generate', methods=['POST'])
def pipeline_generate():
    """Start async asset generation from a text prompt. Returns job_id immediately."""
    data = request.get_json(force=True)
    prompt = data.get("prompt", "").strip()
    if not prompt:
        return jsonify({"error": "prompt is required"}), 400

    project = data.get("project", "default")
    auto_approve = data.get("auto_approve", False)

    try:
        jm = _get_job_manager()
        job_id = jm.create_job(prompt, project=project, auto_approve=auto_approve)
        return jsonify({"job_id": job_id, "status": "queued"}), 202
    except Exception as e:
        logger.exception("Pipeline job creation failed")
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/jobs', methods=['GET'])
def pipeline_list_jobs():
    """List all pipeline jobs (active + recent)."""
    try:
        jm = _get_job_manager()
        jobs = jm.list_jobs()
        return jsonify({"jobs": jobs})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/jobs/<job_id>', methods=['GET'])
def pipeline_get_job(job_id):
    """Get job status/progress/result."""
    try:
        jm = _get_job_manager()
        job = jm.get_job(job_id)
        if job is None:
            return jsonify({"error": "Job not found"}), 404
        return jsonify(job)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/jobs/<job_id>/cancel', methods=['POST'])
def pipeline_cancel_job(job_id):
    """Cancel a running job."""
    try:
        jm = _get_job_manager()
        cancelled = jm.cancel_job(job_id)
        if not cancelled:
            return jsonify({"error": "Job not found or already finished"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/assets', methods=['GET'])
def pipeline_list_assets():
    """List all pipeline assets."""
    project = request.args.get("project")
    try:
        pipeline = _get_pipeline()
        assets = pipeline.list_assets(project)
        return jsonify({"assets": assets})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/asset/<asset_id>', methods=['GET'])
def pipeline_get_asset(asset_id):
    """Get full asset detail."""
    try:
        pipeline = _get_pipeline()
        asset = pipeline.get_asset(asset_id)
        if asset is None:
            return jsonify({"error": "Asset not found"}), 404
        return jsonify(asset)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/validate/<asset_id>', methods=['POST'])
def pipeline_revalidate(asset_id):
    """Re-trigger validation on an asset."""
    try:
        pipeline = _get_pipeline()
        result = pipeline.revalidate(asset_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/approve/<asset_id>', methods=['POST'])
def pipeline_approve(asset_id):
    """Manually approve an asset for UEFN import."""
    try:
        pipeline = _get_pipeline()
        result = pipeline.approve(asset_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/import/<asset_id>', methods=['POST'])
def pipeline_import(asset_id):
    """Import an approved asset to UEFN via MCP bridge."""
    try:
        pipeline = _get_pipeline()
        result = pipeline.import_asset(asset_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/delete/<asset_id>', methods=['DELETE'])
def pipeline_delete(asset_id):
    """Delete a pipeline asset."""
    try:
        pipeline = _get_pipeline()
        deleted = pipeline.delete_asset(asset_id)
        if not deleted:
            return jsonify({"error": "Asset not found"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/ai_assets/<path:filepath>', methods=['GET'])
def serve_ai_asset(filepath):
    """Serve generated GLB/PNG files from data/ai_assets/."""
    from flask import send_from_directory
    ai_assets_dir = DATA_DIR / "ai_assets"
    return send_from_directory(str(ai_assets_dir), filepath)


# ============================================================================
# MODEL AI ROUTES
# ============================================================================

@app.route('/api/model-ai/edit', methods=['POST'])
def model_ai_edit():
    """Start an async AI edit job for a model."""
    try:
        data = request.json or {}
        asset_id = data.get("asset_id")
        edit_prompt = data.get("edit_prompt")
        if not asset_id or not edit_prompt:
            return jsonify({"error": "asset_id and edit_prompt required"}), 400
        jm = _get_job_manager()
        job_id = jm.create_edit_job(asset_id, edit_prompt)
        return jsonify({"job_id": job_id}), 202
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/detect-problems/<asset_id>', methods=['POST'])
def model_ai_detect_problems(asset_id):
    """Use vision AI to detect structural problems in a model."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.editor import detect_problems
        from apps.asset_pipeline.registry import AssetRegistry
        registry = _get_pipeline().registry
        result = detect_problems(asset_id, registry)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/suggest-fixes/<asset_id>', methods=['POST'])
def model_ai_suggest_fixes(asset_id):
    """Generate fix suggestions for detected problems."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.editor import suggest_fixes
        data = request.json or {}
        problems = data.get("problems", [])
        if not problems:
            return jsonify({"error": "problems list required"}), 400
        registry = _get_pipeline().registry
        result = suggest_fixes(asset_id, problems, registry)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/references/<asset_id>', methods=['GET'])
def model_ai_list_references(asset_id):
    """List reference images for an asset."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.references import list_references
        registry = _get_pipeline().registry
        result = list_references(asset_id, registry)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/references/<asset_id>', methods=['POST'])
def model_ai_upload_reference(asset_id):
    """Upload a reference image for an asset."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.references import save_reference
        registry = _get_pipeline().registry

        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({"error": "No filename"}), 400
        image_data = f.read()
        result = save_reference(asset_id, image_data, f.filename, registry)
        if not result.get("success"):
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/references/<asset_id>/<filename>', methods=['DELETE'])
def model_ai_delete_reference(asset_id, filename):
    """Delete a reference image."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.references import delete_reference
        registry = _get_pipeline().registry
        result = delete_reference(asset_id, filename, registry)
        if not result.get("success"):
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/model-ai/analyze-references/<asset_id>', methods=['POST'])
def model_ai_analyze_references(asset_id):
    """Analyze reference images using vision AI."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.references import analyze_references
        registry = _get_pipeline().registry
        result = analyze_references(asset_id, registry)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# PLACEMENT ROUTES
# ============================================================================

@app.route('/api/model-ai/placement-preview/<asset_id>', methods=['POST'])
def model_ai_placement_preview(asset_id):
    """Query scene and suggest placement for an asset."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.placement import query_scene_context, suggest_placement
        registry = _get_pipeline().registry
        record = registry.get(asset_id)
        if not record:
            return jsonify({"error": "Asset not found"}), 404
        scene = query_scene_context()
        suggestion = suggest_placement(record, scene)
        return jsonify({
            "scene_summary": {
                "total_actors": scene["total_actors"],
                "buildings": len(scene["buildings"]),
                "props": len(scene["props"]),
                "devices": len(scene["devices"]),
            },
            "suggestion": suggestion,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pipeline/import-and-place/<asset_id>', methods=['POST'])
def pipeline_import_and_place(asset_id):
    """Import asset to UEFN and place it in the scene in one step."""
    try:
        import sys
        if str(WORKSPACE_ROOT) not in sys.path:
            sys.path.insert(0, str(WORKSPACE_ROOT))
        from apps.asset_pipeline.placement import query_scene_context, suggest_placement, place_asset

        pipeline = _get_pipeline()
        record = pipeline.registry.get(asset_id)
        if not record:
            return jsonify({"error": "Asset not found"}), 404

        # Import first if not already imported
        if record.status != "imported" or not record.uefn_import_path:
            import_result = pipeline.import_asset(asset_id)
            if not import_result.get("success"):
                return jsonify({"error": f"Import failed: {import_result.get('error')}"}), 500
            record = pipeline.registry.get(asset_id)

        # Get placement suggestion from AI
        data = request.json or {}
        if "position" in data:
            # User specified position directly
            position = data["position"]
            rotation = data.get("rotation", {"yaw": 0})
            scale = data.get("scale", 1.0)
            reasoning = "User-specified position"
            placement_type = "manual"
        else:
            # AI suggests placement based on scene
            scene = query_scene_context()
            suggestion = suggest_placement(record, scene)
            position = suggestion["position"]
            rotation = suggestion["rotation"]
            scale = suggestion["scale"]
            reasoning = suggestion["reasoning"]
            placement_type = suggestion["placement_type"]

        # Place in UEFN
        place_result = place_asset(record, position, rotation, scale)
        if not place_result.get("success"):
            return jsonify({"error": place_result.get("error")}), 500

        return jsonify({
            "success": True,
            "asset_id": asset_id,
            "uefn_path": record.uefn_import_path,
            "position": position,
            "rotation": rotation,
            "scale": scale,
            "reasoning": reasoning,
            "placement_type": placement_type,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Server error"}), 500


# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    logger.info(f"Starting UEFN Codex Backend on {BACKEND_HOST}:{BACKEND_PORT}")
    logger.info(f"Tools: {len(tool_registry.tools)}")
    logger.info(f"UEFN Connected: {uefn_bridge.is_connected}")
    logger.info(f"Available categories: {', '.join(tool_registry.categories.keys())}")
    _start_local_model_prewarm()
    
    app.run(host=BACKEND_HOST, port=BACKEND_PORT, debug=False, threaded=True)
